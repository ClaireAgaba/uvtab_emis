from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.shortcuts import get_object_or_404
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from .models import Candidate, Result, CandidateLevel, CandidateModule, CandidatePaper
import io

@login_required
@require_POST
def export_candidates(request):
    """Export selected candidates to Excel with comprehensive data"""
    
    # Get selected candidate IDs from POST data
    candidate_ids = request.POST.getlist('candidate_ids')
    
    if not candidate_ids:
        return HttpResponse("No candidates selected for export", status=400)
    
    # Get candidates with related data
    candidates = Candidate.objects.filter(
        id__in=candidate_ids
    ).select_related(
        'assessment_center', 'occupation', 'district', 'village', 'created_by', 'updated_by'
    ).prefetch_related(
        'nature_of_disability', 'candidatelevel_set__level', 'candidatemodule_set__module',
        'candidatepaper_set__paper', 'result_set'
    ).order_by('reg_number')
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidates Export"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define headers with all comprehensive fields
    headers = [
        # Bio Data
        "Registration Number", "Full Name", "Date of Birth", "Gender", "Nationality",
        "National ID", "Phone Number", "Email", "Next of Kin", "Next of Kin Phone",
        
        # Address Information
        "District", "Village", "Physical Address",
        
        # Educational Background
        "Education Level", "Institution Attended", "Year of Completion",
        
        # Assessment Information
        "Assessment Center", "Occupation", "Registration Category", "Entry Year", "Intake",
        "Assessment Date", "Assessment Series", "Start Date", "Finish Date",
        
        # Disability Information
        "Has Disability", "Nature of Disability",
        
        # Enrollment Information
        "Enrolled Levels", "Enrolled Modules", "Enrolled Papers",
        
        # Financial Information
        "Fees Balance",
        
        # Results Summary
        "Has Results", "Result Count", "Latest Result Date",
        
        # System Information
        "Status", "Created Date", "Created By", "Updated Date", "Updated By"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data for each candidate
    for row, candidate in enumerate(candidates, 2):
        # Get enrollment information
        enrolled_levels = list(candidate.candidatelevel_set.all())
        enrolled_modules = list(candidate.candidatemodule_set.all())
        enrolled_papers = list(candidate.candidatepaper_set.all())
        
        # Get results information
        results = list(candidate.result_set.all())
        latest_result = results[0] if results else None
        
        # Get nature of disability
        disabilities = list(candidate.nature_of_disability.all())
        disability_names = ", ".join([d.name for d in disabilities]) if disabilities else ""
        
        # Prepare data row
        data = [
            # Bio Data
            candidate.reg_number or "",
            candidate.full_name or "",
            candidate.date_of_birth.strftime('%d/%m/%Y') if candidate.date_of_birth else "",
            candidate.get_gender_display() if candidate.gender else "",
            candidate.nationality or "",
            candidate.national_id or "",
            candidate.phone_number or "",
            candidate.email or "",
            candidate.next_of_kin or "",
            candidate.next_of_kin_phone or "",
            
            # Address Information
            candidate.district.name if candidate.district else "",
            candidate.village.name if candidate.village else "",
            candidate.physical_address or "",
            
            # Educational Background
            candidate.education_level or "",
            candidate.institution_attended or "",
            candidate.year_of_completion or "",
            
            # Assessment Information
            candidate.assessment_center.center_name if candidate.assessment_center else "",
            candidate.occupation.name if candidate.occupation else "",
            candidate.registration_category or "",
            candidate.entry_year or "",
            candidate.intake or "",
            candidate.assessment_date.strftime('%d/%m/%Y') if candidate.assessment_date else "",
            candidate.assessment_series.name if candidate.assessment_series else "",
            candidate.start_date.strftime('%d/%m/%Y') if candidate.start_date else "",
            candidate.finish_date.strftime('%d/%m/%Y') if candidate.finish_date else "",
            
            # Disability Information
            "Yes" if candidate.disability else "No",
            disability_names,
            
            # Enrollment Information
            ", ".join([f"{cl.level.name}" for cl in enrolled_levels]) if enrolled_levels else "",
            ", ".join([f"{cm.module.name}" for cm in enrolled_modules]) if enrolled_modules else "",
            ", ".join([f"{cp.paper.name}" for cp in enrolled_papers]) if enrolled_papers else "",
            
            # Financial Information
            float(candidate.fees_balance) if candidate.fees_balance else 0.00,
            
            # Results Summary
            "Yes" if results else "No",
            len(results),
            latest_result.assessment_date.strftime('%d/%m/%Y') if latest_result and latest_result.assessment_date else "",
            
            # System Information
            candidate.status or "Active",
            candidate.created_at.strftime('%d/%m/%Y %H:%M') if candidate.created_at else "",
            candidate.created_by.username if candidate.created_by else "",
            candidate.updated_at.strftime('%d/%m/%Y %H:%M') if candidate.updated_at else "",
            candidate.updated_by.username if candidate.updated_by else ""
        ]
        
        # Write data to worksheet
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            
            # Format specific columns
            if col in [3, 23, 25, 26, 33, 35, 37]:  # Date columns
                cell.alignment = Alignment(horizontal="center")
            elif col in [34, 36]:  # Numeric columns
                cell.alignment = Alignment(horizontal="right")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set minimum and maximum widths
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"EMIS_Candidates_Export_{timestamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response
