@login_required
def candidate_import_dual(request):
    """
    Handle GET (show dual import page) and POST (process Excel + photo zip upload).
    Excel is required, photos are optional and can be added later.
    """
    if request.method == 'GET':
        return render(request, 'candidates/import_dual.html')

    excel_file = request.FILES.get('excel_file')
    photo_zip = request.FILES.get('photo_zip')
    errors = []
    created = 0
    
    if not excel_file:
        errors.append('Excel file is required.')
        return render(request, 'candidates/import_dual.html', {'errors': errors, 'imported_count': 0})
    
    # Load Excel
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
    except Exception:
        errors.append('Invalid Excel file.')
        return render(request, 'candidates/import_dual.html', {'errors': errors, 'imported_count': 0})
    
    headers = [str(cell.value).replace(u'\xa0', ' ').strip().lower() for cell in ws[1] if cell.value]

    # Process photos if ZIP file is provided (optional)
    image_name_map = {}
    tmp_dir = None
    if photo_zip:
        tmp_dir = tempfile.mkdtemp()
        try:
            with zipfile.ZipFile(photo_zip) as zf:
                image_files = [n for n in zf.namelist() if n.lower().endswith(('.jpg', '.jpeg', '.png'))]
                for image_name in image_files:
                    # Extracts '614956.Kajumba Ruth' from '.../614956.Kajumba Ruth.jpg'
                    image_name_only = os.path.splitext(os.path.basename(image_name))[0]
                    if '.' in image_name_only:
                        try:
                            # Extracts 'Kajumba Ruth'
                            name_part = image_name_only.split('.', 1)[1]
                            # Replaces '_' with ' ' -> 'Kajumba Ruth'
                            cleaned_name = name_part.replace('_', ' ').strip()
                            # Map 'kajumba ruth' to '.../614956.Kajumba_Ruth.jpg'
                            image_name_map[cleaned_name.lower()] = image_name
                        except IndexError:
                            # This handles cases like '.DS_Store' or filenames without a name part
                            continue
                zf.extractall(tmp_dir)
        except Exception as e:
            errors.append(f'Invalid ZIP file or unable to extract images: {e}')
            if tmp_dir:
                import shutil
                shutil.rmtree(tmp_dir, ignore_errors=True)
            return render(request, 'candidates/import_dual.html', {'errors': errors, 'imported_count': 0})
    
    # Read Excel rows (always process Excel, regardless of photos)
    rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if not all(cell is None for cell in row)]
    
    # Process each candidate row
    for idx, row in enumerate(rows, start=2):
        data = dict(zip(headers, row))
        candidate_name_full = data.get('full_name', '').strip()

        # Handle cases where Excel has 3 names but image has 2
        name_parts = candidate_name_full.split()
        if len(name_parts) >= 3:
            # Use only the first two names for matching
            candidate_name_for_match = " ".join(name_parts[:2])
        else:
            candidate_name_for_match = candidate_name_full
        
        img_name = image_name_map.get(candidate_name_for_match.lower())

        # Note: If no image is found, we'll still create the candidate without a photo
        if not img_name and photo_zip:
            # Only log as info if photos were provided but this candidate's photo wasn't found
            errors.append(f"Row {idx}: Image not found for candidate '{candidate_name_full}' - candidate will be created without photo.")
        
        # --- (reuse import logic from candidate_import) ---
        form_data = data.copy()
        # Normalize registration_category for modular candidates
        regcat = str(form_data.get('registration_category', '')).strip().capitalize()
        if regcat == 'Modular':
            form_data['registration_category'] = 'Modular'
        
        # Robust date parsing: handle string, datetime, and Excel serial (float/int)
        from openpyxl.utils.datetime import from_excel
        import datetime as dt
        for date_field in ['date_of_birth', 'start_date', 'finish_date', 'assessment_date']:
            val = form_data.get(date_field)
            if not val:
                continue
            if isinstance(val, (dt.date, dt.datetime)):
                form_data[date_field] = val.date() if isinstance(val, dt.datetime) else val
            elif isinstance(val, (float, int)):
                try:
                    form_data[date_field] = from_excel(val).date()
                except Exception:
                    errors.append(f"Row {idx}: Invalid Excel serial date in '{date_field}'.")
                    continue
            elif isinstance(val, str):
                for fmt in ('%d/%m/%Y', '%-d/%-m/%Y', '%Y-%m-%d'):
                    try:
                        form_data[date_field] = dt.datetime.strptime(val, fmt).date()
                        break
                    except Exception:
                        continue
                else:
                    errors.append(f"Row {idx}: Invalid date format in '{date_field}'. Use D/M/YYYY, DD/MM/YYYY, or YYYY-MM-DD.")
                    continue
        
        # DEBUG: Print nationality value before validation
        print(f"[DEBUG] Row {idx} nationality value: '{form_data.get('nationality', '')}'")
        from django_countries import countries
        import re
        def normalize_country(val):
            val = val.lower().replace('&', 'and')
            val = re.sub(r'[^a-z0-9 ]', '', val)
            val = re.sub(r'\s+', ' ', val).strip()
            return val
        
        nat_val = form_data.get('nationality', '')
        if nat_val:
            nat_val_str = str(nat_val).strip()
            if nat_val_str:
                # Try exact match first
                country_found = None
                for code, name in countries:
                    if name.lower() == nat_val_str.lower():
                        country_found = code
                        break
                
                # If no exact match, try normalized matching
                if not country_found:
                    nat_normalized = normalize_country(nat_val_str)
                    for code, name in countries:
                        name_normalized = normalize_country(name)
                        if name_normalized == nat_normalized:
                            country_found = code
                            break
                
                if country_found:
                    form_data['nationality'] = country_found
                    print(f"[DEBUG] Row {idx} nationality mapped: '{nat_val_str}' -> '{country_found}'")
                else:
                    errors.append(f"Row {idx}: Invalid nationality '{nat_val_str}'. Must be a valid country name.")
                    continue
        
        # Handle occupation lookup
        occ_val = form_data.get('occupation')
        if occ_val:
            occ_str = str(occ_val).strip()
            occupation = Occupation.objects.filter(name__iexact=occ_str).first()
            if occupation:
                form_data['occupation'] = occupation.id
            else:
                errors.append(f"Row {idx}: Occupation '{occ_str}' not found.")
                continue
        
        # Handle assessment center lookup
        center_val = form_data.get('assessment_center')
        if center_val:
            center_str = str(center_val).strip()
            center = AssessmentCenter.objects.filter(name__iexact=center_str).first()
            if center:
                form_data['assessment_center'] = center.id
            else:
                errors.append(f"Row {idx}: Assessment Center '{center_str}' not found.")
                continue
        
        # Use CandidateForm for validation, but patch required fields for import
        form = CandidateForm(form_data)
        for f in ['district', 'village']:
            if f in form.fields:
                form.fields[f].required = False
        
        if not form.is_valid():
            print(f"[DEBUG] Row {idx} SKIPPED: CandidateForm errors: {form.errors}")
            errors.append(f"Row {idx}: Form errors: {form.errors}")
            continue
        
        # District and village: truly optional for import (skip if missing or blank)
        for loc_field, model_cls in [('district', District), ('village', Village)]:
            val = form_data.get(loc_field)
            if val is None or str(val).strip() == '':
                form_data[loc_field] = None
            else:
                val_str = str(val).strip()
                obj = model_cls.objects.filter(name__iexact=val_str).first()
                form_data[loc_field] = obj.id if obj else None
        
        # Debug: print after district/village assignment
        print(f"[DEBUG] Row {idx} after district/village assignment: district={form_data.get('district')}, village={form_data.get('village')}")

        # Define date and foreign key fields for use below
        date_fields = ['date_of_birth', 'start_date', 'finish_date', 'assessment_date']
        fk_fields = ['occupation', 'assessment_center', 'district', 'village']
        
        # Coerce every other field except dates and foreign keys to string
        for k in form_data:
            if k not in date_fields + fk_fields:
                v = form_data[k]
                if v is not None and not isinstance(v, str):
                    form_data[k] = str(v)
        
        # Debug: print types of all fields before form creation
        print(f"[IMPORT DEBUG] Row {idx} form_data types: " + ", ".join(f"{k}: {type(v).__name__}" for k,v in form_data.items()))
        
        # Remove reg_number if present
        form_data.pop('reg_number', None)
        
        # Use CandidateForm for validation, but patch required fields for import
        form = CandidateForm(form_data)
        for f in ['district', 'village']:
            if f in form.fields:
                form.fields[f].required = False
        
        if not form.is_valid():
            error_list = '; '.join([f"{k}: {v[0]}" for k, v in form.errors.items()])
            errors.append(f"Row {idx}: {error_list}")
            continue
        
        # Convert date fields in form.cleaned_data from DD/MM/YYYY string to date objects
        for date_field in date_fields:
            val = form.cleaned_data.get(date_field)
            if val and isinstance(val, str):
                for fmt in ('%d/%m/%Y', '%-d/%-m/%Y', '%Y-%m-%d'):
                    try:
                        form.cleaned_data[date_field] = dt.datetime.strptime(val, fmt).date()
                        break
                    except ValueError:
                        continue
        
        # Create candidate
        candidate = Candidate(**form.cleaned_data)
        
        # Attach photo if available
        if img_name and tmp_dir:
            img_path = os.path.join(tmp_dir, img_name)
            if os.path.exists(img_path):
                from PIL import Image
                import io
                with Image.open(img_path) as img:
                    if img.mode != 'RGB':
                        img = img.convert('RGB')
                    buffer = io.BytesIO()
                    img.save(buffer, format='JPEG', quality=85)
                    buffer.seek(0)
                    candidate.passport_photo.save(
                        os.path.splitext(img_name)[0] + '.jpg',
                        File(buffer),
                        save=False
                    )
            else:
                errors.append(f"Row {idx}: Image file '{img_name}' not found after extraction - candidate created without photo.")
        
        candidate.save()
        created += 1
        print(f"[DEBUG] Row {idx} IMPORTED: Candidate '{candidate.full_name}' saved.")
    
    # Cleanup temporary directory if photos were processed
    if tmp_dir:
        import shutil
        shutil.rmtree(tmp_dir, ignore_errors=True)
    
    success_message = None
    if created > 0:
        success_message = f"{created} candidate{'s' if created != 1 else ''} imported successfully."
    
    return render(request, 'candidates/import_dual.html', {
        'errors': errors,
        'imported_count': created,
        'success_message': success_message
    })
