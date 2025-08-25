from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, JsonResponse
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from django.contrib.auth.decorators import login_required
from reportlab.platypus import Image as RLImage
import calendar
import logging
from django.db.models import Count, Q
from .models import Candidate, Occupation, AssessmentCenter, Result
from .models import SupportStaff
from .forms import SupportStaffForm
from .models import Staff
from .forms import StaffForm
from .forms import CenterRepForm
from .models import AssessmentSeries
from .forms import AssessmentSeriesForm
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, BaseDocTemplate, PageTemplate, Frame
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from io import BytesIO
from collections import Counter
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Session Management Utilities
def get_user_staff_info(request):
    """
    Safely get staff information for authenticated users.
    Returns tuple: (staff_object, user_department, is_authenticated)
    """
    if not request.user.is_authenticated:
        return None, None, False
    
    try:
        staff = Staff.objects.get(user=request.user)
        return staff, staff.department, True
    except Staff.DoesNotExist:
        # User is authenticated but not a staff member
        return None, None, True

def require_staff_permissions(request, required_departments=None):
    """
    Check if user has required staff permissions.
    Returns tuple: (has_permission, staff_object, user_department)
    """
    staff, user_department, is_authenticated = get_user_staff_info(request)
    
    if not is_authenticated:
        return False, None, None
    
    # Superuser always has permission
    if request.user.is_superuser:
        return True, staff, user_department
    
    # Check specific department requirements
    if required_departments and user_department not in required_departments:
        return False, staff, user_department
    
    return True, staff, user_department

@login_required
def staff_list(request):
    """List all staff members with their departments"""
    staff_members = Staff.objects.select_related('user').all()
    return render(request, 'users/staff/list_staff.html', {'staff_members': staff_members})

@login_required
def staff_create(request):
    """Create a new staff member with department assignment"""
    if request.method == 'POST':
        print(f"[DEBUG] POST data: {request.POST}")
        form = StaffForm(request.POST)
        print(f"[DEBUG] Form is_valid: {form.is_valid()}")
        
        if not form.is_valid():
            print(f"[DEBUG] Form errors: {form.errors}")
            print(f"[DEBUG] Form non_field_errors: {form.non_field_errors()}")
            
        if form.is_valid():
            try:
                print("[DEBUG] Form is valid, attempting to save...")
                staff = form.save(commit=False)
                print(f"[DEBUG] Staff object created: {staff}")
                staff.created_by = request.user
                staff.updated_by = request.user
                staff.save()
                print(f"[DEBUG] Staff saved successfully: {staff.id}")
                messages.success(request, f'Staff member {staff.name} created successfully!')
                return redirect('staff_list')
            except Exception as e:
                print(f"[DEBUG] Exception during save: {e}")
                import traceback
                traceback.print_exc()
                messages.error(request, f'Error creating staff: {e}')
    else:
        form = StaffForm()
    return render(request, 'users/staff/create_staff.html', {'form': form})
@login_required
def staff_detail(request, pk):
    """View staff member details"""
    staff = get_object_or_404(Staff, pk=pk)
    return render(request, 'users/staff/view_staff.html', {'staff': staff})

@login_required
def staff_edit(request, pk):
    """Edit staff member details and department"""
    staff = get_object_or_404(Staff, pk=pk)
    if request.method == 'POST':
        form = StaffForm(request.POST, instance=staff)
        if form.is_valid():
            staff_obj = form.save(commit=False)
            staff_obj.updated_by = request.user
            staff_obj.save()
            messages.success(request, f'Staff member {staff.name} updated successfully!')
            return redirect('staff_detail', pk=staff.pk)
    else:
        form = StaffForm(instance=staff)
    return render(request, 'users/staff/edit_staff.html', {'form': form, 'staff': staff})

def get_user_department_modules(user):
    """Get modules accessible to user based on their staff department"""
    try:
        staff = Staff.objects.get(user=user)  # Changed from SupportStaff to Staff
        department = staff.department
        
        if department == 'Research':
            return {
                'assessment_centers': {'access': 'create_edit', 'name': 'Assessment Centers'},
                'occupations': {'access': 'view_only', 'name': 'Occupations'},
                'statistics': {'access': 'view_only', 'name': 'Statistics'},
            }
        elif department == 'Data':
            return {
                'candidates': {'access': 'full_access', 'name': 'Candidates'},
                'results': {'access': 'full_access', 'name': 'Results'},
                'reports': {'access': 'full_access', 'name': 'Reports'},
            }
    except Staff.DoesNotExist:  # Changed from SupportStaff to Staff
        pass
    
    return {}

@login_required
def generate_result_list(request):
    months = [
        {"value": "1", "name": "January"},
        {"value": "2", "name": "February"},
        {"value": "3", "name": "March"},
        {"value": "4", "name": "April"},
        {"value": "5", "name": "May"},
        {"value": "6", "name": "June"},
        {"value": "7", "name": "July"},
        {"value": "8", "name": "August"},
        {"value": "9", "name": "September"},
        {"value": "10", "name": "October"},
        {"value": "11", "name": "November"},
        {"value": "12", "name": "December"},
    ]
    years = ["2021", "2022", "2023", "2024", "2025"]

    if request.method == 'POST':
        # Get form data
        month = request.POST.get('assessment_month')
        year = request.POST.get('assessment_year')
        regcat = request.POST.get('registration_category')
        occupation_id = request.POST.get('occupation')
        level_id = request.POST.get('level')
        center_id = request.POST.get('assessment_center')
        # Validate required fields
        errors = []
        if not (month and year and regcat and occupation_id):
            errors.append('Please fill in all required fields.')
        from datetime import datetime
        from .models import Candidate, Result, Occupation, Level
        candidates = []
        result_data = []
        papers_list = []
        if not errors:
            # Check if results have been released for this assessment period
            # Only block center representatives, allow admin staff and support to see results
            is_center_rep = request.user.groups.filter(name='CenterRep').exists()
            
            # Find the assessment series for this period
            from .models import AssessmentSeries, AssessmentCenter
            import calendar
            month_name = calendar.month_name[int(month)]
            series_name = f"{month_name} {year} Series"
            
            assessment_series = AssessmentSeries.objects.filter(
                start_date__year=year,
                start_date__month=month
            ).first()
            
            if not assessment_series:
                # Try to find by name pattern if no exact date match
                assessment_series = AssessmentSeries.objects.filter(name=series_name).first()
            
            # Check if results are released
            results_released = assessment_series and assessment_series.results_released or not is_center_rep
            
            if not results_released:
                # Return a message indicating results are not yet released
                context = {
                    'months': months,
                    'years': years,
                    'occupations': Occupation.objects.all(),
                    'levels': Level.objects.all(),
                    'assessment_centers': AssessmentCenter.objects.all(),
                    'results_not_released': True,
                    'assessment_period': f"{month_name} {year}",
                    'series_name': series_name if assessment_series else f"{month_name} {year} Series"
                }
                return render(request, 'reports/result_list.html', context)
            
            # Build assessment period filter
            period_start = datetime(int(year), int(month), 1)
            period_end = datetime(int(year), int(month), 28)  # crude, will catch all results in month
            # Filter candidates for occupation, level, center
            filters = {
                'occupation_id': occupation_id,
                'registration_category__iexact': regcat,
            }
            # For paper-based, fetch all papers for occupation/level
            structure_type = 'modules'
            if regcat.lower() == 'formal' and occupation_id and level_id:
                from .models import OccupationLevel, Paper
                occ_level = OccupationLevel.objects.filter(occupation_id=occupation_id, level_id=level_id).first()
                if occ_level:
                    structure_type = occ_level.structure_type
                if structure_type == 'papers':
                    papers = Paper.objects.filter(occupation_id=occupation_id, level_id=level_id)
                    papers_list = [{'code': p.code, 'name': p.name, 'type': p.get_grade_type_display()} for p in papers]
                    print(f'[DEBUG] Papers for occupation_id={occupation_id}, level_id={level_id}: {papers_list}')
            if center_id:
                filters['assessment_center_id'] = center_id
            if level_id:
                filters['candidatelevel__level_id'] = level_id
                from .models import CandidateLevel
                candidate_levels = CandidateLevel.objects.filter(level_id=level_id)
                print(f'[DEBUG] CandidateLevel entries for level_id={level_id}: ' + str([{"candidate_id": cl.candidate_id, "level_id": cl.level_id} for cl in candidate_levels]))
                # Print all candidates for this level with their registration_category and occupation_id
                from .models import Candidate
                candidate_ids = [cl.candidate_id for cl in candidate_levels]
                debug_candidates = Candidate.objects.filter(id__in=candidate_ids)
                for c in debug_candidates:
                    print(f'[DEBUG] Candidate id={c.id}, reg_number={c.reg_number}, full_name={c.full_name}, occupation_id={c.occupation_id}, registration_category={c.registration_category}')
            qs = Candidate.objects.filter(**filters).distinct()
            print(f'[DEBUG] Candidate queryset count: {qs.count()}')
            print(f'[DEBUG] Candidate IDs in queryset: {list(qs.values_list("id", flat=True))}')
            print(f'[DEBUG] Candidate filter values: occupation_id={occupation_id}, regcat={regcat}, level_id={level_id}, center_id={center_id}')
            # Only keep candidates with at least one result in period
            for cand in qs:
                print(f'[DEBUG] Processing candidate: id={cand.id}, regno={cand.reg_number}, name={cand.full_name}, center={getattr(cand.assessment_center, "center_name", None)}')
                regcat_lower = regcat.lower()
                if regcat_lower == 'modular':
                    results = Result.objects.filter(
                        candidate=cand,
                        assessment_date__year=year,
                        assessment_date__month=month,
                        result_type='modular',
                    )
                    print(f'[DEBUG] Filtering results for candidate {cand.reg_number} (modular): year={year}, month={month}, result_type=modular')
                elif regcat_lower == 'formal':
                    # Determine structure type for occupation/level
                    structure_type = 'modules'
                    if occupation_id and level_id:
                        from .models import OccupationLevel
                        occ_level = OccupationLevel.objects.filter(occupation_id=occupation_id, level_id=level_id).first()
                        if occ_level:
                            structure_type = occ_level.structure_type
                    if structure_type == 'papers':
                        # Paper-based: fetch paper results with optimized queries
                        results = Result.objects.filter(
                            candidate=cand,
                            assessment_date__year=year,
                            assessment_date__month=month,
                            result_type='formal',
                            paper__isnull=False
                        ).select_related('paper', 'user')
                        print(f'[DEBUG] Filtering results for candidate {cand.reg_number} (formal, papers): year={year}, month={month}, result_type=formal, structure_type=papers')
                    else:
                        # Module-based: keep old logic with optimized queries
                        results = Result.objects.filter(
                            candidate=cand,
                            assessment_date__year=year,
                            assessment_date__month=month,
                            result_type='formal',
                        ).select_related('user')
                        print(f'[DEBUG] Filtering results for candidate {cand.reg_number} (formal, modules): year={year}, month={month}, result_type=formal, structure_type=modules')
                elif regcat_lower in ['informal', "worker's pas", "workers pas"]:
                    results = Result.objects.filter(
                        candidate=cand,
                        assessment_date__year=year,
                        assessment_date__month=month,
                        result_type='informal',
                    ).select_related('user')
                    print(f'[DEBUG] Filtering results for candidate {cand.reg_number} (informal): year={year}, month={month}, result_type=informal')
                else:
                    results = Result.objects.none()
                print(f'[DEBUG] Results found for candidate {cand.reg_number}: {results.count()}')
                if results.exists():
                    candidates.append(cand)
                    # Serialize candidate fields needed by template
                    cand_dict = {
                        'id': cand.id,  # Add id field for module grouping
                        'reg_number': cand.reg_number,
                        'full_name': cand.full_name,
                        'gender': cand.get_gender_display(),
                        'passport_photo_with_regno': cand.passport_photo_with_regno.url if cand.passport_photo_with_regno else None,
                        'passport_photo': cand.passport_photo.url if cand.passport_photo else None,
                        'assessment_center': getattr(cand.assessment_center, 'center_name', None),
                    }
                    # Serialize results fields needed by template
                    results_list = []
                    if regcat_lower == 'formal' and 'structure_type' in locals() and structure_type == 'papers':
                        for r in results:
                            results_list.append({
                                'grade': r.grade,
                                'comment': r.comment,
                                'assessment_type': getattr(r, 'assessment_type', ''),
                                'paper_code': r.paper.code if r.paper else '',
                                'paper_name': r.paper.name if r.paper else '',
                                'paper_type': r.paper.get_grade_type_display() if r.paper else '',
                                'mark': r.mark,
                                'date': r.assessment_date,
                                'status': getattr(r, 'status', ''),
                                'user': getattr(r, 'user', ''),
                            })
                    else:
                        for r in results:
                            results_list.append({
                                'grade': r.grade,
                                'comment': r.comment,
                                'assessment_type': getattr(r, 'assessment_type', ''),
                                'module_code': getattr(r, 'module_code', ''),
                                'module_name': getattr(r, 'module_name', ''),
                                'mark': r.mark,
                                'date': r.assessment_date,
                                'status': getattr(r, 'status', ''),
                                'user': getattr(r, 'user', ''),
                            })
                    is_successful = any(r['comment'] == 'Successful' for r in results_list)
                    result_data.append({
                        'candidate': cand_dict,
                        'results': results_list,
                        'successful': is_successful,
                    })
        # Get occupation and level names for display
        occupation_name = None
        level_name = None
        if occupation_id:
            try:
                occ = Occupation.objects.filter(id=int(occupation_id)).first()
                if occ:
                    occupation_name = occ.name
            except Exception:
                occupation_name = None
        if level_id:
            try:
                print(f"DEBUG: level_id received: {level_id}")
                lvl = Level.objects.filter(id=int(level_id)).first()
                if lvl:
                    level_name = lvl.name
                else:
                    pass
            except Exception as e:
                level_name = None
        else:
            pass
        # Format month as "June, 2025"
        import calendar
        formatted_period = f"{calendar.month_name[int(month)]}, {year}"
        
        # For modular category, group by modules instead of centers (DISABLED - restoring original functionality)
        # Enable module grouping for modular category
        module_result_data = None
        if regcat.lower() == 'modular':
            from collections import defaultdict
            from .models import CandidateModule, Module
            module_result_data = defaultdict(list)
            
            # Get all results for modular category
            modular_results = Result.objects.filter(
                candidate__occupation_id=occupation_id,
                assessment_date__year=year,
                assessment_date__month=month,
                result_type='modular',
            )
            if level_id:
                modular_results = modular_results.filter(candidate__candidatelevel__level=level_id)
            if center_id:
                modular_results = modular_results.filter(candidate__assessment_center=center_id)
            
            # Process all results to group by module
            for result in modular_results:
                candidate = result.candidate
                candidate_id = candidate.id
                
                # Get all modules this candidate is enrolled in
                candidate_modules = CandidateModule.objects.filter(candidate_id=candidate_id).select_related('module')
                
                for cm in candidate_modules:
                    module = cm.module
                    module_key = f"{module.code} - {module.name}"
                    
                    # Check if this result belongs to this module
                    if hasattr(result, 'module') and result.module == module:
                        # Create candidate entry for this module
                        candidate_entry = {
                            'id': candidate.id,
                            'reg_number': candidate.reg_number,
                            'full_name': candidate.full_name,
                            'gender': candidate.get_gender_display(),
                            'assessment_center': getattr(candidate.assessment_center, 'center_name', None),
                        }
                        
                        # Add result data
                        result_entry = {
                            'candidate': candidate_entry,
                            'results': [{
                                'grade': result.grade,
                                'comment': result.comment,
                                'mark': result.mark,
                            }],
                            'successful': result.comment != 'CTR',
                        }
                        
                        module_result_data[module_key].append(result_entry)
            
            # Convert to regular dict
            module_result_data = dict(module_result_data)
                    # Enable module and paper grouping for informal category
        informal_module_data = None
        if regcat.lower() in ['informal', "worker's pas", "workers pas"] and result_data:
            from collections import defaultdict
            from .models import Module, Paper
            informal_module_data = defaultdict(lambda: defaultdict(lambda: {'papers': [], 'candidates': []}))
            
            # Get all modules for this level to ensure we show all modules even if no results
            if level_id and occupation_id:
                all_modules = Module.objects.filter(occupation_id=occupation_id, level_id=level_id)
                for module in all_modules:
                    module_key = f"{module.code} - {module.name}"
                    # Get all papers for this module
                    module_papers = Paper.objects.filter(module=module)
                    # Initialize for all centers that have candidates
                    centers_with_candidates = set(entry['candidate']['assessment_center'] for entry in result_data)
                    for center_name in centers_with_candidates:
                        informal_module_data[center_name][module_key]['papers'] = [{
                            'code': p.code,
                            'name': p.name,
                            'id': p.id
                        } for p in module_papers]
            
            # Process all results to group by center, then by module
            processed_candidates = set()
            for entry in result_data:
                candidate = entry['candidate']
                candidate_id = candidate['id']
                center_name = candidate['assessment_center'] or 'Unknown Center'
                
                # Get all results for this candidate
                candidate_results = Result.objects.filter(
                    candidate_id=candidate_id,
                    assessment_date__year=year,
                    assessment_date__month=month,
                    result_type='informal',
                ).select_related('module', 'paper')
                
                for result in candidate_results:
                    if result.module:
                        module = result.module
                        # Only process results from the selected level
                        if level_id and module.level_id != int(level_id):
                            continue
                        module_key = f"{module.code} - {module.name}"
                        
                        # Check if we already processed this candidate for this module in this center
                        candidate_module_key = f"{candidate_id}_{module.id}_{center_name}"
                        if candidate_module_key not in processed_candidates:
                            processed_candidates.add(candidate_module_key)
                            
                            # Create candidate entry for this module in this center
                            candidate_entry = {
                                'id': candidate['id'],
                                'reg_number': candidate['reg_number'],
                                'full_name': candidate['full_name'],
                                'gender': candidate['gender'],
                                'passport_photo_with_regno': candidate['passport_photo_with_regno'],
                                'passport_photo': candidate['passport_photo'],
                                'assessment_center': candidate['assessment_center'],
                            }
                            
                            # Get all results for this candidate in this module
                            candidate_module_results = candidate_results.filter(module=module)
                            
                            # Create paper results mapping
                            paper_results = {}
                            for r in candidate_module_results:
                                if r.paper:
                                    paper_results[r.paper.id] = {
                                        'grade': r.grade,
                                        'comment': r.comment,
                                        'mark': r.mark,
                                    }
                            
                            candidate_entry['paper_results'] = paper_results
                            candidate_entry['has_ctr'] = any(r.comment == 'CTR' for r in candidate_module_results)
                            
                            informal_module_data[center_name][module_key]['candidates'].append(candidate_entry)
            
            # Convert to regular dict
            informal_module_data = dict(informal_module_data)
            for center_name in informal_module_data:
                informal_module_data[center_name] = dict(informal_module_data[center_name])



        # Group by center if all centers (for non-modular or when no module grouping)
        centered_result_data = None
        if not center_id:
            from collections import defaultdict
            centered_result_data = defaultdict(list)
            for entry in result_data:
                center_name = entry['candidate'].get('assessment_center', 'No Center')
                centered_result_data[center_name].append(entry)
        # Always include occupations, levels, centers for dropdowns
        from .models import AssessmentCenter
        occupations = Occupation.objects.all().order_by('code')
        levels = Level.objects.all()
        centers = AssessmentCenter.objects.all()
        print(f'[DEBUG] centered_result_data: {centered_result_data}')
        print(f'[DEBUG] informal_module_data: {informal_module_data}')

        context = {
            'months': months,
            'years': years,
            'preview': True,
            'form_data': {
                'month': month,
                'year': year,
                'regcat': regcat,
                'occupation_id': occupation_id,
                'occupation_name': occupation_name,
                'level_id': level_id,
                'level_name': level_name,
                'center_id': center_id,
            },
            'candidates': candidates,
            'result_data': result_data,
            'data': result_data,  # Template expects 'data' variable
            'centered_result_data': dict(centered_result_data) if centered_result_data else None,
            'module_result_data': dict(module_result_data) if module_result_data else None,
            'informal_module_data': informal_module_data,
            'formatted_period': formatted_period,
            'logo_path': '/static/images/uvtab_logo.png',
            'errors': errors,
            'occupations': occupations,
            'levels': levels,
            'centers': centers,
            'papers_list': papers_list,
        }
        return render(request, 'reports/result_list.html', context)
    else:
        # GET: show form only
        from .models import AssessmentCenter, Occupation, Level
        occupations = Occupation.objects.all().order_by('code')
        levels = []  # Don't load all levels initially - will be loaded via AJAX
        centers = AssessmentCenter.objects.all()
        context = {
            'preview': False,
            'form_data': {},
            'candidates': [],
            'result_data': [],
            'centered_result_data': None,
            'formatted_period': '',
            'logo_path': '/static/images/uvtab_logo.png',
            'errors': [],
            'occupations': occupations,
            'levels': levels,
            'centers': centers,
            'months': months,
            'years': years,
        }
        return render(request, 'reports/result_list.html', context)
        return render(request, 'reports/result_list.html', context)


@login_required
def results_home(request):
    logger = logging.getLogger(__name__)
    logger.info(f'Results home accessed by user: {request.user}')
    
    # Check if results have been released - filter candidates based on their assessment series
    # Only block center representatives, allow admin staff and support to see results
    from .models import AssessmentSeries
    is_center_rep = request.user.groups.filter(name='CenterRep').exists()
    
    # If user is a center rep, only show candidates whose assessment series has results released
    if is_center_rep:
        # Filter to only include candidates from assessment series with results released
        enrolled_candidates = enrolled_candidates.filter(
            assessment_series__results_released=True
        )
        
        # Check if there are any candidates left after filtering
        if not enrolled_candidates.exists():
            # No candidates with released results
            return render(request, 'results/home.html', {
                'results_not_released': True,
                'current_series': None,
                'candidates_with_status': [],
                'page_obj': None,
                'paginator': None,
                'reg_categories': [],
                'filters': {
                'reg_number': '',
                'name': '',
                'registration_category': '',
            }
        })
    
    # Get filter parameters
    reg_number = request.GET.get('reg_number', '').strip()
    name = request.GET.get('name', '').strip()
    registration_category = request.GET.get('registration_category', '').strip()
    
    # Fetch enrolled candidates with their marks status
    from django.db.models import Q
    from django.core.paginator import Paginator
    
    enrolled_candidates = Candidate.objects.filter(
        status='Active'
    ).filter(
        Q(candidatelevel__isnull=False) | Q(candidatemodule__isnull=False)
    ).distinct().select_related('occupation', 'assessment_center')
    
    # Apply filters
    if reg_number:
        enrolled_candidates = enrolled_candidates.filter(reg_number__icontains=reg_number)
    if name:
        enrolled_candidates = enrolled_candidates.filter(full_name__icontains=name)
    if registration_category:
        enrolled_candidates = enrolled_candidates.filter(registration_category=registration_category)
    
    enrolled_candidates = enrolled_candidates.order_by('reg_number')
    
    # Pagination: 50 per page
    paginator = Paginator(enrolled_candidates, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Add upload status annotation
    candidates_with_status = []
    for candidate in page_obj.object_list:
        # Check if candidate has any results/marks
        has_marks = Result.objects.filter(candidate=candidate).exists()
        
        candidates_with_status.append({
            'candidate': candidate,
            'upload_status': 'Uploaded' if has_marks else 'Not Uploaded'
        })
    
    # Get registration categories for filter dropdown
    reg_categories = [
        ('Formal', 'Formal'),
        ('Modular', 'Modular'),
        ('Informal', "Worker's PAS")
    ]
    
    return render(request, 'results/home.html', {
        'candidates_with_status': candidates_with_status,
        'page_obj': page_obj,
        'paginator': paginator,
        'reg_categories': reg_categories,
        'filters': {
            'reg_number': reg_number,
            'name': name,
            'registration_category': registration_category,
        }
    })

from django.views.decorators.http import require_POST
from django.http import JsonResponse
from openpyxl import Workbook
from openpyxl import load_workbook
from io import BytesIO
from .models import Candidate, Module, Paper, Level, Occupation, CandidateLevel, CandidateModule, CandidatePaper, Result
import datetime


@login_required
@require_POST
def upload_marks(request):
    """
    Bulk upload candidate marks from Excel for all registration categories.
    Expects POST with Excel file and selection params (occupation, level, registration_category, etc).
    """
    
    errors = []
    updated = 0
    file = request.FILES.get('marks_file') or request.FILES.get('excel_file')
    regcat = request.POST.get('registration_category')
    occupation_id = request.POST.get('occupation')
    level_id = request.POST.get('level')
    center_id = request.POST.get('assessment_center')
    modules_param = request.POST.get('modules')
    assessment_series_id = request.POST.get('assessment_series')
    
    # Get assessment series object
    assessment_series = None
    if assessment_series_id:
        assessment_series = AssessmentSeries.objects.filter(pk=assessment_series_id).first()
        if not assessment_series:
            return JsonResponse({'success': False, 'error': 'Invalid assessment series selected.'})
    else:
        return JsonResponse({'success': False, 'error': 'Assessment series is required.'})
    # Normalize registration category
    regcat_normalized = regcat.strip().lower() if regcat else ''
    if regcat_normalized in ["workers_pas", "worker's pas"]:
        regcat_normalized = "informal"
    try:
        wb = load_workbook(file)
        ws = wb.active
    except Exception as e:
        print('Excel file load error:', e)
        return JsonResponse({'success': False, 'error': 'Invalid Excel file.'})
    import re
    def normalize_header(h):
        return re.sub(r'[^a-z0-9]', '', h.strip().lower())
    headers = [str(cell.value).strip() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    normalized_headers = {normalize_header(h): h for h in headers}
    regno_header = None
    for candidate in ['regnumber', 'registrationno', 'registrationnumber', 'regno']:
        if candidate in normalized_headers:
            regno_header = normalized_headers[candidate]
            break
    if not regno_header:
        return JsonResponse({'success': False, 'error': 'Missing registration number (reg_number) column in Excel. Please use a marksheet generated by the system.'})

    # Get occupation, level, center
    occupation = Occupation.objects.filter(pk=occupation_id).first() if occupation_id else None
    level = Level.objects.filter(pk=level_id).first() if level_id else None
    # Determine structure_type
    structure_type = None
    from .models import OccupationLevel
    occ_level = None
    if occupation and level:
        occ_level = OccupationLevel.objects.filter(occupation=occupation, level=level).first()
        if occ_level:
            structure_type = occ_level.structure_type
    # Determine which modules/papers are in the Excel
    code_to_obj = {}
    # Modular: use selected module and 'PRACTICAL' column
    if regcat_normalized == 'modular':
        selected_module_id = modules_param
        selected_module = Module.objects.filter(pk=selected_module_id).first() if selected_module_id else None
    # Formal module-based: use 'THEORY' and 'PRACTICAL' columns
    # Paper-based/informal: use paper codes as columns
    if regcat_normalized in ['formal', 'informal', 'modular']:
        papers = Paper.objects.filter(occupation=occupation, level=level)
        for p in papers:
            code_to_obj[p.code] = p
        modules = Module.objects.filter(occupation=occupation, level=level)
        for m in modules:
            code_to_obj[m.code] = m
    # For each row, update results
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        row_data = dict(zip(headers, row))
        print(f"[DEBUG] Row {idx} raw data: {row_data}")
        regno = str(row_data.get(regno_header, '')).strip()
        print(f"[DEBUG] Row {idx} regno: '{regno}'")
        if not regno:
            errors.append(f"Row {idx}: Missing reg_number.")
            continue
        candidate = Candidate.objects.filter(reg_number__iexact=regno).first()
        if not candidate:
            errors.append(f"Row {idx}: Candidate with reg_number '{regno}' not found.")
            continue
        # Validate candidate registration category, occupation, level
        cand_regcat = getattr(candidate, 'registration_category', '').strip().lower()
        cand_occupation_id = getattr(candidate, 'occupation_id', None)
        if regcat_normalized != cand_regcat:
            errors.append(f"Row {idx}: Candidate '{regno}' registration category mismatch.")
            continue
        if occupation and candidate.occupation_id != occupation.id:
            errors.append(f"Row {idx}: Candidate '{regno}' occupation mismatch.")
            continue
        # For formal/module-based and informal, check level enrollment
        if regcat_normalized in ['formal', 'informal'] and level:
            if not CandidateLevel.objects.filter(candidate=candidate, level=level).exists():
                errors.append(f"Row {idx}: Candidate '{regno}' not enrolled in selected level.")
                continue
        # Assessment date from assessment series
        assessment_day = assessment_series.start_date
        print(f"Assessment date from series: {assessment_day}")
        # Modular: expects PRACTICAL column and selected module
        if regcat_normalized == 'modular':
            mark = row_data.get('PRACTICAL') or row_data.get('Practical') or row_data.get('practical')
            if selected_module is None:
                errors.append(f"Row {idx}: No module selected for modular upload.")
                continue
            if mark is None or str(mark).strip() == '':
                errors.append(f"Row {idx}: Missing practical mark for candidate '{regno}'.")
                continue
            try:
                mark_val = float(mark)
            except Exception:
                errors.append(f"Row {idx}: Invalid practical mark for candidate '{regno}'.")
                continue
            Result.objects.update_or_create(
                candidate=candidate,
                module=selected_module,
                assessment_date=assessment_day,
                assessment_series=assessment_series,
                result_type='modular',
                defaults={
                    'assessment_type': 'practical',
                    'mark': mark_val,
                    'user': request.user,
                    'status': ''
                }
            )
            updated += 1
        # Formal module-based: expects THEORY and PRACTICAL columns
        elif regcat_normalized == 'formal' and structure_type == 'modules':
            theory_mark = row_data.get('THEORY') or row_data.get('Theory') or row_data.get('theory')
            practical_mark = row_data.get('PRACTICAL') or row_data.get('Practical') or row_data.get('practical')
            # Save theory result (if present)
            if theory_mark is not None and str(theory_mark).strip() != '':
                try:
                    theory_mark_val = float(theory_mark)
                except Exception:
                    errors.append(f"Row {idx}: Invalid THEORY mark for candidate '{regno}'.")
                    theory_mark_val = None
                if theory_mark_val is not None:
                    Result.objects.update_or_create(
                        candidate=candidate,
                        level=level,
                        assessment_date=assessment_day,
                        assessment_series=assessment_series,
                        assessment_type='theory',
                        result_type='formal',
                        defaults={
                            'mark': theory_mark_val,
                            'user': request.user,
                            'status': ''
                        }
                    )
                    updated += 1
            # Save practical result (if present)
            if practical_mark is not None and str(practical_mark).strip() != '':
                try:
                    practical_mark_val = float(practical_mark)
                except Exception:
                    errors.append(f"Row {idx}: Invalid PRACTICAL mark for candidate '{regno}'.")
                    practical_mark_val = None
                if practical_mark_val is not None:
                    Result.objects.update_or_create(
                        candidate=candidate,
                        level=level,
                        assessment_date=assessment_day,
                        assessment_series=assessment_series,
                        assessment_type='practical',
                        result_type='formal',
                        defaults={
                            'mark': practical_mark_val,
                            'user': request.user,
                            'status': ''
                        }
                    )
                    updated += 1
        # Formal paper-based and informal: expects paper code columns
        else:
            for code, obj in code_to_obj.items():
                mark = row_data.get(code)
                if mark is None or str(mark).strip() == '':
                    continue
                try:
                    mark_val = float(mark)
                except Exception:
                    errors.append(f"Row {idx}: Invalid mark for {code} (candidate '{regno}').")
                    continue
                # Allow -1 for "Missing" (Ms) status, otherwise marks must be between 0 and 100
                if not (mark_val == -1 or (0 <= mark_val <= 100)):
                    errors.append(f"Row {idx}: Mark for {code} (candidate '{regno}') must be between 0 and 100, or -1 for Missing. Got {mark_val}.")
                    continue
                # Paper-based formal
                if regcat_normalized == 'formal' and structure_type == 'papers':
                    Result.objects.update_or_create(
                        candidate=candidate,
                        level=level,
                        paper=obj,
                        assessment_date=assessment_day,
                        assessment_series=assessment_series,
                        result_type='formal',
                        defaults={
                            'assessment_type': obj.grade_type if hasattr(obj, 'grade_type') else 'practical',
                            'mark': mark_val,
                            'user': request.user,
                            'status': ''
                        }
                    )
                    updated += 1
                # Informal/worker's PAS
                elif regcat_normalized == 'informal':
                    candidate_paper = CandidatePaper.objects.filter(candidate=candidate, paper=obj, level=level).first()
                    module = candidate_paper.module if candidate_paper else None
                    Result.objects.update_or_create(
                        candidate=candidate,
                        level=level,
                        module=module,
                        paper=obj,
                        assessment_type='practical',
                        assessment_date=assessment_day,
                        assessment_series=assessment_series,
                        result_type='informal',
                        defaults={
                            'mark': mark_val,
                            'user': request.user,
                            'status': ''
                        }
                    )
                    updated += 1

    if errors:
        return JsonResponse({'success': False, 'updated_count': updated, 'errors': errors})
    return JsonResponse({'success': True, 'updated_count': updated, 'errors': []})

@login_required
@require_POST
def generate_marksheet(request):
    """
    Accepts POST with params, returns JSON with download URL for marksheet.
    Now supports 'modules' param for Modular marksheet generation.
    """
    # Extract params
    month = request.POST.get('assessment_month')
    year = request.POST.get('assessment_year')
    regcat = request.POST.get('registration_category')
    occupation = request.POST.get('occupation')
    level = request.POST.get('level')
    center = request.POST.get('assessment_center')
    modules = request.POST.getlist('modules')  # Accept multiple module IDs

    # Determine structure_type for occupation/level
    structure_type = None
    if occupation and level:
        from .models import OccupationLevel
        occ_level = OccupationLevel.objects.filter(occupation_id=occupation, level_id=level).first()
        if occ_level:
            structure_type = occ_level.structure_type
    print(f'[DEBUG] structure_type for occupation {occupation}, level {level}:', structure_type)
    from django.urls import reverse
    import urllib.parse
    params = {
        'assessment_month': month,
        'assessment_year': year,
        'registration_category': regcat,
        'occupation': occupation,
        'level': level,
        'assessment_center': center,
    }
    # Add modules to params for Modular
    if regcat and regcat.lower() == 'modular' and modules:
        params['modules'] = ','.join(modules)
    # Remove empty params
    params = {k: v for k, v in params.items() if v}

    # Validate required params for occupation and level
    from .models import Occupation, Level, Module, Candidate
    occ = None
    lvl = None
    if occupation:
        try:
            occ = Occupation.objects.get(pk=occupation)
        except Occupation.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Selected occupation not found.'})
    if level:
        try:
            lvl = Level.objects.get(pk=level)
        except Level.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Selected level not found.'})

    # Check for enrolled candidates
    candidates = Candidate.objects.all()
    print('[DEBUG] Initial candidates:', candidates.count())
    print('[DEBUG] Initial candidate IDs and regcat:', list(candidates.values_list('id', 'registration_category')))
    if regcat and regcat.lower() == 'modular':
        candidates = candidates.filter(occupation__has_modular=True)
        print('[DEBUG] After has_modular filter:', candidates.count())
        print('[DEBUG] After has_modular IDs and regcat:', list(candidates.values_list('id', 'registration_category')))
        # Filter by selected modules if provided
        if modules:
            module_objs = Module.objects.filter(id__in=modules)
            from .models import CandidateModule
            enrolled_ids = set(CandidateModule.objects.filter(module__in=module_objs).values_list('candidate_id', flat=True))
            print('[DEBUG] Modular enrolled_ids:', enrolled_ids)
            candidates = candidates.filter(id__in=enrolled_ids)
            print('[DEBUG] After module enrollment filter:', candidates.count())
            print('[DEBUG] After module enrollment IDs and regcat:', list(candidates.values_list('id', 'registration_category')))
    elif regcat and regcat.lower() == 'formal' and structure_type == 'papers':
        candidates = candidates.filter(occupation__has_modular=False)
        print('[DEBUG] After NOT has_modular filter (paper-based):', candidates.count())
        print('[DEBUG] After NOT has_modular IDs and regcat (paper-based):', list(candidates.values_list('id', 'registration_category')))
    else:
        print('[DEBUG] Skipping occupation__has_modular filter for formal module-based marksheet.')
    # Normalize 'workers_pas' to 'Informal' for DB filtering
    regcat_db = regcat
    if regcat and regcat.lower() in ['workers_pas', "worker's pas"]:
        regcat_db = 'Informal'
    if regcat_db:
        candidates = candidates.filter(registration_category__iexact=regcat_db)
        print(f'[DEBUG] After regcat ({regcat_db}) filter:', candidates.count())
        print(f'[DEBUG] After regcat IDs and regcat:', list(candidates.values_list('id', 'registration_category')))
    print('[DEBUG] Candidate IDs and occupation IDs:', list(candidates.values_list('id', 'occupation_id')))
    # For formal (module-based), filter by CandidateLevel FIRST, do NOT apply occupation filter before this!
    if regcat and regcat.lower() == 'formal' and level and occupation:
        from .models import CandidateLevel
        candidate_levels_qs = CandidateLevel.objects.filter(level=level, candidate__occupation=occupation)
        print(f"[DEBUG] CandidateLevel entries for occupation {occupation}, level {level}: {[{'candidate_id': cl.candidate.id, 'candidate_occupation_id': cl.candidate.occupation_id} for cl in candidate_levels_qs]}")
        enrolled_ids = set(candidate_levels_qs.values_list('candidate_id', flat=True))
        print(f"[DEBUG] Enrolled candidate IDs for occupation {occupation}, level {level}: {enrolled_ids}")
        candidates = candidates.filter(id__in=enrolled_ids)
        print('[DEBUG] After CandidateLevel filter:', candidates.count())
    elif occupation:
        candidates = candidates.filter(occupation_id=occupation)
        print(f'[DEBUG] After occupation ({occupation}) filter:', candidates.count())
    if center:
        candidates = candidates.filter(assessment_center=center)
        print(f'[DEBUG] After center ({center}) filter:', candidates.count())
    if not occ:
        return JsonResponse({'success': False, 'error': 'Please select an occupation.'})
    if lvl is None and regcat and regcat.lower() != 'modular':
        return JsonResponse({'success': False, 'error': 'Please select a level.'})
    if regcat and regcat.lower() == 'modular' and (not modules or not Module.objects.filter(id__in=modules).exists()):
        return JsonResponse({'success': False, 'error': 'Please select at least one module for Modular marksheet generation.'})
    print("DEBUG: Candidate IDs after all filters:", list(candidates.values_list('id', flat=True)))
    print("DEBUG: Level param:", level)
    print("DEBUG: Registration category param:", regcat)
    print("DEBUG: Occupation param:", occupation)
    if not candidates.exists():
        return JsonResponse({'success': False, 'error': 'No enrolled candidates found for the selected parameters.'})

    url = reverse('download_marksheet') + '?' + urllib.parse.urlencode(params)
    return JsonResponse({'success': True, 'download_url': url})


@login_required
@require_POST
def print_marksheet(request):
    """
    Accepts POST with params, validates them, and returns a JSON response with a URL
    to download the printable PDF marksheet. This logic mirrors `generate_marksheet`.
    """
    from .models import Occupation, Level, Module, Candidate, OccupationLevel, CandidateLevel, CandidateModule
    from django.urls import reverse
    import urllib.parse

    # Extract params
    month = request.POST.get('assessment_month')
    year = request.POST.get('assessment_year')
    regcat = request.POST.get('registration_category')
    occupation_id = request.POST.get('occupation')
    level_id = request.POST.get('level')
    center_id = request.POST.get('assessment_center')
    modules = request.POST.getlist('modules')

    # Determine structure_type for validation
    structure_type = None
    if occupation_id and level_id:
        occ_level = OccupationLevel.objects.filter(occupation_id=occupation_id, level_id=level_id).first()
        if occ_level:
            structure_type = occ_level.structure_type

    # --- Validation --- #
    if not all([month, year, regcat, occupation_id]):
        return JsonResponse({'success': False, 'error': 'Missing required fields (Month, Year, Category, Occupation).'})

    if regcat.lower() != 'modular' and not level_id:
        return JsonResponse({'success': False, 'error': 'Level is required for this registration category.'})

    if regcat.lower() == 'modular' and not modules:
        return JsonResponse({'success': False, 'error': 'Please select at least one module for Modular registration.'})

    # Check if any candidates match the criteria
    candidates = Candidate.objects.all()
    regcat_filter = 'Informal' if regcat.lower() in ['workers_pas', "worker's pas"] else regcat
    candidates = candidates.filter(registration_category__iexact=regcat_filter)
    candidates = candidates.filter(occupation_id=occupation_id)

    if level_id:
        candidates = candidates.filter(candidatelevel__level_id=level_id)
    if center_id:
        candidates = candidates.filter(assessment_center_id=center_id)
    if regcat.lower() == 'modular' and modules:
        candidates = candidates.filter(candidatemodule__module_id__in=modules)

    if not candidates.exists():
        return JsonResponse({'success': False, 'error': 'No enrolled candidates found for the selected criteria.'})

    # --- URL Generation --- #
    params = {
        'assessment_month': month,
        'assessment_year': year,
        'registration_category': regcat,
        'occupation': occupation_id,
        'level': level_id,
        'assessment_center': center_id,
    }
    if regcat.lower() == 'modular' and modules:
        params['modules'] = ','.join(modules)

    params = {k: v for k, v in params.items() if v}
    url = reverse('download_printed_marksheet') + '?' + urllib.parse.urlencode(params)
    return JsonResponse({'success': True, 'download_url': url})


@login_required
def download_printed_marksheet(request):
    """
    Generates a printable PDF marksheet with candidates grouped by assessment center.
    The logic mirrors the filtering of `print_marksheet` to ensure consistency.
    """
    # 1. Imports and setup
    from io import BytesIO
    from .models import Candidate, Occupation, Level, Module, Paper, Result, AssessmentCenter, OccupationLevel
    from collections import defaultdict
    import calendar
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib import colors

    logger = logging.getLogger(__name__)

    # 2. Get and log params from request.GET
    month = request.GET.get('assessment_month')
    year = request.GET.get('assessment_year')
    regcat = request.GET.get('registration_category')
    occupation_id = request.GET.get('occupation')
    occupation = Occupation.objects.get(id=occupation_id)   
    level_id = request.GET.get('level')
    center_id = request.GET.get('assessment_center')
    module_ids_raw = request.GET.get('modules', '')
    module_ids = [mid for mid in module_ids_raw.split(',') if mid]

    logger.info(f"--- PDF Marksheet: regcat='{regcat}', occupation='{occupation_id}', level='{level_id}', modules='{module_ids}' ---")

    # 3. Filter candidates (mirroring the logic from print_marksheet)
    candidates = Candidate.objects.select_related('occupation', 'assessment_center').order_by('assessment_center__center_name', 'full_name')
    
    regcat_filter = 'Informal' if regcat and regcat.lower() in ['workers_pas', "worker's pas"] else regcat
    if regcat_filter:
        candidates = candidates.filter(registration_category__iexact=regcat_filter)
    if occupation_id:
        candidates = candidates.filter(occupation_id=occupation_id)
    if level_id:
        candidates = candidates.filter(candidatelevel__level_id=level_id)
    if center_id:
        candidates = candidates.filter(assessment_center_id=center_id)
    if regcat.lower() == 'modular' and module_ids:
        candidates = candidates.filter(candidatemodule__module_id__in=module_ids).distinct()

    logger.info(f"Found {candidates.count()} candidates after filtering.")

    # 4. Prepare PDF document
    response = HttpResponse(content_type='application/pdf')
    filename = f'marksheet_{regcat_filter}_{occupation.name}_{year}_{month}.pdf'
    pdf_title = f"Marksheet: {occupation.name} ({regcat.replace('_', ' ').title()}) - {month}/{year}"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=30, bottomMargin=30, title=pdf_title)
    elements = []
    styles = getSampleStyleSheet()

    # Main Title
    title_style = styles['h1']
    title_style.alignment = 1 # Center
    elements.append(Paragraph(f"UGANDA VOCATIONAL AND TECHNICAL ASSESSMENT BOARD", title_style))
    elements.append(Paragraph(f"CANDIDATES MARKSHEETS - {calendar.month_name[int(month)]} {year}", title_style))   
    elements.append(Spacer(1, 24))

    # 5. Handle no candidates
    if not candidates.exists():
        logger.warning("No candidates found, returning PDF with 'No data' message.")
        elements.append(Paragraph("No data found for the selected criteria.", styles['h2']))
        doc.build(elements)
        response.write(buffer.getvalue())
        buffer.close()
        return response

    # 6. Get related objects
    occupation = Occupation.objects.get(pk=occupation_id)
    level = Level.objects.get(pk=level_id) if level_id else None
    
    structure_type = None
    if occupation and level:
        occ_level = OccupationLevel.objects.filter(occupation=occupation, level=level).first()
        if occ_level:
            structure_type = occ_level.structure_type
    logger.info(f"Processing with structure_type: '{structure_type}'")

    # 7. Group candidates by center
    candidates_by_center = defaultdict(list)
    for c in candidates:
        center_key = c.assessment_center.id if c.assessment_center else 'unassigned'
        candidates_by_center[center_key].append(c)

    # 8. Build PDF content
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ])

    first_center = True
    for center_id_key, candidates_in_center in sorted(candidates_by_center.items()):
        if not first_center:
            elements.append(PageBreak())
        first_center = False
        
        center = AssessmentCenter.objects.get(pk=center_id_key) if center_id_key != 'unassigned' else None
        center_name = center.center_name if center else "Unassigned Center"
        center_number = center.center_number if center else "N/A"

        # --- Modular ---
        if regcat.lower() == 'modular':
            modules_to_print = Module.objects.filter(id__in=module_ids)
            for print_module in modules_to_print:
                elements.append(Paragraph(f"<b>Centre:</b> {center_number} - {center_name}", styles['h2']))
                elements.append(Paragraph(f"<b>Occupation:</b> {occupation.name}", styles['Normal']))
                elements.append(Paragraph(f"<b>Module:</b> {print_module.name}", styles['Normal']))
                elements.append(Spacer(1, 12))
                data = [["SN", "Regno", "Names", "Mark"]]
                dynamic_styles = []
                for i, c in enumerate(candidates_in_center, 1):
                    res = Result.objects.filter(candidate=c, module=print_module, assessment_date__year=year, assessment_date__month=month).first()
                    mark_val = f"{res.mark:.0f}" if res and res.mark is not None else ''
                    if res and res.mark == -1:
                        row_index = i
                        col_index = 3 # Mark column
                        dynamic_styles.append(('BACKGROUND', (col_index, row_index), (col_index, row_index), colors.lightpink))
                        dynamic_styles.append(('TEXTCOLOR', (col_index, row_index), (col_index, row_index), colors.red))
                    data.append([i, c.reg_number, c.full_name, mark_val])
                
                final_table_style = TableStyle(table_style.getCommands() + dynamic_styles)
                elements.append(Table(data, style=final_table_style))
                elements.append(Spacer(1, 24))

        # --- Paper-Based (Formal & Informal) ---
        elif structure_type == 'papers':
            papers = Paper.objects.filter(occupation=occupation, level=level)
            elements.append(Paragraph(f"<b>Centre:</b> {center_number} - {center_name}", styles['h2']))
            elements.append(Paragraph(f"<b>Occupation:</b> {occupation.name} - {level.name if level else ''}", styles['Normal']))
            elements.append(Spacer(1, 12))
            for paper in papers:
                elements.append(Paragraph(f"<b>Paper:</b> {paper.name} ({paper.grade_type.title()})", styles['Normal']))
                data = [["SN", "Regno", "Names", "Mark"]]
                dynamic_styles = []
                for i, c in enumerate(candidates_in_center, 1):
                    res = Result.objects.filter(candidate=c, paper=paper, assessment_date__year=year, assessment_date__month=month).first()
                    mark_val = f"{res.mark:.0f}" if res and res.mark is not None else ''
                    if res and res.mark == -1:
                        row_index = i
                        col_index = 3 # Mark column
                        dynamic_styles.append(('BACKGROUND', (col_index, row_index), (col_index, row_index), colors.lightpink))
                        dynamic_styles.append(('TEXTCOLOR', (col_index, row_index), (col_index, row_index), colors.red))
                    data.append([i, c.reg_number, c.full_name, mark_val])

                final_table_style = TableStyle(table_style.getCommands() + dynamic_styles)
                elements.append(Table(data, style=final_table_style))
                elements.append(Spacer(1, 12))

        # --- Module-Based (Theory/Practical) ---
        elif structure_type == 'modules':
            elements.append(Paragraph(f"<b>Centre:</b> {center_number} - {center_name}", styles['h2']))
            elements.append(Paragraph(f"<b>Occupation:</b> {occupation.name} - {level.name if level else ''}", styles['Normal']))
            elements.append(Spacer(1, 12))
            data = [["SN", "Regno", "Names", "Theory", "Practical"]]
            dynamic_styles = []
            for i, c in enumerate(candidates_in_center, 1):
                theory = Result.objects.filter(candidate=c, level=level, assessment_type='theory', assessment_date__year=year, assessment_date__month=month).first()
                practical = Result.objects.filter(candidate=c, level=level, assessment_type='practical', assessment_date__year=year, assessment_date__month=month).first()
                
                theory_val = f"{theory.mark:.0f}" if theory and theory.mark is not None else ''
                practical_val = f"{practical.mark:.0f}" if practical and practical.mark is not None else ''
                
                row_index = i
                if theory and theory.mark == -1:
                    dynamic_styles.append(('BACKGROUND', (3, row_index), (3, row_index), colors.lightpink))
                    dynamic_styles.append(('TEXTCOLOR', (3, row_index), (3, row_index), colors.red))
                
                if practical and practical.mark == -1:
                    dynamic_styles.append(('BACKGROUND', (4, row_index), (4, row_index), colors.lightpink))
                    dynamic_styles.append(('TEXTCOLOR', (4, row_index), (4, row_index), colors.red))

                data.append([i, c.reg_number, c.full_name, theory_val, practical_val])

            final_table_style = TableStyle(table_style.getCommands() + dynamic_styles)
            elements.append(Table(data, style=final_table_style))
            elements.append(Spacer(1, 24))
        
        else:
             logger.error(f"Unknown structure type '{structure_type}' or category '{regcat}' for PDF generation.")
             elements.append(Paragraph(f"Error: Could not determine marksheet structure for {occupation.name}.", styles['Normal']))

    # 9. Build and return PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response

@login_required
def download_marksheet(request):
    """
    Generates and streams the Excel marksheet in memory based on GET params.
    Supports all registration categories and structures (modular, formal module-based, formal paper-based, informal/worker's PAS).
    """
    from .models import Candidate, OccupationLevel, Module, Paper, Result, AssessmentCenter, Occupation, Level, CandidateModule, CandidateLevel, CandidatePaper
    from openpyxl import Workbook
    from io import BytesIO

    # Extract params
    month = request.GET.get('assessment_month')
    year = request.GET.get('assessment_year')
    regcat = request.GET.get('registration_category')
    # Normalize workers_pas and worker's pas to Informal for both filtering and logic
    if regcat and regcat.strip().lower() in ["workers_pas", "worker's pas"]:
        regcat = "Informal"
    occupation_id = request.GET.get('occupation')
    level_id = request.GET.get('level')
    center_id = request.GET.get('assessment_center')
    modules_param = request.GET.get('modules')  # comma-separated module IDs

    # Get occupation, level, center
    occupation = Occupation.objects.filter(pk=occupation_id).first() if occupation_id else None
    level = Level.objects.filter(pk=level_id).first() if level_id else None
    center = AssessmentCenter.objects.filter(pk=center_id).first() if center_id else None

    regcat_normalized = regcat.strip().lower() if regcat else ''
    modules = []
    papers = []
    enrolled_ids = set()
    enrolled_paper_ids = set()
    structure_type = None
    informal = False

    # --- Fetch modules/papers FIRST ---
    occ_level = None
    if occupation and level:
        occ_level = OccupationLevel.objects.filter(occupation=occupation, level=level).first()
        if occ_level:
            structure_type = occ_level.structure_type  # 'modules' or 'papers'

    # Modular: Only PR per selected module(s)
    if regcat_normalized == 'modular':
        # Use only selected modules if provided (filter by occupation only, not level)
        if modules_param:
            module_ids = [int(mid) for mid in modules_param.split(',') if mid.isdigit()]
            modules = list(Module.objects.filter(id__in=module_ids, occupation=occupation))
        else:
            modules = list(Module.objects.filter(occupation=occupation))
        module_ids = [m.id for m in modules]
        # Candidates enrolled in at least one of these modules
        enrolled_modules = CandidateModule.objects.filter(module__in=modules)
        enrolled_ids = set(enrolled_modules.values_list('candidate_id', flat=True))
        print(f"[DEBUG] Modular: modules={module_ids}, enrolled_ids={enrolled_ids}")

    # --- Candidate Query ---
    candidates = Candidate.objects.all()
    if regcat:
        candidates = candidates.filter(registration_category__iexact=regcat)
    # For formal (module-based), filter by CandidateLevel FIRST, do not apply occupation filter before this!
    if regcat_normalized == 'formal' and structure_type == 'modules' and level and occupation:
        from .models import CandidateLevel
        enrolled_ids = set(
            CandidateLevel.objects.filter(level=level, candidate__occupation=occupation)
            .values_list('candidate_id', flat=True)
        )
        candidates = candidates.filter(id__in=enrolled_ids)
    elif occupation_id:
        candidates = candidates.filter(occupation_id=occupation_id)
    elif regcat_normalized != 'modular' and hasattr(Candidate, 'level') and level:
        candidates = candidates.filter(candidatelevel__level_id=level_id)
    if center:
        candidates = candidates.filter(assessment_center=center)

    # --- Enrollment Filtering ---
    enrolled_ids = set()
    informal = False
    regcat_normalized = regcat.strip().lower() if regcat else ''
    # Modular: Only PR per module
    if regcat_normalized == 'modular':
        enrolled_ids = set(CandidateModule.objects.filter(module__in=modules).values_list('candidate_id', flat=True))
        candidates = candidates.filter(id__in=enrolled_ids)
        structure_type = 'modules'
    # Formal: Module-based (theory/practical per module)
    elif regcat_normalized == 'formal' and structure_type == 'modules':
        # For formal (level module-based): candidates enroll for a level, not modules
        candidate_levels_qs = CandidateLevel.objects.filter(level=level, candidate__occupation=occupation)
        print(f"[DEBUG] CandidateLevel entries for occupation {occupation.id if occupation else None}, level {level.id if level else None}: {[{'candidate_id': cl.candidate.id, 'candidate_occupation_id': cl.candidate.occupation_id} for cl in candidate_levels_qs]}")
        enrolled_level_ids = set(candidate_levels_qs.values_list('candidate_id', flat=True))
        print(f"[DEBUG] Enrolled candidate IDs for occupation {occupation.id if occupation else None}, level {level.id if level else None}: {enrolled_level_ids}")
        candidates = candidates.filter(id__in=enrolled_level_ids)
        modules = list(Module.objects.filter(occupation=occupation, level=level))
    # Formal: Paper-based (dynamic paper codes per occupation/level)
    elif regcat_normalized == 'formal' and structure_type == 'papers':
        enrolled_ids = set(
            CandidateLevel.objects.filter(level=level, candidate__occupation=occupation)
            .values_list('candidate_id', flat=True)
        )
        print(f"[DEBUG] Paper-based: enrolled_ids for occupation {occupation.id if occupation else None}, level {level.id if level else None}: {enrolled_ids}")
        candidates = candidates.filter(id__in=enrolled_ids)
        # Fetch all papers for this occupation/level
        papers = list(Paper.objects.filter(occupation=occupation, level=level))
        print(f"[DEBUG] Paper-based: papers for occupation {occupation.id if occupation else None}, level {level.id if level else None}: {[p.code for p in papers]}")
    # Informal/Worker's PAS: All papers for occupation/level (flatten, ignore module grouping)
    elif regcat_normalized in ['informal', "worker's pas"]:
        informal = True
        # Fetch modules for this occupation and level
        modules = list(Module.objects.filter(occupation_id=occupation.id, level_id=level.id))
        # Fetch papers both ways: directly by occupation/level, and by modules
        papers_direct = list(Paper.objects.filter(occupation_id=occupation.id, level_id=level.id))
        papers_via_modules = list(Paper.objects.filter(module__in=modules))
        # Combine and deduplicate by paper ID
        papers_dict = {p.id: p for p in papers_direct + papers_via_modules}
        papers = list(papers_dict.values())
        print(f"[DEBUG] Informal: papers_direct for occupation {occupation.id}, level {level.id}: {[{'id': p.id, 'code': p.code} for p in papers_direct]}")
        print(f"[DEBUG] Informal: papers_via_modules for modules {[m.id for m in modules]}: {[{'id': p.id, 'code': p.code} for p in papers_via_modules]}")
        print(f"[DEBUG] Informal: final papers for occupation {occupation.id}, level {level.id}: {[{'id': p.id, 'code': p.code} for p in papers]}")
        # All candidates enrolled for this level in this occupation
        enrolled_ids = set(CandidateLevel.objects.filter(level_id=level.id, candidate__occupation_id=occupation.id).values_list('candidate_id', flat=True))
        print(f"[DEBUG] Informal: enrolled_ids for occupation {occupation.id if occupation else None}, level {level.id if level else None}: {enrolled_ids}")
        candidates = candidates.filter(id__in=enrolled_ids)
    else:
        # fallback: no structure or unknown regcat
        modules = []
        papers = []

    print(f"Enrolled candidate IDs: {enrolled_ids}")
    print(f"Modules: {modules}")
    print(f"Papers: {papers}")

    if regcat_normalized == 'modular' and modules:
        # --- Build Modular Marksheet ---
        wb = Workbook()
        ws = wb.active
        ws.title = 'Marksheet'
        # Header
        ws.append(['SN', 'REGISTRATION NO.', 'FULL NAME', 'OCCUPATION CODE', 'CATEGORY', 'MODULE CODE', 'PRACTICAL'])
        sn = 1
        for module in modules:
            # Candidates enrolled in this module
            enrolled_cands = candidates.filter(id__in=CandidateModule.objects.filter(module=module).values_list('candidate_id', flat=True))
            for cand in enrolled_cands:
                ws.append([
                    sn,
                    cand.reg_number,
                    cand.full_name,
                    cand.occupation.code if cand.occupation else '',
                    'Modular',
                    module.code,
                    ''  # PR mark cell (to be filled)
                ])
                sn += 1
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        #filename = f"marksheet_{regcat_normalized or 'all'}_{month or ''}_{year or ''}.xlsx"
        occupation_name = occupation.name if occupation else 'Occupation'
        regcat_title = regcat_normalized.title() if regcat_normalized else 'All'
        month_str = calendar.month_name[int(month)] if month and month.isdigit() and int(month) in range(1, 13) else (month if month else '')
        year_str = year if year else ''
        filename = f"{occupation_name} Marksheet {month_str} {year_str} {regcat_title}.xlsx"
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # --- Formal (module-based): SN, REGISTRATION NO., FULL NAME, OCCUPATION CODE, CATEGORY, LEVEL, THEORY, PRACTICAL ---
    if regcat_normalized == 'formal' and structure_type == 'modules':
        wb = Workbook()
        ws = wb.active
        ws.title = 'Marksheet'
        ws.append(['SN', 'REGISTRATION NO.', 'FULL NAME', 'OCCUPATION CODE', 'CATEGORY', 'LEVEL', 'THEORY', 'PRACTICAL'])
        sn = 1
        for cand in candidates:
            ws.append([
                sn,
                cand.reg_number,
                cand.full_name,
                cand.occupation.code if cand.occupation else '',
                'Formal',
                level.name if level else '',
                '',  # Theory mark
                '',  # Practical mark
            ])
            sn += 1
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        #filename = f"marksheet_{regcat_normalized or 'all'}_{month or ''}_{year or ''}.xlsx"
        occupation_name = occupation.name if occupation else 'Occupation'
        regcat_title = regcat_normalized.title() if regcat_normalized else 'All'
        month_str = calendar.month_name[int(month)] if month and month.isdigit() and int(month) in range(1, 13) else (month if month else '')
        year_str = year if year else ''
        filename = f"{occupation_name} Marksheet {month_str} {year_str} {regcat_title}.xlsx"    
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    if not modules and not papers:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Marksheet'
        ws.append(['No modules or papers found for the selected occupation and level.'])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        #filename = f"marksheet_{regcat_normalized or 'all'}_{month or ''}_{year or ''}.xlsx"
        occupation_name = occupation.name if occupation else 'Occupation'
        regcat_title = regcat_normalized.title() if regcat_normalized else 'All'
        month_str = calendar.month_name[int(month)] if month and month.isdigit() and int(month) in range(1, 13) else (month if month else '')
        year_str = year if year else ''
        filename = f"{occupation_name} Marksheet {month_str} {year_str} {regcat_title}.xlsx"
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    if not candidates.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = 'Marksheet'
        ws.append(['No enrolled candidates found for the selected parameters.'])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"marksheet_{regcat_normalized or 'all'}_{month or ''}_{year or ''}.xlsx"
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    # --- Build Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = 'Marksheet'

    # --- Build Headers ---
    # For paper-based marksheets, use a minimal set of columns and always include LEVEL
    if regcat_normalized == 'formal' and structure_type == 'papers':
        base_headers = ['SN', 'REGISTRATION NO.', 'FULL NAME', 'OCCUPATION CODE', 'LEVEL', 'CATEGORY']
    else:
        # For informal/worker's PAS, match the Excel sample: include LEVEL after CATEGORY
        if informal:
            base_headers = ['SN', 'REGISTRATION NO.', 'FULL NAME', 'OCCUPATION CODE', 'CATEGORY', 'LEVEL']
        else:
            base_headers = ['REGISTRATION NO.', 'FULL NAME', 'OCCUPATION CODE', 'LEVEL', 'CATEGORY', 'ASSESSMENT CENTER', 'MONTH', 'YEAR']
    dynamic_headers = []
    if regcat_normalized == 'modular':
        for module in modules:
            dynamic_headers.append(f"{module.code} (PR)")
    elif regcat_normalized == 'formal' and structure_type == 'modules':
        for module in modules:
            dynamic_headers.append(f"{module.code} (TH)")
            dynamic_headers.append(f"{module.code} (PR)")
    elif (regcat_normalized == 'formal' and structure_type == 'papers') or informal:
        for paper in papers:
            dynamic_headers.append(f"{paper.code}")
    else:
        dynamic_headers.append('MARK')
    ws.append(base_headers + dynamic_headers)

    # --- Populate Rows ---
    sn = 1
    for candidate in candidates:
        # For paper-based marksheets, always include level name and only required columns
        if regcat_normalized == 'formal' and structure_type == 'papers':
            row = [
                sn,
                candidate.reg_number,
                candidate.full_name,
                getattr(candidate.occupation, 'code', '') if candidate.occupation else '',
                level.name if level else '',
                getattr(candidate, 'registration_category', ''),
            ]
        elif informal:
            row = [
                sn,
                candidate.reg_number,
                candidate.full_name,
                getattr(candidate.occupation, 'code', '') if candidate.occupation else '',
                getattr(candidate, 'registration_category', ''),
                level.name if level else '',
            ]
        else:
            row = [
                candidate.reg_number,
                candidate.full_name,
                getattr(candidate.occupation, 'code', '') if candidate.occupation else '',
                str(candidate.level) if hasattr(candidate, 'level') and candidate.level else '',
                getattr(candidate, 'registration_category', ''),
                str(candidate.assessment_center) if candidate.assessment_center else '',
                month or '',
                year or ''
            ]
        marks = []
        if regcat_normalized == 'modular':
            for module in modules:
                marks.append('')
        elif regcat_normalized == 'formal' and structure_type == 'modules':
            for module in modules:
                marks.append('')  # Theory
                marks.append('')  # Practical
        elif (regcat_normalized == 'formal' and structure_type == 'papers') or informal:
            for paper in papers:
                marks.append('')
        else:
            marks.append('')

        # Append the row and get the row number for formatting
        ws.append(row + marks)
        current_row = ws.max_row

        # --- Highlight enrolled papers for informal/worker's PAS ---
        if informal:
            from openpyxl.styles import PatternFill
            # Get enrolled paper IDs for this candidate (for this level)
            enrolled_paper_ids = set()
            for cp in CandidatePaper.objects.filter(candidate=candidate, level=level):
                enrolled_paper_ids.add(cp.paper_id)
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            # The paper columns start after the base headers
            paper_col_offset = len(row)
            for idx, paper in enumerate(papers):
                if paper.id in enrolled_paper_ids:
                    col_letter = ws.cell(row=current_row, column=paper_col_offset + idx + 1).column_letter
                    ws[f"{col_letter}{current_row}"].fill = yellow_fill

        if regcat_normalized == 'formal' and structure_type == 'papers':
            sn += 1
        elif informal:
            sn += 1

    # --- Stream to memory ---
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    #filename = f"marksheet_{regcat_normalized or 'all'}_{month or ''}_{year or ''}.xlsx"
    occupation_name = occupation.name if occupation else 'Occupation'
    regcat_title = regcat_normalized.title() if regcat_normalized else 'All'
    month_str = calendar.month_name[int(month)] if month and month.isdigit() and int(month) in range(1, 13) else (month if month else '')
    year_str = year if year else ''
    filename = f"{occupation_name} Marksheet {month_str} {year_str} {regcat_title}.xlsx"    
    
    response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse

import copy
import openpyxl
from openpyxl import Workbook
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.contrib import messages
from django.utils.html import format_html
from django.urls import reverse
from django.core.files.storage import FileSystemStorage
from .models import Candidate, Occupation, AssessmentCenter, District, Village
from .forms import CandidateForm
from datetime import datetime
import tempfile
import os
import zipfile
from django.core.files import File
def candidate_import_dual(request):
    """
    Handle GET (show dual import page) and POST (process Excel + photo zip upload).
    Excel is required, photos are optional and can be added later.
    """
    if request.method == 'GET':
        return render(request, 'candidates/import_dual.html')

    excel_file = request.FILES.get('excel_file')
    errors = []
    created = 0
    
    if not excel_file:
        errors.append('Excel file is required.')
        return render(request, 'candidates/import_dual.html', {'errors': errors, 'imported_count': 0})
    
    # Load Excel
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
    except Exception:
        errors.append('Invalid Excel file.')
        return render(request, 'candidates/import_dual.html', {'errors': errors, 'imported_count': 0})
    
    headers = [str(cell.value).replace(u'\xa0', ' ').strip().lower() for cell in ws[1] if cell.value]

    # Process photos if ZIP file is provided (optional)
    image_name_map = {}
    tmp_dir = None
    photo_zip = request.FILES.get('photo_zip')
    if photo_zip:
        tmp_dir = tempfile.mkdtemp()
        try:
            with zipfile.ZipFile(photo_zip) as zf:
                image_files = [n for n in zf.namelist() if n.lower().endswith(('.jpg', '.jpeg', '.png'))]
                for image_name in zf.namelist():
                    # Skip directories and non-image files
                    if image_name.endswith('/') or not image_name.lower().endswith(('.jpg', '.jpeg', '.png')):
                        continue
                        
                    # Extract filename without extension
                    image_name_only = os.path.splitext(os.path.basename(image_name))[0]
                    
                    # Smart photo name parsing to handle multiple formats:
                    # Format 1: '540255.Afoyo_Vani' -> extract 'Afoyo Vani' (old system with numeric prefix)
                    # Format 2: '614956.Kajumba Ruth' -> extract 'Kajumba Ruth' (numbered format)
                    # Format 3: 'Afoyo_Vani' -> extract 'Afoyo Vani' (simple format)
                    
                    cleaned_name = None
                    
                    if '.' in image_name_only:
                        # Split by first dot to separate potential prefix from name
                        prefix, name_part = image_name_only.split('.', 1)
                        
                        # Check if prefix is purely numeric (old system format)
                        if prefix.isdigit():
                            # Old system format: '540255.Afoyo_Vani' -> 'Afoyo Vani'
                            cleaned_name = name_part.replace('_', ' ').strip()
                            print(f"[PHOTO DEBUG] Old system format detected: '{image_name_only}' -> '{cleaned_name}'")
                        else:
                            # Mixed format: '614956.Kajumba Ruth' -> 'Kajumba Ruth'
                            cleaned_name = name_part.replace('_', ' ').strip()
                            print(f"[PHOTO DEBUG] Mixed format detected: '{image_name_only}' -> '{cleaned_name}'")
                    else:
                        # Simple format: 'Afoyo_Vani' -> 'Afoyo Vani'
                        cleaned_name = image_name_only.replace('_', ' ').strip()
                        print(f"[PHOTO DEBUG] Simple format detected: '{image_name_only}' -> '{cleaned_name}'")
                    
                    if cleaned_name:
                        image_name_map[cleaned_name.lower()] = image_name
                
                # Debug: Show all parsed photo names
                print(f"[PHOTO PARSING DEBUG] Total photos parsed: {len(image_name_map)}")
                for parsed_name, original_file in image_name_map.items():
                    print(f"[PHOTO PARSING DEBUG] '{parsed_name}' -> '{original_file}'")
                
                zf.extractall(tmp_dir)
        except Exception as e:
            errors.append(f'Invalid ZIP file or unable to extract images: {e}')
            if tmp_dir:
                import shutil
                shutil.rmtree(tmp_dir, ignore_errors=True)
            return render(request, 'candidates/import_dual.html', {'errors': errors, 'imported_count': 0})
    
    # Read Excel rows (always process Excel, regardless of photos)
    rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if not all(cell is None for cell in row)]
    
    # If photos are provided, analyze photo-candidate matching first
    photo_analysis = []
    unmatched_candidates = []
    unmatched_photos = []
    
    if photo_zip and image_name_map:
        candidate_names = []
        for row in rows:
            data = dict(zip(headers, row))
            candidate_name_full = (data.get('full_name') or '').strip()
            if candidate_name_full:
                candidate_names.append(candidate_name_full)
        
        # Analyze matching using multi-strategy approach
        for candidate_name in candidate_names:
            name_parts = candidate_name.split()
            matched_photo = None
            
            # Strategy 1: Try full name first
            if candidate_name.lower() in image_name_map:
                matched_photo = image_name_map[candidate_name.lower()]
            
            # Strategy 2: If no match and 3+ parts, try first two names
            elif len(name_parts) >= 3:
                candidate_name_for_match = " ".join(name_parts[:2])
                if candidate_name_for_match.lower() in image_name_map:
                    matched_photo = image_name_map[candidate_name_for_match.lower()]
            
            # Strategy 3: If still no match and 3+ parts, try first and last names
            if not matched_photo and len(name_parts) >= 3:
                candidate_name_for_match = f"{name_parts[0]} {name_parts[-1]}"
                if candidate_name_for_match.lower() in image_name_map:
                    matched_photo = image_name_map[candidate_name_for_match.lower()]
            
            if matched_photo:
                photo_analysis.append(f"✅ '{candidate_name}' → matches photo '{matched_photo}'")
            else:
                unmatched_candidates.append(candidate_name)
        
        # Find photos that don't match any candidate
        matched_photos = set(image_name_map.values())
        all_photos = set()
        try:
            with zipfile.ZipFile(photo_zip) as zf:
                all_photos = {n for n in zf.namelist() if n.lower().endswith(('.jpg', '.jpeg', '.png')) and not n.endswith('/')}
        except:
            pass
        unmatched_photos = list(all_photos - matched_photos)
        
        # If there are significant mismatches, provide detailed analysis
        if unmatched_candidates or unmatched_photos:
            errors.append("📋 PHOTO MATCHING ANALYSIS:")
            
            if photo_analysis:
                errors.append("\n✅ SUCCESSFUL MATCHES:")
                for match in photo_analysis:
                    errors.append(match)
            
            if unmatched_candidates:
                errors.append("\n❌ CANDIDATES WITHOUT MATCHING PHOTOS:")
                for candidate in unmatched_candidates:
                    errors.append(f"   • '{candidate}' - no matching photo found")
                errors.append("\n💡 TROUBLESHOOTING TIPS:")
                errors.append("   • Check photo filenames match candidate names exactly")
                errors.append("   • Use underscores or spaces: 'John_Doe.jpg' or 'John Doe.jpg'")
                errors.append("   • For 3-part names, photos should match first 2 names")
                errors.append("   • Ensure photos are in JPG or PNG format")
            
            if unmatched_photos:
                errors.append("\n📷 PHOTOS WITHOUT MATCHING CANDIDATES:")
                for photo in unmatched_photos[:10]:  # Show max 10
                    errors.append(f"   • '{photo}' - no matching candidate found")
                if len(unmatched_photos) > 10:
                    errors.append(f"   ... and {len(unmatched_photos) - 10} more")
            
            # If more than 50% of candidates don't have photos, suggest fixing
            mismatch_rate = len(unmatched_candidates) / len(candidate_names) if candidate_names else 0
            if mismatch_rate > 0.5:
                errors.append("\n⚠️  HIGH MISMATCH RATE: Consider fixing photo names before importing.")
                return render(request, 'candidates/import_dual.html', {'errors': errors, 'imported_count': 0})
    
    for idx, row in enumerate(rows, start=2):
        data = dict(zip(headers, row))
        candidate_name_full = (data.get('full_name') or '').strip()

        # Try multiple name matching strategies
        name_parts = candidate_name_full.split()
        img_name = None
        
        # Strategy 1: Try full name first
        img_name = image_name_map.get(candidate_name_full.lower())
        
        # Strategy 2: If no match and 3+ parts, try first two names
        if not img_name and len(name_parts) >= 3:
            candidate_name_for_match = " ".join(name_parts[:2])
            img_name = image_name_map.get(candidate_name_for_match.lower())
        
        # Strategy 3: If still no match and 3+ parts, try first and last names
        if not img_name and len(name_parts) >= 3:
            candidate_name_for_match = f"{name_parts[0]} {name_parts[-1]}"
            img_name = image_name_map.get(candidate_name_for_match.lower())
        
        # For debugging, show what we tried
        candidate_name_for_match = candidate_name_full if img_name else "(no match found)"
        
        # Debug photo matching
        print(f"[PHOTO MATCH DEBUG] Row {idx}: Candidate '{candidate_name_full}' -> Looking for '{candidate_name_for_match}' -> Found: {img_name}")
        if photo_zip and not img_name:
            print(f"[PHOTO MATCH DEBUG] Available photos: {list(image_name_map.keys())}")

        # If photos were provided but this candidate has no match, log it
        if not img_name and photo_zip:
            errors.append(f"Row {idx}: No photo found for '{candidate_name_full}' (searched for '{candidate_name_for_match}')")
        
        # --- (reuse import logic from candidate_import) ---
        form_data = data.copy()
        # Normalize registration_category for modular candidates
        regcat = str(form_data.get('registration_category', '')).strip().capitalize()
        if regcat == 'Modular':
            form_data['registration_category'] = 'Modular'
        
        # Robust date parsing: handle string, datetime, and Excel serial (float/int)
        from openpyxl.utils.datetime import from_excel
        import datetime as dt
        for date_field in ['date_of_birth', 'start_date', 'finish_date', 'assessment_date']:
            val = form_data.get(date_field)
            if not val:
                continue
            if isinstance(val, (dt.date, dt.datetime)):
                form_data[date_field] = val.date() if isinstance(val, dt.datetime) else val
            elif isinstance(val, (float, int)):
                try:
                    form_data[date_field] = from_excel(val).date()
                except Exception:
                    errors.append(f"Row {idx}: Invalid Excel serial date in '{date_field}'.")
                    continue
            elif isinstance(val, str):
                for fmt in ('%d/%m/%Y', '%-d/%-m/%Y', '%Y-%m-%d'):
                    try:
                        form_data[date_field] = dt.datetime.strptime(val, fmt).date()
                        break
                    except Exception:
                        continue
                else:
                    errors.append(f"Row {idx}: Invalid date format in '{date_field}'. Use D/M/YYYY, DD/MM/YYYY, or YYYY-MM-DD.")
                    continue
        
        # DEBUG: Print nationality value before validation
        print(f"[DEBUG] Row {idx} nationality value: '{form_data.get('nationality', '')}'")
        from django_countries import countries
        import re
        def normalize_country(val):
            val = val.lower().replace('&', 'and')
            val = re.sub(r'[^a-z0-9 ]', '', val)
            val = re.sub(r'\s+', ' ', val).strip()
            return val
        
        nat_val = form_data.get('nationality', '')
        if nat_val:
            nat_val_str = str(nat_val).strip()
            if nat_val_str:
                # Try exact match first
                country_found = None
                for code, name in countries:
                    if name.lower() == nat_val_str.lower():
                        country_found = code
                        break
                
                # If no exact match, try normalized matching
                if not country_found:
                    nat_normalized = normalize_country(nat_val_str)
                    for code, name in countries:
                        name_normalized = normalize_country(name)
                        if name_normalized == nat_normalized:
                            country_found = code
                            break
                
                if country_found:
                    form_data['nationality'] = country_found
                    print(f"[DEBUG] Row {idx} nationality mapped: '{nat_val_str}' -> '{country_found}'")
                else:
                    errors.append(f"Row {idx}: Invalid nationality '{nat_val_str}'. Must be a valid country name.")
                    continue
        
        # Handle occupation lookup (try both code and name)
        occ_val = form_data.get('occupation')
        if occ_val:
            occ_str = str(occ_val).strip()
            # Try to find by code first, then by name
            occupation = Occupation.objects.filter(code__iexact=occ_str).first()
            if not occupation:
                occupation = Occupation.objects.filter(name__iexact=occ_str).first()
            
            if occupation:
                form_data['occupation'] = occupation.id
            else:
                errors.append(f"Row {idx}: Occupation '{occ_str}' not found.")
                continue
        
        # Enhanced assessment center/branch lookup (by code/number)
        center_val = form_data.get('assessment_center')
        if center_val:
            center_str = str(center_val).strip()
            
            # First try to find a branch by branch_code
            branch_found = None
            center_found = None
            
            try:
                # Try to find branch by branch_code first
                from .models import AssessmentCenterBranch
                branch_found = AssessmentCenterBranch.objects.select_related('assessment_center').filter(branch_code__iexact=center_str).first()
                if branch_found:
                    center_found = branch_found.assessment_center
                    form_data['assessment_center'] = center_found.id
                    form_data['assessment_center_branch'] = branch_found.id
                else:
                    # If no branch found, try to find assessment center by center_number or center_name
                    center_found = AssessmentCenter.objects.filter(center_number__iexact=center_str).first()
                    if not center_found:
                        center_found = AssessmentCenter.objects.filter(center_name__iexact=center_str).first()
                    
                    if center_found:
                        form_data['assessment_center'] = center_found.id
                        form_data['assessment_center_branch'] = None  # No branch specified
                    else:
                        errors.append(f"Row {idx}: Assessment Center or Branch '{center_str}' not found. Please use a valid center number, center name, or branch code.")
                        continue
            except Exception as e:
                errors.append(f"Row {idx}: Error looking up assessment center/branch '{center_str}': {str(e)}")
                continue
        
        # Use CandidateForm for validation, but patch required fields for import
        form = CandidateForm(form_data)
        for f in ['district', 'village']:
            if f in form.fields:
                form.fields[f].required = False
        
        if not form.is_valid():
            print(f"[DEBUG] Row {idx} SKIPPED: CandidateForm errors: {form.errors}")
            errors.append(f"Row {idx}: Form errors: {form.errors}")
            continue
        
        # District and village: truly optional for import (skip if missing or blank)
        for loc_field, model_cls in [('district', District), ('village', Village)]:
            val = form_data.get(loc_field)
            if val is None or str(val).strip() == '':
                form_data[loc_field] = None
            else:
                val_str = str(val).strip()
                obj = model_cls.objects.filter(name__iexact=val_str).first()
                form_data[loc_field] = obj.id if obj else None
        
        # Debug: print after district/village assignment
        print(f"[DEBUG] Row {idx} after district/village assignment: district={form_data.get('district')}, village={form_data.get('village')}")

        # Define date and foreign key fields for use below
        date_fields = ['date_of_birth', 'start_date', 'finish_date', 'assessment_date']
        fk_fields = ['occupation', 'assessment_center', 'assessment_center_branch', 'district', 'village']
        
        # Coerce every other field except dates and foreign keys to string
        for k in form_data:
            if k not in date_fields + fk_fields:
                v = form_data[k]
                if v is not None and not isinstance(v, str):
                    form_data[k] = str(v)
        
        # Debug: print types of all fields before form creation
        print(f"[IMPORT DEBUG] Row {idx} form_data types: " + ", ".join(f"{k}: {type(v).__name__}" for k,v in form_data.items()))
        
        # Remove reg_number if present
        form_data.pop('reg_number', None)
        
        # Use CandidateForm for validation, but patch required fields for import
        form = CandidateForm(form_data)
        for f in ['district', 'village']:
            if f in form.fields:
                form.fields[f].required = False
        
        if not form.is_valid():
            error_list = '; '.join([f"{k}: {v[0]}" for k, v in form.errors.items()])
            errors.append(f"Row {idx}: {error_list}")
            continue
        
        # Convert date fields in form.cleaned_data from DD/MM/YYYY string to date objects
        for date_field in date_fields:
            val = form.cleaned_data.get(date_field)
            if val and isinstance(val, str):
                for fmt in ('%d/%m/%Y', '%-d/%-m/%Y', '%Y-%m-%d'):
                    try:
                        form.cleaned_data[date_field] = dt.datetime.strptime(val, fmt).date()
                        break
                    except ValueError:
                        continue
        
        # Create candidate (handle many-to-many fields separately)
        cleaned_data = form.cleaned_data.copy()
        
        # Extract many-to-many fields that need to be set after creation
        m2m_fields = {}
        if 'nature_of_disability' in cleaned_data:
            m2m_fields['nature_of_disability'] = cleaned_data.pop('nature_of_disability')
        
        # Create candidate with non-m2m fields
        candidate = Candidate(**cleaned_data)
        
        # Attach photo if available
        if img_name and tmp_dir:
            img_path = os.path.join(tmp_dir, img_name)
            if os.path.exists(img_path):
                from PIL import Image, ExifTags
                import io
                with Image.open(img_path) as img:
                    # Handle EXIF orientation to prevent rotation issues
                    try:
                        for orientation in ExifTags.TAGS.keys():
                            if ExifTags.TAGS[orientation] == 'Orientation':
                                break
                        exif = img._getexif()
                        if exif is not None:
                            orientation_value = exif.get(orientation)
                            if orientation_value == 3:
                                img = img.rotate(180, expand=True)
                            elif orientation_value == 6:
                                img = img.rotate(270, expand=True)
                            elif orientation_value == 8:
                                img = img.rotate(90, expand=True)
                    except (AttributeError, KeyError, TypeError):
                        # No EXIF data or orientation info, continue without rotation
                        pass
                    
                    if img.mode != 'RGB':
                        img = img.convert('RGB')
                    buffer = io.BytesIO()
                    img.save(buffer, format='JPEG', quality=85)
                    buffer.seek(0)
                    # Use only the base filename, not the full path from ZIP
                    base_filename = os.path.basename(img_name)
                    clean_filename = os.path.splitext(base_filename)[0] + '.jpg'
                    candidate.passport_photo.save(
                        clean_filename,
                        File(buffer),
                        save=False
                    )
                    print(f"[PHOTO SAVE DEBUG] Saved photo as: {clean_filename}")
            else:
                errors.append(f"Row {idx}: Image file '{img_name}' not found after extraction - candidate created without photo.")
        
        candidate.save()
        
        # Set many-to-many fields after candidate is saved
        for field_name, field_value in m2m_fields.items():
            if field_value:
                getattr(candidate, field_name).set(field_value)
        
        created += 1
        print(f"[DEBUG] Row {idx} IMPORTED: Candidate '{candidate.full_name}' saved.")
    
    # Cleanup temporary directory if photos were processed
    if tmp_dir:
        import shutil
        shutil.rmtree(tmp_dir, ignore_errors=True)
    
    success_message = None
    if created > 0:
        success_message = f"{created} candidate{'s' if created != 1 else ''} imported successfully."
    
    return render(request, 'candidates/import_dual.html', {
        'errors': errors,
        'imported_count': created,
        'success_message': success_message
    })

@login_required
def candidate_import_template(request):
    """
    Serve an Excel template for candidate import including all required fields and a sample row.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"
    # Define headers (excluding RegNo and images)
    headers = [
        'full_name', 'gender', 'nationality', 'date_of_birth', 'occupation', 'registration_category',
        'assessment_center', 'entry_year', 'intake', 'start_date', 'finish_date', 'assessment_date'
    ]
    ws.append(headers)
    # Add a sample row
    sample_row = [
        'Jane Doe', 'F', 'Ugandan', '20/06/2000', 'OCC001', 'modular', 'CTR001', '2025', 'Jan', '01/01/2025', '31/12/2025', '15/06/2025'
    ]
    ws.append(sample_row)
    # Add notes row (for user guidance)
    ws.append([
        'Sample: Use DD/MM/YYYY for all dates. Use occupation code and center code as in system. RegNo will be generated.'
    ] + [''] * (len(headers)-1))
    # Save to temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        data = tmp.read()
    response = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="candidate_import_template.xlsx"'
    os.unlink(tmp.name)
    return response

@login_required
def candidate_import(request):
    """
    Handle GET (show import page) and POST (process Excel upload) for candidate import.
    """
    if request.method == 'GET':
        return render(request, 'candidates/import.html')

    # POST: process uploaded Excel
    file = request.FILES.get('excel_file')
    errors = []
    created = 0
    if not file:
        errors.append('No file uploaded.')
        return render(request, 'candidates/import.html', {'errors': errors, 'imported_count': 0})
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
    except Exception:
        errors.append('Invalid Excel file.')
        return render(request, 'candidates/import.html', {'errors': errors, 'imported_count': 0})

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    expected_headers = [
        'full_name', 'gender', 'nationality', 'date_of_birth', 'occupation', 'registration_category',
        'assessment_center', 'entry_year', 'intake', 'start_date', 'finish_date', 'assessment_date'
    ]
    if headers != expected_headers:
        errors.append('Excel headers do not match template. Please download the latest template.')
        return render(request, 'candidates/import.html', {'errors': errors, 'imported_count': 0})

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None for cell in row):
            continue  # skip empty rows
        data = dict(zip(headers, row))
        # Convert and clean data for import
        form_data = data.copy()
        # Dates: convert DD/MM/YYYY to date objects
        for date_field in ['date_of_birth', 'start_date', 'finish_date', 'assessment_date']:
            val = form_data.get(date_field)
            if val:
                try:
                    form_data[date_field] = datetime.strptime(str(val), '%d/%m/%Y').date()
                except Exception:
                    errors.append(f"Row {idx}: Invalid date format in '{date_field}'. Use DD/MM/YYYY.")
                    continue
        # Nationality: must be a valid country name (case-insensitive)
        from django_countries import countries
        nat_val = form_data.get('nationality', '')
        if not isinstance(nat_val, str):
            nat_val = str(nat_val)
        nat_val = nat_val.strip()
        valid_countries = [c[0].lower() for c in countries] + [c[1].lower() for c in countries]
        if nat_val.lower() not in valid_countries:
            errors.append(f"Row {idx}: Nationality '{form_data.get('nationality')}' is not a valid country. Please use a country name from the dropdown.")
            continue
        # Store as country name (display value)
        for code, name in countries:
            if nat_val.lower() in (code.lower(), name.lower()):
                form_data['nationality'] = name
                break
        # Occupation and assessment center: lookup by code
        occ_code = str(form_data.get('occupation')) if form_data.get('occupation') is not None else ''
        center_code = str(form_data.get('assessment_center')) if form_data.get('assessment_center') is not None else ''
        try:
            form_data['occupation'] = Occupation.objects.get(code=occ_code).id
        except Exception:
            errors.append(f"Row {idx}: Invalid occupation code '{occ_code}'.")
            continue
        
        # Enhanced assessment center/branch lookup
        # First try to find a branch by branch_code
        branch_found = None
        center_found = None
        
        try:
            # Try to find branch by branch_code first
            from .models import AssessmentCenterBranch
            branch_found = AssessmentCenterBranch.objects.select_related('assessment_center').get(branch_code=center_code)
            center_found = branch_found.assessment_center
            form_data['assessment_center'] = center_found.id
            form_data['assessment_center_branch'] = branch_found.id
        except AssessmentCenterBranch.DoesNotExist:
            # If no branch found, try to find assessment center by center_number
            try:
                center_found = AssessmentCenter.objects.get(center_number=center_code)
                form_data['assessment_center'] = center_found.id
                form_data['assessment_center_branch'] = None  # No branch specified
            except AssessmentCenter.DoesNotExist:
                errors.append(f"Row {idx}: Invalid assessment center code or branch code '{center_code}'. Please use either a valid center number or branch code.")
                continue
        except Exception as e:
            errors.append(f"Row {idx}: Error looking up assessment center/branch '{center_code}': {str(e)}")
            continue
        # District and village: optional for import
        for loc_field, model_cls in [('district', District), ('village', Village)]:
            val = form_data.get(loc_field)
            if val:
                obj = model_cls.objects.filter(name__iexact=str(val).strip()).first()
                form_data[loc_field] = obj.id if obj else None
            else:
                form_data[loc_field] = None
        # Coerce all other values to string except dates and foreign keys
        for k in form_data:
            if k not in ['date_of_birth', 'start_date', 'finish_date', 'assessment_date', 'occupation', 'assessment_center', 'assessment_center_branch', 'district', 'village']:
                v = form_data[k]
                if v is not None:
                    form_data[k] = str(v)
        # Remove reg_number if present
        form_data.pop('reg_number', None)
        # Use CandidateForm for validation, but patch required fields for import
        form = CandidateForm(form_data)
        for f in ['district', 'village']:
            if f in form.fields:
                form.fields[f].required = False
        if not form.is_valid():
            error_list = '; '.join([f"{k}: {v[0]}" for k, v in form.errors.items()])
            errors.append(f"Row {idx}: {error_list}")
            continue
        candidate = form.save(commit=False)
        candidate.reg_number = None  # Regenerate
        candidate.save()
        created += 1
    return render(request, 'candidates/import.html', {
        'errors': errors,
        'imported_count': created
    })

from django.views.decorators.http import require_POST

@require_POST
def bulk_candidate_modules(request):
    import json
    from .models import Candidate, Module, Level
    try:
        data = json.loads(request.body.decode()) if request.body else request.POST
        ids = data.get('candidate_ids')
        if isinstance(ids, str):
            ids = ids.split(',')
        candidate_ids = [int(i) for i in ids if str(i).isdigit()]
    except Exception:
        return JsonResponse({'success': False, 'error': 'Invalid request.'}, status=400)

    candidates = Candidate.objects.filter(id__in=candidate_ids)
    if not candidates.exists():
        return JsonResponse({'success': False, 'error': 'No candidates found.'}, status=400)

    occupations = set(c.occupation_id for c in candidates)
    regcats = set(c.registration_category for c in candidates)
    print("[DEBUG] Candidate IDs:", candidate_ids)
    print("[DEBUG] Occupations:", occupations)
    print("[DEBUG] Registration categories:", regcats)
    for c in candidates:
        print(f"[DEBUG] Candidate {c.id}: occupation_id={c.occupation_id}, registration_category={c.registration_category!r}")
    if len(occupations) != 1 or len(regcats) != 1 or list(regcats)[0] != 'Modular':
        return JsonResponse({'success': False, 'error': 'All candidates must be Modular and have the same occupation.'}, status=400)

    occupation_id = occupations.pop()
    occupation = Occupation.objects.get(id=occupation_id)
    level_name = f"Level 1 {occupation.code}"
    level = Level.objects.filter(name=level_name, occupation_id=occupation_id).first()
    print(f"[DEBUG] Occupation: {occupation.name} ({occupation.code}), Looking for level: '{level_name}', Found: {level.id if level else None}")
    if not level:
        return JsonResponse({'success': False, 'error': 'Level 1 not found.'}, status=400)
    modules = Module.objects.filter(occupation_id=occupation_id, level=level)
    module_list = [{'id': m.id, 'name': m.name, 'code': getattr(m, 'code', '')} for m in modules]
    print(f"[DEBUG] Returning module_list: {module_list}")
    return JsonResponse({'success': True, 'modules': module_list, 'level_id': level.id})

@login_required
@require_POST
def bulk_candidate_action(request):
    import json
    try:
        data = json.loads(request.body.decode()) if request.body else request.POST
        print(f"[DEBUG] bulk_candidate_action raw data: {data}")
        action = data.get('action')
        ids = data.get('candidate_ids')
        module_ids = data.get('module_ids')
        level_id = data.get('level_id')
        paper_ids = data.get('paper_ids')
        print(f"[DEBUG] action: {action}, candidate_ids: {ids}, module_ids: {module_ids}, level_id: {level_id}, paper_ids: {paper_ids}")
        if isinstance(ids, str):
            ids = ids.split(',')
        candidate_ids = [int(i) for i in ids if str(i).isdigit()]
    except Exception:
        return JsonResponse({'success': False, 'error': 'Invalid request.'}, status=400)

    candidates = Candidate.objects.filter(id__in=candidate_ids)
    if not candidates.exists():
        return JsonResponse({'success': False, 'error': 'No candidates found.'}, status=400)

    if action == 'enroll':
        # Bulk enrollment for formal/informal/worker's pas/modular
        level_id = data.get('level_id')
        paper_ids = data.get('paper_ids', [])
        module_ids = data.get('module_ids', [])
        assessment_series_id = data.get('assessment_series_id')
    
        # Enforce same occupation and regcat for all candidates
        occupations = set(c.occupation_id for c in candidates)
        regcats = set((c.registration_category or '').strip().lower() for c in candidates)
        if len(occupations) != 1 or len(regcats) != 1:
            return JsonResponse({'success': False, 'error': 'All candidates must have the same occupation and registration category to enroll together.'}, status=400)
    
        regcat = regcats.pop()
    
        # Validate Assessment Series selection
        if not assessment_series_id:
            return JsonResponse({'success': False, 'error': 'Assessment Series is required for bulk enrollment.'}, status=400)
    
        try:
            assessment_series = AssessmentSeries.objects.get(id=assessment_series_id)
        except AssessmentSeries.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Invalid Assessment Series selected.'}, status=400)
    
        from .models import Level, Module, Paper, CandidateLevel, CandidateModule, CandidatePaper
        enrolled = 0
    
         # --- FORMAL ---
        if regcat == 'formal':
            if not level_id:
                return JsonResponse({'success': False, 'error': 'Level is required.'})
            level = Level.objects.filter(id=level_id).first()
            if not level:
                return JsonResponse({'success': False, 'error': 'Invalid level.'})
            for c in candidates:
                CandidateLevel.objects.get_or_create(candidate=c, level=level)
                # Update candidate's assessment series
                c.assessment_series = assessment_series
                # Calculate and set billing fees
                if hasattr(c, 'calculate_fees_balance'):
                    calculated_fees = c.calculate_fees_balance()
                    c.fees_balance = calculated_fees
                c.save()
                enrolled += 1
            return JsonResponse({'success': True, 'message': f'Enrolled and billed {enrolled} candidates in {level.name}.'})
        
        # --- WORKER'S PAS / INFORMAL (Cross-level paper selection) ---
        elif regcat in ['informal', "worker's pas", "workers pas"]:
            # For cross-level enrollment, we don't require a specific level_id
            # Instead, we process the selected papers and determine levels/modules from them
            if not paper_ids:
                return JsonResponse({'success': False, 'error': 'At least one paper must be selected for Worker\'s PAS/Informal enrollment.'})
            
            # Check if any candidate has had previous results to determine if this is first sitting
            has_any_previous_results = Result.objects.filter(candidate__in=candidates).exists()
            

            
            # Validate paper count based on whether this is first sitting or subsequent sitting
            if has_any_previous_results:
                # Subsequent sitting: No minimum restriction, only maximum of 4 papers
                if len(paper_ids) > 4:
                    return JsonResponse({'success': False, 'error': 'Worker\'s PAS/Informal candidates can select a maximum of 4 papers at any given sitting.'})
                # Allow any number of papers (including 1) for re-enrollment
            else:
                # First sitting: Enforce minimum of 2 and maximum of 4 papers
                if len(paper_ids) < 2:
                    return JsonResponse({'success': False, 'error': 'Worker\'s PAS/Informal candidates must select a minimum of 2 papers for their first sitting.'})
                elif len(paper_ids) > 4:
                    return JsonResponse({'success': False, 'error': 'Worker\'s PAS/Informal candidates can select a maximum of 4 papers at any given sitting.'})
            
            # Get selected papers with their levels and modules
            papers = Paper.objects.filter(id__in=paper_ids).select_related('module', 'level', 'occupation')
            if papers.count() != len(paper_ids):
                return JsonResponse({'success': False, 'error': 'Invalid paper selection.'})
            
            # Validate that papers are from the same occupation as candidates
            occupation_id = occupations.pop()
            paper_occupations = set(paper.occupation_id for paper in papers)
            if len(paper_occupations) != 1 or occupation_id not in paper_occupations:
                return JsonResponse({'success': False, 'error': 'Selected papers must be from the same occupation as candidates.'})
            
            # Validate one paper per module rule
            module_ids_in_papers = [paper.module_id for paper in papers]
            if len(module_ids_in_papers) != len(set(module_ids_in_papers)):
                return JsonResponse({'success': False, 'error': 'You can only select one paper per module.'})
            
            # Validate per-assessment series enrollment eligibility for each candidate
            ineligible_candidates = []
            for c in candidates:
                # Get candidate's enrollment and result history
                enrolled_papers = set(CandidatePaper.objects.filter(candidate=c).values_list('paper_id', flat=True))
                
                # Get papers with failed or missing results (eligible for re-enrollment)
                failed_or_missing_papers = set(Result.objects.filter(
                    candidate=c,
                    paper_id__in=enrolled_papers
                ).filter(
                    # Failed: CTR comment OR Missing: Ms grade
                    Q(comment='CTR') | Q(grade='Ms')
                ).values_list('paper_id', flat=True))
                
                # Check eligibility for each selected paper
                ineligible_papers = []
                for paper in papers:
                    if paper.id in enrolled_papers:
                        # Candidate has enrolled for this paper before
                        if paper.id not in failed_or_missing_papers:
                            # Paper was passed - not eligible for re-enrollment
                            ineligible_papers.append(paper.name)
                
                if ineligible_papers:
                    ineligible_candidates.append({
                        'candidate': c.full_name,
                        'reg_number': c.reg_number,
                        'ineligible_papers': ineligible_papers
                    })
            
            # If any candidates have ineligible papers, return error
            if ineligible_candidates:
                error_details = []
                for item in ineligible_candidates[:3]:  # Show first 3 candidates
                    error_details.append(f"{item['candidate']} ({item['reg_number']}): {', '.join(item['ineligible_papers'])}")
                
                error_msg = f"Some candidates cannot enroll for selected papers (already passed): {'; '.join(error_details)}"
                if len(ineligible_candidates) > 3:
                    error_msg += f" and {len(ineligible_candidates) - 3} more candidates."
                
                return JsonResponse({'success': False, 'error': error_msg})
            
            # Process enrollment for each candidate
            for c in candidates:
                # For retakes, preserve all previous enrollments and results
                # Only clear enrollments for the CURRENT assessment series to avoid duplicates
                current_series = AssessmentSeries.objects.filter(is_current=True).first()
                if current_series:
                    CandidatePaper.objects.filter(
                        candidate=c, 
                        enrolled_at__gte=current_series.start_date,
                        enrolled_at__lte=current_series.end_date
                    ).delete()
                
                # Track levels and modules we need to enroll in
                levels_to_enroll = set()
                modules_to_enroll = set()
                
                # Process each selected paper
                for paper in papers:
                    levels_to_enroll.add(paper.level)
                    modules_to_enroll.add(paper.module)
                    
                    # Create CandidatePaper record
                    CandidatePaper.objects.create(
                        candidate=c, 
                        module=paper.module, 
                        paper=paper, 
                        level=paper.level
                    )
                
                # Enroll in all required levels (avoid duplicates for retakes)
                for level in levels_to_enroll:
                    CandidateLevel.objects.get_or_create(candidate=c, level=level)
                
                # Enroll in all required modules (avoid duplicates for retakes)
                for module in modules_to_enroll:
                    CandidateModule.objects.get_or_create(candidate=c, module=module)
                
                # Update candidate's assessment series
                c.assessment_series = assessment_series
                
                # Calculate and set billing fees after enrollment
                if hasattr(c, 'calculate_fees_balance'):
                    calculated_fees = c.calculate_fees_balance()
                    c.fees_balance = calculated_fees
                c.save()
                enrolled += 1
            
            # Create success message
            level_names = list(set(paper.level.name for paper in papers))
            paper_names = [paper.name for paper in papers]
            return JsonResponse({
                'success': True, 
                'message': f'Enrolled and billed {enrolled} candidates across {len(level_names)} level(s): {", ".join(level_names)}. Selected {len(paper_names)} paper(s): {", ".join(paper_names)}'
            })
        
        # --- MODULAR ---
        elif regcat == 'modular':
            print(f"[DEBUG] Modular enrollment - module_ids: {module_ids}, len: {len(module_ids)}")
            if not (1 <= len(module_ids) <= 2):
                return JsonResponse({'success': False, 'error': 'Select 1 or 2 modules.'}, status=400)
            occupation_id = occupations.pop()
            print(f"[DEBUG] Occupation ID: {occupation_id}, Level ID: {level_id}")
            level = Level.objects.filter(id=level_id).first() if level_id else Level.objects.filter(name__icontains='1').first()
            print(f"[DEBUG] Found level: {level}")
            if not level:
                return JsonResponse({'success': False, 'error': 'Level not found.'}, status=400)
            modules = Module.objects.filter(id__in=module_ids, occupation_id=occupation_id, level=level)
            print(f"[DEBUG] Found modules: {modules.count()}/{len(module_ids)} - {list(modules.values_list('name', flat=True))}")
            if modules.count() != len(module_ids):
                return JsonResponse({'success': False, 'error': 'Invalid module selection.'}, status=400)
            for c in candidates:
                CandidateModule.objects.filter(candidate=c).delete()
                for m in modules:
                    CandidateModule.objects.create(
                        candidate=c, 
                        module=m,
                        assessment_series=assessment_series
                    )
                # Update candidate's assessment series
                c.assessment_series = assessment_series
                # Calculate and set billing fees after module enrollment
                if hasattr(c, 'calculate_fees_balance'):
                    calculated_fees = c.calculate_fees_balance()
                    c.fees_balance = calculated_fees
                c.save()
                enrolled += 1
            return JsonResponse({'success': True, 'message': f'Successfully enrolled and billed {enrolled} candidates in {modules.count()} module(s).'})
        else:
            return JsonResponse({'success': False, 'error': 'Bulk enroll only supported for Formal, Modular, or Worker\'s PAS/Informal registration categories.'}, status=400)
    if action == 'add_regno_photo':
        results = []
        from PIL import Image as PILImage, ImageDraw, ImageFont
        import glob
        from django.core.files import File
        for c in candidates:
            try:
                if not c.passport_photo or not c.reg_number:
                    results.append({'id': c.id, 'success': False, 'error': 'Missing photo or regno.'})
                    continue
                image_path = c.passport_photo.path
                img = PILImage.open(image_path).convert("RGBA")
                draw = ImageDraw.Draw(img)
                # Remove old regno-stamped images
                base_dir = os.path.dirname(image_path)
                base_name = os.path.splitext(os.path.basename(image_path))[0]
                regno_img_pattern = os.path.join(base_dir, f"{base_name}_regno*.png")
                for old in glob.glob(regno_img_pattern):
                    try:
                        os.remove(old)
                    except Exception:
                        pass
                text = c.reg_number
                width, height = img.size
                max_width = width - 32
                # Try to use system TrueType font (DejaVuSans-Bold), fallback to PIL default
                font = None
                truetype_paths = [
                    '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
                    '/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf',
                    '/usr/local/share/fonts/DejaVuSans-Bold.ttf',
                    '/usr/share/fonts/truetype/freefont/FreeSansBold.ttf',
                ]
                font_size = 40
                min_font_size = 10
                for path in truetype_paths:
                    if os.path.exists(path):
                        try:
                            font = ImageFont.truetype(path, size=font_size)
                            break
                        except Exception:
                            font = None
                if font is None:
                    font = ImageFont.load_default()
                    font_size = 12
                    min_font_size = 8
                # Dynamically shrink font size if needed
                while font_size >= min_font_size:
                    try:
                        text_w, text_h = font.getsize(text)
                    except AttributeError:
                        bbox = font.getbbox(text)
                        text_w, text_h = bbox[2] - bbox[0], bbox[3] - bbox[1]
                    if text_w <= max_width:
                        break
                    font_size -= 2
                    if hasattr(font, 'path'):
                        try:
                            font = ImageFont.truetype(font.path, size=font_size)
                        except Exception:
                            font = ImageFont.load_default()
                            break
                # If still too wide, truncate
                if text_w > max_width:
                    ellipsis = '...'
                    for i in range(len(text)-1, 0, -1):
                        truncated = text[:i] + ellipsis
                        try:
                            text_w, text_h = font.getsize(truncated)
                        except AttributeError:
                            bbox = font.getbbox(truncated)
                            text_w, text_h = bbox[2] - bbox[0], bbox[3] - bbox[1]
                        if text_w <= max_width:
                            text = truncated
                            break
                padding_v = max(10, text_h // 4)
                strip_h = text_h + 2 * padding_v
                strip_y = height - strip_h
                overlay = PILImage.new('RGBA', (width, strip_h), (255, 255, 255, 255))  # Solid white
                img = img.convert('RGBA')
                img.alpha_composite(overlay, (0, strip_y))
                draw = ImageDraw.Draw(img)
                x = (width - text_w) // 2
                y = strip_y + (strip_h - text_h) // 2
                draw.text((x, y), text, font=font, fill=(0,0,0,255))
                new_filename = os.path.join(base_dir, f"{base_name}_regno.png")
                img.save(new_filename)
                with open(new_filename, 'rb') as f:
                    c.passport_photo_with_regno.save(os.path.basename(new_filename), File(f), save=True)
                results.append({'id': c.id, 'success': True})
            except Exception as e:
                results.append({'id': c.id, 'success': False, 'error': str(e)})
        return JsonResponse({'success': True, 'results': results})


    elif action == 'regenerate':
        updated = 0
        for c in candidates:
            c.reg_number = None
            c.save()
            updated += 1
        return JsonResponse({'success': True, 'message': f'Regenerated registration numbers for {updated} candidates.'})
    elif action == 'change_reg_cat':
        # Bulk change registration category with validation
        new_cat = data.get('registration_category')
        if not new_cat:
            return JsonResponse({'success': False, 'error': 'No registration category provided.'}, status=400)
        # Validate all candidates
        for candidate in candidates:
            occupation = candidate.occupation
            occ_cat = occupation.category.name if occupation and occupation.category else None
            has_modular = getattr(occupation, 'has_modular', False)
            if new_cat == 'Modular':
                if not has_modular:
                    return JsonResponse({'success': False, 'error': f"Candidate {candidate.full_name} occupation does not support Modular registration."}, status=400)
            elif new_cat == 'Formal':
                if occ_cat != 'Formal':
                    return JsonResponse({'success': False, 'error': f"Candidate {candidate.full_name} occupation is not in the 'Formal' category."}, status=400)
            elif new_cat == 'Informal':
                if occ_cat != "Worker's PAS":
                    return JsonResponse({'success': False, 'error': f"Candidate {candidate.full_name} occupation is not in the 'Worker's PAS' category."}, status=400)
            else:
                return JsonResponse({'success': False, 'error': 'Invalid registration category selected.'}, status=400)
        # If all valid, update
        updated = 0
        for candidate in candidates:
            candidate.registration_category = new_cat
            candidate.reg_number = None
            candidate.save(update_fields=["registration_category", "reg_number"])
            updated += 1
        return JsonResponse({'success': True, 'message': f'Changed registration category for {updated} candidates.'})
    elif action == 'mark_disabled':
        # Bulk mark as disabled
        nature_ids = data.get('nature_of_disability_ids', [])
        if isinstance(nature_ids, str):
            nature_ids = [nature_ids]
        if not isinstance(nature_ids, list):
            return JsonResponse({'success': False, 'error': 'Invalid nature_of_disability_ids.'}, status=400)
        if not nature_ids:
            return JsonResponse({'success': False, 'error': 'Please select at least one nature of disability.'}, status=400)
        from .models import NatureOfDisability
        natures = NatureOfDisability.objects.filter(id__in=nature_ids)
        if natures.count() != len(nature_ids):
            return JsonResponse({'success': False, 'error': 'One or more selected natures are invalid.'}, status=400)
        updated = 0
        for c in candidates:
            c.disability = True
            c.save()
            c.nature_of_disability.set(natures)
            updated += 1
        return JsonResponse({'success': True, 'message': f'Marked {updated} candidates as disabled with selected nature(s).'})
    
    elif action == 'change_occupation':
        # Bulk change occupation for candidates
        occupation_id = data.get('occupation_id')
        if not occupation_id:
            return JsonResponse({'success': False, 'error': 'Occupation is required.'}, status=400)
        
        try:
            new_occupation = Occupation.objects.get(id=occupation_id)
        except Occupation.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Invalid occupation selected.'}, status=400)
        
        # Check if any candidates are enrolled (cannot change occupation if enrolled)
        enrolled_candidates = []
        for candidate in candidates:
            if candidate.is_enrolled():
                enrolled_candidates.append(candidate.reg_number or f"ID:{candidate.id}")
        
        if enrolled_candidates:
            enrolled_list = ', '.join(enrolled_candidates[:5])  # Show first 5
            if len(enrolled_candidates) > 5:
                enrolled_list += f" and {len(enrolled_candidates) - 5} more"
            return JsonResponse({
                'success': False, 
                'error': f'Cannot change occupation for enrolled candidates: {enrolled_list}'
            }, status=400)
        
        # Update occupation and regenerate registration numbers
        updated = 0
        for candidate in candidates:
            candidate.occupation = new_occupation
            candidate.build_reg_number()  # Regenerate registration number
            candidate.save(update_fields=['occupation', 'reg_number'])
            updated += 1
        
        return JsonResponse({
            'success': True, 
            'message': f'Successfully changed occupation for {updated} candidate{"s" if updated != 1 else ""}. Registration numbers have been updated.'
        })
    
    elif action == 'change_center':
        # Bulk change assessment center for candidates
        center_id = data.get('assessment_center_id') or data.get('assessment_center')
        if not center_id:
            return JsonResponse({'success': False, 'error': 'Assessment Center is required.'}, status=400)
        
        try:
            new_center = AssessmentCenter.objects.get(id=center_id)
        except AssessmentCenter.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Invalid Assessment Center selected.'}, status=400)
        
        # Block if any candidate is enrolled
        enrolled_candidates = []
        for candidate in candidates:
            if candidate.is_enrolled():
                enrolled_candidates.append(candidate.reg_number or f"ID:{candidate.id}")
        if enrolled_candidates:
            enrolled_list = ', '.join(enrolled_candidates[:5])
            if len(enrolled_candidates) > 5:
                enrolled_list += f" and {len(enrolled_candidates) - 5} more"
            return JsonResponse({
                'success': False,
                'error': f'Cannot change assessment center for enrolled candidates: {enrolled_list}'
            }, status=400)
        
        # Update center and regenerate reg numbers
        updated = 0
        for candidate in candidates:
            candidate.assessment_center = new_center
            candidate.build_reg_number()  # Rebuild reg_number since center code changes
            candidate.save(update_fields=['assessment_center', 'reg_number'])
            updated += 1
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully changed assessment center to {new_center.center_name} for {updated} candidate' + ('' if updated == 1 else 's') + '. Registration numbers have been updated.'
        })
    
    elif action == 'verify':
        # Bulk verify candidates - only admin/staff can do this
        if not (request.user.is_staff or (request.user.groups.exists() and 'CenterRep' not in [g.name for g in request.user.groups.all()])):
            return JsonResponse({'success': False, 'error': 'Permission denied. Only admin/staff can verify candidates.'}, status=403)
        
        from django.utils import timezone
        
        verified_count = 0
        already_verified_count = 0
        
        for candidate in candidates:
            if candidate.verification_status == 'verified':
                already_verified_count += 1
            else:
                candidate.verification_status = 'verified'
                candidate.verification_date = timezone.now()
                candidate.verified_by = request.user
                candidate.decline_reason = None  # Clear any previous decline reason
                candidate.save(update_fields=['verification_status', 'verification_date', 'verified_by', 'decline_reason'])
                verified_count += 1
        
        message_parts = []
        if verified_count > 0:
            message_parts.append(f'Successfully verified {verified_count} candidate{"s" if verified_count != 1 else ""}')
        if already_verified_count > 0:
            message_parts.append(f'{already_verified_count} candidate{"s" if already_verified_count != 1 else ""} {"were" if already_verified_count != 1 else "was"} already verified')
        
        message = '. '.join(message_parts) + '.'
        
        return JsonResponse({
            'success': True,
            'message': message
        })
    
    elif action == 'decline':
        # Bulk decline candidates - only admin/staff can do this
        if not (request.user.is_staff or (request.user.groups.exists() and 'CenterRep' not in [g.name for g in request.user.groups.all()])):
            return JsonResponse({'success': False, 'error': 'Permission denied. Only admin/staff can decline candidates.'}, status=403)
        
        decline_reason = data.get('decline_reason', '').strip()
        if not decline_reason:
            return JsonResponse({'success': False, 'error': 'Decline reason is required.'}, status=400)
        
        from django.utils import timezone
        
        declined_count = 0
        already_declined_count = 0
        
        for candidate in candidates:
            if candidate.verification_status == 'declined':
                already_declined_count += 1
            else:
                candidate.verification_status = 'declined'
                candidate.verification_date = timezone.now()
                candidate.verified_by = request.user
                candidate.decline_reason = decline_reason
                candidate.save(update_fields=['verification_status', 'verification_date', 'verified_by', 'decline_reason'])
                declined_count += 1
        
        message_parts = []
        if declined_count > 0:
            message_parts.append(f'Successfully declined {declined_count} candidate{"s" if declined_count != 1 else ""}')
        if already_declined_count > 0:
            message_parts.append(f'{already_declined_count} candidate{"s" if already_declined_count != 1 else ""} {"were" if already_declined_count != 1 else "was"} already declined')
        
        message = '. '.join(message_parts) + '.'
        
        return JsonResponse({
            'success': True,
            'message': message
        })
    
    elif action == 'mark_refugee':
        # Bulk mark candidates as refugees
        refugee_numbers = data.get('refugee_numbers', {})
        updated_count = 0
        already_refugee_count = 0
        
        for candidate in candidates:
            if candidate.is_refugee:
                already_refugee_count += 1
            else:
                candidate.is_refugee = True
                # Set refugee number if provided
                candidate_id_str = str(candidate.id)
                if candidate_id_str in refugee_numbers:
                    candidate.refugee_number = refugee_numbers[candidate_id_str]
                candidate.save(update_fields=['is_refugee', 'refugee_number'])
                updated_count += 1
        
        message_parts = []
        if updated_count > 0:
            message_parts.append(f'Successfully marked {updated_count} candidate{"s" if updated_count != 1 else ""} as refugee{"s" if updated_count != 1 else ""}')
        if already_refugee_count > 0:
            message_parts.append(f'{already_refugee_count} candidate{"s" if already_refugee_count != 1 else ""} {"were" if already_refugee_count != 1 else "was"} already marked as refugee{"s" if already_refugee_count != 1 else ""}')
        
        message = '. '.join(message_parts) + '.'
        
        return JsonResponse({
            'success': True,
            'message': message
        })
    
    else:
        return JsonResponse({'success': False, 'error': 'Unknown action.'}, status=400)

@login_required
def export_candidates(request):
    """Export selected candidates or all filtered candidates to Excel with comprehensive data"""
    from django.views.decorators.http import require_POST
    from datetime import datetime
    
    # Get selected candidate IDs from POST data
    candidate_ids = request.POST.getlist('candidate_ids')
    export_all = request.POST.get('export_all', 'false').lower() == 'true'
    
    if not candidate_ids and not export_all:
        return HttpResponse("No candidates selected for export", status=400)
    
    # If export_all is true, get all candidates with current filters applied
    if export_all:
        # Get current filters from session
        current_filters = request.session.get('candidate_filters', {})
        
        # Start with base query
        candidates = Candidate.objects.select_related('occupation', 'occupation__sector', 'assessment_center').order_by('-created_at')
        
        # Restrict for Center Representatives
        if request.user.groups.filter(name='CenterRep').exists():
            from .models import CenterRepresentative
            try:
                center_rep = CenterRepresentative.objects.get(user=request.user)
                candidates = candidates.filter(assessment_center=center_rep.center)
            except CenterRepresentative.DoesNotExist:
                candidates = candidates.none()
        
        # Apply all current filters
        if current_filters.get('reg_number'):
            candidates = candidates.filter(reg_number__icontains=current_filters.get('reg_number'))
        if current_filters.get('search'):
            candidates = candidates.filter(full_name__icontains=current_filters.get('search'))
        if current_filters.get('occupation'):
            candidates = candidates.filter(occupation_id=current_filters.get('occupation'))
        if current_filters.get('registration_category'):
            candidates = candidates.filter(registration_category=current_filters.get('registration_category'))
        if current_filters.get('assessment_center'):
            candidates = candidates.filter(assessment_center_id=current_filters.get('assessment_center'))
        if current_filters.get('sector'):
            candidates = candidates.filter(occupation__sector_id=current_filters.get('sector'))
        if current_filters.get('gender'):
            candidates = candidates.filter(gender=current_filters.get('gender'))
        if current_filters.get('disability'):
            disability_filter = current_filters.get('disability').lower() == 'true'
            candidates = candidates.filter(disability=disability_filter)
        if current_filters.get('is_refugee'):
            refugee_filter = current_filters.get('is_refugee').lower() == 'true'
            candidates = candidates.filter(is_refugee=refugee_filter)
        if current_filters.get('assessment_year'):
            try:
                year = int(current_filters.get('assessment_year'))
                candidates = candidates.filter(assessment_date__year=year)
            except (ValueError, TypeError):
                pass
        if current_filters.get('assessment_month'):
            try:
                month = int(current_filters.get('assessment_month'))
                candidates = candidates.filter(assessment_date__month=month)
            except (ValueError, TypeError):
                pass
        
        # Add additional prefetch for export
        candidates = candidates.select_related(
            'assessment_center', 'occupation', 'occupation__sector', 'district', 'village', 'created_by', 'updated_by'
        ).prefetch_related(
            'nature_of_disability', 'candidatelevel_set__level', 'candidatemodule_set__module',
            'candidatepaper_set__paper', 'result_set'
        ).order_by('reg_number')
    else:
        # Get candidates with related data for selected IDs
        candidates = Candidate.objects.filter(
            id__in=candidate_ids
        ).select_related(
            'assessment_center', 'occupation', 'occupation__sector', 'district', 'village', 'created_by', 'updated_by'
        ).prefetch_related(
            'nature_of_disability', 'candidatelevel_set__level', 'candidatemodule_set__module',
            'candidatepaper_set__paper', 'result_set'
        ).order_by('reg_number')
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Candidates Export"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define headers with all comprehensive fields
    headers = [
        # Bio Data
        "Registration Number", "Full Name", "Date of Birth", "Gender", "Nationality",
        "Contact", 
        
        # Address Information
        "District", "Village",
        
        # Assessment Information
        "Assessment Center", "Occupation", "Sector", "Registration Category", "Entry Year", "Intake",
        "Assessment Date", "Assessment Series", "Start Date", "Finish Date",
        
        # Disability Information
        "Has Disability", "Nature of Disability", "Disability Specification",
        
        # Enrollment Information
        "Enrolled Levels", "Enrolled Modules", "Enrolled Papers",
        
        # Financial Information
        "Fees Balance",
        
        # Results Summary
        "Has Results", "Result Count", "Latest Result Date",
        
        # System Information
        "Status", "Created Date", "Created By", "Updated Date", "Updated By"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data for each candidate
    for row, candidate in enumerate(candidates, 2):
        # Get enrollment information
        enrolled_levels = list(candidate.candidatelevel_set.all())
        enrolled_modules = list(candidate.candidatemodule_set.all())
        enrolled_papers = list(candidate.candidatepaper_set.all())
        
        # Get results information
        results = list(candidate.result_set.all())
        latest_result = results[0] if results else None
        
        # Get nature of disability
        disabilities = list(candidate.nature_of_disability.all())
        disability_names = ", ".join([d.name for d in disabilities]) if disabilities else ""
        
        # Prepare data row
        data = [
            # Bio Data
            candidate.reg_number or "",
            candidate.full_name or "",
            candidate.date_of_birth.strftime('%d/%m/%Y') if candidate.date_of_birth else "",
            candidate.get_gender_display() if candidate.gender else "",
            candidate.nationality or "",
            candidate.contact or "",
            
            # Address Information
            candidate.district.name if candidate.district else "",
            candidate.village.name if candidate.village else "",
            
            # Assessment Information
            candidate.assessment_center.center_name if candidate.assessment_center else "",
            candidate.occupation.name if candidate.occupation else "",
            candidate.occupation.sector.name if candidate.occupation and candidate.occupation.sector else "Unknown",
            candidate.registration_category or "",
            candidate.entry_year or "",
            candidate.intake or "",
            candidate.assessment_date.strftime('%d/%m/%Y') if candidate.assessment_date else "",
            candidate.assessment_series.name if candidate.assessment_series else "",
            candidate.start_date.strftime('%d/%m/%Y') if candidate.start_date else "",
            candidate.finish_date.strftime('%d/%m/%Y') if candidate.finish_date else "",
            
            # Disability Information
            "Yes" if candidate.disability else "No",
            disability_names,
            candidate.disability_specification or "",
            
            # Enrollment Information
            ", ".join([f"{cl.level.name}" for cl in enrolled_levels]) if enrolled_levels else "",
            ", ".join([f"{cm.module.name}" for cm in enrolled_modules]) if enrolled_modules else "",
            ", ".join([f"{cp.paper.name}" for cp in enrolled_papers]) if enrolled_papers else "",
            
            # Financial Information
            float(candidate.fees_balance) if candidate.fees_balance else 0.00,
            
            # Results Summary
            "Yes" if results else "No",
            len(results),
            latest_result.assessment_date.strftime('%d/%m/%Y') if latest_result and latest_result.assessment_date else "",
            
            # System Information
            candidate.status or "Active",
            candidate.created_at.strftime('%d/%m/%Y %H:%M') if candidate.created_at else "",
            candidate.created_by.username if candidate.created_by else "",
            candidate.updated_at.strftime('%d/%m/%Y %H:%M') if candidate.updated_at else "",
            candidate.updated_by.username if candidate.updated_by else ""
        ]
        
        # Write data to worksheet
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            
            # Format specific columns
            if col in [3, 10, 12, 13, 18, 20, 22]:  # Date columns (DOB, Assessment Date, Start Date, Finish Date, Latest Result Date, Created Date, Updated Date)
                cell.alignment = Alignment(horizontal="center")
            elif col in [17]:  # Numeric columns (Fees Balance)
                cell.alignment = Alignment(horizontal="right")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        # Set minimum and maximum widths
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"EMIS_Candidates_Export_{timestamp}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response

from django.urls import reverse
from .models import AssessmentCenter, AssessmentCenterBranch, Candidate, Occupation, AssessmentCenterCategory, Level, Module, Paper, CandidateLevel, CandidateModule, Village, District
from . import views_api
from .forms import AssessmentCenterForm, AssessmentCenterBranchForm, OccupationForm, ModuleForm, PaperForm, CandidateForm, EnrollmentForm, DistrictForm, VillageForm
from django.contrib import messages 
from reportlab.lib import colors    
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from io import BytesIO
from PIL import Image as PILImage
import os
from django.conf import settings
from django.contrib.auth.decorators import login_required
from .forms import ChangeOccupationForm, ChangeCenterForm
from .models import NatureOfDisability
from .forms import NatureOfDisabilityForm

from django.urls import reverse

@login_required
def natureofdisability_list(request):
    # Get filter parameters
    name = request.GET.get('name', '').strip()
    description = request.GET.get('description', '').strip()
    
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    disabilities = NatureOfDisability.objects.all()
    
    # Apply filters
    if name:
        disabilities = disabilities.filter(name__icontains=name)
    if description:
        disabilities = disabilities.filter(description__icontains=description)
    
    disabilities = disabilities.order_by('name')
    
    # Pagination: 20 per page
    paginator = Paginator(disabilities, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'configurations/natureofdisability_list.html', {
        'page_obj': page_obj,
        'paginator': paginator,
        'filters': {
            'name': name,
            'description': description,
        }
    })

@login_required
def natureofdisability_create(request):
    if request.method == 'POST':
        form = NatureOfDisabilityForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('natureofdisability_list')
    else:
        form = NatureOfDisabilityForm()
    return render(request, 'configurations/natureofdisability_form.html', {'form': form, 'create': True})

@login_required
def natureofdisability_view(request, pk):
    entry = get_object_or_404(NatureOfDisability, pk=pk)
    return render(request, 'configurations/natureofdisability_view.html', {'entry': entry})

@login_required
def natureofdisability_edit(request, pk):
    entry = get_object_or_404(NatureOfDisability, pk=pk)
    if request.method == 'POST':
        form = NatureOfDisabilityForm(request.POST, instance=entry)
        if form.is_valid():
            form.save()
            return redirect('natureofdisability_view', pk=pk)
    else:
        form = NatureOfDisabilityForm(instance=entry)
    return render(request, 'configurations/natureofdisability_form.html', {'form': form, 'entry': entry, 'edit': True})
@login_required
def dashboard(request):
    logger = logging.getLogger(__name__)
    logger.info(f'Dashboard accessed by user: {request.user}')
    group_names = list(request.user.groups.values_list('name', flat=True))
    
    # Get user department if they are staff using session management utility
    staff, user_department, is_authenticated = get_user_staff_info(request)
    
    context = {
        'group_names': group_names,
        'user_department': user_department
    }
    
    return render(request, 'dashboard.html', context)
    
def district_villages_api(request, district_id):
    villages = Village.objects.filter(district_id=district_id).values('id', 'name')
    # returns [{"id": 3, "name": "Ntare"}, ...]
    return JsonResponse(list(villages), safe=False)


def assessment_center_list(request):
    centers = AssessmentCenter.objects.all().order_by('center_name')
    
    # Get filter parameters
    center_number = request.GET.get('center_number', '').strip()
    center_name = request.GET.get('center_name', '').strip()
    district = request.GET.get('district', '').strip()
    village = request.GET.get('village', '').strip()
    category = request.GET.get('category', '').strip()
    
    # Apply filters
    if center_number:
        centers = centers.filter(center_number__icontains=center_number)
    if center_name:
        centers = centers.filter(center_name__icontains=center_name)
    if district:
        centers = centers.filter(district__name__icontains=district)
    if village:
        centers = centers.filter(village__name__icontains=village)
    if category:
        centers = centers.filter(category_id=category)
    
    # Get all categories for the filter dropdown
    from .models import AssessmentCenterCategory
    categories = AssessmentCenterCategory.objects.all().order_by('name')
    
    # Pagination: 100 per page
    from django.core.paginator import Paginator
    paginator = Paginator(centers, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'assessment_centers/list.html', {
        'centers': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'categories': categories,
    })



def assessment_center_view(request, id):
    center = get_object_or_404(AssessmentCenter, id=id)
    candidates = Candidate.objects.filter(assessment_center=center)
    
    # Get unique occupations from candidates in this assessment center
    occupations = Occupation.objects.filter(candidate__assessment_center=center).distinct()

    return render(request, 'assessment_centers/view.html', {
        'center': center,
        'candidates': candidates,
        'occupations': occupations,
    })





def assessment_center_create(request):
    if request.method == 'POST':
        form = AssessmentCenterForm(request.POST)
        if form.is_valid():
            # Check for center name duplicates and show warning (form validation handles center_number duplicates)
            center_name = form.cleaned_data['center_name']
            if AssessmentCenter.objects.filter(center_name__iexact=center_name).exists():
                messages.warning(request, f'An Assessment Center with the name "{center_name}" already exists. Please verify this is not a duplicate.')
            
            try:
                center = form.save()
                messages.success(request, f'Assessment Center "{center.center_name}" ({center.center_number}) has been created successfully!')
                return redirect('assessment_center_list')
            except Exception as e:
                messages.error(request, f'Error creating assessment center: {str(e)}')
                return render(request, 'assessment_centers/create.html', {'form': form})
        else:
            # Form validation failed - errors will be displayed automatically
            messages.error(request, 'Please correct the errors below and try again.')
    else:
        form = AssessmentCenterForm()

    return render(request, 'assessment_centers/create.html', {'form': form})


def edit_assessment_center(request, id):
    center = get_object_or_404(AssessmentCenter, id=id)
    if request.method == 'POST':
        form = AssessmentCenterForm(request.POST, instance=center)
        if form.is_valid():
            form.save()
            return redirect('assessment_center_view', id=center.id)
    else:
        form = AssessmentCenterForm(instance=center)
    
    return render(request, 'assessment_centers/edit.html', {
        'form': form,
        'center': center,
    })


# Assessment Center Branch Management Views
def assessment_center_branches(request, center_id):
    """List all branches for a specific assessment center"""
    center = get_object_or_404(AssessmentCenter, id=center_id)
    branches = center.branches.all().order_by('branch_code')
    
    return render(request, 'assessment_centers/branches.html', {
        'center': center,
        'branches': branches,
    })


def assessment_center_branch_create(request, center_id):
    """Create a new branch for an assessment center"""
    center = get_object_or_404(AssessmentCenter, id=center_id)
    
    # Check if center has branches enabled
    if not center.has_branches:
        messages.error(request, 'This assessment center does not have branches enabled. Please enable branches first.')
        return redirect('assessment_center_view', id=center.id)
    
    if request.method == 'POST':
        form = AssessmentCenterBranchForm(request.POST, assessment_center=center)
        if form.is_valid():
            try:
                branch = form.save()
                messages.success(request, f'Branch "{branch.branch_code}" has been created successfully!')
                return redirect('assessment_center_branches', center_id=center.id)
            except Exception as e:
                messages.error(request, f'Error creating branch: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below and try again.')
    else:
        form = AssessmentCenterBranchForm(assessment_center=center)
    
    return render(request, 'assessment_centers/branch_create.html', {
        'form': form,
        'center': center,
    })


def assessment_center_branch_edit(request, center_id, branch_id):
    """Edit an existing branch"""
    center = get_object_or_404(AssessmentCenter, id=center_id)
    branch = get_object_or_404(AssessmentCenterBranch, id=branch_id, assessment_center=center)
    
    if request.method == 'POST':
        form = AssessmentCenterBranchForm(request.POST, instance=branch, assessment_center=center)
        if form.is_valid():
            try:
                branch = form.save()
                messages.success(request, f'Branch "{branch.branch_code}" has been updated successfully!')
                return redirect('assessment_center_branches', center_id=center.id)
            except Exception as e:
                messages.error(request, f'Error updating branch: {str(e)}')
        else:
            messages.error(request, 'Please correct the errors below and try again.')
    else:
        form = AssessmentCenterBranchForm(instance=branch, assessment_center=center)
    
    return render(request, 'assessment_centers/branch_edit.html', {
        'form': form,
        'center': center,
        'branch': branch,
    })


def assessment_center_branch_delete(request, center_id, branch_id):
    """Delete a branch"""
    center = get_object_or_404(AssessmentCenter, id=center_id)
    branch = get_object_or_404(AssessmentCenterBranch, id=branch_id, assessment_center=center)
    
    # Check if branch has candidates assigned
    candidates_count = branch.candidate_set.count()
    
    if request.method == 'POST':
        if candidates_count > 0:
            messages.error(request, f'Cannot delete branch "{branch.branch_code}" because it has {candidates_count} candidate(s) assigned to it.')
        else:
            branch_code = branch.branch_code
            branch.delete()
            messages.success(request, f'Branch "{branch_code}" has been deleted successfully!')
        return redirect('assessment_center_branches', center_id=center.id)
    
    return render(request, 'assessment_centers/branch_delete.html', {
        'center': center,
        'branch': branch,
        'candidates_count': candidates_count,
    })


def occupation_list(request):
    occupations = Occupation.objects.select_related('category', 'sector').all().order_by('code')
    code = request.GET.get('code', '').strip()
    name = request.GET.get('name', '').strip()
    category = request.GET.get('category', '').strip()
    sector = request.GET.get('sector', '').strip()

    if code:
        occupations = occupations.filter(code__icontains=code)
    if name:
        occupations = occupations.filter(name__icontains=name)
    if category:
        occupations = occupations.filter(category_id=category)
    if sector:
        occupations = occupations.filter(sector_id=sector)

    # Get all categories and sectors for the filter dropdowns
    from .models import OccupationCategory, Sector
    categories = OccupationCategory.objects.all().order_by('name')
    sectors = Sector.objects.all().order_by('name')

    # Pagination: 100 per page
    from django.core.paginator import Paginator
    paginator = Paginator(occupations, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'occupations/list.html', {
        'occupations': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'categories': categories,
        'sectors': sectors,
    })

def occupation_create(request):
    from .models import Level, OccupationLevel
    error = None
    if request.method == 'POST':
        form = OccupationForm(request.POST)
        if form.is_valid():
            occupation = form.save(commit=False)
            occupation.created_by = request.user
            occupation.updated_by = request.user
            occupation.save()
            # Handle dynamic levels
            num_levels = int(request.POST.get('num_levels', 0))
            level_names = set()
            for i in range(1, num_levels + 1):
                lname = request.POST.get(f'level_name_{i}', '').strip()
                stype = request.POST.get(f'structure_type_{i}', 'modules')
                if not lname:
                    error = 'All level names are required.'
                    break
                if lname in level_names:
                    error = f"Duplicate level name: {lname}"
                    break
                level_names.add(lname)
                # Enforce: If has_modular is checked, Level 1 must be modules
                if occupation.has_modular and lname.lower() == 'level 1' and stype == 'papers':
                    error = "If 'Has Modular' is checked, Level 1 must be set to Modules."
                    break
            if not error:
                for i in range(1, num_levels + 1):
                    lname = request.POST.get(f'level_name_{i}', '').strip()
                    stype = request.POST.get(f'structure_type_{i}', 'modules')
                    if lname:
                        # Compose level name as 'Level X OCC' for clarity
                        occ_code = occupation.code.upper() if occupation.code else ''
                        display_name = f"{lname} {occ_code}" if occ_code and occ_code not in lname else lname
                        level = Level.objects.create(name=display_name, occupation=occupation)
                        OccupationLevel.objects.create(occupation=occupation, level=level, structure_type=stype)
                return redirect('occupation_list')
    else:
        form = OccupationForm()
    return render(request, 'occupations/create.html', {'form': form, 'error': error})
    if request.method == 'POST':
        form = OccupationForm(request.POST)
        if form.is_valid():
            occupation = form.save(commit=False)
            occupation.created_by = request.user
            occupation.updated_by = request.user
            occupation.save()
            # Handle OccupationLevel creation
            selected_level_ids = request.POST.getlist('levels')
            # Enforce restriction: If has_modular is checked, Level 1 must be modules
            has_modular = occupation.has_modular
            level1 = levels.filter(name__icontains='1').first()
            error = None
            for level in levels:
                if str(level.id) in selected_level_ids:
                    if has_modular and level1 and level.id == level1.id:
                        # If user tries to set Level 1 to papers, error
                        if request.POST.get(f'structure_type_{level.id}') == 'papers':
                            error = "If 'Has Modular' is checked, Level 1 must be set to Modules."
                            break
                        structure_type = 'modules'
                    else:
                        structure_type = request.POST.get(f'structure_type_{level.id}', 'modules')
                    OccupationLevel.objects.create(
                        occupation=occupation,
                        level=level,
                        structure_type=structure_type
                    )
            if error:
                occupation.delete()
                return render(request, 'occupations/create.html', {'form': form, 'levels': levels, 'error': error})
            return redirect('occupation_list')
    else:
        form = OccupationForm()
    return render(request, 'occupations/create.html', {'form': form, 'levels': levels})


def occupation_view(request, pk):
    occupation = get_object_or_404(Occupation, pk=pk)
    return render(request, 'occupations/view.html', {'occupation': occupation})


def occupation_detail(request, pk):
    occupation = get_object_or_404(Occupation, pk=pk)
    from .models import OccupationLevel, Module, Paper
    occupation_levels = OccupationLevel.objects.filter(occupation=occupation).select_related('level')
    levels = [ol.level for ol in occupation_levels]

    level_data = []
    for ol in occupation_levels:
        if ol.structure_type == 'modules':
            modules = Module.objects.filter(occupation=occupation, level=ol.level)
            # Attach papers to each module
            for module in modules:
                module_papers = list(Paper.objects.filter(module=module))
                print(f"Module: {module.id} {module.name} has papers: {[p.code for p in module_papers]}")
                module.papers = module_papers
            content = modules
        else:
            content = Paper.objects.filter(occupation=occupation, level=ol.level)
        level_data.append({'level': ol.level, 'structure_type': ol.structure_type, 'content': content})

    return render(request, 'occupations/view.html', {
        'occupation': occupation,
        'levels': levels,
        'level_data': level_data
    })

@login_required
def update_occupation_fees(request, pk):
    """Update fees for all levels in an occupation via AJAX"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Method not allowed'})
    
    try:
        occupation = get_object_or_404(Occupation, pk=pk)
        from .models import OccupationLevel, Level
        from decimal import Decimal, InvalidOperation
        
        # Get all levels for this occupation
        occupation_levels = OccupationLevel.objects.filter(occupation=occupation).select_related('level')
        levels = [ol.level for ol in occupation_levels]
        
        updated_count = 0
        
        # Update modular fees (if occupation supports modular)
        if occupation.has_modular and levels:
            # All levels in an occupation share the same modular fees
            try:
                modular_single = request.POST.get('modular_fee_single')
                modular_double = request.POST.get('modular_fee_double')
                
                if modular_single is not None:
                    modular_single = Decimal(str(modular_single))
                if modular_double is not None:
                    modular_double = Decimal(str(modular_double))
                
                # Update modular fees for all levels in this occupation
                for level in levels:
                    if modular_single is not None:
                        level.modular_fee_single = modular_single
                    if modular_double is not None:
                        level.modular_fee_double = modular_double
                    level.save()
                    updated_count += 1
            except (InvalidOperation, ValueError) as e:
                return JsonResponse({'success': False, 'error': f'Invalid modular fee amount: {e}'})
        
        # Check occupation category to determine fee structure
        if occupation.category.name == 'Formal':
            # Update individual level fees for Formal occupations
            for level in levels:
                try:
                    formal_fee = request.POST.get(f'formal_fee_{level.id}')
                    
                    if formal_fee is not None:
                        level.formal_fee = Decimal(str(formal_fee))
                    
                    level.save()
                    
                except (InvalidOperation, ValueError) as e:
                    return JsonResponse({'success': False, 'error': f'Invalid fee amount for {level.name}: {e}'})
        
        elif occupation.category.name == 'Informal' or 'Worker' in occupation.category.name or 'PAS' in occupation.category.name:
            # Update single per-module fee for Worker's PAS occupations
            try:
                workers_pas_module_fee = request.POST.get('workers_pas_module_fee')
                
                if workers_pas_module_fee is not None:
                    module_fee = Decimal(str(workers_pas_module_fee))
                    # Apply the same per-module fee to all levels in this occupation
                    for level in levels:
                        level.workers_pas_module_fee = module_fee
                        level.save()
                        
            except (InvalidOperation, ValueError) as e:
                return JsonResponse({'success': False, 'error': f'Invalid per-module fee amount: {e}'})
        
        # Update candidate fees balances for all candidates in this occupation
        from .models import Candidate
        candidates_updated = 0
        for candidate in Candidate.objects.filter(occupation=occupation, candidatelevel__isnull=False).distinct():
            candidate.update_fees_balance()
            candidates_updated += 1
        
        return JsonResponse({
            'success': True, 
            'message': f'Successfully updated fees for {len(levels)} levels and {candidates_updated} candidate balances'
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


#Add Module View

from django.http import JsonResponse
from .models import Occupation, Level, OccupationLevel

@login_required
def api_occupations(request):
    # Returns all occupations as JSON (id, name, has_modular, category)
    occupations = Occupation.objects.select_related('category').all()
    def map_category(cat):
        if not cat:
            return ''
        name = cat.name.lower()
        if name in ['informal', "worker's pas", "workers' pas", "workers pas", "worker pas"]:
            return 'informal'
        if name in ['formal', 'modular']:
            return 'formal'
        return name
    occ_list = [
        {
            'id': o.id,
            'name': o.name,
            'has_modular': o.has_modular,
            'category': map_category(o.category) if o.category else ''
        }
        for o in occupations
    ]
    return JsonResponse(occ_list, safe=False)

def api_levels_for_occupation(request):
    occ_id = request.GET.get('occupation_id')
    levels = []
    if occ_id:
        levels = list(Level.objects.filter(occupation_id=occ_id).values('id', 'name'))
    return JsonResponse({'levels': levels})

def api_levels_for_papers(request):
    """
    Returns levels for occupation, excluding Level 1 for modular occupations
    (since Level 1 is always module-based for modular occupations)
    """
    occ_id = request.GET.get('occupation_id')
    levels = []
    if occ_id:
        try:
            occupation = Occupation.objects.get(id=occ_id)
            level_filter = Level.objects.filter(occupation_id=occ_id)
            
            # For modular occupations, exclude Level 1 (it's always module-based)
            if occupation.has_modular:
                level_filter = level_filter.exclude(name__icontains='Level 1')
            
            levels = list(level_filter.values('id', 'name'))
        except Occupation.DoesNotExist:
            levels = []
    
    return JsonResponse({'levels': levels})

def api_levels(request):
    # Returns all levels as JSON (id, name, occupation_id)
    occ_levels = OccupationLevel.objects.select_related('level', 'occupation').all()
    levels = [
        {'id': ol.level.id, 'name': ol.level.name, 'occupation_id': ol.occupation.id}
        for ol in occ_levels
    ]
    return JsonResponse(levels, safe=False)

@login_required
def api_centers(request):
    # Returns all assessment centers as JSON (id, name)
    from .models import AssessmentCenter
    centers = AssessmentCenter.objects.all().order_by('center_name')
    data = [
        {'id': c.id, 'name': str(c)} for c in centers
    ]
    return JsonResponse(data, safe=False)

def api_assessment_center_branches(request, center_id):
    """Returns branches for a specific assessment center as JSON"""
    try:
        center = get_object_or_404(AssessmentCenter, id=center_id)
        
        data = {
            'has_branches': center.has_branches,
            'branches': []
        }
        
        if center.has_branches:
            branches = center.branches.select_related('district', 'village').order_by('branch_code')
            data['branches'] = [
                {
                    'id': branch.id,
                    'branch_code': branch.branch_code,
                    'district_name': branch.district.name,
                    'village_name': branch.village.name,
                    'full_name': branch.get_full_name()
                }
                for branch in branches
            ]
        
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# --- AJAX API for dynamic paper creation form ---
from .models import OccupationLevel, Module
from django.views.decorators.http import require_GET

@require_GET
def api_occupation_level_structure(request):
    occupation_id = request.GET.get('occupation_id')
    level_id = request.GET.get('level_id')
    if not occupation_id or not level_id:
        return JsonResponse({'error': 'Missing occupation_id or level_id'}, status=400)
    occ_level = OccupationLevel.objects.filter(occupation_id=occupation_id, level_id=level_id).first()
    if not occ_level:
        return JsonResponse({'error': 'Not found'}, status=404)
    return JsonResponse({'structure_type': occ_level.structure_type})

@require_GET
def api_modules(request):
    occupation_id = request.GET.get('occupation_id')
    level_id = request.GET.get('level_id')
    if not occupation_id:
        return JsonResponse({'modules': []})
    if level_id:
        modules = Module.objects.filter(occupation_id=occupation_id, level_id=level_id)
    else:
        modules = Module.objects.filter(occupation_id=occupation_id)
    data = [{'id': m.id, 'code': m.code, 'name': m.name} for m in modules]
    return JsonResponse({'modules': data})

@require_GET
def api_occupation_category(request):
    occupation_id = request.GET.get('occupation_id')
    if not occupation_id:
        return JsonResponse({'error': 'occupation_id is required'})
    
    try:
        occupation = Occupation.objects.select_related('category').get(id=occupation_id)
        return JsonResponse({
            'category_name': occupation.category.name,
            'category_id': occupation.category.id
        })
    except Occupation.DoesNotExist:
        return JsonResponse({'error': 'Occupation not found'})

def module_list(request):
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # Get current filters from session
    current_filters = request.session.get('module_filters', {})
    
    # Handle filter actions
    if 'apply_filters' in request.GET:
        current_filters = {
            'occupation': request.GET.get('occupation', '').strip(),
            'level': request.GET.get('level', '').strip(),
            'search': request.GET.get('search', '').strip(),
        }
        request.session['module_filters'] = current_filters
    elif 'clear_filters' in request.GET:
        current_filters = {}
        request.session['module_filters'] = {}
    
    # Get items per page from request or default to 25
    items_per_page = int(request.GET.get('items_per_page', 25))
    
    # Base queryset
    modules = Module.objects.select_related('occupation', 'level').all()
    occupations = Occupation.objects.all().order_by('code')
    levels = Level.objects.all()
    
    # Apply filters
    if current_filters.get('occupation'):
        modules = modules.filter(occupation_id=current_filters['occupation'])
    
    if current_filters.get('level'):
        modules = modules.filter(level_id=current_filters['level'])
    
    if current_filters.get('search'):
        search_term = current_filters['search']
        modules = modules.filter(
            Q(name__icontains=search_term) | 
            Q(code__icontains=search_term)
        )
    
    # Order by occupation, level, then name
    modules = modules.order_by('occupation__name', 'level__name', 'name')
    
    # Add filter display names for template
    filter_names = {}
    if current_filters.get('level'):
        try:
            level_obj = Level.objects.get(id=current_filters['level'])
            filter_names['level_name'] = level_obj.name
        except Level.DoesNotExist:
            pass
    
    if current_filters.get('occupation'):
        try:
            occupation_obj = Occupation.objects.get(id=current_filters['occupation'])
            filter_names['occupation_name'] = f"{occupation_obj.code} - {occupation_obj.name}"
        except Occupation.DoesNotExist:
            pass
    
    # Pagination
    paginator = Paginator(modules, items_per_page)
    page_number = request.GET.get('page', 1)
    
    try:
        modules_page = paginator.page(page_number)
    except:
        modules_page = paginator.page(1)
    
    # Merge filters with display names
    current_filters.update(filter_names)
    
    context = {
        'modules': modules_page,
        'occupations': occupations,
        'levels': levels,
        'filters': current_filters,
        'items_per_page': items_per_page,
    }
    
    return render(request, 'modules/list.html', context)

@login_required
def module_create(request):
    if request.method == 'POST':
        form = ModuleForm(request.POST)
        if form.is_valid():
            module = form.save()
            messages.success(request, f'Module "{module.name}" was created successfully.')
            return redirect('module_list')
    else:
        form = ModuleForm()
    return render(request, 'modules/create.html', {'form': form})

@login_required
def module_detail(request, pk):
    module = get_object_or_404(Module, pk=pk)
    return render(request, 'modules/detail.html', {'module': module})

@login_required
def module_edit(request, pk):
    module = get_object_or_404(Module, pk=pk)
    if request.method == 'POST':
        form = ModuleForm(request.POST, instance=module)
        if form.is_valid():
            module = form.save()
            module_url = reverse('module_detail', args=[module.pk])
            success_message = format_html(
                'Module "{}" was updated successfully. <a href="{}">View Details</a>',
                module.name,
                module_url
            )
            messages.success(request, success_message)
            return redirect('module_list')
    else:
        form = ModuleForm(instance=module)
    return render(request, 'modules/edit.html', {
        'form': form,
        'module': module
    })

@login_required
def module_delete(request, pk):
    module = get_object_or_404(Module, pk=pk)
    if request.method == 'POST':
        module.delete()
        messages.success(request, 'Module deleted successfully.')
        return redirect('module_list')
    return render(request, 'modules/delete.html', {'module': module})


@login_required
def module_bulk_delete(request):
    if request.method == 'POST':
        selected_modules = request.POST.getlist('selected_modules')
        if selected_modules:
            try:
                modules_to_delete = Module.objects.filter(pk__in=selected_modules)
                count = modules_to_delete.count()
                modules_to_delete.delete()
                messages.success(request, f'Successfully deleted {count} module{"s" if count != 1 else ""}.')
            except Exception as e:
                messages.error(request, f'Error deleting modules: {str(e)}')
        else:
            messages.warning(request, 'No modules were selected for deletion.')
    return redirect('module_list')


def paper_list(request):
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # Get current filters from session
    current_filters = request.session.get('paper_filters', {})
    
    # Handle filter actions
    if 'apply_filters' in request.GET:
        current_filters = {
            'search': request.GET.get('search', '').strip(),
            'occupation': request.GET.get('occupation', '').strip(),
            'level': request.GET.get('level', '').strip(),
        }
        request.session['paper_filters'] = current_filters
    elif 'clear_filters' in request.GET:
        current_filters = {}
        request.session['paper_filters'] = {}
    
    # Get items per page from request or default to 25
    items_per_page = int(request.GET.get('items_per_page', 25))
    
    # Base queryset
    papers = Paper.objects.select_related('level', 'occupation', 'module').all()
    occupations = Occupation.objects.all().order_by('code')
    levels = Level.objects.all()
    
    # Apply filters
    if current_filters.get('search'):
        search_term = current_filters['search']
        papers = papers.filter(
            Q(name__icontains=search_term) | 
            Q(code__icontains=search_term)
        )
    
    if current_filters.get('occupation'):
        papers = papers.filter(occupation_id=current_filters['occupation'])
    
    if current_filters.get('level'):
        papers = papers.filter(level_id=current_filters['level'])
    
    # Order by occupation, level, then name
    papers = papers.order_by('occupation__name', 'level__name', 'name')
    
    # Add filter display names for template
    filter_names = {}
    if current_filters.get('level'):
        try:
            level_obj = Level.objects.get(id=current_filters['level'])
            filter_names['level_name'] = level_obj.name
        except Level.DoesNotExist:
            pass
    
    if current_filters.get('occupation'):
        try:
            occupation_obj = Occupation.objects.get(id=current_filters['occupation'])
            filter_names['occupation_name'] = f"{occupation_obj.code} - {occupation_obj.name}"
        except Occupation.DoesNotExist:
            pass
    
    # Pagination
    paginator = Paginator(papers, items_per_page)
    page_number = request.GET.get('page', 1)
    
    try:
        papers_page = paginator.page(page_number)
    except:
        papers_page = paginator.page(1)
    
    # Merge filters with display names
    current_filters.update(filter_names)
    
    context = {
        'papers': papers_page,
        'occupations': occupations,
        'levels': levels,
        'filters': current_filters,
        'items_per_page': items_per_page,
    }
    
    return render(request, 'papers/list.html', context)

def paper_create(request):
    if request.method == 'POST':
        form = PaperForm(request.POST)
        if form.is_valid():
            paper = form.save()
            return redirect('paper_list')
    else:
        form = PaperForm()
    return render(request, 'papers/create.html', {'form': form})

def paper_detail(request, pk):
    paper = get_object_or_404(Paper, pk=pk)
    return render(request, 'papers/detail.html', {'paper': paper})


def paper_edit(request, pk):
    paper = get_object_or_404(Paper, pk=pk)
    if request.method == 'POST':
        form = PaperForm(request.POST, instance=paper)
        if form.is_valid():
            paper = form.save()
            paper_url = reverse('paper_detail', args=[paper.pk])
            success_message = format_html(
                'Paper "{}" was updated successfully. <a href="{}">View Details</a>',
                paper.name,
                paper_url
            )
            messages.success(request, success_message)
            return redirect('paper_list')
    else:
        form = PaperForm(instance=paper)
    # Add current module id as data-initial for JS selection
    if paper.module:
        form.fields['module'].widget.attrs['data-initial'] = str(paper.module.id)
    return render(request, 'papers/edit.html', {'form': form, 'paper': paper})


@login_required
def paper_delete(request, pk):
    paper = get_object_or_404(Paper, pk=pk)
    if request.method == 'POST':
        paper.delete()
        messages.success(request, 'Paper deleted successfully.')
        return redirect('paper_list')
    return render(request, 'papers/delete.html', {'paper': paper})


@login_required
def paper_bulk_delete(request):
    if request.method == 'POST':
        selected_papers = request.POST.getlist('selected_papers')
        if selected_papers:
            try:
                papers_to_delete = Paper.objects.filter(pk__in=selected_papers)
                count = papers_to_delete.count()
                papers_to_delete.delete()
                messages.success(request, f'Successfully deleted {count} paper{"s" if count != 1 else ""}.')
            except Exception as e:
                messages.error(request, f'Error deleting papers: {str(e)}')
        else:
            messages.warning(request, 'No papers were selected for deletion.')
    return redirect('paper_list')


def report_list(request):
    """Main reports dashboard showing available reports"""
    group_names = list(request.user.groups.values_list('name', flat=True))
    return render(request, 'reports/list.html', {'group_names': group_names})

from reportlab.platypus import Image, Paragraph, Spacer, Table, TableStyle, PageBreak, KeepInFrame
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from io import BytesIO
from django.conf import settings
import os
from PIL import Image as PILImage

def _create_photo_cell_content(candidate, styles, photo_width=0.8*inch, photo_height=0.8*inch):
    """Creates a list of flowables for the candidate photo cell, including image, name, and details."""
    cell_elements = []
    photo_image = None

    # Photo Styles
    photo_name_style = ParagraphStyle(
        'PhotoName',
        parent=styles['Normal'],
        fontSize=6,
        alignment=TA_CENTER,
        spaceBefore=1,
        leading=6
    )
    photo_detail_style = ParagraphStyle(
        'PhotoDetail',
        parent=styles['Normal'],
        fontSize=6,
        alignment=TA_CENTER,
        spaceBefore=1,
        leading=6
    )

    # Prefer regno-stamped photo if available
    photo_path = None
    if hasattr(candidate, 'passport_photo_with_regno') and candidate.passport_photo_with_regno and hasattr(candidate.passport_photo_with_regno, 'path') and os.path.exists(candidate.passport_photo_with_regno.path):
        photo_path = candidate.passport_photo_with_regno.path
    elif hasattr(candidate, 'passport_photo') and candidate.passport_photo and hasattr(candidate.passport_photo, 'path') and os.path.exists(candidate.passport_photo.path):
        photo_path = candidate.passport_photo.path
    if photo_path:
        try:
            img = PILImage.open(photo_path)
            scale_factor = min(photo_width / img.width, photo_height / img.height)
            scaled_width = img.width * scale_factor
            scaled_height = img.height * scale_factor
            if img.mode != 'RGB':
                img = img.convert('RGB')
                temp_path = os.path.join(settings.MEDIA_ROOT, f'temp_photo_cell_{candidate.id}.jpg')
                img.save(temp_path, 'JPEG')
                photo_image = Image(temp_path, width=scaled_width, height=scaled_height)
            else:
                photo_image = Image(photo_path, width=scaled_width, height=scaled_height)
        except Exception as e:
            print(f"Error processing photo for cell (ID: {candidate.id}): {e}")
            photo_image = Paragraph("[No Photo]", photo_detail_style)
    else:
        photo_image = Paragraph("[No Photo]", photo_detail_style)
    
    cell_elements.append(photo_image)
    # cell_elements.append(Paragraph(candidate.reg_number.upper(), photo_name_style))
    
    occupation_code = candidate.occupation.code if candidate.occupation else 'N/A'
    reg_category_short = candidate.registration_category.upper() if candidate.registration_category else 'N/A'
    # cell_elements.append(Paragraph(f"{reg_category_short} | {occupation_code}", photo_detail_style))
    
    return cell_elements

@login_required
def generate_album(request):
    logger = logging.getLogger(__name__)
    logger.info("generate_album view started.")

    centers = AssessmentCenter.objects.all()
    occupations = Occupation.objects.all().order_by('code')
    levels = Level.objects.all() # Though not directly used in this version's header/table structure as per screenshot

    if request.method == 'POST':
        try:
            center_id = request.POST.get('center')
            occupation_id = request.POST.get('occupation')
            reg_category_form = request.POST.get('registration_category', '') # Name from form
            level_id = request.POST.get('level') # Keep for filtering logic if needed
            assessment_month_str = request.POST.get('assessment_month')
            assessment_year_str = request.POST.get('assessment_year')

            logger.info(f"POST request received with params: center={center_id}, occupation={occupation_id}, category={reg_category_form}, level={level_id}, month={assessment_month_str}, year={assessment_year_str}")


            if not all([center_id, occupation_id, reg_category_form, assessment_month_str, assessment_year_str]):
                logger.warning("Missing required filter parameters.")
                return HttpResponse("All filter parameters are required.", status=400)

            try:
                assessment_month = int(assessment_month_str)
                assessment_year = int(assessment_year_str)
                center = AssessmentCenter.objects.get(id=center_id)
                occupation = Occupation.objects.get(id=occupation_id)
            except (ValueError, AssessmentCenter.DoesNotExist, Occupation.DoesNotExist) as e:
                logger.error(f"Invalid parameter provided: {e}")
                return HttpResponse(f"Invalid parameter: {e}", status=400)

            # Candidate Querying with branch optimization
            logger.info("Querying candidates...")
            candidate_qs = Candidate.objects.select_related(
                'occupation', 'assessment_center', 'assessment_center_branch'
            ).prefetch_related('nature_of_disability').filter(
                assessment_center=center,
                occupation=occupation,
                registration_category__iexact=reg_category_form, # Use form value for filtering
                assessment_date__year=assessment_year,
                assessment_date__month=assessment_month
            )

            # Optional level filtering (if applicable for the registration category)
            if reg_category_form.lower() in ['formal', 'informal', 'workers pas'] and level_id:
                logger.info(f"Filtering by level_id: {level_id}")
                candidate_qs = candidate_qs.filter(
                    id__in=CandidateLevel.objects.filter(level_id=level_id).values('candidate_id')
                )
            
            # Group candidates by branch for centers with branches
            if center.has_branches:
                logger.info("Center has branches - organizing candidates by branch")
                # Order by branch first, then by reg_number within each branch
                candidate_qs = candidate_qs.order_by(
                    'assessment_center_branch__branch_code',  # Branches first (nulls will be last)
                    'reg_number'
                )
                
                # Group candidates by branch
                from collections import defaultdict
                branch_candidates = defaultdict(list)
                
                for candidate in candidate_qs:
                    if candidate.assessment_center_branch:
                        branch_key = candidate.assessment_center_branch.branch_code
                        branch_name = f"{candidate.assessment_center_branch.branch_code} - {candidate.assessment_center_branch.village.name}"
                    else:
                        branch_key = "main_center"
                        branch_name = "Main Center"
                    
                    branch_candidates[branch_key].append({
                        'candidate': candidate,
                        'branch_name': branch_name,
                        'branch_obj': candidate.assessment_center_branch
                    })
                
                total_candidates = sum(len(candidates) for candidates in branch_candidates.values())
                logger.info(f"Found {total_candidates} candidates across {len(branch_candidates)} branches/sections.")
                
                if not branch_candidates:
                    logger.warning("No candidates found matching the criteria.")
                    return HttpResponse("No candidates found matching the criteria.", status=404)
                    
            else:
                # No branches - use existing logic
                final_candidates = list(candidate_qs.order_by('reg_number'))
                logger.info(f"Found {len(final_candidates)} candidates.")
                if not final_candidates:
                    logger.warning("No candidates found matching the criteria.")
                    return HttpResponse("No candidates found matching the criteria.", status=404)
                    
                # Convert to branch format for unified processing
                branch_candidates = {
                    "main_center": [{
                        'candidate': candidate,
                        'branch_name': center.center_name,
                        'branch_obj': None
                    } for candidate in final_candidates]
                }

            # PDF Generation with branch support
            logger.info("Starting PDF generation...")
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=landscape(letter),
                                    title="UVTAB",
                                    rightMargin=0.4*inch, leftMargin=0.4*inch,
                                    topMargin=0.3*inch, bottomMargin=0.3*inch)
            elements = []
            styles = getSampleStyleSheet()

            # Define Styles
            contact_style = ParagraphStyle('ContactInfo', parent=styles['Normal'], fontSize=9, leading=11)
            board_title_style = ParagraphStyle('BoardTitle', parent=styles['h1'], fontSize=14, alignment=TA_CENTER, spaceBefore=6, spaceAfter=6, textColor=colors.HexColor('#000000'))
            report_title_style = ParagraphStyle('ReportTitle', parent=styles['h2'], fontSize=12, alignment=TA_CENTER, spaceAfter=4)
            center_info_style = ParagraphStyle('CenterInfo', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, spaceAfter=10)
            details_label_style = ParagraphStyle('DetailsLabel', parent=styles['Normal'], fontSize=9, alignment=TA_LEFT, spaceAfter=2)
            branch_title_style = ParagraphStyle('BranchTitle', parent=styles['h3'], fontSize=11, alignment=TA_CENTER, spaceBefore=10, spaceAfter=8, textColor=colors.HexColor('#2E5984'))
            
            # Helper function to create header section
            def create_header_section(branch_name=None):
                header_elements = []
                
                # Logo setup
                logo_path = None
                possible_paths = [
                    os.path.join(settings.BASE_DIR, 'eims', 'static', 'images', 'uvtab logo.png'),
                    os.path.join(settings.BASE_DIR, 'static', 'images', 'uvtab logo.png'),
                    os.path.join(settings.BASE_DIR, 'emis', 'static', 'images', 'uvtab logo.png'),
                    os.path.join(settings.STATIC_ROOT or '', 'images', 'uvtab logo.png')
                ]
                for path in possible_paths:
                    if path and os.path.exists(path):
                        logo_path = path
                        break
                
                logo_image = Image(logo_path, width=1*inch, height=1*inch) if logo_path else Paragraph(" ", styles['Normal'])
                
                # Header table
                header_table_data = [
                    [Paragraph("P.O.Box 1499<br/>Email: info@uvtab.go.ug", contact_style), 
                     logo_image, 
                     Paragraph("Tel: 256 414 289786", contact_style)]
                ]
                header_table = Table(header_table_data, colWidths=[3*inch, 3*inch, 3*inch])
                header_table.setStyle(TableStyle([
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                    ('ALIGN', (0,0), (0,0), 'LEFT'),
                    ('ALIGN', (1,0), (1,0), 'CENTER'),
                    ('ALIGN', (2,0), (2,0), 'RIGHT'),
                ]))
                header_elements.append(header_table)
                header_elements.append(Spacer(1, 0.1*inch))
                
                # Board title
                header_elements.append(Paragraph("UGANDA VOCATIONAL AND TECHNICAL ASSESSMENT BOARD", board_title_style))
                
                # Assessment period
                month_names = ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December']
                assessment_period_str = f"{month_names[assessment_month-1]} {assessment_year}"
                header_elements.append(Paragraph(f"Registered Candidates for {assessment_period_str} Assessment", report_title_style))
                
                # Center info with branch
                if branch_name and branch_name != center.center_name:
                    center_info_text = f"Assessment Center: {center.center_number} - {center.center_name}<br/>Branch: {branch_name}"
                else:
                    center_info_text = f"Assessment Center: {center.center_number} - {center.center_name}"
                header_elements.append(Paragraph(center_info_text, center_info_style))
                
                # Occupation details
                level_text = f" - Level: {Level.objects.get(id=level_id).name.upper()}" if reg_category_form.lower() in ['formal', 'informal'] and level_id else ''
                occupation_details = f"Occupation Name: {occupation.name.upper()}<br/>Occupation Code: {occupation.code.upper()}<br/>Registration Category: {reg_category_form.upper()}{level_text}"
                header_elements.append(Paragraph(occupation_details, details_label_style))
                header_elements.append(Spacer(1, 0.2*inch))
                
                return header_elements

            # Helper function to create candidate table for a branch
            def create_candidate_table(candidates_data, start_sn=1):
                table_header_style = ParagraphStyle('TableHeader', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, alignment=TA_CENTER, textColor=colors.white)
                table_cell_style = ParagraphStyle('TableCell', parent=styles['Normal'], fontSize=8, alignment=TA_LEFT, leading=10)
                table_cell_center_style = ParagraphStyle('TableCellCenter', parent=table_cell_style, alignment=TA_CENTER)
                
                data = []
                # Table Headers
                header_row = [Paragraph(h, table_header_style) for h in ['S/N', 'PHOTO', 'REG NO.', 'FULL NAME', 'OCCUPATION', 'REG TYPE', 'SPECIAL NEEDS', 'SIGNATURE']]
                data.append(header_row)
                
                for i, cand_data in enumerate(candidates_data):
                    cand = cand_data['candidate']
                    logger.debug(f"Processing candidate {start_sn + i}: {cand.reg_number}")
                    photo_cell_flowables = _create_photo_cell_content(cand, styles)
                    
                    # Generate Special Needs text
                    if cand.disability:
                        # Get nature of disability names
                        nature_names = list(cand.nature_of_disability.values_list('name', flat=True))
                        nature_text = ', '.join(nature_names) if nature_names else 'Not specified'
                        
                        # Add disability specification if available
                        specification = cand.disability_specification or ''
                        if specification:
                            special_needs_text = f"Yes ({nature_text} - {specification})"
                        else:
                            special_needs_text = f"Yes ({nature_text})"
                    else:
                        special_needs_text = "No"
                    
                    row = [
                        Paragraph(str(start_sn + i), table_cell_center_style),
                        photo_cell_flowables, # This is a list of flowables
                        Paragraph(cand.reg_number or 'N/A', table_cell_style),
                        Paragraph(cand.full_name.upper(), table_cell_style),
                        Paragraph(cand.occupation.name.upper() if cand.occupation else 'N/A', table_cell_style),
                        Paragraph(cand.registration_category.upper() if cand.registration_category else 'N/A', table_cell_style),
                        Paragraph(special_needs_text, table_cell_style),
                        Paragraph('', table_cell_style) # Empty for signature
                    ]
                    data.append(row)
                
                # Column widths (adjust as needed, total should be around 10.2 inch for landscape letter with 0.4 margins)
                col_widths = [0.4*inch, 1.2*inch, 1.4*inch, 2.2*inch, 1.2*inch, 1.0*inch, 1.8*inch, 1.2*inch]
                
                candidate_table = Table(data, colWidths=col_widths, repeatRows=1)
                candidate_table.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4F81BD')), # Header background
                    ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
                    ('ALIGN', (0,0), (-1,0), 'CENTER'), # Header text alignment
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), # All cells middle aligned vertically
                    ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0,0), (-1,0), 8),
                    ('BOTTOMPADDING', (0,0), (-1,0), 6),
                    ('TOPPADDING', (0,0), (-1,0), 6),
                    
                    ('GRID', (0,0), (-1,-1), 0.5, colors.black),
                    ('BOX', (0,0), (-1,-1), 1, colors.black),
                    
                    # Data row styling
                    ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
                    ('FONTSIZE', (0,1), (-1,-1), 8),
                    ('ALIGN', (0,1), (0,-1), 'CENTER'), # S/N centered
                    ('ALIGN', (1,1), (1,-1), 'CENTER'), # Photo centered horizontally
                    ('ALIGN', (2,1), (2,-1), 'LEFT'), # Reg No left
                    ('ALIGN', (3,1), (3,-1), 'LEFT'), # Full Name left
                    ('ALIGN', (4,1), (4,-1), 'LEFT'), # Occupation left
                    ('ALIGN', (5,1), (5,-1), 'CENTER'), # Reg Type center
                    ('ALIGN', (6,1), (6,-1), 'LEFT'), # Special Needs left
                    ('ALIGN', (7,1), (7,-1), 'CENTER'), # Signature center
                    ('TOPPADDING', (0,1), (-1,-1), 2), # Reduced padding
                    ('BOTTOMPADDING', (0,1), (-1,-1), 2), # Reduced padding
                ]))
                
                return candidate_table
            
            # Generate branch-organized content
            logger.info("Generating branch-organized PDF content.")
            current_sn = 1
            branch_keys = sorted(branch_candidates.keys(), key=lambda x: (x != 'main_center', x))  # Main center first, then branches alphabetically
            
            for branch_index, branch_key in enumerate(branch_keys):
                candidates_data = branch_candidates[branch_key]
                branch_name = candidates_data[0]['branch_name'] if candidates_data else "Unknown Branch"
                
                logger.info(f"Processing branch: {branch_name} ({len(candidates_data)} candidates)")
                
                # Add page break for subsequent branches (not for the first one)
                if branch_index > 0:
                    elements.append(PageBreak())
                
                # Add header section for this branch
                header_elements = create_header_section(branch_name)
                elements.extend(header_elements)
                
                # Add candidate table for this branch
                if candidates_data:
                    candidate_table = create_candidate_table(candidates_data, current_sn)
                    elements.append(candidate_table)
                    current_sn += len(candidates_data)
                else:
                    elements.append(Paragraph("No candidates found for this branch.", styles['Normal']))

            # --- First pass to count total pages ---
            logger.info("Starting first pass for page count.")
            # Use a deep copy of elements for the first pass to avoid consuming them
            first_pass_elements = copy.deepcopy(elements)
            count_buffer = BytesIO() # Temporary buffer for counting
            # Ensure all SimpleDocTemplate parameters match the final document for accurate page count
            count_doc = SimpleDocTemplate(count_buffer, pagesize=landscape(letter),
                                          rightMargin=0.4*inch, leftMargin=0.4*inch,
                                          topMargin=0.3*inch, bottomMargin=0.3*inch)
            
            def _count_pages_callback(canvas, doc): # Minimal callback for the first pass
                pass # We only care about the page count

            try:
                count_doc.build(first_pass_elements, onFirstPage=_count_pages_callback, onLaterPages=_count_pages_callback)
                total_pages = count_doc.page
                logger.info(f"Page count successful. Total pages: {total_pages}")
            except Exception as e:
                # Handle potential errors during the first pass, though less likely with a simple callback
                logger.error(f"Error during page count pass: {e}", exc_info=True)
                # Fallback to simple page numbering if counting fails
                total_pages = 0 # Indicates an issue, or use a flag

            # --- End of first pass ---

            # Helper function to add page numbers (includes total_pages if available)
            def _add_page_numbers(canvas, doc, total_pages_count):
                canvas.saveState()
                canvas.setFont('Helvetica', 9)
                if total_pages_count > 0:
                    page_number_text = f"Page {doc.page} of {total_pages_count}"
                else:
                    page_number_text = f"Page {doc.page}" # Fallback if total_pages is not available
                
                page_width = doc.pagesize[0] # doc.pagesize[0] is width for landscape
                canvas.drawCentredString(page_width / 2.0, 0.2 * inch, page_number_text)
                canvas.restoreState()

            # Build PDF (Second pass - actual PDF generation with 'Page X of Y')
            # The original 'doc', 'buffer', and 'elements' are used for the final output.
            logger.info("Starting second pass for final PDF build.")
            doc.build(elements, 
                      onFirstPage=lambda c, d: _add_page_numbers(c, d, total_pages), 
                      onLaterPages=lambda c, d: _add_page_numbers(c, d, total_pages))
            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="candidate_album_{center.center_number}_{occupation.code}_{assessment_year}_{assessment_month}.pdf"'
            logger.info("PDF generated successfully. Returning response.")
            return response
        except Exception as e:
            logger.critical(f"Unhandled exception in generate_album view: {e}", exc_info=True)
            import traceback
            traceback.print_exc()
            return HttpResponse(f"Error generating PDF: {e}", status=500)

    # GET request or if form not submitted properly
    logger.info("GET request received, rendering form.")
    return render(request, 'reports/albums.html', {
        'centers': centers,
        'occupations': occupations,
        'levels': levels,
        'form_action': reverse('generate_album')
    })

def add_module(request, level_id):
    level = get_object_or_404(Level, id=level_id)
    occupation = level.occupation  # If Level has ForeignKey to Occupation

    if request.method == 'POST':
        form = ModuleForm(request.POST)
        if form.is_valid():
            module = form.save(commit=False)
            module.level = level
            module.occupation = occupation
            module.save()
            return redirect('occupation_detail', pk=occupation.pk)
    else:
        form = ModuleForm()

    return render(request, 'modules/create.html', {'form': form, 'level': level})


# Add Paper View
def add_paper(request, level_id):
    level = get_object_or_404(Level, id=level_id)
    occupation = level.occupation  # If Level has ForeignKey to Occupation

    if request.method == 'POST':
        form = PaperForm(request.POST)
        if form.is_valid():
            paper = form.save(commit=False)
            paper.level = level
            paper.occupation = occupation
            paper.save()
            return redirect('occupation_detail', pk=occupation.pk)
    else:
        form = PaperForm()

    return render(request, 'papers/create.html', {'form': form, 'level': level})

from django.core.paginator import Paginator

@login_required
def candidate_list(request):
    # Get current filters from session or initialize
    current_filters = request.session.get('candidate_filters', {})
    
    # Handle URL-based filters (for statistics clickable links)
    url_filters = {}
    if request.GET:
        # Extract filters from URL parameters
        url_filters = {
            'reg_number': request.GET.get('reg_number', '').strip(),
            'search': request.GET.get('search', '').strip(),
            'occupation': request.GET.get('occupation', '').strip(),
            'registration_category': request.GET.get('registration_category', '').strip(),
            'assessment_center': request.GET.get('assessment_center', '').strip(),
            'sector': request.GET.get('sector', '').strip(),
            'gender': request.GET.get('gender', '').strip(),
            'disability': request.GET.get('disability', '').strip(),
            'is_refugee': request.GET.get('is_refugee', '').strip(),
            'assessment_year': request.GET.get('assessment_year', '').strip(),
            'assessment_month': request.GET.get('assessment_month', '').strip(),
        }
        # Remove empty values
        url_filters = {k: v for k, v in url_filters.items() if v}
        
        # If URL filters exist, use them and update session
        if url_filters:
            current_filters.update(url_filters)
            request.session['candidate_filters'] = current_filters

    # If form is submitted, update filters in session
    if 'apply_filters' in request.GET:
        current_filters = {
            'reg_number': request.GET.get('reg_number', '').strip(),
            'search': request.GET.get('search', '').strip(),
            'occupation': request.GET.get('occupation', '').strip(),
            'registration_category': request.GET.get('registration_category', '').strip(),
            'assessment_center': request.GET.get('assessment_center', '').strip(),
            'sector': request.GET.get('sector', '').strip(),
            'assessment_month': request.GET.get('assessment_month', '').strip(),
            'assessment_year': request.GET.get('assessment_year', '').strip(),
            'disability': request.GET.get('disability', '').strip(),
            'is_refugee': request.GET.get('is_refugee', '').strip(),
        }
        
        # IMPORTANT: Preserve assessment series context from existing session or URL
        if current_filters.get('assessment_year'):
            current_filters['assessment_year'] = current_filters.get('assessment_year')
        if current_filters.get('assessment_month'):
            current_filters['assessment_month'] = current_filters.get('assessment_month')
        
        # Also check if assessment series parameters are in the current request
        if request.GET.get('assessment_year'):
            current_filters['assessment_year'] = request.GET.get('assessment_year')
        if request.GET.get('assessment_month'):
            current_filters['assessment_month'] = request.GET.get('assessment_month')
            
        # Remove empty values so they don't clutter the URL
        current_filters = {k: v for k, v in current_filters.items() if v}
        request.session['candidate_filters'] = current_filters
        # Redirect to a clean URL. The view will then use session filters.
        return redirect('candidate_list')

    # If reset is requested, clear filters and redirect
    if 'reset_filters' in request.GET:
        if 'candidate_filters' in request.session:
            del request.session['candidate_filters']
        return redirect('candidate_list')

    candidates = Candidate.objects.select_related('occupation', 'occupation__sector', 'assessment_center').order_by('-created_at')
    
    # Restrict for Center Representatives
    if request.user.groups.filter(name='CenterRep').exists():
        from .models import CenterRepresentative
        try:
            center_rep = CenterRepresentative.objects.get(user=request.user)
            candidates = candidates.filter(assessment_center=center_rep.center)
        except CenterRepresentative.DoesNotExist:
            candidates = candidates.none()

    # Enhanced filtering logic from session filters
    if current_filters.get('reg_number'):
        candidates = candidates.filter(reg_number__icontains=current_filters.get('reg_number'))
    if current_filters.get('search'):
        candidates = candidates.filter(full_name__icontains=current_filters.get('search'))
    if current_filters.get('occupation'):
        candidates = candidates.filter(occupation_id=current_filters.get('occupation'))
    if current_filters.get('registration_category'):
        candidates = candidates.filter(registration_category=current_filters.get('registration_category'))
    if current_filters.get('assessment_center'):
        candidates = candidates.filter(assessment_center_id=current_filters.get('assessment_center'))
    if current_filters.get('sector'):
        candidates = candidates.filter(occupation__sector_id=current_filters.get('sector'))
    
    # NEW: Additional filters for statistics integration
    if current_filters.get('gender'):
        candidates = candidates.filter(gender=current_filters.get('gender'))
    if current_filters.get('disability'):
        # Convert string to boolean
        disability_filter = current_filters.get('disability').lower() == 'true'
        candidates = candidates.filter(disability=disability_filter)
    
    if current_filters.get('is_refugee'):
        # Convert string to boolean
        refugee_filter = current_filters.get('is_refugee').lower() == 'true'
        candidates = candidates.filter(is_refugee=refugee_filter)
    
    # Assessment date filtering by month and year
    if current_filters.get('assessment_year'):
        try:
            year = int(current_filters.get('assessment_year'))
            candidates = candidates.filter(assessment_date__year=year)
        except (ValueError, TypeError):
            pass
    
    if current_filters.get('assessment_month'):
        try:
            month = int(current_filters.get('assessment_month'))
            candidates = candidates.filter(assessment_date__month=month)
        except (ValueError, TypeError):
            pass

    def build_candidate_filter_url(base_filters=None, **additional_filters):
        """Build a URL for filtering candidates with given parameters"""
        from django.urls import reverse
        import urllib.parse
        
        filters = base_filters.copy() if base_filters else {}
        filters.update(additional_filters)
    
        # Remove empty values
        filters = {k: v for k, v in filters.items() if v}
    
        if filters:
            query_string = urllib.parse.urlencode(filters)
            return f"{reverse('candidate_list')}?{query_string}"
        else:
            return reverse('candidate_list')

    
    from .models import Occupation, AssessmentCenter, Sector
    occupations = Occupation.objects.select_related('sector').all().order_by('code')
    centers = AssessmentCenter.objects.all()
    sectors = Sector.objects.all().order_by('name')
    
    # Get available assessment years for the dropdown
    from django.db.models import Q
    import datetime
    current_year = datetime.datetime.now().year
    # Get years from existing assessment dates, plus current and next year
    assessment_years = list(range(current_year - 2, current_year + 3))
    assessment_years.sort(reverse=True)

    # Pagination: 100 per page
    paginator = Paginator(candidates, 100)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    import urllib
    filter_params = urllib.parse.urlencode(current_filters)

    from .models import NatureOfDisability
    return render(request, 'candidates/list.html', {
        'candidates': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'total_candidates': paginator.count,
        'occupations': occupations,
        'centers': centers,
        'sectors': sectors,
        'filters': current_filters,
        'filter_params': filter_params,
        'nature_of_disabilities': NatureOfDisability.objects.all(),
        'assessment_years': assessment_years,
    })


from django.template.loader import render_to_string

from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from .models import Candidate, Result, Module, CandidateLevel, CandidateModule, OccupationLevel
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO
import os
from django.conf import settings

import qrcode

@login_required
def generate_transcript(request, id):
    """
    Generate a PDF transcript for a level-module-based candidate, following strict eligibility logic. Adds transcript serial and QR code.
    """
    from .models import CandidateLevel
    candidate = Candidate.objects.select_related('occupation', 'assessment_center').get(id=id)
    reg_cat = getattr(candidate, 'registration_category', '').lower()

    # --- Transcript Serial Number ---
    org_code = candidate.assessment_center.center_number if candidate.assessment_center and hasattr(candidate.assessment_center, 'center_number') else "ORG"
    occ_code = candidate.occupation.code if candidate.occupation and hasattr(candidate.occupation, 'code') else "XX"
    serial_number = f"{org_code}/TR {occ_code}{str(candidate.id).zfill(6)}"

    # --- QR Code Data ---
    level_name = getattr(getattr(candidate, 'level', None), 'name', '')
    if not level_name:
        cl = CandidateLevel.objects.filter(candidate=candidate).first()
        if cl and cl.level:
            level_name = cl.level.name
    qr_data = {
        "name": candidate.full_name,
        "regno": candidate.reg_number,
        "occupation": candidate.occupation.name if candidate.occupation else '',
        "level": level_name,
    }
    import json
    qr_str = json.dumps(qr_data, ensure_ascii=False)

    # --- QR Code Generation ---
    import qrcode
    from reportlab.platypus import Image as RLImage
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(qr_str)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white")
    qr_buffer = BytesIO()
    qr_img.save(qr_buffer, format='PNG')
    qr_buffer.seek(0)
    qr_rl_img = RLImage(qr_buffer, width=1.0*inch, height=1.0*inch)
    qr_rl_img.hAlign = 'LEFT'

    # Only allow for module-based levels
    occ_level = None
    level = getattr(candidate, 'level', None)
    if not level:
        cl = CandidateLevel.objects.filter(candidate=candidate).first()
        if cl:
            level = cl.level
    # Only allow for module-based levels
    occ_level = None
    level = getattr(candidate, 'level', None)
    if not level:
        cl = CandidateLevel.objects.filter(candidate=candidate).first()
        if cl:
            level = cl.level
    if candidate.occupation and level:
        occ_level = OccupationLevel.objects.filter(occupation=candidate.occupation, level=level).first()
    # PDF generation - both pages portrait orientation with landscape content layout
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter,
                            title="Transcript",
                            rightMargin=0.4*inch, leftMargin=0.4*inch,
                            topMargin=0.3*inch, bottomMargin=0.3*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    normal = styles['Normal']
    normal.fontSize = 10
    normal.leading = 12
    normal.spaceAfter = 0
    normal.spaceBefore = 0
    bold = ParagraphStyle('Bold', parent=normal, fontName='Helvetica-Bold')
    center = ParagraphStyle('Center', parent=normal, alignment=TA_CENTER)
    red_center = ParagraphStyle('RedCenter', parent=center, textColor=colors.red, fontSize=12, spaceAfter=6)

    # 0. Leave space for pre-printed header (about 1.5 inches)
    elements.append(Spacer(1, 1.5*inch))

    # 1. Top Heading (now lower, logic fixed)
    # Always define level_name safely
    level_name = getattr(level, 'name', None)
    if not level_name:
        level_name = str(level) if level else ''
    transcript_heading = ""
    if reg_cat == 'modular':
        transcript_heading = "Modular Transcript"
    elif level_name and level_name.lower() != 'none':
        # Avoid double 'Level' in heading
        if level_name.lower().startswith('level'):
            transcript_heading = f"{level_name} Transcript"
        else:
            transcript_heading = f"Level {level_name} Transcript"
    else:
        transcript_heading = "Transcript"

    # --- Top Row: Transcript Heading (left), QR code, Serial, and Photo (right, stacked) ---
    heading_para = Paragraph(transcript_heading, red_center)
    # Prepare QR code and serial number
    serial_style = ParagraphStyle('SerialSmall', parent=bold, fontSize=8, textColor=colors.blue, alignment=TA_RIGHT)
    serial_para = Paragraph(serial_number, serial_style)
    qr_rl_img.drawWidth = 0.65*inch
    qr_rl_img.drawHeight = 0.65*inch
    qr_rl_img.hAlign = 'RIGHT'

    # Stack QR and serial only in right col (photo will be shown with biodata)
    right_col = [qr_rl_img, Spacer(1, 0.02*inch), serial_para]
    top_row = Table([
        [heading_para, right_col]
    ], colWidths=[4.5*inch, 1.5*inch], hAlign='RIGHT')
    top_row.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (0,0), 'LEFT'),
        ('ALIGN', (1,0), (1,0), 'RIGHT'),
        ('LEFTPADDING', (0,0), (-1,-1), 0),
        ('RIGHTPADDING', (0,0), (-1,-1), 0),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(Spacer(1, 0.18*inch))
    elements.append(top_row)
    elements.append(Spacer(1, 0.035*inch))

    # Use verified results biodata layout - clean and professional
    from django_countries import countries
    nationality = candidate.nationality if getattr(candidate, 'nationality', None) else ""
    if nationality and len(nationality) == 2 and nationality.isupper():
        code_to_name = dict(countries)
        nationality = code_to_name.get(nationality, nationality)
    
    # Format dates
    from datetime import datetime
    birthdate = candidate.date_of_birth.strftime('%d %b, %Y') if getattr(candidate, 'date_of_birth', None) else ""
    print_date = datetime.now().strftime('%d-%b-%Y')
    
    # Get level information for Formal candidates
    candidate_level = None
    reg_cat = getattr(candidate, 'registration_category', '').lower()
    if reg_cat == 'formal':
        from .models import CandidateLevel
        candidate_levels = CandidateLevel.objects.filter(candidate=candidate).select_related('level')
        if candidate_levels.exists():
            candidate_level = candidate_levels.first().level
    
    # Bio data layout (same as verified results)
    bio_left_content = [
        f"<b>NAME:</b> {candidate.full_name}",
        f"<b>REG NO:</b> {candidate.reg_number}",
        f"<b>GENDER:</b> {candidate.get_gender_display() if hasattr(candidate, 'get_gender_display') else candidate.gender}",
        f"<b>CENTER NAME:</b> {candidate.assessment_center.center_name if candidate.assessment_center else ''}",
        f"<b>REGISTRATION CATEGORY:</b> {get_registration_category_display(candidate.registration_category) if hasattr(candidate, 'registration_category') else ''}",
        f"<b>OCCUPATION:</b> {candidate.occupation.name if candidate.occupation else ''}"
    ]
    
    # Add level information for Formal candidates
    if reg_cat == 'formal' and candidate_level:
        bio_left_content.append(f"<b>LEVEL:</b> {candidate_level.name}")
    elif reg_cat == 'formal':
        bio_left_content.append(f"<b>LEVEL:</b> N/A")
    
    bio_right_content = [
        f"<b>NATIONALITY:</b> {nationality}",
        f"<b>BIRTHDATE:</b> {birthdate}",
        f"<b>PRINTDATE:</b> {print_date}",
        "",
        "",
        ""
    ]
    
    # Create bio paragraphs with smaller font
    bio_left = Paragraph("<br/>".join(bio_left_content), 
                        ParagraphStyle('BioLeft', parent=normal, fontSize=9, leading=11))
    bio_right = Paragraph("<br/>".join(bio_right_content), 
                         ParagraphStyle('BioRight', parent=normal, fontSize=9, leading=11))
    
    # Prepare photo for biodata section
    photo = None
    if getattr(candidate, 'passport_photo', None) and candidate.passport_photo.name:
        try:
            photo_path = candidate.passport_photo.path
            photo = RLImage(photo_path, width=1.0*inch, height=1.2*inch)
            photo.hAlign = 'LEFT'
        except Exception:
            photo = None
    
    # Layout with photo on left side (same as verified results)
    if photo:
        # Photo and bio data in same row
        bio_table = Table([
            [photo, bio_left, bio_right]
        ], colWidths=[1.2*inch, 3.0*inch, 2.8*inch])
        bio_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (1,0), (1,0), 'LEFT'),
            ('ALIGN', (2,0), (2,0), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
    else:
        # No photo - just bio data in two columns
        bio_table = Table([
            [bio_left, bio_right]
        ], colWidths=[3.5*inch, 3.5*inch])
        bio_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (1,0), (1,0), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
    
    elements.append(bio_table)
    elements.append(Spacer(1, 0.2*inch))


    # 4. Assessment Result
    elements.append(Paragraph("<b>ASSESSMENT RESULT</b>", center))
    elements.append(Spacer(1, 0.025*inch))

    reg_cat = getattr(candidate, 'registration_category', '').lower()
    transcript_heading = None
    if reg_cat == 'modular':
        from .models import CandidateModule, Result
        candidate_modules = CandidateModule.objects.filter(candidate=candidate)
        result_headers = [Paragraph("Module", bold), Paragraph("Grade", bold)]
        result_table_data = [result_headers]
        all_passed = True
        results = Result.objects.filter(candidate=candidate).select_related('level', 'module', 'paper', 'assessment_series')
        series_level_enrollments = {}
        for candidate_module in candidate_modules:
            module = candidate_module.module
            result = next((r for r in results if r.comment == 'Successful'), None)
            if not result:
                result = results.first() if results else None
            grade = result.grade if result else ''
            if not result or result.comment != 'Successful':
                all_passed = False
            result_table_data.append([module.name, grade])
        result_table = Table(result_table_data, hAlign='CENTER', colWidths=[3.5*inch, 1.5*inch])
        result_table.setStyle(TableStyle([
            ('BOX', (0,0), (-1,-1), 1, colors.black),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('LEFTPADDING', (0,0), (-1,-1), 3),
            ('RIGHTPADDING', (0,0), (-1,-1), 3),
            ('TOPPADDING', (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ]))
        elements.append(result_table)
        elements.append(Spacer(1, 0.04*inch))
        overall_comp = "Successful" if all_passed else "Unsuccessful"
        elements.append(Paragraph(f"Overall Competence: {overall_comp}", bold))
        elements.append(Spacer(1, 0.03*inch))
        elements.append(Spacer(1, 0.015*inch))




    else:
        # Level-based: use the Level name (e.g., 'Level 1 Transcript')
        level_name = ''
        if level:
            level_name = str(level)
        elif occ_level and occ_level.level:
            level_name = str(occ_level.level)
        if level_name:
            transcript_heading = f'{level_name} Transcript'
        else:
            transcript_heading = 'Level Transcript'
        from .models import Paper, Result
        papers = Paper.objects.filter(level=level, occupation=candidate.occupation)
        is_module_based = occ_level and occ_level.structure_type == 'modules'
        is_paper_based = occ_level and occ_level.structure_type == 'papers'
        level_number = None
        if hasattr(level, 'name') and level.name:
            try:
                level_number = int(''.join(filter(str.isdigit, str(level.name))))
            except Exception:
                level_number = None
        if is_module_based:
            result_headers = [Paragraph("Paper", bold), Paragraph("Grade", bold)]
            result_table_data = [result_headers]
            theory_paper = papers.filter(grade_type='theory').first()
            practical_paper = papers.filter(grade_type='practical').first()
            theory_result = Result.objects.filter(candidate=candidate, paper=theory_paper).order_by('-assessment_date').first() if theory_paper else None
            practical_result = Result.objects.filter(candidate=candidate, paper=practical_paper).order_by('-assessment_date').first() if practical_paper else None
            eligible = False
            overall_comp = ''
            if level_number and level_number >= 3:
                if practical_result and practical_result.comment == 'Successful' and theory_result and theory_result.comment == 'Successful':
                    eligible = True
                    overall_comp = "Successful"
            else:
                if practical_result and practical_result.comment == 'Successful':
                    eligible = True
                    if not theory_result or theory_result.comment != 'Successful':
                        overall_comp = "Successful in Practical Only"
                    else:
                        overall_comp = "Successful"
            # Improved eligibility and display logic for level module-based candidates
            # Level 1/2: transcript issued if Practical is 'Successful' (Theory can be failed/CTR)
            # Level 3/4: both must be 'Successful'
            # If Practical is 'CTR', block transcript
            # Always show a row for each passed component as appropriate
            practical_status = practical_result.comment if practical_result else None
            theory_status = theory_result.comment if theory_result else None

            eligible = False
            overall_comp = ''
            show_theory = False
            show_practical = False

            if level_number and level_number >= 3:
                # Level 3/4: both must be 'Successful'
                if practical_status == 'Successful' and theory_status == 'Successful':
                    eligible = True
                    show_theory = True
                    show_practical = True
                    overall_comp = 'Successful'
            else:
                # Level 1/2: Practical must be 'Successful', Theory can be anything
                if practical_status == 'Successful':
                    eligible = True
                    show_practical = True
                    if theory_status == 'Successful':
                        show_theory = True
                        overall_comp = 'Successful'
                    else:
                        overall_comp = 'Successful in Practical Only'
                elif practical_status == 'CTR':
                    eligible = False
                else:
                    eligible = False

            if not eligible:
                return HttpResponse("Candidate doesn't qualify.")

            if show_theory:
                result_table_data.append([
                    Paragraph(theory_paper.code + " - " + theory_paper.name, normal),
                    Paragraph(theory_result.grade, normal)
                ])
            if show_practical:
                result_table_data.append([
                    Paragraph(practical_paper.code + " - " + practical_paper.name, normal),
                    Paragraph(practical_result.grade, normal)
                ])

            col_widths = [3.5*inch, 1.0*inch]
            result_table = Table(result_table_data, colWidths=col_widths, hAlign='LEFT')
            result_table.setStyle(TableStyle([
                ('BOX', (0,0), (-1,-1), 1, colors.black),
                ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('FONTSIZE', (0,0), (-1,-1), 8),
                ('LEFTPADDING', (0,0), (-1,-1), 3),
                ('RIGHTPADDING', (0,0), (-1,-1), 3),
                ('TOPPADDING', (0,0), (-1,-1), 2),
                ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ]))
            elements.append(result_table)
            elements.append(Spacer(1, 0.1*inch))
            elements.append(Paragraph(f"Overall Competence: {overall_comp}", bold))
            elements.append(Spacer(1, 0.18*inch))  # Add extra space before grading table
        elif is_paper_based:
            result_headers = [Paragraph("Paper", bold), Paragraph("Grade", bold)]
            result_table_data = [result_headers]
            # For each paper, find the latest successful result (if any), else latest result
            from collections import defaultdict
            paper_results = defaultdict(list)
            for r in Result.objects.filter(candidate=candidate, paper__level=level, paper__occupation=candidate.occupation).order_by('-assessment_date'):
                if r.paper:
                    paper_results[r.paper.id].append(r)
            practical_results = []
            theory_results = []
            for paper_id, results in paper_results.items():
                paper = results[0].paper
                latest_success = next((r for r in results if r.comment == 'Successful'), None)
                result = latest_success if latest_success else results[0]
                if paper.grade_type == 'practical':
                    practical_results.append(result)
                elif paper.grade_type == 'theory':
                    theory_results.append(result)
            all_practical_successful = all(r.comment == 'Successful' for r in practical_results) and practical_results
            all_theory_successful = all(r.comment == 'Successful' for r in theory_results) and theory_results
            eligible = False
            overall_comp = ''
            if level_number and level_number >= 3:
                if all_practical_successful and all_theory_successful:
                    eligible = True
                    overall_comp = "Successful"
            else:
                if all_practical_successful:
                    eligible = True
                    if not all_theory_successful:
                        overall_comp = "Successful in Practical Only"
                    else:
                        overall_comp = "Successful"
            if not eligible:
                return HttpResponse("Candidate doesn't qualify.")
            for paper_id, results in paper_results.items():
                paper = results[0].paper
                latest_success = next((r for r in results if r.comment == 'Successful'), None)
                result = latest_success if latest_success else results[0]
                result_table_data.append([
                    Paragraph(f"{paper.code} - {paper.name}", normal),
                    Paragraph(result.grade, normal)
                ])
            col_widths = [3.5*inch, 1.0*inch]
            result_table = Table(result_table_data, colWidths=col_widths, hAlign='LEFT')
            result_table.setStyle(TableStyle([
                ('BOX', (0,0), (-1,-1), 1, colors.black),
                ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('FONTSIZE', (0,0), (-1,-1), 8),
                ('LEFTPADDING', (0,0), (-1,-1), 3),
                ('RIGHTPADDING', (0,0), (-1,-1), 3),
                ('TOPPADDING', (0,0), (-1,-1), 2),
                ('BOTTOMPADDING', (0,0), (-1,-1), 2),
            ]))
            elements.append(result_table)
            elements.append(Spacer(1, 0.1*inch))
            elements.append(Paragraph(f"Overall Competence: {overall_comp}", bold))
            elements.append(Spacer(1, 0.18*inch))  # Add extra space before grading table

# Only one grading key table and pass mark note, based on candidate type
    theory_bands = [
        ('85-100', 'A+'),
        ('80-84', 'A'),
        ('70-79', 'B'),
        ('60-69', 'B-'),
        ('50-59', 'C'),
        ('40-49', 'C-'),
        ('30-39', 'D'),
        ('0-29', 'E'),
    ]
    practical_bands = [
        ('90-100', 'A+'),
        ('85-89', 'A'),
        ('75-84', 'B+'),
        ('65-74', 'B'),
        ('60-64', 'B-'),
        ('55-59', 'C'),
        ('50-54', 'C-'),
        ('40-49', 'D'),
        ('30-39', 'D-'),
        ('0-29', 'E'),
    ]
    grading_key_data = [
        [Paragraph('<b>KEY : GRADING</b>', normal), '', '', ''],
        [Paragraph('<b>THEORY SCORES</b>', normal), Paragraph('Score %', normal), Paragraph('<b>PRACTICAL SCORES</b>', normal), Paragraph('Score %', normal)],
        [Paragraph('Grade', normal), Paragraph('Scores%', normal), Paragraph('Grade', normal), Paragraph('Scores%', normal)],
    ]
    # Render the bands row by row, filling empty cells if needed
    max_rows = max(len(theory_bands), len(practical_bands))
    for i in range(max_rows):
        t_score, t_grade = theory_bands[i] if i < len(theory_bands) else ('', '')
        p_score, p_grade = practical_bands[i] if i < len(practical_bands) else ('', '')
        grading_key_data.append([
            Paragraph(f"{t_score}", normal), Paragraph(f"{t_grade}", normal),
            Paragraph(f"{p_score}", normal), Paragraph(f"{p_grade}", normal)
        ])

    grading_key_table = Table(grading_key_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch], hAlign='LEFT')
    grading_key_table.setStyle(TableStyle([
        ('SPAN', (0,0), (3,0)),  # Span the first row across all columns
        ('SPAN', (0,1), (1,1)),  # Span the second row across first two columns
        ('TOPPADDING', (0,0), (-1,-1), 0.5),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0.5),
        ('SPAN', (2,1), (3,1)),  # Span the second row across last two columns
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('LEFTPADDING', (0,0), (-1,-1), 3),
        ('RIGHTPADDING', (0,0), (-1,-1), 3),
        ('TOPPADDING', (0,0), (-1,-1), 2),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        # No BOX or INNERGRID: no visible borders
    ]))
    # Move key/grades to back page with vertical portrait-style content layout
    elements.append(PageBreak())
    
    # Back page: Single-page vertical layout optimized for space
    elements.append(Spacer(1, 0.5*inch))  # Reduced top margin
    
    # 1. UVTAB Logo - centered at top (smaller size for single page)
    try:
        from django.contrib.staticfiles import finders
        from django.conf import settings
        import os
        
        # Try to find the logo using Django's static file system
        logo_static_path = finders.find('images/uvtab_logo.png')
        
        if logo_static_path and os.path.exists(logo_static_path):
            # Use the found static file path
            uvtab_logo = RLImage(logo_static_path, width=1.5*inch, height=1.5*inch)
            uvtab_logo.hAlign = 'CENTER'
        else:
            # Fallback: try direct path in static directory
            static_root = getattr(settings, 'STATIC_ROOT', None)
            if static_root:
                fallback_path = os.path.join(static_root, 'images', 'uvtab_logo.png')
                if os.path.exists(fallback_path):
                    uvtab_logo = RLImage(fallback_path, width=1.5*inch, height=1.5*inch)
                    uvtab_logo.hAlign = 'CENTER'
                else:
                    raise FileNotFoundError("Logo not found in static files")
            else:
                raise FileNotFoundError("STATIC_ROOT not configured")
    except:
        # Fallback if logo loading fails
        uvtab_logo = Paragraph("UVTAB<br/>LOGO", 
                             ParagraphStyle('LogoPlaceholder', parent=bold, fontSize=12, alignment=TA_CENTER))
    
    # Center the logo
    logo_table = Table([[uvtab_logo]], colWidths=[7.7*inch])
    logo_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (0,0), 'CENTER'),
        ('VALIGN', (0,0), (0,0), 'MIDDLE'),
    ]))
    elements.append(logo_table)
    elements.append(Spacer(1, 0.2*inch))  # Reduced spacing
    
    # 2. Main Institution Header (smaller font for single page)
    elements.append(Paragraph("UGANDA VOCATIONAL AND TECHNICAL ASSESSMENT BOARD", 
                             ParagraphStyle('MainHeader', parent=bold, fontSize=14, alignment=TA_CENTER, spaceAfter=8)))
    
    # 3. Sub-heading (smaller font)
    elements.append(Paragraph("KEY : GRADING AND QUALIFICATIONS BOARD", 
                             ParagraphStyle('SubHeader', parent=bold, fontSize=11, alignment=TA_CENTER, spaceAfter=6)))
    
    # 4. KEY : GRADING title (smaller font)
    elements.append(Paragraph("KEY : GRADING", 
                             ParagraphStyle('KeyTitle', parent=bold, fontSize=12, alignment=TA_CENTER, 
                                          textColor=colors.blue, spaceAfter=12)))
    
    # 5. Create separate Theory and Practical grading tables (stacked vertically)
    
    # Theory Scores Table with proper column headers
    theory_table_data = [
        [Paragraph('<b>THEORY SCORES</b>', ParagraphStyle('TheoryHeader', parent=bold, fontSize=12, alignment=TA_CENTER)), ''],
        [Paragraph('<b>Letter Grade</b>', ParagraphStyle('SubHeader', parent=bold, fontSize=10, alignment=TA_CENTER)), 
         Paragraph('<b>Marks Boundary</b>', ParagraphStyle('SubHeader', parent=bold, fontSize=10, alignment=TA_CENTER))]
    ]
    
    # Add theory grading bands
    for score_range, grade in theory_bands:
        theory_table_data.append([
            Paragraph(f"<b>{grade}</b>", ParagraphStyle('GradeData', parent=normal, fontSize=10, alignment=TA_CENTER)),
            Paragraph(f"{score_range}", ParagraphStyle('ScoreData', parent=normal, fontSize=10, alignment=TA_CENTER))
        ])
    
    theory_table = Table(theory_table_data, colWidths=[1.5*inch, 2.0*inch])
    theory_table.setStyle(TableStyle([
        ('SPAN', (0,0), (1,0)),  # Span header
        ('BACKGROUND', (0,0), (-1,1), colors.lightgrey),  # Header background
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ('BOX', (0,0), (-1,-1), 1, colors.black),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0,2), (-1,-1), [colors.white, colors.lightblue]),
    ]))
    
    # Practical Scores Table with proper column headers
    practical_table_data = [
        [Paragraph('<b>PRACTICAL SCORES</b>', ParagraphStyle('PracticalHeader', parent=bold, fontSize=12, alignment=TA_CENTER)), ''],
        [Paragraph('<b>Letter Grade</b>', ParagraphStyle('SubHeader', parent=bold, fontSize=10, alignment=TA_CENTER)), 
         Paragraph('<b>Marks Boundary</b>', ParagraphStyle('SubHeader', parent=bold, fontSize=10, alignment=TA_CENTER))]
    ]
    
    # Add practical grading bands
    for score_range, grade in practical_bands:
        practical_table_data.append([
            Paragraph(f"<b>{grade}</b>", ParagraphStyle('GradeData', parent=normal, fontSize=10, alignment=TA_CENTER)),
            Paragraph(f"{score_range}", ParagraphStyle('ScoreData', parent=normal, fontSize=10, alignment=TA_CENTER))
        ])
    
    practical_table = Table(practical_table_data, colWidths=[1.5*inch, 2.0*inch])
    practical_table.setStyle(TableStyle([
        ('SPAN', (0,0), (1,0)),  # Span header
        ('BACKGROUND', (0,0), (-1,1), colors.lightgrey),  # Header background
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('LEFTPADDING', (0,0), (-1,-1), 4),
        ('RIGHTPADDING', (0,0), (-1,-1), 4),
        ('BOX', (0,0), (-1,-1), 1, colors.black),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0,2), (-1,-1), [colors.white, colors.lightblue]),
    ]))
    
    # 6. Create horizontal side-by-side layout for both tables
    horizontal_tables = Table([
        [theory_table, practical_table]
    ], colWidths=[3.5*inch, 3.5*inch])  # Equal width for both tables
    
    horizontal_tables.setStyle(TableStyle([
        ('ALIGN', (0,0), (0,0), 'CENTER'),  # Theory table center
        ('ALIGN', (1,0), (1,0), 'CENTER'),  # Practical table center
        ('VALIGN', (0,0), (-1,-1), 'TOP'),   # Top alignment for both tables
        ('LEFTPADDING', (0,0), (-1,-1), 10),
        ('RIGHTPADDING', (0,0), (-1,-1), 10),
        ('TOPPADDING', (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 0),
    ]))
    
    # Center the horizontal tables layout
    tables_container = Table([[horizontal_tables]], colWidths=[7.7*inch])
    tables_container.setStyle(TableStyle([
        ('ALIGN', (0,0), (0,0), 'CENTER'),
        ('VALIGN', (0,0), (0,0), 'MIDDLE'),
    ]))
    elements.append(tables_container)
    elements.append(Spacer(1, 0.15*inch))  # Reduced bottom spacing
    elements.append(Paragraph('NOTE: Pass mark is 50% in theory and 65% in practical assessment', 
                             ParagraphStyle('PassMark', parent=normal, fontSize=12, alignment=TA_CENTER, fontName='Helvetica-Bold')))
    
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="transcript_{candidate.reg_number}.pdf"'
    return response


def get_registration_category_display(registration_category):
    """
    Convert registration category to proper display terminology:
    - Formal -> Full Occupation
    - Informal -> Worker's Pass
    - Modular -> Modular (unchanged)
    """
    if not registration_category:
        return ''
    
    category = registration_category.strip().lower()
    if category == 'formal':
        return 'Full Occupation'
    elif category == 'informal':
        return "Worker's Pass"
    elif category == 'modular':
        return 'Modular'
    else:
        # Return original if not recognized
        return registration_category

def get_formal_level_info(candidate):
    """
    Get level information for Formal candidates and determine if level is module-based or paper-based
    Returns: (candidate_level, is_module_based)
    """
    candidate_level = None
    is_module_based = False
    
    reg_cat = getattr(candidate, 'registration_category', '').lower()
    
    if reg_cat == 'formal':
        # Get the level this candidate enrolled for and was assessed in
        from .models import CandidateLevel
        candidate_levels = CandidateLevel.objects.filter(candidate=candidate).select_related('level')
        if candidate_levels.exists():
            candidate_level = candidate_levels.first().level
            
            # For formal candidates, check if results have modules vs papers
            from .models import Result
            results = Result.objects.filter(candidate=candidate)
            has_modules = results.filter(module__isnull=False).exists()
            has_papers = results.filter(paper__isnull=False).exists()
            
            # Determine if module-based:
            # - If has modules and no papers: definitely module-based
            # - If has both modules and papers: prefer module-based (mixed case)
            # - If has only papers: paper-based
            # - If has neither: default to module-based for formal candidates
            if has_modules:
                is_module_based = True  # Prefer module-based if modules exist
            elif not has_papers and not has_modules:
                is_module_based = True  # Default to module-based for formal candidates with no results
            else:
                is_module_based = False  # Only papers, use paper-based
    
    return candidate_level, is_module_based


@login_required
def generate_verified_results(request, id):
    """
    Generate a PDF verification of results document for a candidate.
    Matches the reference layout with UVTAB logo, clean bio data, and organized results table.
    """
    candidate = Candidate.objects.select_related('occupation', 'assessment_center').get(id=id)
    
    # Get level information for Formal candidates
    candidate_level, is_module_based = get_formal_level_info(candidate)
    
    # Get all results for this candidate
    results = Result.objects.filter(candidate=candidate).select_related(
        'level', 'module', 'paper', 'assessment_series', 'user'
    ).order_by('assessment_date', 'level', 'module', 'paper')
    
    # PDF generation with fixed footer
    buffer = BytesIO()
    
    # Create custom page template with fixed footer
    def create_footer_canvas(canvas, doc):
        """Draw fixed footer only on first page"""
        canvas.saveState()
        
        # Only show footer on page 1 (page numbers start from 1)
        if canvas.getPageNumber() == 1:
            # Footer content - same as current footer but positioned absolutely
            from django.contrib.staticfiles import finders
            
            # Footer text elements
            footer_text1 = "THIS IS NOT A TRANSCRIPT"
            footer_text2 = "OFFICIAL TRANSCRIPT SHALL BE ISSUED AS SOON AS IT IS READY"
            footer_text3 = "*The medium of instruction is ENGLISH*"
            footer_text4 = "ANY ALTERATIONS WHATSOEVER RENDERS THIS VERIFICATION INVALID"
            footer_text5 = "See Reverse for Key Grades"
            
            # Position footer at bottom of page (1 inch from bottom)
            footer_y = 1*inch
            left_margin = 0.5*inch
            
            # Left side footer text
            canvas.setFont("Helvetica", 7)
            canvas.drawString(left_margin, footer_y + 40, footer_text1)
            canvas.drawString(left_margin, footer_y + 30, footer_text2)
            
            canvas.setFont("Helvetica-Oblique", 7)
            canvas.drawString(left_margin, footer_y + 20, footer_text3)
            
            canvas.setFont("Helvetica-Bold", 7)
            canvas.drawString(left_margin, footer_y + 10, footer_text4)
            
            canvas.setFont("Helvetica", 7)
            canvas.drawString(left_margin, footer_y, footer_text5)
            
            # Right side signature
            signature_x = 6*inch
            canvas.setFont("Helvetica", 6)
            canvas.drawString(signature_x, footer_y + 20, "EXECUTIVE SECRETARY")
            canvas.drawString(signature_x, footer_y + 10, "Not Valid Without Official Stamp")
            
            # Try to add signature image
            try:
                signature_path = finders.find('images/es_signature.jpg')
                if signature_path:
                    canvas.drawImage(signature_path, signature_x, footer_y + 25, width=1.2*inch, height=0.6*inch)
            except:
                pass
        
        canvas.restoreState()
    
    # Use BaseDocTemplate with custom page template
    doc = BaseDocTemplate(buffer, pagesize=letter, title="Verification of Results")
    
    # Create frame for main content (leaving space for fixed footer)
    main_frame = Frame(0.5*inch, 1.5*inch, 7*inch, 9*inch, 
                       leftPadding=0, rightPadding=0, topPadding=0, bottomPadding=0)
    
    # Create page template with fixed footer
    page_template = PageTemplate(id='main', frames=[main_frame], 
                                onPage=create_footer_canvas)
    doc.addPageTemplates([page_template])
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], 
                                fontSize=16, spaceAfter=12, alignment=TA_CENTER,
                                fontName='Helvetica-Bold')
    header_style = ParagraphStyle('Header', parent=styles['Heading2'], 
                                 fontSize=14, spaceAfter=8, alignment=TA_CENTER,
                                 fontName='Helvetica-Bold')
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], 
                                 fontSize=10, spaceAfter=6)
    bold_style = ParagraphStyle('Bold', parent=normal_style, 
                               fontName='Helvetica-Bold')
    bio_style = ParagraphStyle('Bio', parent=normal_style, 
                              fontSize=11, spaceAfter=3, fontName='Helvetica-Bold')
    # Styles for highlighting failed results
    failed_style = ParagraphStyle('Failed', parent=normal_style, 
                                textColor=colors.red, fontName='Helvetica-Bold')
    normal_text_style = ParagraphStyle('NormalText', parent=normal_style)
    
    # Helper function to check if a result is failed
    def is_failed_result(grade, comment):
        """Check if a result represents a failure based on grade and comment"""
        failed_comments = ['CTR', 'Missing', 'Absent']
        failed_grades = ['C-', 'D', 'D+', 'D-', 'E', 'F', 'Ms', 'Fail']
        
        return (comment in failed_comments or 
                grade in failed_grades or
                (grade and grade.endswith('-') and grade[0] in ['C', 'D']))
    
    # Main title first - above everything
    elements.append(Paragraph("UGANDA VOCATIONAL AND TECHNICAL ASSESSMENT BOARD", 
                             ParagraphStyle('MainTitle', parent=bold_style, fontSize=14, alignment=TA_CENTER)))
    elements.append(Spacer(1, 0.2*inch))
    
    # UVTAB Header with Logo and contact info (matching image 2)
    # Left contact info
    left_contact = Paragraph("Plot 7, Valley Drive, Ntinda-Kyambogo Road<br/>Email: info@uvtab.go.ug", 
                            ParagraphStyle('LeftContact', parent=normal_style, fontSize=10, alignment=TA_LEFT))
    
    # Right contact info  
    right_contact = Paragraph("P.O.Box 1499, Kampala,<br/>Tel: +256 414 289786", 
                             ParagraphStyle('RightContact', parent=normal_style, fontSize=10, alignment=TA_RIGHT))
    
    # Try to add UVTAB logo
    uvtab_logo = None
    try:
        from django.conf import settings
        import os
        logo_path = os.path.join(settings.BASE_DIR, 'eims', 'static', 'images', 'uvtab_logo.png')
        if not os.path.exists(logo_path):
            raise FileNotFoundError('Logo file not found')
        uvtab_logo = RLImage(logo_path, width=0.8*inch, height=0.8*inch)
        uvtab_logo.hAlign = 'CENTER'
    except Exception:
        uvtab_logo = None
    
    # Create header table
    if uvtab_logo:
        header_table = Table([
            [left_contact, uvtab_logo, right_contact]
        ], colWidths=[2.5*inch, 1.5*inch, 2.5*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (1,0), (1,0), 'CENTER'),
            ('ALIGN', (2,0), (2,0), 'RIGHT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
        elements.append(header_table)
    else:
        # Fallback without logo
        header_table = Table([
            [left_contact, right_contact]
        ], colWidths=[3.5*inch, 3.5*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (1,0), (1,0), 'RIGHT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
        elements.append(header_table)
    
    elements.append(Spacer(1, 0.2*inch))
    
    # Main heading
    elements.append(Paragraph("VERIFICATION OF RESULTS", header_style))
    elements.append(Spacer(1, 0.1*inch))
    elements.append(Paragraph("TO WHOM IT MAY CONCERN", 
                             ParagraphStyle('Concern', parent=normal_style, alignment=TA_CENTER)))
    elements.append(Spacer(1, 0.1*inch))
    
    # Verification text
    verification_text = "This is to verify that this candidate registered and sat for UVTAB assessments with the following particulars and obtained the following"
    elements.append(Paragraph(f"<i>{verification_text}</i>", 
                             ParagraphStyle('Italic', parent=normal_style, alignment=TA_CENTER)))
    elements.append(Spacer(1, 0.2*inch))
    
    # Candidate photo (if available) - positioned on left like image 2
    photo = None
    if getattr(candidate, 'passport_photo', None) and candidate.passport_photo.name:
        try:
            photo_path = candidate.passport_photo.path
            photo = RLImage(photo_path, width=1.0*inch, height=1.2*inch)
            photo.hAlign = 'LEFT'
        except Exception:
            photo = None
    
    # Candidate details - Clean layout like image 2
    from django_countries import countries
    nationality = candidate.nationality if getattr(candidate, 'nationality', None) else ""
    if nationality and len(nationality) == 2 and nationality.isupper():
        code_to_name = dict(countries)
        nationality = code_to_name.get(nationality, nationality)
    
    # Format dates - use today's date for print date
    from datetime import datetime
    birthdate = candidate.date_of_birth.strftime('%d %b, %Y') if getattr(candidate, 'date_of_birth', None) else ""
    print_date = datetime.now().strftime('%d-%b-%Y')  # Today's date
    
    # Determine programme/occupation label based on registration category
    reg_cat = getattr(candidate, 'registration_category', '').lower()
    if reg_cat == 'informal':
        programme_label = "OCCUPATION:"
    programme_label = "OCCUPATION:"
    
    # Get level information for Formal candidates
    candidate_level = None
    reg_cat = getattr(candidate, 'registration_category', '').lower()
    if reg_cat == 'formal':
        # Get the level this candidate enrolled for and was assessed in
        from .models import CandidateLevel
        candidate_levels = CandidateLevel.objects.filter(candidate=candidate).select_related('level')
        if candidate_levels.exists():
            candidate_level = candidate_levels.first().level
    
    # Bio data layout with photo on left (like image 2)
    bio_left_content = [
        f"<b>NAME:</b> {candidate.full_name}",
        f"<b>REG NO:</b> {candidate.reg_number}",
        f"<b>GENDER:</b> {candidate.get_gender_display() if hasattr(candidate, 'get_gender_display') else candidate.gender}",
        f"<b>CENTER NAME:</b> {candidate.assessment_center.center_name if candidate.assessment_center else ''}",
        f"<b>REGISTRATION CATEGORY:</b> {get_registration_category_display(candidate.registration_category) if hasattr(candidate, 'registration_category') else ''}",
        f"<b>{programme_label}</b> {candidate.occupation.name if candidate.occupation else ''}"
    ]
    
    # Add level information for Formal candidates
    reg_cat = getattr(candidate, 'registration_category', '').lower()
    if reg_cat == 'formal' and candidate_level:
        bio_left_content.append(f"<b>LEVEL:</b> {candidate_level.name}")
    elif reg_cat == 'formal':
        bio_left_content.append(f"<b>LEVEL:</b> N/A")
    
    bio_right_content = [
        f"<b>NATIONALITY:</b> {nationality}",
        f"<b>BIRTHDATE:</b> {birthdate}",
        f"<b>PRINTDATE:</b> {print_date}",
        "",
        "",
        ""
    ]
    
    # Create bio paragraphs with smaller font like image 2
    bio_left = Paragraph("<br/>".join(bio_left_content), 
                        ParagraphStyle('BioLeft', parent=normal_style, fontSize=9, leading=11))
    bio_right = Paragraph("<br/>".join(bio_right_content), 
                         ParagraphStyle('BioRight', parent=normal_style, fontSize=9, leading=11))
    
    # Layout with photo on left side like image 2
    if photo:
        # Photo and bio data in same row like image 2
        bio_table = Table([
            [photo, bio_left, bio_right]
        ], colWidths=[1.2*inch, 3.0*inch, 2.8*inch])
        bio_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (1,0), (1,0), 'LEFT'),
            ('ALIGN', (2,0), (2,0), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
    else:
        # No photo - just bio data in two columns
        bio_table = Table([
            [bio_left, bio_right]
        ], colWidths=[3.5*inch, 3.5*inch])
        bio_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
    
    elements.append(bio_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Results table - Only this table should have borders
    if results.exists():
        # Group results by assessment series and organize by registration category
        reg_cat = getattr(candidate, 'registration_category', '').lower()
        
        # Determine overall success
        passed_results = results.filter(comment='Successful').count()
        total_results = results.count()
        overall_success = passed_results == total_results and total_results > 0
        
        # Check if formal candidate has only Theory/Practical (no actual modules or papers)
        formal_theory_practical_only = False
        if reg_cat == 'formal':
            has_actual_modules = results.filter(module__isnull=False).exists()
            has_actual_papers = results.filter(paper__isnull=False).exists()
            formal_theory_practical_only = not has_actual_modules and not has_actual_papers
        
        if formal_theory_practical_only:
            # Simplified layout for Formal candidates with only Theory/Practical assessments
            elements.append(Paragraph("<b>ASSESSMENT RESULTS</b>", 
                                     ParagraphStyle('ResultsHeader', parent=bold_style, alignment=TA_CENTER)))
            elements.append(Spacer(1, 0.1*inch))
            
            # Create simplified results table (only Assessment Type, Grade, Comment)
            results_data = [
                [Paragraph("<b>ASSESSMENT TYPE</b>", bold_style),
                 Paragraph("<b>GRADE</b>", bold_style),
                 Paragraph("<b>COMMENT</b>", bold_style)]
            ]
            
            for result in results:
                assessment_type = result.get_assessment_type_display() if hasattr(result, 'get_assessment_type_display') else "Practical"
                grade = result.grade
                comment = result.comment if result.comment else ""
                
                # Apply conditional styling for failed results
                is_failed = is_failed_result(grade, comment)
                grade_cell = Paragraph(grade, failed_style if is_failed else normal_text_style)
                comment_cell = Paragraph(comment, failed_style if is_failed else normal_text_style)
                
                results_data.append([assessment_type, grade_cell, comment_cell])
        
        elif reg_cat == 'modular' or (reg_cat == 'formal' and is_module_based):
            # Module-based layout - for Modular candidates and Formal module-based levels
            elements.append(Paragraph("<b>ASSESSMENT RESULTS</b>", 
                                     ParagraphStyle('ResultsHeader', parent=bold_style, alignment=TA_CENTER)))
            elements.append(Spacer(1, 0.1*inch))
            
            # Create module-based results table (no Paper Code/Name columns)
            results_data = [
                [Paragraph("<b>MODULE NAME</b>", bold_style), 
                 Paragraph("<b>ASSESSMENT TYPE</b>", bold_style),
                 Paragraph("<b>GRADE</b>", bold_style),
                 Paragraph("<b>COMMENT</b>", bold_style)]
            ]
            
            for result in results:
                module_name = result.module.name if result.module else ""
                assessment_type = result.get_assessment_type_display() if hasattr(result, 'get_assessment_type_display') else "Practical"
                grade = result.grade
                comment = result.comment if result.comment else ""
                
                # Apply conditional styling for failed results
                is_failed = is_failed_result(grade, comment)
                grade_cell = Paragraph(grade, failed_style if is_failed else normal_text_style)
                comment_cell = Paragraph(comment, failed_style if is_failed else normal_text_style)
                
                results_data.append([module_name, assessment_type, grade_cell, comment_cell])
        
        else:
            # Paper-based layout - for Formal paper-based levels and Informal candidates
            elements.append(Paragraph("<b>ASSESSMENT RESULTS</b>", 
                                     ParagraphStyle('ResultsHeader', parent=bold_style, alignment=TA_CENTER)))
            elements.append(Spacer(1, 0.1*inch))
            
            results_data = [
                [Paragraph("<b>Paper Code</b>", bold_style), 
                 Paragraph("<b>Paper Name</b>", bold_style),
                 Paragraph("<b>Level</b>", bold_style),
                 Paragraph("<b>Assessment<br/>Type</b>", bold_style), 
                 Paragraph("<b>Grade</b>", bold_style),
                 Paragraph("<b>Comment</b>", bold_style)]
            ]
            
            for result in results:
                paper_code = result.paper.code if result.paper else (result.module.code if result.module else "")
                paper_name = result.paper.name if result.paper else (result.module.name if result.module else "")
                level_name = result.level.name if result.level else ""
                assessment_type = result.get_assessment_type_display()
                grade = result.grade
                comment = result.comment if result.comment else ""
                
                # Apply conditional styling for failed results
                is_failed = is_failed_result(grade, comment)
                grade_cell = Paragraph(grade, failed_style if is_failed else normal_text_style)
                comment_cell = Paragraph(comment, failed_style if is_failed else normal_text_style)
                
                results_data.append([paper_code, Paragraph(paper_name, normal_style), Paragraph(level_name, normal_style), assessment_type, grade_cell, comment_cell])
        
        # Create results table with appropriate column widths based on layout (fixed to prevent overlap)
        if formal_theory_practical_only:
            # Simplified layout (theory/practical only) (3 columns) - Total: 6.0 inches
            results_table = Table(results_data, colWidths=[2.5*inch, 1.0*inch, 2.5*inch])
        elif reg_cat == 'modular' or (reg_cat == 'formal' and is_module_based):
            # Module-based layout (modular or formal module-based) (4 columns) - Total: 6.0 inches
            results_table = Table(results_data, colWidths=[2.2*inch, 1.3*inch, 0.8*inch, 1.7*inch])
        else:
            # Paper-based layout (formal paper-based or informal) (6 columns) - Total: 6.0 inches
            results_table = Table(results_data, colWidths=[0.6*inch, 2.2*inch, 0.8*inch, 1.0*inch, 0.6*inch, 0.8*inch])
        
        results_table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            
            # Table borders - only for results table
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('RIGHTPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        
        elements.append(results_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Success/Failure summary comment - skip for modular candidates
        if not (reg_cat == 'modular' or (reg_cat == 'formal' and is_module_based) or reg_cat == 'informal'):
            if overall_success:
                comment_text = "Overall Assessment Comment: Successful"
                comment_style = ParagraphStyle('Success', parent=bold_style, 
                                             textColor=colors.green, alignment=TA_CENTER)
            else:
                comment_text = "Overall Assessment Comment: Not successful"
                comment_style = ParagraphStyle('Failure', parent=bold_style, 
                                             textColor=colors.red, alignment=TA_CENTER)
            
            elements.append(Paragraph(comment_text, comment_style))
        
    else:
        elements.append(Paragraph("<b>No results recorded for this candidate.</b>", bold_style))
    
    # Footer is now handled by fixed page template - no need to add as elements
    
    # Add page break for grading system back page
    from reportlab.platypus import PageBreak
    elements.append(PageBreak())
    
    # Back page: UVTAB header and grading system
    elements.append(Paragraph("UGANDA VOCATIONAL AND TECHNICAL ASSESSMENT BOARD", 
                             ParagraphStyle('BackPageTitle', parent=bold_style, fontSize=14, alignment=TA_CENTER)))
    elements.append(Spacer(1, 0.3*inch))
    
    # Add UVTAB logo on back page
    if uvtab_logo:
        try:
            uvtab_logo_back = RLImage(logo_path, width=0.8*inch, height=0.8*inch)
            uvtab_logo_back.hAlign = 'CENTER'
            elements.append(uvtab_logo_back)
            elements.append(Spacer(1, 0.3*inch))
        except:
            pass
    
    # Grading system tables
    elements.append(Paragraph("<b>KEY : GRADING</b>", 
                             ParagraphStyle('GradingTitle', parent=bold_style, fontSize=12, alignment=TA_CENTER)))
    elements.append(Spacer(1, 0.2*inch))
    
    # Theory and Practical grading bands
    theory_bands = [
        ('85-100', 'A+'),
        ('80-84', 'A'),
        ('70-79', 'B'),
        ('60-69', 'B-'),
        ('50-59', 'C'),
        ('40-49', 'C-'),
        ('30-39', 'D'),
        ('0-29', 'E'),
    ]
    practical_bands = [
        ('90-100', 'A+'),
        ('85-89', 'A'),
        ('75-84', 'B+'),
        ('65-74', 'B'),
        ('60-64', 'B-'),
        ('55-59', 'C'),
        ('50-54', 'C-'),
        ('40-49', 'D'),
        ('30-39', 'D-'),
        ('0-29', 'E'),
    ]
    
    # Create grading table
    grading_data = [
        [Paragraph('<b>THEORY SCORES</b>', bold_style), '', Paragraph('<b>PRACTICAL SCORES</b>', bold_style), ''],
        [Paragraph('<b>Grade</b>', bold_style), Paragraph('<b>Scores%</b>', bold_style), 
         Paragraph('<b>Grade</b>', bold_style), Paragraph('<b>Scores%</b>', bold_style)],
    ]
    
    # Add grading rows
    max_rows = max(len(theory_bands), len(practical_bands))
    for i in range(max_rows):
        t_score, t_grade = theory_bands[i] if i < len(theory_bands) else ('', '')
        p_score, p_grade = practical_bands[i] if i < len(practical_bands) else ('', '')
        grading_data.append([
            Paragraph(t_grade, normal_style), Paragraph(t_score, normal_style),
            Paragraph(p_grade, normal_style), Paragraph(p_score, normal_style)
        ])
    
    grading_table = Table(grading_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    grading_table.setStyle(TableStyle([
        ('SPAN', (0,0), (1,0)),  # Span THEORY SCORES
        ('SPAN', (2,0), (3,0)),  # Span PRACTICAL SCORES
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (-1,1), colors.lightgrey),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    
    elements.append(grading_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Pass mark information
    elements.append(Paragraph('<b>Pass mark is 50% in theory and 65% in practical assessment</b>', 
                             ParagraphStyle('PassMark', parent=bold_style, fontSize=11, alignment=TA_CENTER)))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="verified_results_{candidate.reg_number}.pdf"'
    return response


def generate_testimonial(request, id):
    """
    Generate a testimonial PDF for center users - identical to verified results but with different heading and no footer
    """
    candidate = get_object_or_404(Candidate, id=id)
    
    # Get level information for Formal candidates
    candidate_level, is_module_based = get_formal_level_info(candidate)
    
    # Security check: Prevent testimonial generation if results haven't been released
    is_center_rep = request.user.groups.filter(name='CenterRep').exists()
    candidate_series = candidate.assessment_series
    results_released = candidate_series and candidate_series.results_released or not is_center_rep
    
    if not results_released:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden("Testimonial cannot be generated. Results have not been released for this assessment series.")
    
    results = candidate.result_set.all().order_by('assessment_series__name', 'level__name', 'module__name')
    
    # Create PDF buffer
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch,
                          leftMargin=0.75*inch, rightMargin=0.75*inch)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], 
                                fontSize=16, spaceAfter=12, alignment=TA_CENTER, fontName='Helvetica-Bold')
    header_style = ParagraphStyle('CustomHeader', parent=styles['Heading1'], 
                                 fontSize=14, spaceAfter=6, alignment=TA_CENTER, fontName='Helvetica-Bold')
    normal_style = styles['Normal']
    bold_style = ParagraphStyle('Bold', parent=normal_style, fontName='Helvetica-Bold')
    bio_style = ParagraphStyle('Bio', parent=normal_style, 
                              fontSize=11, spaceAfter=3, fontName='Helvetica-Bold')
    
    # Main title first - above everything
    elements.append(Paragraph("UGANDA VOCATIONAL AND TECHNICAL ASSESSMENT BOARD", 
                             ParagraphStyle('MainTitle', parent=bold_style, fontSize=14, alignment=TA_CENTER)))
    elements.append(Spacer(1, 0.2*inch))
    
    # UVTAB Header with Logo and contact info (matching verified results)
    # Left contact info
    left_contact = Paragraph("Plot 7, Valley Drive, Ntinda-Kyambogo Road<br/>Email: info@uvtab.go.ug", 
                            ParagraphStyle('LeftContact', parent=normal_style, fontSize=10, alignment=TA_LEFT))
    
    # Right contact info  
    right_contact = Paragraph("P.O.Box 1499, Kampala,<br/>Tel: +256 414 289786", 
                             ParagraphStyle('RightContact', parent=normal_style, fontSize=10, alignment=TA_RIGHT))
    
    # Try to add UVTAB logo
    uvtab_logo = None
    try:
        from django.conf import settings
        import os
        logo_path = os.path.join(settings.BASE_DIR, 'eims', 'static', 'images', 'uvtab_logo.png')
        if not os.path.exists(logo_path):
            raise FileNotFoundError('Logo file not found')
        uvtab_logo = RLImage(logo_path, width=0.8*inch, height=0.8*inch)
        uvtab_logo.hAlign = 'CENTER'
    except Exception:
        uvtab_logo = None
    
    # Create header table
    if uvtab_logo:
        header_table = Table([
            [left_contact, uvtab_logo, right_contact]
        ], colWidths=[2.5*inch, 1.5*inch, 2.5*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (1,0), (1,0), 'CENTER'),
            ('ALIGN', (2,0), (2,0), 'RIGHT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
        elements.append(header_table)
    else:
        # Fallback without logo
        header_table = Table([
            [left_contact, right_contact]
        ], colWidths=[3.5*inch, 3.5*inch])
        header_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (1,0), (1,0), 'RIGHT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
        elements.append(header_table)
    
    elements.append(Spacer(1, 0.2*inch))
    
    # Main heading - "TESTIMONIAL" instead of "VERIFICATION OF RESULTS"
    elements.append(Paragraph("TESTIMONIAL", header_style))
    elements.append(Spacer(1, 0.2*inch))
    
    # Candidate photo (if available) - positioned on left like verified results
    photo = None
    if getattr(candidate, 'passport_photo', None) and candidate.passport_photo.name:
        try:
            photo_path = candidate.passport_photo.path
            photo = RLImage(photo_path, width=1.0*inch, height=1.2*inch)
            photo.hAlign = 'LEFT'
        except Exception:
            photo = None
    
    # Candidate details - Clean layout like verified results
    from django_countries import countries
    nationality = candidate.nationality if getattr(candidate, 'nationality', None) else ""
    if nationality and len(nationality) == 2 and nationality.isupper():
        code_to_name = dict(countries)
        nationality = code_to_name.get(nationality, nationality)
    
    # Format dates - use today's date for print date
    from datetime import datetime
    birthdate = candidate.date_of_birth.strftime('%d %b, %Y') if getattr(candidate, 'date_of_birth', None) else ""
    print_date = datetime.now().strftime('%d-%b-%Y')  # Today's date
    
    # Determine programme/occupation label based on registration category
    reg_cat = getattr(candidate, 'registration_category', '').lower()
    if reg_cat == 'informal':
        programme_label = "OCCUPATION:"
    else:
        programme_label = "OCCUPATION:"
    
    # Bio data layout with photo on left (same as verified results)
    bio_left_content = [
        f"<b>NAME:</b> {candidate.full_name}",
        f"<b>REG NO:</b> {candidate.reg_number}",
        f"<b>GENDER:</b> {candidate.get_gender_display() if hasattr(candidate, 'get_gender_display') else candidate.gender}",
        f"<b>CENTER NAME:</b> {candidate.assessment_center.center_name if candidate.assessment_center else ''}",
        f"<b>REGISTRATION CATEGORY:</b> {get_registration_category_display(candidate.registration_category) if hasattr(candidate, 'registration_category') else ''}",
        f"<b>{programme_label}</b> {candidate.occupation.name if candidate.occupation else ''}"
    ]
    
    # Add level information for Formal candidates
    reg_cat = getattr(candidate, 'registration_category', '').lower()
    if reg_cat == 'formal' and candidate_level:
        bio_left_content.append(f"<b>LEVEL:</b> {candidate_level.name}")
    elif reg_cat == 'formal':
        bio_left_content.append(f"<b>LEVEL:</b> N/A")
    
    bio_right_content = [
        f"<b>NATIONALITY:</b> {nationality}",
        f"<b>BIRTHDATE:</b> {birthdate}",
        f"<b>PRINTDATE:</b> {print_date}",
        "",
        "",
        ""
    ]
    
    # Create bio paragraphs with smaller font like verified results
    bio_left = Paragraph("<br/>".join(bio_left_content), 
                        ParagraphStyle('BioLeft', parent=normal_style, fontSize=9, leading=11))
    bio_right = Paragraph("<br/>".join(bio_right_content), 
                         ParagraphStyle('BioRight', parent=normal_style, fontSize=9, leading=11))
    
    # Layout with photo on left side like verified results
    if photo:
        # Photo and bio data in same row like verified results
        bio_table = Table([
            [photo, bio_left, bio_right]
        ], colWidths=[1.2*inch, 3.0*inch, 2.8*inch])
        bio_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (1,0), (1,0), 'LEFT'),
            ('ALIGN', (2,0), (2,0), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
    else:
        # No photo - just bio data in two columns
        bio_table = Table([
            [bio_left, bio_right]
        ], colWidths=[3.5*inch, 3.5*inch])
        bio_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
    
    elements.append(bio_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Results table - Only this table should have borders (same as verified results)
    if results.exists():
        # Group results by assessment series and organize by registration category
        reg_cat = getattr(candidate, 'registration_category', '').lower()
        
        # Determine overall success
        passed_results = results.filter(comment='Successful').count()
        total_results = results.count()
        overall_success = passed_results == total_results and total_results > 0
        
        # Check if formal candidate has only Theory/Practical (no actual modules or papers) - same as verified results
        formal_theory_practical_only = False
        if reg_cat == 'formal':
            has_actual_modules = results.filter(module__isnull=False).exists()
            has_actual_papers = results.filter(paper__isnull=False).exists()
            formal_theory_practical_only = not has_actual_modules and not has_actual_papers
        
        if formal_theory_practical_only:
            # Simplified layout for Formal candidates with only Theory/Practical assessments
            elements.append(Paragraph("<b>ASSESSMENT RESULTS</b>", 
                                     ParagraphStyle('ResultsHeader', parent=bold_style, alignment=TA_CENTER)))
            elements.append(Spacer(1, 0.1*inch))
            
            # Create simplified results table (only Assessment Type, Grade, Comment)
            results_data = [
                [Paragraph("<b>ASSESSMENT TYPE</b>", bold_style),
                 Paragraph("<b>GRADE</b>", bold_style),
                 Paragraph("<b>COMMENT</b>", bold_style)]
            ]
            
            for result in results:
                assessment_type = result.get_assessment_type_display() if hasattr(result, 'get_assessment_type_display') else "Practical"
                grade = result.grade
                comment = result.comment if result.comment else ""
                
                # Apply conditional styling for failed results
                is_failed = is_failed_result(grade, comment)
                grade_cell = Paragraph(grade, failed_style if is_failed else normal_text_style)
                comment_cell = Paragraph(comment, failed_style if is_failed else normal_text_style)
                
                results_data.append([assessment_type, grade_cell, comment_cell])
        
        elif reg_cat == 'modular':
            # Modular layout - Module structure
            elements.append(Paragraph("<b>ASSESSMENT RESULTS</b>", 
                                     ParagraphStyle('ResultsHeader', parent=bold_style, alignment=TA_CENTER)))
            elements.append(Spacer(1, 0.1*inch))
            
            # Group by assessment series (year) and modules
            from collections import defaultdict
            year_modules = defaultdict(lambda: defaultdict(list))
            
            for result in results:
                year = result.assessment_series.name if result.assessment_series else "Unknown Year"
                module_name = f"{result.module.code} - {result.module.name}" if result.module else "Unknown Module"
                year_modules[year][module_name].append(result)
            
            # Create modular results table
            results_data = []
            
            # Add table headers once
            results_data.append([Paragraph("<b>MODULE NAME</b>", bold_style), 
                               Paragraph("<b>ASSESSMENT TYPE</b>", bold_style),
                               Paragraph("<b>GRADE</b>", bold_style),
                               Paragraph("<b>COMMENT</b>", bold_style)])
            
            for year, modules in year_modules.items():
                for module_name, module_results in modules.items():
                    for result in module_results:
                        module_name = result.module.name if result.module else ""
                        assessment_type = result.get_assessment_type_display() if hasattr(result, 'get_assessment_type_display') else "Practical"
                        grade = result.grade
                        comment = result.comment if result.comment else ""
                        
                        # Apply conditional styling for failed results
                is_failed = is_failed_result(grade, comment)
                grade_cell = Paragraph(grade, failed_style if is_failed else normal_text_style)
                comment_cell = Paragraph(comment, failed_style if is_failed else normal_text_style)
                
                results_data.append([module_name, assessment_type, grade_cell, comment_cell])
        
        else:
            # Formal/Informal layout - Paper-based
            elements.append(Paragraph("<b>ASSESSMENT RESULTS</b>", 
                                     ParagraphStyle('ResultsHeader', parent=bold_style, alignment=TA_CENTER)))
            elements.append(Spacer(1, 0.1*inch))
            
            results_data = [
                [Paragraph("<b>Paper Code</b>", bold_style), 
                 Paragraph("<b>Paper Name</b>", bold_style),
                 Paragraph("<b>Level</b>", bold_style),
                 Paragraph("<b>Assessment<br/>Type</b>", bold_style), 
                 Paragraph("<b>Grade</b>", bold_style),
                 Paragraph("<b>Comment</b>", bold_style)]
            ]
            
            for result in results:
                paper_code = result.paper.code if result.paper else (result.module.code if result.module else "")
                paper_name = result.paper.name if result.paper else (result.module.name if result.module else "")
                level_name = result.level.name if result.level else ""
                assessment_type = result.get_assessment_type_display()
                grade = result.grade
                comment = result.comment if result.comment else ""
                
                # Apply conditional styling for failed results
                is_failed = is_failed_result(grade, comment)
                grade_cell = Paragraph(grade, failed_style if is_failed else normal_text_style)
                comment_cell = Paragraph(comment, failed_style if is_failed else normal_text_style)
                
                results_data.append([paper_code, Paragraph(paper_name, normal_style), Paragraph(level_name, normal_style), assessment_type, grade_cell, comment_cell])
        
        # Create and style the results table (same as verified results)
        if formal_theory_practical_only:
            # Simplified layout for Theory/Practical-only Formal candidates (3 columns)
            results_table = Table(results_data, colWidths=[2.0*inch, 1.0*inch, 3.0*inch])
        elif reg_cat == 'modular' or (reg_cat == 'formal' and is_module_based):
            # Module-based layout (modular or formal module-based) (4 columns)
            results_table = Table(results_data, colWidths=[2.5*inch, 1.5*inch, 0.8*inch, 1.2*inch])
        else:
            # Paper-based layout (formal paper-based or informal) (5 columns)
            results_table = Table(results_data, colWidths=[0.8*inch, 3.0*inch, 1.0*inch, 1.2*inch, 0.7*inch, 1.3*inch])
        
        results_table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            
            # Table borders - only for results table
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('LEFTPADDING', (0,0), (-1,-1), 4),
            ('RIGHTPADDING', (0,0), (-1,-1), 4),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ]))
        
        elements.append(results_table)
        elements.append(Spacer(1, 0.2*inch))
        
        # Success/Failure summary comment - skip for modular candidates
        if not (reg_cat == 'modular' or (reg_cat == 'formal' and is_module_based) or reg_cat == 'informal'):
            if overall_success:
                comment_text = "Comment: Successful"
                comment_style = ParagraphStyle('Success', parent=bold_style, 
                                             textColor=colors.green, alignment=TA_CENTER)
            else:
                comment_text = "Comment: Not successful"
                comment_style = ParagraphStyle('NotSuccess', parent=bold_style, 
                                             textColor=colors.red, alignment=TA_CENTER)
            
            elements.append(Paragraph(comment_text, comment_style))
    else:
        elements.append(Paragraph("<b>No results recorded for this candidate.</b>", bold_style))
    
    # Push footer to bottom of page 1 - static positioning to ensure signature always appears
    elements.append(Spacer(1, 1.8*inch))  # Reduced spacer to ensure footer fits on page
    
    # Footer section (like image 2) - smaller fonts and better alignment
    # Add footer text with smaller fonts
    footer_text1 = Paragraph("THIS IS NOT A TRANSCRIPT", 
                            ParagraphStyle('FooterBold', parent=bold_style, fontSize=8, alignment=TA_LEFT, leading=10))
    footer_text2 = Paragraph("OFFICIAL TRANSCRIPT SHALL BE ISSUED AS SOON AS IT IS READY", 
                            ParagraphStyle('FooterItalic', parent=normal_style, fontSize=7, fontName='Helvetica-Oblique', alignment=TA_LEFT, leading=9))
    footer_text3 = Paragraph("*The medium of instruction is ENGLISH*", 
                            ParagraphStyle('FooterItalic', parent=normal_style, fontSize=7, fontName='Helvetica-Oblique', alignment=TA_LEFT, leading=9))
    footer_text4 = Paragraph("ANY ALTERATIONS WHATSOEVER RENDERS THIS VERIFICATION INVALID", 
                            ParagraphStyle('FooterBold', parent=bold_style, fontSize=7, alignment=TA_LEFT, leading=9))
    footer_text5 = Paragraph("See Reverse for Key Grades", 
                            ParagraphStyle('FooterReverse', parent=normal_style, fontSize=7, alignment=TA_LEFT, leading=9))
    
    # ES Signature - smaller size to match image 2
    es_signature = None
    try:
        # NEW:
        from django.contrib.staticfiles import finders
        signature_path = finders.find('images/es_signature.jpg')
        if signature_path:
            es_signature = RLImage(signature_path, width=1.2*inch, height=0.6*inch)
            es_signature.hAlign = 'RIGHT'
    except Exception:
        es_signature = None
    
    # Create footer layout like image 2
    if es_signature:
        # Left footer text block
        footer_left = [
            footer_text1,
            footer_text2, 
            footer_text3,
            Spacer(1, 0.05*inch),
            footer_text4,
            footer_text5
        ]
        
        # Right signature block - well aligned signature and text
        signature_text = Paragraph("EXECUTIVE SECRETARY<br/>Not Valid Without Official Stamp", 
                                  ParagraphStyle('SignatureText', parent=normal_style, fontSize=6, alignment=TA_CENTER, leading=8))
        
        # Create signature section with proper alignment
        signature_section = Table([
            [es_signature],
            [signature_text]
        ], colWidths=[2.2*inch])
        
        signature_section.setStyle(TableStyle([
            ('ALIGN', (0,0), (0,0), 'CENTER'),  # Center signature image
            ('ALIGN', (0,1), (0,1), 'CENTER'),  # Center signature text
            ('VALIGN', (0,0), (0,0), 'BOTTOM'), # Align signature to bottom
            ('VALIGN', (0,1), (0,1), 'TOP'),    # Align text to top
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ]))
        
        # Create table with proper alignment like image 2
        footer_table = Table([
            [footer_left, signature_section]
        ], colWidths=[4.8*inch, 2.2*inch])
        
        footer_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (1,0), (1,0), 'RIGHT'),
            ('VALIGN', (0,0), (0,0), 'BOTTOM'),  # Align footer text to bottom
            ('VALIGN', (1,0), (1,0), 'BOTTOM'),  # Align signature to bottom
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
        
        elements.append(footer_table)
    else:
        # Fallback without signature image - smaller fonts
        elements.append(footer_text1)
        elements.append(footer_text2)
        elements.append(footer_text3)
        elements.append(Spacer(1, 0.05*inch))
        elements.append(footer_text4)
        elements.append(footer_text5)
        elements.append(Spacer(1, 0.1*inch))
        elements.append(Paragraph("EXECUTIVE SECRETARY<br/>Not Valid Without Official Stamp", 
                                 ParagraphStyle('SignatureText', parent=normal_style, fontSize=6, alignment=TA_CENTER)))
    
    # Add page break for grading system back page
    from reportlab.platypus import PageBreak
    elements.append(PageBreak())
    
    # Back page: UVTAB header and grading system
    elements.append(Paragraph("UGANDA VOCATIONAL AND TECHNICAL ASSESSMENT BOARD", 
                             ParagraphStyle('BackPageTitle', parent=bold_style, fontSize=14, alignment=TA_CENTER)))
    elements.append(Spacer(1, 0.3*inch))
    
    # Add UVTAB logo on back page
    if uvtab_logo:
        try:
            uvtab_logo_back = RLImage(logo_path, width=0.8*inch, height=0.8*inch)
            uvtab_logo_back.hAlign = 'CENTER'
            elements.append(uvtab_logo_back)
            elements.append(Spacer(1, 0.3*inch))
        except:
            pass
    
    # Grading system tables
    elements.append(Paragraph("<b>KEY : GRADING</b>", 
                             ParagraphStyle('GradingTitle', parent=bold_style, fontSize=12, alignment=TA_CENTER)))
    elements.append(Spacer(1, 0.2*inch))
    
    # Theory and Practical grading bands
    theory_bands = [
        ('85-100', 'A+'),
        ('80-84', 'A'),
        ('70-79', 'B'),
        ('60-69', 'B-'),
        ('50-59', 'C'),
        ('40-49', 'C-'),
        ('30-39', 'D'),
        ('0-29', 'E'),
    ]
    practical_bands = [
        ('90-100', 'A+'),
        ('85-89', 'A'),
        ('75-84', 'B+'),
        ('65-74', 'B'),
        ('60-64', 'B-'),
        ('55-59', 'C'),
        ('50-54', 'C-'),
        ('40-49', 'D'),
        ('30-39', 'D-'),
        ('0-29', 'E'),
    ]
    
    # Create grading table
    grading_data = [
        [Paragraph('<b>THEORY SCORES</b>', bold_style), '', Paragraph('<b>PRACTICAL SCORES</b>', bold_style), ''],
        [Paragraph('<b>Grade</b>', bold_style), Paragraph('<b>Scores%</b>', bold_style), 
         Paragraph('<b>Grade</b>', bold_style), Paragraph('<b>Scores%</b>', bold_style)],
    ]
    
    # Add grading rows
    max_rows = max(len(theory_bands), len(practical_bands))
    for i in range(max_rows):
        t_score, t_grade = theory_bands[i] if i < len(theory_bands) else ('', '')
        p_score, p_grade = practical_bands[i] if i < len(practical_bands) else ('', '')
        grading_data.append([
            Paragraph(t_grade, normal_style), Paragraph(t_score, normal_style),
            Paragraph(p_grade, normal_style), Paragraph(p_score, normal_style)
        ])
    
    grading_table = Table(grading_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
    grading_table.setStyle(TableStyle([
        ('SPAN', (0,0), (1,0)),  # Span THEORY SCORES
        ('SPAN', (2,0), (3,0)),  # Span PRACTICAL SCORES
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('BACKGROUND', (0,0), (-1,1), colors.lightgrey),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    
    elements.append(grading_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Pass mark information
    elements.append(Paragraph('<b>Pass mark is 50% in theory and 65% in practical assessment</b>', 
                             ParagraphStyle('PassMark', parent=bold_style, fontSize=11, alignment=TA_CENTER)))
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="testimonial_{candidate.reg_number}.pdf"'
    return response


def candidate_create(request):
    reg_cat = request.GET.get('registration_category')
    form_kwargs = {'user': request.user}
    if reg_cat:
        form_kwargs['initial'] = {'registration_category': reg_cat}
        if request.method == 'POST':
            form = CandidateForm(request.POST, request.FILES, **form_kwargs)
        else:
            form = CandidateForm(**form_kwargs)
    else:
        if request.method == 'POST':
            form = CandidateForm(request.POST, request.FILES, user=request.user)
        else:
            form = CandidateForm(user=request.user)
    # AJAX: return only occupation field HTML for dynamic update
    if request.headers.get('x-requested-with') == 'XMLHttpRequest' and reg_cat:
        occupation_field_html = render_to_string('partials/occupation_field.html', {'form': form})
        return JsonResponse({'occupation_field_html': occupation_field_html})
    # --- Handle form submission ---
    if request.method == 'POST':
        if form.is_valid():
            candidate = form.save(commit=False)
            candidate.created_by = request.user
            candidate.updated_by = request.user
            candidate.save()
            form.save_m2m()  # Ensure ManyToMany fields like nature_of_disability are saved
            return redirect('candidate_view', id=candidate.id)
    return render(request, 'candidates/create.html', {'form': form})



from django.contrib.auth.decorators import user_passes_test

@login_required
def edit_result(request, id):
    from .models import Candidate, Result, Module, CandidateLevel, OccupationLevel, Paper, AssessmentSeries
    from .forms import ModularResultsForm, ResultForm, PaperResultsForm, WorkerPASPaperResultsForm
    candidate = get_object_or_404(Candidate, id=id)
    
    # Check if user has edit permissions (superuser, IT, or Admin) using session management utility
    has_permission, staff, user_department = require_staff_permissions(request, ['IT', 'Admin'])
    
    if not has_permission:
        return redirect('candidate_view', id=candidate.id)
    reg_cat = getattr(candidate, 'registration_category', '').lower().strip()
    print('DEBUG edit_result: candidate', candidate, 'reg_cat', reg_cat)
    
    # Get all assessment series for the dropdown, ordered by current first, then by start date
    assessment_series = AssessmentSeries.objects.all().order_by('-is_current', '-start_date')
    context = {'candidate': candidate, 'edit_mode': True, 'assessment_series': assessment_series}

    # --- Modular candidate edit logic ---
    if reg_cat == 'modular':
        from .models import Module
        enrolled_modules = Module.objects.filter(candidatemodule__candidate=candidate)
        if not enrolled_modules.exists():
            context['form'] = None
            context['error'] = 'Candidate is not enrolled in any modules.'
            return render(request, 'candidates/add_result.html', context)
        existing_results = Result.objects.filter(candidate=candidate, result_type='modular')
        if request.method == 'POST':
            print(f'DEBUG modular edit POST: request.POST = {request.POST}')
            form = ModularResultsForm(request.POST, candidate=candidate)
            print(f'DEBUG modular edit: form created, is_valid = {form.is_valid()}')
            if not form.is_valid():
                print(f'DEBUG modular edit: form errors = {form.errors}')
            if form.is_valid():
                month = int(form.cleaned_data['month'])
                year = int(form.cleaned_data['year'])
                assessment_date = f"{year}-{month:02d}-01"
                
                # Find the matching assessment series for the month/year
                assessment_series = AssessmentSeries.objects.filter(
                    start_date__year=year,
                    start_date__month=month
                ).first()
                
                # If no matching series found, use candidate's current assessment series
                if not assessment_series:
                    assessment_series = candidate.assessment_series
                
                print(f'DEBUG modular edit: month={month}, year={year}, assessment_series={assessment_series}')
                
                for module in form.modules:
                    mark = form.cleaned_data.get(f'mark_{module.id}')
                    # Only create a new result if the mark has changed
                    existing_result = Result.objects.filter(
                        candidate=candidate,
                        module=module,
                        assessment_date=assessment_date,
                        result_type='modular',
                    ).first()
                    if mark is not None and (not existing_result or existing_result.mark != mark):
                        # Determine edit status for audit trail
                        edit_status = 'Updated' if existing_result else ''
                        if existing_result:
                            existing_date_str = existing_result.assessment_date.strftime('%Y-%m-%d') if existing_result.assessment_date else None
                            if existing_date_str != assessment_date:
                                edit_status = 'Retake'  # Different assessment date = retake
                        
                        Result.objects.create(
                            candidate=candidate,
                            module=module,
                            assessment_date=assessment_date,
                            assessment_series=assessment_series,
                            result_type='modular',
                            assessment_type='practical',
                            mark=mark,
                            user=request.user,
                            status=edit_status
                        )
                        print(f'DEBUG modular edit: created result for module {module.code}, mark={mark}, status={edit_status}')
                return redirect('candidate_view', id=candidate.id)
        else:
            # Prepopulate with existing marks if present
            initial = {}
            results_by_module = {r.module_id: r for r in existing_results}
            for module in enrolled_modules:
                result = results_by_module.get(module.id)
                if result:
                    initial[f'mark_{module.id}'] = result.mark
            # Prepopulate month/year from latest modular result
            latest_result = existing_results.order_by('-assessment_date').first()
            if latest_result and latest_result.assessment_date:
                initial['month'] = str(latest_result.assessment_date.month)
                initial['year'] = str(latest_result.assessment_date.year)
            form = ModularResultsForm(candidate=candidate, initial=initial)
        context['form'] = form
        context['is_modular'] = True
        context['module_mark_fields'] = [(module, form[f'mark_{module.id}']) for module in form.modules]
        return render(request, 'candidates/add_result.html', context)

    # --- Paper-based FORMAL edit logic ---
    level = getattr(candidate, 'level', None)
    from django.contrib import messages
    cl = CandidateLevel.objects.filter(candidate=candidate).first()
    if not level and cl:
        level = cl.level
    if not cl:
        messages.error(request, f'Candidate "{candidate}" is not enrolled. Please enroll candidate to add marks.')
        return redirect('candidate_view', id=candidate.id)
    is_paper_based = False
    if level:
        occ_level = OccupationLevel.objects.filter(occupation=candidate.occupation, level=level, structure_type='papers').first()
        if occ_level:
            is_paper_based = True
    if is_paper_based:
        # Fetch existing paper results
        papers = Paper.objects.filter(occupation=candidate.occupation, level=level)
        initial = {}
        for paper in papers:
            result = Result.objects.filter(candidate=candidate, paper=paper, result_type='formal').order_by('-assessment_date').first()
            if result:
                initial[f'mark_{paper.id}'] = result.mark
        # Prepopulate month/year from latest result
        latest_result = Result.objects.filter(candidate=candidate, paper__in=papers, result_type='formal').order_by('-assessment_date').first()
        if latest_result and latest_result.assessment_date:
            initial['month'] = latest_result.assessment_date.month
            initial['year'] = latest_result.assessment_date.year
        if request.method == 'POST':
            form = PaperResultsForm(request.POST, candidate=candidate, initial=initial)
            if form.is_valid():
                assessment_date = form.cleaned_data['assessment_date']
                for paper in form.papers:
                    mark = form.cleaned_data.get(f'mark_{paper.id}')
                    # Only create a new result if the mark has changed
                    existing_result = Result.objects.filter(candidate=candidate, paper=paper, result_type='formal').order_by('-assessment_date').first()
                    if mark is not None and (not existing_result or existing_result.mark != mark):
                        Result.objects.create(
                            candidate=candidate,
                            level=level,
                            paper=paper,
                            assessment_type=paper.grade_type,
                            assessment_date=assessment_date,
                            result_type='formal',
                            mark=mark,
                            user=request.user,
                            status='Updated'
                        )
                return redirect('candidate_view', id=candidate.id)
        else:
            form = PaperResultsForm(candidate=candidate, initial=initial)
        context['form'] = form
        context['is_paper_based'] = True
        # Debug: Check if form has papers
        print('DEBUG edit_result paper-based: form.papers:', getattr(form, 'papers', 'No papers attribute'))
        if hasattr(form, 'papers') and form.papers:
            context['paper_mark_fields'] = [(paper, form[f'mark_{paper.id}']) for paper in form.papers]
            print('DEBUG edit_result paper-based: paper_mark_fields created for', len(form.papers), 'papers')
        else:
            print('DEBUG edit_result paper-based: No papers found, falling back to formal fields')
            # Fallback to formal fields if no papers found
            context['is_paper_based'] = False
            context['is_modular'] = False
        return render(request, 'candidates/add_result.html', context)
    # --- END paper-based FORMAL edit logic ---

    # --- Informal/Worker's PAS edit logic ---
    if reg_cat == 'informal':
        from .forms import WorkerPASPaperResultsForm
        from .models import CandidatePaper, Level
        # Get level from GET param if present, else use first enrolled level
        level_id = request.GET.get('level')
        level = None
        if level_id:
            try:
                level = Level.objects.get(id=level_id)
            except Level.DoesNotExist:
                level = None
        if not level:
            cl = CandidateLevel.objects.filter(candidate=candidate).first()
            if cl:
                level = cl.level
        if not level:
            context['form'] = None
            context['error'] = 'Candidate is not enrolled in any level.'
            return render(request, 'candidates/add_result.html', context)
        # Get enrolled papers for this candidate/level
        enrolled_papers = CandidatePaper.objects.filter(candidate=candidate, level=level).select_related('paper', 'module')
        if not enrolled_papers.exists():
            context['form'] = None
            context['error'] = 'Candidate has no enrolled papers.'
            return render(request, 'candidates/add_result.html', context)
        # Prepopulate marks for each enrolled paper
        initial = {}
        for cp in enrolled_papers:
            result = Result.objects.filter(candidate=candidate, paper=cp.paper, level=level, result_type='informal').order_by('-assessment_date').first()
            if result:
                initial[f'mark_{cp.paper.id}'] = result.mark
        # Prepopulate month/year from latest result
        latest_result = Result.objects.filter(candidate=candidate, paper__in=[cp.paper for cp in enrolled_papers], level=level, result_type='informal').order_by('-assessment_date').first()
        if latest_result and latest_result.assessment_date:
            initial['month'] = latest_result.assessment_date.month
            initial['year'] = latest_result.assessment_date.year
        if request.method == 'POST':
            form = WorkerPASPaperResultsForm(request.POST, candidate=candidate, level=level, initial=initial)
            if form.is_valid():
                assessment_date = form.cleaned_data['assessment_date']
                for cp in enrolled_papers:
                    paper = cp.paper
                    mark = form.cleaned_data.get(f'mark_{paper.id}')
                    # Only update/create if mark is changed
                    existing_result = Result.objects.filter(candidate=candidate, paper=paper, level=level, result_type='informal').order_by('-assessment_date').first()
                    if mark is not None and (not existing_result or existing_result.mark != mark):
                        Result.objects.create(
                            candidate=candidate,
                            level=level,
                            module=cp.module,
                            paper=paper,
                            assessment_type='practical',
                            assessment_date=assessment_date,
                            result_type='informal',
                            mark=mark,
                            user=request.user,
                            status='Updated'
                        )
                return redirect('candidate_view', id=candidate.id)
        else:
            form = WorkerPASPaperResultsForm(candidate=candidate, level=level, initial=initial)
        context['form'] = form
        context['is_worker_pas'] = True
        context['paper_mark_fields'] = [(cp.paper, form[f'mark_{cp.paper.id}']) for cp in enrolled_papers]
        return render(request, 'candidates/add_result.html', context)
    # --- END Informal/Worker's PAS edit logic ---

    elif reg_cat == 'formal':
        from .forms import FormalResultsForm
        from .models import Level, CandidateLevel
        # Find the candidate's enrolled level
        level_enrollment = CandidateLevel.objects.filter(candidate=candidate).first()
        level = level_enrollment.level if level_enrollment else None
        existing_theory = None
        existing_practical = None
        # Fetch latest theory/practical results for this candidate (not filtered by level)
        theory_result = Result.objects.filter(candidate=candidate, assessment_type='theory', result_type='formal').order_by('-assessment_date').first()
        practical_result = Result.objects.filter(candidate=candidate, assessment_type='practical', result_type='formal').order_by('-assessment_date').first()
        if theory_result:
            existing_theory = theory_result.mark
        if practical_result:
            existing_practical = practical_result.mark
        initial = {}
        # Prepopulate marks
        if existing_theory is not None:
            initial['theory_mark'] = existing_theory
        if existing_practical is not None:
            initial['practical_mark'] = existing_practical
        # Prepopulate month/year from the latest result (theory or practical)
        assessment_date = None
        if theory_result and getattr(theory_result, 'assessment_date', None):
            assessment_date = theory_result.assessment_date
        elif practical_result and getattr(practical_result, 'assessment_date', None):
            assessment_date = practical_result.assessment_date
        if assessment_date:
            initial['month'] = str(assessment_date.month)
            initial['year'] = str(assessment_date.year)
        
        print('DEBUG edit_result: initial values:', initial)
        print('DEBUG edit_result: existing_theory:', existing_theory)
        print('DEBUG edit_result: existing_practical:', existing_practical)
        if request.method == 'POST':
            form = FormalResultsForm(request.POST, candidate=candidate)
            if form.is_valid() and level:
                theory_mark = form.cleaned_data.get('theory_mark')
                practical_mark = form.cleaned_data.get('practical_mark')
                month = int(form.cleaned_data.get('month')) if 'month' in form.cleaned_data else None
                year = int(form.cleaned_data.get('year')) if 'year' in form.cleaned_data else None
                if month and year:
                    assessment_date_str = f"{year}-{month:02d}-01"
                else:
                    assessment_date_str = None
                # Determine if this is an update (same assessment date) or retake (different assessment date)
                def get_edit_status(existing_result, new_assessment_date_str):
                    if not existing_result:
                        return ''  # New result, no status needed
                    
                    existing_date_str = existing_result.assessment_date.strftime('%Y-%m-%d') if existing_result.assessment_date else None
                    if existing_date_str == new_assessment_date_str:
                        return 'Updated'  # Same assessment series - mark correction/update
                    else:
                        return 'Retake'   # Different assessment series - actual retake
                
                # Only create a new Result if the mark has changed
                if theory_result and theory_result.mark != theory_mark:
                    edit_status = get_edit_status(theory_result, assessment_date_str)
                    Result.objects.create(
                        candidate=candidate,
                        level=level,
                        assessment_type='theory',
                        result_type='formal',
                        mark=theory_mark,
                        assessment_date=assessment_date_str,
                        status=edit_status,
                        user=request.user
                    )
                elif not theory_result and theory_mark is not None:
                    Result.objects.create(
                        candidate=candidate,
                        level=level,
                        assessment_type='theory',
                        result_type='formal',
                        mark=theory_mark,
                        assessment_date=assessment_date_str,
                        status='',
                        user=request.user
                    )
                if practical_result and practical_result.mark != practical_mark:
                    edit_status = get_edit_status(practical_result, assessment_date_str)
                    Result.objects.create(
                        candidate=candidate,
                        level=level,
                        assessment_type='practical',
                        result_type='formal',
                        mark=practical_mark,
                        assessment_date=assessment_date_str,
                        status=edit_status,
                        user=request.user
                    )
                elif not practical_result:
                    Result.objects.create(
                        candidate=candidate,
                        level=level,
                        assessment_type='practical',
                        result_type='formal',
                        mark=practical_mark,
                        assessment_date=assessment_date_str,
                        status='',
                        user=request.user
                    )
                return redirect('candidate_view', id=candidate.id)
        else:
            form = FormalResultsForm(candidate=candidate, initial=initial)
        context['form'] = form
        context['is_modular'] = False
        # Note: template now uses form.field_name directly instead of formal_mark_fields array
        return render(request, 'candidates/add_result.html', context)
    else:
        return render(request, 'candidates/add_result.html', context)

    # Fallback for unknown categories
    context['form'] = None
    context['error'] = f'Unknown or unsupported registration category: {reg_cat!r} for candidate {candidate}'
    print('DEBUG edit_result: unknown registration_category', reg_cat, 'for candidate', candidate)
    return render(request, 'candidates/add_result.html', context)


def add_result(request, id):
    from .models import Candidate, Result, Module, Paper, Level, OccupationLevel, CandidateLevel, CandidatePaper, AssessmentSeries
    from .forms import ResultForm, ModularResultsForm, WorkerPASPaperResultsForm
    candidate = get_object_or_404(Candidate, id=id)
    reg_cat = getattr(candidate, 'registration_category', '').strip().lower()
    
    # Get all assessment series for the dropdown, ordered by current first, then by start date
    assessment_series = AssessmentSeries.objects.all().order_by('-is_current', '-start_date')
    context = {'candidate': candidate, 'assessment_series': assessment_series}

    # Worker PAS/informal: mark per enrolled paper
    if reg_cat == "Informal":
        # Get level from GET param if present, else use first enrolled level
        level_id = request.GET.get('level')
        level = None
        if level_id:
            try:
                from .models import Level
                level = Level.objects.get(id=level_id)
            except Level.DoesNotExist:
                level = None
        if not level:
            cl = CandidateLevel.objects.filter(candidate=candidate).first()
            if cl:
                level = cl.level
        if not level:
            context['form'] = None
            context['error'] = 'Candidate is not enrolled in any level.'
            return render(request, 'candidates/add_result.html', context)
        # Get enrolled papers for this candidate/level
        enrolled_papers = CandidatePaper.objects.filter(candidate=candidate, level=level).select_related('paper', 'module')
        if not enrolled_papers.exists():
            context['form'] = None
            context['error'] = 'Candidate has no enrolled papers.'
            return render(request, 'candidates/add_result.html', context)
        if request.method == 'POST':
            # Get assessment series from form
            assessment_series_id = request.POST.get('assessment_series')
            assessment_series = None
            if assessment_series_id:
                try:
                    assessment_series = AssessmentSeries.objects.get(id=assessment_series_id)
                except AssessmentSeries.DoesNotExist:
                    assessment_series = candidate.assessment_series
            else:
                assessment_series = candidate.assessment_series
            
            # Use assessment series date or current date
            assessment_date = assessment_series.start_date if assessment_series else timezone.now().date()
            
            # Process marks for each enrolled paper
            for cp in enrolled_papers:
                paper = cp.paper
                mark_field_name = f'mark_{paper.id}'
                mark = request.POST.get(mark_field_name)
                if mark is not None and mark != '':
                    try:
                        mark = float(mark)
                        Result.objects.update_or_create(
                            candidate=candidate,
                            level=level,
                            module=cp.module,
                            paper=paper,
                            assessment_type='practical',
                            assessment_date=assessment_date,
                            result_type='informal',
                            defaults={
                                'assessment_series': assessment_series,
                                'mark': mark,
                                'user': request.user,
                                'status': ''
                            }
                        )
                    except ValueError:
                        pass  # Skip invalid marks
                return redirect('candidate_view', id=candidate.id)
        else:
            # Default assessment month/year to candidate's enrolled period for this level
            initial = {}
            cl = CandidateLevel.objects.filter(candidate=candidate, level=level).first()
            if cl and cl.start_date:
                initial['month'] = str(cl.start_date.month)
                initial['year'] = str(cl.start_date.year)
            form = WorkerPASPaperResultsForm(candidate=candidate, level=level, initial=initial)
        context['form'] = form
        context['is_worker_pas'] = True
        context['paper_mark_fields'] = [(cp.paper, form[f'mark_{cp.paper.id}']) for cp in enrolled_papers]
        return render(request, 'candidates/add_result.html', context)



    if reg_cat == 'modular':
        if request.method == 'POST':
            # Get assessment series from form
            assessment_series_id = request.POST.get('assessment_series')
            assessment_series = None
            if assessment_series_id:
                try:
                    assessment_series = AssessmentSeries.objects.get(id=assessment_series_id)
                except AssessmentSeries.DoesNotExist:
                    assessment_series = candidate.assessment_series
            else:
                assessment_series = candidate.assessment_series
            
            # Use assessment series date or current date
            assessment_date = assessment_series.start_date if assessment_series else timezone.now().date()
            
            # Get candidate's enrolled modules
            from .models import CandidateModule
            enrolled_modules = CandidateModule.objects.filter(candidate=candidate).select_related('module')
            
            for cm in enrolled_modules:
                module = cm.module
                mark_field_name = f'mark_{module.id}'
                mark = request.POST.get(mark_field_name)
                if mark is not None and mark != '':
                    try:
                        mark = float(mark)
                        result, created = Result.objects.update_or_create(
                            candidate=candidate,
                            module=module,
                            assessment_date=assessment_date,
                            result_type='modular',
                            defaults={
                                'level': None,  # Modular results are not tied to a level
                                'assessment_series': assessment_series,
                                'assessment_type': 'practical',
                                'mark': mark,
                                'user': request.user,
                                'status': ''
                            }
                        )
                    except ValueError:
                        pass  # Skip invalid marks
            return redirect('candidate_view', id=candidate.id)
        else:
            form = ModularResultsForm(candidate=candidate)
        context['form'] = form
        context['is_modular'] = True
        # Pass (module, field) pairs for template rendering
        context['module_mark_fields'] = [(module, form[f'mark_{module.id}']) for module in form.modules]
    else:
        from .forms import FormalResultsForm, PaperResultsForm
        from .models import OccupationLevel
        # Determine if candidate's level is paper-based
        from .models import CandidateLevel
        level = getattr(candidate, 'level', None)
        if not level:
            cl = CandidateLevel.objects.filter(candidate=candidate).first()
            if cl:
                level = cl.level
        is_paper_based = False
        if level:
            occ_level = OccupationLevel.objects.filter(occupation=candidate.occupation, level=level, structure_type='papers').first()
            if occ_level:
                is_paper_based = True
        if is_paper_based:
            if request.method == 'POST':
                form = PaperResultsForm(request.POST, candidate=candidate)
                if form.is_valid():
                    assessment_date = form.cleaned_data['assessment_date']
                    for paper in form.papers:
                        mark = form.cleaned_data.get(f'mark_{paper.id}')
                        if mark is not None:
                            from .models import Result
                            Result.objects.create(
                                candidate=candidate,
                                level=level,
                                paper=paper,
                                assessment_type=paper.grade_type,
                                assessment_date=assessment_date,
                                result_type='formal',
                                mark=mark,
                                user=request.user,
                                status=''
                            )
                    return redirect('candidate_view', id=candidate.id)
            else:
                form = PaperResultsForm(candidate=candidate)
            context['form'] = form
            context['is_paper_based'] = True
            context['paper_mark_fields'] = [(paper, form[f'mark_{paper.id}']) for paper in getattr(form, 'papers', [])]
        else:
            if request.method == 'POST':
                form = FormalResultsForm(request.POST, candidate=candidate)
                print(f"DEBUG: Form data received: {request.POST}")
                print(f"DEBUG: Form is_valid: {form.is_valid()}")
                if not form.is_valid():
                    print(f"DEBUG: Form errors: {form.errors}")
                if form.is_valid():
                    from datetime import date
                    today = date.today()
                    assessment_date = today.strftime("%Y-%m-01")
                    theory_mark = form.cleaned_data.get('theory_mark')
                    practical_mark = form.cleaned_data.get('practical_mark')
                    print(f"DEBUG: theory_mark = {theory_mark}, practical_mark = {practical_mark}")
                    print(f"DEBUG: assessment_date = {assessment_date}")
                    
                    # Save theory result (combined for the level)
                    if theory_mark is not None:
                        result, created = Result.objects.update_or_create(
                            candidate=candidate,
                            level=level,
                            assessment_date=assessment_date,
                            assessment_type='theory',
                            result_type='formal',
                            defaults={'mark': theory_mark, 'user': request.user}
                        )
                        print(f"DEBUG: Theory result {'created' if created else 'updated'}: {result}")
                    
                    # Save practical result (combined for the level)
                    if practical_mark is not None:
                        result, created = Result.objects.update_or_create(
                            candidate=candidate,
                            level=level,
                            assessment_date=assessment_date,
                            assessment_type='practical',
                            result_type='formal',
                            defaults={'mark': practical_mark, 'user': request.user}
                        )
                        print(f"DEBUG: Practical result {'created' if created else 'updated'}: {result}")
                    
                    return redirect('candidate_view', id=candidate.id)
            else:
                form = FormalResultsForm(candidate=candidate)
            context['form'] = form
            context['is_modular'] = False
            # Pass just the two fields for the template
            context['formal_mark_fields'] = [form['theory_mark'], form['practical_mark']]
    return render(request, 'candidates/add_result.html', context)



def edit_candidate(request, id):
    candidate = get_object_or_404(Candidate, id=id)

    if request.method == 'POST':
        form = CandidateForm(request.POST, request.FILES, instance=candidate, edit=True)
        if form.is_valid():
            candidate = form.save(commit=False)
            candidate.updated_by = request.user
            candidate.save()
            form.save_m2m()  # Ensure ManyToMany fields like nature_of_disability are saved
            return redirect('candidate_view', id=candidate.id)
    else:
        form = CandidateForm(instance=candidate, edit=True)

    return render(request, 'candidates/edit.html', {'form': form, 'candidate': candidate})


def enroll_candidate_view(request, id):
    candidate = get_object_or_404(Candidate, id=id)

    # Only enroll if POST and _enroll=1
    if request.method == 'POST' and request.POST.get('_enroll') == '1':
        form = EnrollmentForm(request.POST, candidate=candidate)
        if form.is_valid():
            # Get the selected assessment series
            assessment_series = form.cleaned_data.get('assessment_series')     

            # Update candidate's assessment series
            if assessment_series:
                candidate.assessment_series = assessment_series
                candidate.save() 

            registration_category = candidate.registration_category
            

            # Handle formal registration (level only)
            if registration_category == 'Formal':
                level = form.cleaned_data.get('level')
                if not level:
                    form.add_error('level', 'Please select a level.')
                    return render(request, 'candidates/enroll.html', {
                        'form': form,
                        'candidate': candidate,
                    })
                # Check if already enrolled in this level
                if CandidateLevel.objects.filter(candidate=candidate, level=level).exists():
                    messages.error(request, "Candidate already enrolled for this level.")
                    return render(request, 'candidates/enroll.html', {
                        'form': form,
                        'candidate': candidate,
                    })
                # Not already enrolled: enroll the candidate
                CandidateLevel.objects.create(candidate=candidate, level=level)
                # Update fees balance after enrollment
                candidate.update_fees_balance()
                messages.success(request, f'{candidate.full_name} successfully enrolled in {level.name}.')
            
            # Handle modular registration (progressive enrollment system)
            elif registration_category == 'Modular':
                modules = form.cleaned_data['modules']
                
                # Validate enrollment constraints
                if len(modules) > 2:
                    messages.error(request, "You can only select up to 2 modules at a time.")
                    return render(request, 'candidates/enroll.html', {
                        'form': form,
                        'candidate': candidate,
                    })
                
                # Check if candidate can enroll in more modules
                if not candidate.can_enroll_in_more_modules():
                    messages.error(request, "You have reached the maximum concurrent enrollments (2 modules). Complete current modules before enrolling in new ones.")
                    return render(request, 'candidates/enroll.html', {
                        'form': form,
                        'candidate': candidate,
                    })
                
                # Check for duplicate enrollments
                already_enrolled = []
                for module in modules:
                    if CandidateModule.objects.filter(candidate=candidate, module=module).exists():
                        already_enrolled.append(module.name)
                
                if already_enrolled:
                    messages.error(request, f"Already enrolled in: {', '.join(already_enrolled)}")
                    return render(request, 'candidates/enroll.html', {
                        'form': form,
                        'candidate': candidate,
                    })
                
                # Enroll candidate in selected modules (progressive enrollment - don't delete existing)
                enrolled_modules = []
                for module in modules:
                    candidate_module = CandidateModule.objects.create(
                        candidate=candidate, 
                        module=module,
                        assessment_series=assessment_series,
                        status='enrolled'
                    )
                    enrolled_modules.append(module.name)
                
                # Update fees balance after enrollment
                candidate.update_fees_balance()
                
                # Check if candidate has completed all modules for qualification
                completion_status = candidate.get_modular_completion_status()
                if completion_status and completion_status['is_qualified']:
                    messages.success(request, f"{candidate.full_name} enrolled in {', '.join(enrolled_modules)}. Congratulations! You have completed all modules and are qualified for Level 1!")
                else:
                    remaining = completion_status['remaining_modules'] if completion_status else 0
                    messages.success(request, f"{candidate.full_name} enrolled in {', '.join(enrolled_modules)}. {remaining} module(s) remaining for Level 1 qualification.")

            # Handle informal registration (cross-level paper selection)
            elif registration_category in ['Informal', "Worker's PAS", 'Workers PAS', 'informal', "worker's pas"]:
                from .models import CandidatePaper, Level
                selected_papers = form.cleaned_data.get('selected_papers', [])
                
                # For retakes, preserve all previous enrollments and results
                # Only clear enrollments for the CURRENT assessment series to avoid duplicates
                current_series = AssessmentSeries.objects.filter(is_current=True).first()
                if current_series:
                    CandidatePaper.objects.filter(
                        candidate=candidate, 
                        enrolled_at__gte=current_series.start_date,
                        enrolled_at__lte=current_series.end_date
                    ).delete()
                
                # Track levels and modules we need to enroll in
                levels_to_enroll = set()
                modules_to_enroll = set()
                papers_enrolled = []
                
                # Process each selected paper
                for paper_data in selected_papers:
                    paper = paper_data['paper']
                    level_id = paper_data['level_id']
                    module_id = paper_data['module_id']
                    
                    # Get level and module objects
                    try:
                        level = Level.objects.get(id=level_id)
                        module = Module.objects.get(id=module_id)
                        
                        levels_to_enroll.add(level)
                        modules_to_enroll.add(module)
                        papers_enrolled.append(paper)
                        
                        # Create CandidatePaper record
                        CandidatePaper.objects.create(
                            candidate=candidate, 
                            module=module, 
                            paper=paper, 
                            level=level
                        )
                        
                    except (Level.DoesNotExist, Module.DoesNotExist) as e:
                        messages.error(request, f"Error processing paper selection: {e}")
                        return render(request, 'candidates/enroll.html', {
                            'form': form,
                            'candidate': candidate,
                        })
                
                # Enroll in all required levels (avoid duplicates for retakes)
                for level in levels_to_enroll:
                    CandidateLevel.objects.get_or_create(candidate=candidate, level=level)
                
                # Enroll in all required modules (avoid duplicates for retakes)
                for module in modules_to_enroll:
                    CandidateModule.objects.get_or_create(candidate=candidate, module=module)
                
                # Update fees balance after enrollment
                candidate.update_fees_balance()
                
                # Create success message
                level_names = [level.name for level in levels_to_enroll]
                paper_names = [paper.name for paper in papers_enrolled]
                messages.success(
                    request, 
                    f"{candidate.full_name} enrolled across {len(levels_to_enroll)} level(s): {', '.join(level_names)}. "
                    f"Selected {len(papers_enrolled)} paper(s): {', '.join(paper_names)}"
                )

            messages.success(request, "Candidate enrolled successfully.")
    else:
        # Support dynamic module filtering by selected level (GET param)
        form = EnrollmentForm(request.GET, candidate=candidate)
    return render(request, 'candidates/enroll.html', {
        'form': form,
        'candidate': candidate,
    })


from .models import Candidate, CandidateLevel, CandidateModule, Occupation, CandidatePaper

from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.views.decorators.http import require_POST

@login_required
@require_POST
def clear_enrollment(request, id):
    if not request.user.is_superuser:
        return redirect('candidate_view', id=id)
    candidate = get_object_or_404(Candidate, id=id)
    from .models import Result
    from django.contrib import messages
    if Result.objects.filter(candidate=candidate).exists():
        messages.error(request, 'Candidate has marks/results and cannot be de-enrolled.')
        return redirect('candidate_view', id=id)
    CandidateLevel.objects.filter(candidate=candidate).delete()
    CandidateModule.objects.filter(candidate=candidate).delete()
    CandidatePaper.objects.filter(candidate=candidate).delete()
    # Also clear the assessment series assignment
    candidate.assessment_series = None
    # Reset fees balance to zero after clearing enrollment
    candidate.fees_balance = 0.00
    candidate.save()
    messages.success(request, 'All enrollment records and assessment series assignment for this candidate have been cleared.')
    return redirect('candidate_view', id=id)


# Helper
def _blocked_if_enrolled(request, candidate, action_name):
    if candidate.is_enrolled():
        messages.error(
            request,
            f"Candidate is already enrolled – cannot {action_name.lower()}."
        )
        return True
    return False

import json

@require_POST
def change_occupation(request, id):
    candidate = get_object_or_404(Candidate, id=id)

    if candidate.is_enrolled():
        return JsonResponse({"success": False,
                             "error": "Enrolled candidates cannot be changed."})

    try:
        data       = json.loads(request.body or "{}")
        new_occ_id = int(data["occupation"])
        new_occ    = Occupation.objects.get(pk=new_occ_id)
    except (KeyError, ValueError, Occupation.DoesNotExist):
        return HttpResponseBadRequest("Invalid occupation")

    candidate.occupation = new_occ
    candidate.build_reg_number()     # helper that rebuilds reg_number
    candidate.save(update_fields=["occupation", "reg_number"])

    return JsonResponse({
        "success": True,
        "occupation_name": new_occ.name,
        "reg_number": candidate.reg_number
    })

@require_POST
def change_center(request, id):
    import json
    candidate = get_object_or_404(Candidate, id=id)

    if candidate.is_enrolled():
        return JsonResponse({"success": False,
                             "error": "Enrolled candidates cannot be changed."})

    try:
        data = json.loads(request.body or "{}")
        new_center_id = int(data["assessment_center"])
        new_center = AssessmentCenter.objects.get(pk=new_center_id)
    except (KeyError, ValueError, AssessmentCenter.DoesNotExist):
        return HttpResponseBadRequest("Invalid center")

    candidate.assessment_center = new_center
    candidate.build_reg_number()     # helper that rebuilds reg_number
    candidate.save(update_fields=["assessment_center", "reg_number"])

    return JsonResponse({
        "success": True,
        "center_name": new_center.center_name,
        "reg_number": candidate.reg_number
    })

@require_POST
def change_registration_category(request, id):
    candidate = get_object_or_404(Candidate, id=id)

    if hasattr(candidate, 'is_enrolled') and callable(candidate.is_enrolled):
        if candidate.is_enrolled():
            return JsonResponse({"success": False, "error": "Enrolled candidates cannot be changed."})
    elif getattr(candidate, 'is_enrolled', False):
        return JsonResponse({"success": False, "error": "Enrolled candidates cannot be changed."})

    try:
        data = json.loads(request.body or '{}')
        new_reg_cat = data["registration_category"]
    except (KeyError, ValueError, TypeError):
        return HttpResponseBadRequest("Invalid registration category")

    occupation = candidate.occupation
    occ_cat = occupation.category.name if occupation and occupation.category else None
    has_modular = getattr(occupation, 'has_modular', False)

    # Validate compatibility
    if new_reg_cat == 'Modular':
        if not has_modular:
            return JsonResponse({"success": False, "error": "This occupation does not support Modular registration."})
    elif new_reg_cat == 'Formal':
        if occ_cat != 'Formal':
            return JsonResponse({"success": False, "error": "Only occupations in the 'Formal' category can be registered as Formal."})
    elif new_reg_cat == 'Informal':
        if occ_cat != "Worker's PAS":
            return JsonResponse({"success": False, "error": "Only occupations in the 'Worker's PAS' category can be registered as Informal."})
    else:
        return JsonResponse({"success": False, "error": "Invalid registration category selected."})

    candidate.registration_category = new_reg_cat
    candidate.reg_number = None
    candidate.save(update_fields=["registration_category", "reg_number"])
    # Regenerate reg_number as in regenerate_candidate_reg_number
    candidate.reg_number = None
    candidate.save(update_fields=["reg_number"])

    return JsonResponse({
        "success": True,
        "registration_category": new_reg_cat,
        "reg_number": candidate.reg_number
    })

@login_required
def candidate_view(request, id):
    print('DEBUG: candidate_view (ACTIVE) called for candidate', id)
    from .models import AssessmentCenter, Occupation, Result, CandidateLevel, CandidateModule, Paper, Module, CandidatePaper, AssessmentSeries
    candidate = get_object_or_404(Candidate, id=id)

    # Check if results have been released for the candidate's assessment series
    # Only block center representatives, allow admin staff and support to see results
    is_center_rep = request.user.groups.filter(name='CenterRep').exists()
    candidate_series = candidate.assessment_series
    results_released = candidate_series and candidate_series.results_released or not is_center_rep

    reg_cat = candidate.registration_category
    reg_cat_normalized = reg_cat.strip().lower() if reg_cat else ''
    level_enrollment = CandidateLevel.objects.filter(candidate=candidate).first()

    # Initialize module_enrollments for all registration categories to prevent UnboundLocalError
    module_enrollments = CandidateModule.objects.filter(candidate=candidate)
    
    # Multi-level enrollment summary for worker's PAS/informal/modular
    enrollment_summary = []
    if reg_cat_normalized in ["informal", "worker's pas", "workers pas"]:
        level_enrollments = CandidateLevel.objects.filter(candidate=candidate)
        for lvl_enroll in level_enrollments:
            modules = CandidateModule.objects.filter(candidate=candidate, module__level=lvl_enroll.level)
            module_list = []
            for mod_enroll in modules:
                # Attach all papers for this candidate/module/level (may be >1 after bulk)
                papers = list(
                    CandidatePaper.objects.filter(candidate=candidate, module=mod_enroll.module, level=lvl_enroll.level).select_related('paper')
                )
                module_list.append({
                    'module': mod_enroll.module,
                    'papers': [cp.paper for cp in papers]
                })
            enrollment_summary.append({
                'level': lvl_enroll.level,
                'modules': module_list
            })
    # For modular candidates, use simple results display like formal candidates
    # No enrollment_summary needed
    else:
        # For formal/modular, keep old logic - module_enrollments already initialized above
        for mod_enroll in module_enrollments:
            mod_enroll.papers = list(Paper.objects.filter(module=mod_enroll.module))

    # Attach papers for each module_enrollment (for backward compatibility)
    if reg_cat_normalized in ["worker's pas", "workers pas", "informal"] and level_enrollment:
        for mod_enroll in module_enrollments:
            candidate_paper = CandidatePaper.objects.filter(
                candidate_id=candidate.id,
                module_id=mod_enroll.module.id,
                level_id=level_enrollment.level.id
            ).select_related('paper').first()
            mod_enroll.papers = [candidate_paper.paper] if candidate_paper else []
    else:
        for mod_enroll in module_enrollments:
            mod_enroll.papers = list(Paper.objects.filter(module=mod_enroll.module))

    # Only show results if they have been released - show ALL historical results, not just current enrollment
    if results_released:
        if reg_cat_normalized in ["informal", "worker's pas", "workers pas"]:
            # Show ALL results for this candidate across all assessment series and enrollments
            results = Result.objects.filter(candidate=candidate)
        elif reg_cat_normalized == "modular":
            # For modular candidates, get results for all enrolled modules
            enrolled_module_ids = list(CandidateModule.objects.filter(candidate=candidate).values_list('module_id', flat=True))
            results = Result.objects.filter(candidate=candidate, module_id__in=enrolled_module_ids)
        else:
            results = Result.objects.filter(candidate=candidate).order_by('assessment_date', 'level', 'module', 'paper')
    else:
        # Results not released - show empty results
        results = Result.objects.none()
    level_has_results = {}
    
    # For Informal/Worker's PAS: Show Add/Edit Marks for ALL enrolled levels AND levels with results
    if reg_cat_normalized in ["informal", "worker's pas", "workers pas"]:
        # Get all enrolled levels (for Add Marks button)
        enrolled_levels = CandidateLevel.objects.filter(candidate=candidate).values_list('level_id', flat=True)
        
        # Get all levels that have results (for Edit Marks button)
        levels_with_results = results.values_list('level_id', flat=True).distinct()
        
        # Combine both enrolled levels and levels with results
        all_relevant_levels = set(enrolled_levels) | set(levels_with_results)
        
        for level_id in all_relevant_levels:
            # Always show button for enrolled levels or levels with results
            level_has_results[str(level_id)] = True
    else:
        # For other registration categories: use enrollment-based logic
        for row in enrollment_summary:
            lvl = row['level']
            enrolled_paper_ids = set()
            for mod in row['modules']:
                for paper in mod['papers']:
                    enrolled_paper_ids.add(paper.id)
            result_paper_ids = set(results.filter(level_id=lvl.id).values_list('paper_id', flat=True))
            
            # Only True if ALL enrolled papers have results for this level
            level_has_results[str(lvl.id)] = bool(enrolled_paper_ids) and enrolled_paper_ids == result_paper_ids and len(result_paper_ids) > 0
            
            print(f"DEBUG level_has_results for level {lvl.id} (key='{str(lvl.id)}'): enrolled_papers={enrolled_paper_ids}, result_papers={result_paper_ids}, results_with_marks={results_with_marks}, final={level_has_results[str(lvl.id)]}")
    
    print(f"DEBUG: Final level_has_results dictionary: {level_has_results}")
    # Convert all keys to string for template consistency
    # (already done above for lvl.id)

    # For Informal/Worker's PAS: Create comprehensive results summary including all levels with results
    if reg_cat_normalized in ["informal", "worker's pas", "workers pas"]:
        # Get all levels that have results, not just enrolled levels
        all_results_levels = results.values('level_id', 'level__name').distinct()
        
        # Create comprehensive results summary including historical results
        comprehensive_results_summary = []
        
        # Group by level_id to avoid duplicates
        levels_processed = set()
        for level_data in all_results_levels:
            level_id = level_data['level_id']
            
            # Skip if we've already processed this level
            if level_id in levels_processed:
                continue
            levels_processed.add(level_id)
            
            level = Level.objects.get(id=level_id)
            
            # Get all results for this level across all assessment series
            level_results = results.filter(level_id=level_id)
            
            # Group results by module, then by paper
            modules_with_results = {}
            for result in level_results:
                module_id = result.module.id
                if module_id not in modules_with_results:
                    modules_with_results[module_id] = {
                        'module': result.module,
                        'papers': set(),  # Use set to avoid duplicate papers
                        'results': []
                    }
                modules_with_results[module_id]['papers'].add(result.paper)
                modules_with_results[module_id]['results'].append(result)
            
            # Convert paper sets back to lists
            for module_data in modules_with_results.values():
                module_data['papers'] = list(module_data['papers'])
            
            comprehensive_results_summary.append({
                'level': level,
                'modules': list(modules_with_results.values())
            })
        
        # Use comprehensive summary for results display, but also include enrolled levels without results
        results_summary_for_display = comprehensive_results_summary
        
        # Always add ALL enrolled levels to results_summary (whether they have results or not)
        enrolled_levels = CandidateLevel.objects.filter(candidate=candidate).select_related('level')
        
        # Get levels that already have results (to avoid duplicates)
        levels_with_results = set(result.level.id for result in results)
        
        for level_enrollment in enrolled_levels:
            level = level_enrollment.level
            
            # Skip if this level already has results (already in comprehensive_results_summary)
            if level.id in levels_with_results:
                continue
                
            # Get enrolled modules for this level
            enrolled_modules = CandidateModule.objects.filter(
                candidate=candidate, 
                module__level=level
            ).select_related('module')
            
            modules_data = []
            for module_enrollment in enrolled_modules:
                # Get enrolled papers for this module
                enrolled_papers = CandidatePaper.objects.filter(
                    candidate=candidate,
                    level=level,
                    module=module_enrollment.module
                ).select_related('paper')
                
                if enrolled_papers.exists():
                    modules_data.append({
                        'module': module_enrollment.module,
                        'papers': [cp.paper for cp in enrolled_papers],
                        'results': []  # No results yet
                    })
            
            if modules_data:
                results_summary_for_display.append({
                    'level': level,
                    'modules': modules_data
                })
        
        # Create enrollment history - use both enrollment records AND results
        series_level_enrollments = {}
        
        # First, add enrollment records (for candidates with enrollments but no results)
        candidate_papers = CandidatePaper.objects.filter(candidate=candidate).select_related(
            'level', 'module', 'paper'
        ).order_by('-enrolled_at')
        
        for cp in candidate_papers:
            # Use candidate's assessment series if available, otherwise try to determine from date
            if candidate.assessment_series:
                series_key = f"{candidate.assessment_series.name} ({candidate.assessment_series.start_date.strftime('%b %Y')})"
            else:
                # Fallback: determine from enrollment date
                series_key = None
                enrolled_date = cp.enrolled_at.date() if hasattr(cp.enrolled_at, 'date') else cp.enrolled_at
                
                for series in AssessmentSeries.objects.all().order_by('-start_date'):
                    if enrolled_date >= series.start_date and enrolled_date <= series.end_date:
                        series_key = f"{series.name} ({series.start_date.strftime('%b %Y')})"
                        break
                
                if not series_key:
                    series_key = f"Unknown Series ({enrolled_date.strftime('%b %Y')})"
            
            level_key = f"{series_key}|{cp.level.id}"
            
            if level_key not in series_level_enrollments:
                series_level_enrollments[level_key] = {
                    'series_name': series_key,
                    'level': cp.level,
                    'modules': {}
                }
            
            module_id = cp.module.id
            if module_id not in series_level_enrollments[level_key]['modules']:
                series_level_enrollments[level_key]['modules'][module_id] = {
                    'module': cp.module,
                    'papers': []
                }
            
            # Add paper if not already in list
            if cp.paper not in series_level_enrollments[level_key]['modules'][module_id]['papers']:
                series_level_enrollments[level_key]['modules'][module_id]['papers'].append(cp.paper)
        
        # Then, add any additional papers from results (for historical completeness)
        for result in results:
            # Use assessment series from result
            if result.assessment_series:
                series_key = f"{result.assessment_series.name} ({result.assessment_series.start_date.strftime('%b %Y')})"
            else:
                # Fallback for results without assessment series
                series_key = f"Unknown Series ({result.assessment_date.strftime('%b %Y') if result.assessment_date else 'Unknown'})"
            
            level_key = f"{series_key}|{result.level.id}"
            
            if level_key not in series_level_enrollments:
                series_level_enrollments[level_key] = {
                    'series_name': series_key,
                    'level': result.level,
                    'modules': {}
                }
            
            module_id = result.module.id
            if module_id not in series_level_enrollments[level_key]['modules']:
                series_level_enrollments[level_key]['modules'][module_id] = {
                    'module': result.module,
                    'papers': []
                }
            
            # Add paper if not already in list
            if result.paper not in series_level_enrollments[level_key]['modules'][module_id]['papers']:
                series_level_enrollments[level_key]['modules'][module_id]['papers'].append(result.paper)
        
        # Convert to list format for template
        enrollment_history = []
        for enrollment_data in series_level_enrollments.values():
            enrollment_data['modules'] = list(enrollment_data['modules'].values())
            enrollment_history.append(enrollment_data)
        
    else:
        # For other registration categories, use enrollment summary
        results_summary_for_display = enrollment_summary
        enrollment_history = enrollment_summary

    # Get user department for access control using session management utility
    staff, user_department, is_authenticated = get_user_staff_info(request)

    # Determine transcript/certificate access control
    can_generate_transcript = False
    can_generate_certificate = False
    
    # Only formal candidates (module/paper-based) can have transcripts and certificates
    if reg_cat_normalized == 'formal':
        # Check if candidate has passed all papers (no "Ms" or "CTR" in comments)
        all_results = results.filter(candidate=candidate)
        
        if all_results.exists():
            # Check if any result has failed status (Ms or CTR in comment)
            from django.db.models import Q
            failed_results = all_results.filter(
                Q(comment__icontains='Ms') | 
                Q(comment__icontains='CTR') |
                Q(comment__icontains='ms') |
                Q(comment__icontains='ctr')
            )
            
            # Enable transcript/certificate only if no failed results
            can_generate_transcript = not failed_results.exists()
            can_generate_certificate = not failed_results.exists()
    
    # Worker's Pass/Informal candidates: transcript and certificate buttons are hidden (always False)
    # Modular candidates: can generate transcript/certificate if they have results (no failure check needed)
    elif reg_cat_normalized == 'modular':
        can_generate_transcript = results.exists()
        can_generate_certificate = results.exists()

    context = {
        "candidate":          candidate,
        "level_enrollment":   level_enrollment,
        "module_enrollments": module_enrollments,
        "results":            results,
        "occupations": Occupation.objects.exclude(pk=candidate.occupation_id).order_by('code'),
        "centers":     AssessmentCenter.objects.exclude(pk=candidate.assessment_center_id),
        "enrollment_summary": enrollment_summary,
        "enrollment_history": enrollment_history,  # New comprehensive enrollment history
        "results_summary": results_summary_for_display,  # New comprehensive results summary
        "level_has_results": level_has_results,
        "results_released": results_released,
        "user_department": user_department,  # Add user department for access control
        "is_center_rep": is_center_rep,  # Add center rep status for template access control
        "can_generate_transcript": can_generate_transcript,  # Access control for transcript
        "can_generate_certificate": can_generate_certificate,  # Access control for certificate
    }
    return render(request, "candidates/view.html", context)


@login_required
def regenerate_candidate_reg_number(request, id):
    candidate = get_object_or_404(Candidate, id=id)
    # Clear reg_number and save to regenerate
    candidate.reg_number = None
    candidate.save()
    messages.success(request, f"Registration number regenerated: {candidate.reg_number}")
    return redirect('candidate_view', id=candidate.id)

def district_list(request):
    # Get filter parameters
    name = request.GET.get('name', '').strip()
    region = request.GET.get('region', '').strip()
    
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    districts = District.objects.all()
    
    # Apply filters
    if name:
        districts = districts.filter(name__icontains=name)
    if region:
        districts = districts.filter(region__icontains=region)
    
    districts = districts.order_by('name')
    
    # Pagination: 20 per page
    paginator = Paginator(districts, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'configurations/district_list.html', {
        'page_obj': page_obj,
        'paginator': paginator,
        'filters': {
            'name': name,
            'region': region,
        }
    })

def district_create(request):
    form = DistrictForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('district_list')
    return render(request, 'configurations/district_form.html', {'form': form})

def village_list(request):
    # Get filter parameters
    name = request.GET.get('name', '').strip()
    district_name = request.GET.get('district_name', '').strip()
    region = request.GET.get('region', '').strip()
    district_id = request.GET.get('district')
    
    from django.core.paginator import Paginator
    from django.db.models import Q, Count
    
    villages = Village.objects.select_related('district').all()
    
    # Apply filters
    if name:
        villages = villages.filter(name__icontains=name)
    if district_name:
        villages = villages.filter(district__name__icontains=district_name)
    if region:
        villages = villages.filter(district__region__icontains=region)
    
    if district_id:
        villages = villages.filter(district_id=district_id)
        district = District.objects.get(id=district_id)
        
        # Pagination for specific district view
        villages = villages.order_by('name')
        paginator = Paginator(villages, 20)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'page_obj': page_obj,
            'paginator': paginator,
            'current_district': district,
            'filters': {
                'name': name,
                'district_name': district_name,
                'region': region,
            }
        }
    else:
        # Get districts with village counts
        districts = District.objects.annotate(
            village_count=Count('village')
        ).order_by('name')
        
        context = {
            'villages': villages,
            'districts': districts,
            'filters': {
                'name': name,
                'district_name': district_name,
                'region': region,
            }
        }
    
    return render(request, 'configurations/village_list.html', context)

def village_create(request):
    district_id = request.GET.get('district')
    initial = {}
    
    if district_id:
        try:
            district = District.objects.get(id=district_id)
            initial['district'] = district
        except District.DoesNotExist:
            pass
    
    form = VillageForm(request.POST or None, initial=initial)
    if form.is_valid():
        village = form.save()
        if district_id:
            return redirect(f'{reverse("village_list")}?district={district_id}')
        return redirect('village_list')
    
    context = {'form': form}
    if district_id and 'district' in initial:
        context['district'] = initial['district']
    
    return render(request, 'configurations/village_form.html', context)

def config_home(request):
    """Configuration home page showing available settings"""
    return render(request, 'configurations/config_home.html')

from django.contrib.auth.decorators import login_required

@login_required
def occupation_edit(request, pk):
    from .models import Level, OccupationLevel
    occupation = get_object_or_404(Occupation, pk=pk)
    levels = Level.objects.filter(occupation=occupation)
    add_level_error = request.GET.get('add_level_error')
    if request.method == 'POST':
        form = OccupationForm(request.POST, instance=occupation)
        if form.is_valid():
            occupation = form.save(commit=False)
            occupation.updated_by = request.user
            occupation.save()
            # Update OccupationLevel assignments
            selected_level_ids = request.POST.getlist('levels')
            # Track selected_levels for error restoration
            selected_levels = {}
            # Remove old OccupationLevel objects for this occupation
            OccupationLevel.objects.filter(occupation=occupation).delete()
            # Enforce restriction: If has_modular is checked, Level 1 must be modules
            has_modular = occupation.has_modular
            level1 = levels.filter(name__icontains='1').first()
            error = None
            for level in levels:
                if str(level.id) in selected_level_ids:
                    if has_modular and level1 and level.id == level1.id:
                        # If user tries to set Level 1 to papers, error
                        if request.POST.get(f'structure_type_{level.id}') == 'papers':
                            error = "If 'Has Modular' is checked, Level 1 must be set to Modules."
                            selected_levels[str(level.id)] = 'papers'
                            break
                        structure_type = 'modules'
                    else:
                        structure_type = request.POST.get(f'structure_type_{level.id}', 'modules')
                    selected_levels[str(level.id)] = structure_type
                    OccupationLevel.objects.create(
                        occupation=occupation,
                        level=level,
                        structure_type=structure_type
                    )
            if error:
                # Restore old OccupationLevel assignments
                OccupationLevel.objects.filter(occupation=occupation).delete()
                for lid, stype in selected_levels.items():
                    OccupationLevel.objects.create(
                        occupation=occupation,
                        level=Level.objects.get(id=lid),
                        structure_type=stype
                    )
                return render(request, 'occupations/edit.html', {
                    'form': form,
                    'occupation': occupation,
                    'levels': levels,
                    'selected_levels': selected_levels,
                    'error': error
                })
            return redirect('occupation_detail', pk=occupation.pk)
    else:
        form = OccupationForm(instance=occupation)
    # For edit, build levels with stype attribute for template
    occupation_levels = OccupationLevel.objects.filter(occupation=occupation)
    level_stype_map = {ol.level_id: ol.structure_type for ol in occupation_levels}
    levels_with_stype = []
    for level in levels:
        level.stype = level_stype_map.get(level.id, '')
        levels_with_stype.append(level)
    # Always provide selected_levels for template
    selected_levels = {str(level.id): level.stype for level in levels_with_stype if level.stype}
    return render(request, 'occupations/edit.html', {
        'form': form,
        'occupation': occupation,
        'levels': levels_with_stype,
        'selected_levels': selected_levels
    })

def create_center_rep(request):
    if request.method == 'POST':
        form = CenterRepForm(request.POST)
        if form.is_valid():
            center_rep = form.save(commit=False)
            center_rep.created_by = request.user
            center_rep.updated_by = request.user
            center_rep.save()
            return redirect('view_center_reps')
    else:
        form = CenterRepForm()
    return render(request, 'users/center_representatives/create_center_rep.html', {'form': form})

from .forms import SupportStaffForm

def create_support_staff(request):
    if request.method == 'POST':
        form = SupportStaffForm(request.POST)
        if form.is_valid():
            staff = form.save(commit=False)
            staff.created_by = request.user
            staff.updated_by = request.user
            staff.save()
            return redirect('view_support_staff')
    else:
        form = SupportStaffForm()
    return render(request, 'users/support_staff/create_support_staff.html', {'form': form})


def user_home(request):
    return render(request, 'users/user_home.html')


def view_center_reps(request):
    from .models import CenterRepresentative
    center_reps = CenterRepresentative.objects.select_related('user', 'center').all()
    return render(request, 'users/center_representatives/list_center_rep.html', {'center_reps': center_reps})

def view_center_rep_detail(request, pk):
    from .models import CenterRepresentative
    rep = CenterRepresentative.objects.select_related('user', 'center').get(pk=pk)
    return render(request, 'users/center_representatives/view_center_rep.html', {'rep': rep})

def edit_center_rep(request, pk):
    from .models import CenterRepresentative
    from .forms import CenterRepForm
    rep = CenterRepresentative.objects.select_related('user', 'center').get(pk=pk)
    if request.method == 'POST':
        form = CenterRepForm(request.POST, instance=rep)
        if form.is_valid():
            center_rep = form.save(commit=False)
            center_rep.updated_by = request.user
            center_rep.save()
            return redirect('view_center_rep', pk=rep.pk)
    else:
        form = CenterRepForm(instance=rep)
    return render(request, 'users/center_representatives/edit_center_rep.html', {'form': form, 'rep': rep})

def view_support_staff(request):
    from .models import SupportStaff
    support_staff = SupportStaff.objects.select_related('user').all()
    return render(request, 'users/support_staff/list_support_staff.html', {'support_staff': support_staff})

def view_support_staff_detail(request, pk):
    from .models import SupportStaff
    staff = SupportStaff.objects.select_related('user').get(pk=pk)
    return render(request, 'users/support_staff/view_support_staff.html', {'staff': staff})

def edit_support_staff(request, pk):
    from .models import SupportStaff
    from .forms import SupportStaffForm, CenterRepForm
    staff = SupportStaff.objects.select_related('user').get(pk=pk)
    if request.method == 'POST':
        form = SupportStaffForm(request.POST, instance=staff)
        if form.is_valid():
            staff_obj = form.save(commit=False)
            staff_obj.updated_by = request.user
            staff_obj.save()
            return redirect('view_support_staff_detail', pk=staff.pk)
    else:
        form = SupportStaffForm(instance=staff)
    return render(request, 'users/support_staff/edit_support_staff.html', {'form': form, 'staff': staff})


@login_required
def profile(request):
    """User profile view with password change functionality"""
    from django.contrib.auth import authenticate, logout
    from django.contrib import messages
    
    user = request.user
    
    # Get user's staff profile if exists
    staff_profile = None
    try:
        staff_profile = user.staff_profile
    except:
        try:
            staff_profile = user.supportstaff
        except:
            pass
    
    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        # Validate old password
        if not authenticate(username=user.username, password=old_password):
            messages.error(request, 'Current password is incorrect.')
            return render(request, 'profile.html', {'staff_profile': staff_profile})
        
        # Validate new passwords match
        if new_password != confirm_password:
            messages.error(request, 'New passwords do not match.')
            return render(request, 'profile.html', {'staff_profile': staff_profile})
        
        # Validate password length
        if len(new_password) < 4:
            messages.error(request, 'New password must be at least 4 characters long.')
            return render(request, 'profile.html', {'staff_profile': staff_profile})
        
        # Change password
        user.set_password(new_password)
        user.save()
        
        # Log out user so they can login with new password
        logout(request)
        messages.success(request, 'Password changed successfully. Please login with your new password.')
        return redirect('login')
    
    return render(request, 'profile.html', {'staff_profile': staff_profile})


@login_required
def fix_all_photos(request):
    """Fix orientation for existing candidate photos that were imported with rotation issues"""
    from django.contrib import messages
    from PIL import Image, ExifTags
    from django.core.files.base import ContentFile
    import io
    import os
    
    if request.method == 'GET':
        return render(request, 'admin/fix_photos.html')
    
    if request.method == 'POST':
        action = request.POST.get('action')
        dry_run = request.POST.get('dry_run') == '1'
        
        if action == 'fix_single':
            candidate_id = request.POST.get('candidate_id')
            if not candidate_id:
                messages.error(request, 'Please provide a candidate ID.')
                return render(request, 'admin/fix_photos.html')
            
            try:
                candidate = Candidate.objects.get(id=candidate_id)
                if fix_candidate_photo_orientation(candidate, dry_run):
                    if dry_run:
                        messages.info(request, f'Dry run: Would fix photo orientation for candidate {candidate.id}: {candidate.full_name}')
                    else:
                        messages.success(request, f'Successfully fixed photo orientation for candidate {candidate.id}: {candidate.full_name}')
                else:
                    messages.warning(request, f'No orientation fix needed for candidate {candidate.id}: {candidate.full_name}')
            except Candidate.DoesNotExist:
                messages.error(request, f'Candidate with ID {candidate_id} not found.')
            except Exception as e:
                messages.error(request, f'Error fixing photo for candidate {candidate_id}: {str(e)}')
        
        elif action == 'fix_all':
            try:
                candidates = Candidate.objects.exclude(passport_photo='')
                total_candidates = candidates.count()
                fixed_count = 0
                error_count = 0
                
                for candidate in candidates:
                    try:
                        if fix_candidate_photo_orientation(candidate, dry_run):
                            fixed_count += 1
                    except Exception as e:
                        error_count += 1
                
                if dry_run:
                    messages.info(request, f'Dry run completed: Would fix {fixed_count} photos out of {total_candidates} candidates checked. Errors: {error_count}')
                else:
                    messages.success(request, f'Photo orientation fix completed: Fixed {fixed_count} photos out of {total_candidates} candidates checked. Errors: {error_count}')
                    
            except Exception as e:
                messages.error(request, f'Error during batch photo fix: {str(e)}')
    
    return render(request, 'admin/fix_photos.html')


def fix_candidate_photo_orientation(candidate, dry_run=False):
    """Fix orientation for a single candidate's photo"""
    if not candidate.passport_photo:
        return False
        
    try:
        # Open the existing photo
        photo_path = candidate.passport_photo.path
        if not os.path.exists(photo_path):
            return False
            
        with Image.open(photo_path) as img:
            # Check if image has EXIF orientation data
            orientation_applied = False
            
            try:
                for orientation in ExifTags.TAGS.keys():
                    if ExifTags.TAGS[orientation] == 'Orientation':
                        break
                
                exif = img._getexif()
                if exif is not None:
                    orientation_value = exif.get(orientation)
                    
                    if orientation_value and orientation_value != 1:  # 1 means normal orientation
                        if dry_run:
                            return True
                        
                        # Apply rotation based on EXIF orientation
                        # Handle all 8 EXIF orientation values
                        if orientation_value == 2:
                            img = img.transpose(Image.FLIP_LEFT_RIGHT)
                            orientation_applied = True
                        elif orientation_value == 3:
                            img = img.rotate(180, expand=True)
                            orientation_applied = True
                        elif orientation_value == 4:
                            img = img.transpose(Image.FLIP_TOP_BOTTOM)
                            orientation_applied = True
                        elif orientation_value == 5:
                            img = img.rotate(90, expand=True).transpose(Image.FLIP_LEFT_RIGHT)
                            orientation_applied = True
                        elif orientation_value == 6:
                            img = img.rotate(270, expand=True)
                            orientation_applied = True
                        elif orientation_value == 7:
                            img = img.rotate(270, expand=True).transpose(Image.FLIP_LEFT_RIGHT)
                            orientation_applied = True
                        elif orientation_value == 8:
                            img = img.rotate(90, expand=True)
                            orientation_applied = True
                        
                        if orientation_applied:
                            # Convert to RGB if needed
                            if img.mode != 'RGB':
                                img = img.convert('RGB')
                            
                            # Save the corrected image
                            buffer = io.BytesIO()
                            img.save(buffer, format='JPEG', quality=85)
                            buffer.seek(0)
                            
                            # Get the original filename
                            original_name = os.path.basename(candidate.passport_photo.name)
                            
                            # Save the corrected image back to the same field
                            candidate.passport_photo.save(
                                original_name,
                                ContentFile(buffer.getvalue()),
                                save=True
                            )
                            
                            return True
                            
            except (AttributeError, KeyError, TypeError):
                # No EXIF data
                pass
                
    except Exception as e:
        raise Exception(f'Error processing photo: {str(e)}')
        
    return False


from PIL import Image as PILImage, ImageDraw, ImageFont
from django.conf import settings
from django.shortcuts import get_object_or_404, redirect
from django.http import HttpResponse
import os

from django.views.decorators.http import require_POST
from .models import Occupation, Level

@require_POST
def add_level(request, occupation_id):
    occupation = get_object_or_404(Occupation, pk=occupation_id)
    name = request.POST.get('level_name')
    error = None
    if name:
        if Level.objects.filter(name=name, occupation=occupation).exists():
            error = f"Level '{name}' already exists for this occupation."
        else:
            Level.objects.create(name=name, occupation=occupation)
    else:
        error = "Level name is required."
    if error:
        from django.urls import reverse
        return redirect(f"{reverse('occupation_edit', kwargs={'pk': occupation_id})}?add_level_error={error}")
    return redirect('occupation_edit', pk=occupation_id)

def add_regno_to_photo(request, id):
    candidate = get_object_or_404(Candidate, id=id)

    if not candidate.passport_photo or not candidate.reg_number:
        return HttpResponse("Candidate must have both a photo and regno.", status=400)

    image_path = candidate.passport_photo.path
    # Use PILImage to avoid import shadowing with ReportLab's Image
    img = PILImage.open(image_path).convert("RGBA")

    # Overlay text (regno)
    draw = ImageDraw.Draw(img)
    # Use default PIL font for portability
    text = candidate.reg_number
    width, height = img.size
    max_width = width - 32
    font = None
    truetype_paths = [
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
        '/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf',
        '/usr/local/share/fonts/DejaVuSans-Bold.ttf',
        '/usr/share/fonts/truetype/freefont/FreeSansBold.ttf',
    ]
    font_size = 40
    min_font_size = 10
    for path in truetype_paths:
        if os.path.exists(path):
            try:
                font = ImageFont.truetype(path, size=font_size)
                break
            except Exception:
                font = None
    if font is None:
        font = ImageFont.load_default()
        font_size = 12
        min_font_size = 8
    # Dynamically shrink font size if needed
    while font_size >= min_font_size:
        try:
            text_w, text_h = font.getsize(text)
        except AttributeError:
            bbox = font.getbbox(text)
            text_w, text_h = bbox[2] - bbox[0], bbox[3] - bbox[1]
        if text_w <= max_width:
            break
        font_size -= 2
        if hasattr(font, 'path'):
            try:
                font = ImageFont.truetype(font.path, size=font_size)
            except Exception:
                font = ImageFont.load_default()
                break
    # If still too wide, truncate
    if text_w > max_width:
        ellipsis = '...'
        for i in range(len(text)-1, 0, -1):
            truncated = text[:i] + ellipsis
            try:
                text_w, text_h = font.getsize(truncated)
            except AttributeError:
                bbox = font.getbbox(truncated)
                text_w, text_h = bbox[2] - bbox[0], bbox[3] - bbox[1]
            if text_w <= max_width:
                text = truncated
                break
    padding_v = max(10, text_h // 4)
    strip_h = text_h + 2 * padding_v
    strip_y = height - strip_h
    overlay = PILImage.new('RGBA', (width, strip_h), (255, 255, 255, 255))  # Solid white
    img = img.convert('RGBA')
    img.alpha_composite(overlay, (0, strip_y))
    draw = ImageDraw.Draw(img)
    x = (width - text_w) // 2
    y = strip_y + (strip_h - text_h) // 2
    draw.text((x, y), text, font=font, fill=(0,0,0,255))
    # Save with a new filename
    base_dir = os.path.dirname(image_path)
    base_name = os.path.splitext(os.path.basename(image_path))[0]
    new_filename = os.path.join(base_dir, f"{base_name}_regno.png")
    img.save(new_filename)
    # Save the stamped image to a separate field (passport_photo_with_regno)
    from django.core.files import File
    from django.http import JsonResponse
    with open(new_filename, 'rb') as f:
        candidate.passport_photo_with_regno.save(os.path.basename(new_filename), File(f), save=True)
    url = candidate.passport_photo_with_regno.url
    return JsonResponse({'success': True, 'url': url})


@login_required
def download_result_list_pdf(request):
    """
    Generate a PDF download for the result list that matches the HTML preview exactly
    by converting the rendered HTML to PDF for pixel-perfect fidelity.
    """
    from django.template.loader import render_to_string
    from django.http import HttpResponse
    from .models import Candidate, Result, Occupation, OccupationLevel, AssessmentCenter
    from collections import defaultdict
    import calendar
    import os
    logger = logging.getLogger(__name__)
    
    try:
        import weasyprint
    except ImportError:
        # Fallback to xhtml2pdf if WeasyPrint is not available
        try:
            from xhtml2pdf import pisa
        except ImportError:
            # If neither is available, return an error
            return HttpResponse('PDF generation libraries not available. Please install WeasyPrint or xhtml2pdf.', status=500)
    
    # Use the same logic as the HTML preview to get the exact same data
    # Get parameters from either POST (form submission) or GET (download link)
    assessment_month = request.POST.get('assessment_month') or request.GET.get('assessment_month')
    assessment_year = request.POST.get('assessment_year') or request.GET.get('assessment_year')
    regcat = (request.POST.get('registration_category') or request.GET.get('registration_category', '')).lower()
    occupation_id = request.POST.get('occupation') or request.GET.get('occupation')
    level_id = request.POST.get('level') or request.GET.get('level')
    center_id = request.POST.get('assessment_center') or request.GET.get('assessment_center')
    
    # Validate required parameters
    if not assessment_month or not assessment_year or not occupation_id:
        return HttpResponse("Missing required parameters: assessment_month, assessment_year, and occupation are required.", status=400)
    
    # Convert empty strings to None for optional parameters
    level_id = level_id if level_id else None
    center_id = center_id if center_id else None
    
    logger.info(f"PDF Download - regcat: {regcat}, occupation: {occupation_id}, level: {level_id}, center: {center_id}")
    
    # Find logo path
    logo_path = None
    possible_paths = [
        os.path.join(settings.BASE_DIR, 'eims', 'static', 'images', 'uvtab_logo.png'),
        os.path.join(settings.BASE_DIR, 'static', 'images', 'uvtab_logo.png'),
        os.path.join(settings.BASE_DIR, 'emis', 'static', 'images', 'uvtab_logo.png'),
        os.path.join(settings.STATIC_ROOT or '', 'images', 'uvtab_logo.png')
    ]
    for path in possible_paths:
        if path and os.path.exists(path):
            logo_path = path
            logger.info(f"Found logo at: {logo_path}")
            print(f"DEBUG: Logo found at: {logo_path}")
            print(f"DEBUG: Logo exists: {os.path.exists(logo_path)}")
            break
    
    if not logo_path:
        print("DEBUG: Logo not found in any of the expected locations")
        logger.warning("Logo not found in any of the expected locations")
    
    # Get occupation for filename
    try:
        occupation = Occupation.objects.get(id=occupation_id)
        occupation_name = occupation.name.replace(' ', '_').replace('/', '_')
    except Occupation.DoesNotExist:
        occupation_name = 'Unknown'
    
    # Format registration category for filename
    regcat_display = regcat.replace('_', ' ').title() if regcat else 'All'
    regcat_filename = regcat_display.replace(' ', '_').replace("'", '')
    
    # Create filename
    assessment_period = f"{assessment_month}_{assessment_year}"
    filename = f"{regcat_filename}_{occupation_name}_{assessment_period}.pdf"
    
    # Query the data directly instead of using mock requests
    # Get the occupation and level objects
    try:
        occupation = Occupation.objects.get(id=occupation_id)
    except Occupation.DoesNotExist:
        return HttpResponse("Occupation not found", status=404)
    
    level = None
    if level_id:
        try:
            from .models import Level
            level = Level.objects.get(id=level_id)
        except Level.DoesNotExist:
            pass
    
    # Get assessment center if specified
    center = None
    if center_id:
        try:
            center = AssessmentCenter.objects.get(id=center_id)
        except AssessmentCenter.DoesNotExist:
            pass
    
    # Build the candidate filter
    candidate_filter = {
        'occupation': occupation,
        'registration_category__iexact': regcat,
    }
    if level:
        candidate_filter['candidatelevel__level'] = level
    if center:
        candidate_filter['assessment_center'] = center
    
    # Get candidates
    candidates = Candidate.objects.filter(**candidate_filter).order_by('reg_number')
    
    # Build result filter
    result_filter = {
        'assessment_date__year': int(assessment_year),
        'assessment_date__month': int(assessment_month),
    }
    
    # Add result_type filter based on registration category
    if regcat == 'formal':
        # For formal, check structure type
        if level and hasattr(level, 'structure_type'):
            if level.structure_type == 'modules':
                result_filter['result_type'] = 'modular'
            else:
                result_filter['result_type'] = 'formal'
        else:
            result_filter['result_type'] = 'formal'
    else:
        result_filter['result_type'] = regcat
    
    # Get results for these candidates
    candidate_ids = list(candidates.values_list('id', flat=True))
    results = Result.objects.filter(
        candidate_id__in=candidate_ids,
        **result_filter
    ).select_related('candidate', 'paper')
    
    # Group results by candidate and center
    centered_result_data = defaultdict(lambda: defaultdict(list))
    
    for result in results:
        candidate = result.candidate
        center_name = candidate.assessment_center.center_name if candidate.assessment_center else 'Unknown Center'
        
        # Add candidate photo URLs
        if hasattr(candidate, 'passport_photo_with_regno') and candidate.passport_photo_with_regno:
            candidate.photo_url = candidate.passport_photo_with_regno.url
        elif hasattr(candidate, 'passport_photo') and candidate.passport_photo:
            candidate.photo_url = candidate.passport_photo.url
        else:
            candidate.photo_url = None
        
        centered_result_data[center_name][candidate.reg_number].append(result)
    
    # Convert to regular dict
    centered_result_data = dict(centered_result_data)
    for center in centered_result_data:
        centered_result_data[center] = dict(centered_result_data[center])
    
    # For modular category, create module-based grouping for PDF
    module_result_data = None
    if regcat.lower() == 'modular':
        from collections import defaultdict
        from .models import CandidateModule, Module
        module_result_data = defaultdict(list)
        
        # Process all results to group by module
        for result in results:
            candidate = result.candidate
            candidate_id = candidate.id
            
            # Get all modules this candidate is enrolled in
            candidate_modules = CandidateModule.objects.filter(candidate_id=candidate_id).select_related('module')
            
            for cm in candidate_modules:
                module = cm.module
                module_key = f"{module.code} - {module.name}"
                
                # Check if this result belongs to this module
                if hasattr(result, 'module') and result.module == module:
                    # Create candidate entry for this module
                    candidate_entry = {
                        'id': candidate.id,
                        'reg_number': candidate.reg_number,
                        'full_name': candidate.full_name,
                        'gender': candidate.get_gender_display(),
                        'assessment_center': getattr(candidate.assessment_center, 'center_name', None),
                    }
                    
                    # Add result data
                    result_entry = {
                        'candidate': candidate_entry,
                        'results': [{
                            'grade': result.grade,
                            'comment': result.comment,
                            'mark': result.mark,
                        }],
                        'successful': result.comment != 'CTR',
                    }
                    
                    module_result_data[module_key].append(result_entry)
        
        # Convert to regular dict
        module_result_data = dict(module_result_data)
            # Enable module and paper grouping for informal category (same logic as HTML preview)
    informal_module_data = None
    if regcat.lower() in ['informal', "worker's pas", "workers pas"] and results:
        from collections import defaultdict
        from .models import Module, Paper
        
        informal_module_data = defaultdict(lambda: defaultdict(lambda: {'papers': [], 'candidates': []}))
        
        # Get all modules for this level to ensure we show all modules even if no results
        if level_id and occupation_id:
            all_modules = Module.objects.filter(occupation_id=occupation_id, level_id=level_id)
            for module in all_modules:
                module_key = f"{module.code} - {module.name}"
                # Get all papers for this module
                module_papers = Paper.objects.filter(module=module)
                # Initialize for all centers that have candidates
                centers_with_candidates = set(result.candidate.assessment_center.center_name if result.candidate.assessment_center else 'Unknown Center' for result in results)
                for center_name in centers_with_candidates:
                    informal_module_data[center_name][module_key]['papers'] = [{
                        'code': p.code,
                        'name': p.name,
                        'id': p.id
                    } for p in module_papers]
        
        # Process all results to group by center, then by module
        processed_candidates = set()
        for result in results:
            candidate = result.candidate
            candidate_id = candidate.id
            center_name = candidate.assessment_center.center_name if candidate.assessment_center else 'Unknown Center'
            
            # Get all results for this candidate
            candidate_results = results.filter(candidate=candidate)
            
            for candidate_result in candidate_results:
                if candidate_result.module:
                    module = candidate_result.module
                    # Only process results from the selected level
                    if level_id and module.level_id != int(level_id):
                        continue
                    module_key = f"{module.code} - {module.name}"
                    
                    # Check if we already processed this candidate for this module in this center
                    candidate_module_key = f"{candidate_id}_{module.id}_{center_name}"
                    if candidate_module_key not in processed_candidates:
                        processed_candidates.add(candidate_module_key)
                        
                        # Create candidate entry for this module in this center
                        candidate_entry = {
                            'id': candidate.id,
                            'reg_number': candidate.reg_number,
                            'full_name': candidate.full_name,
                            'gender': candidate.gender,
                            'passport_photo_with_regno': candidate.passport_photo_with_regno.url if candidate.passport_photo_with_regno else None,
                            'passport_photo': candidate.passport_photo.url if candidate.passport_photo else None,
                            'assessment_center': center_name,
                        }
                        
                        # Get all results for this candidate in this module
                        candidate_module_results = candidate_results.filter(module=module)
                        
                        # Create paper results mapping
                        paper_results = {}
                        for r in candidate_module_results:
                            if r.paper:
                                paper_results[r.paper.id] = {
                                    'grade': r.grade,
                                    'comment': r.comment,
                                    'mark': r.mark,
                                }
                        
                        candidate_entry['paper_results'] = paper_results
                        candidate_entry['has_ctr'] = any(r.comment == 'CTR' for r in candidate_module_results)
                        
                        informal_module_data[center_name][module_key]['candidates'].append(candidate_entry)
        
        # Convert to regular dict
        informal_module_data = dict(informal_module_data)
        for center_name in informal_module_data:
            informal_module_data[center_name] = dict(informal_module_data[center_name])
       
    # Get papers list for formal paper-based occupations (using same logic as HTML preview)
    papers_list = []
    if regcat.lower() == 'formal':
        # Determine structure type for occupation/level
        structure_type = 'modules'
        if occupation_id and level_id:
            from .models import OccupationLevel, Paper
            occ_level = OccupationLevel.objects.filter(occupation_id=occupation_id, level_id=level_id).first()
            if occ_level:
                structure_type = occ_level.structure_type
            if structure_type == 'papers':
                papers = Paper.objects.filter(occupation_id=occupation_id, level_id=level_id)
                papers_list = [p.code for p in papers]  # Use just codes for PDF
    
    # Build context
    context = {
        'preview': True,
        'centered_result_data': centered_result_data,
        'module_result_data': module_result_data,
        'informal_module_data': informal_module_data,
        'form_data': {
            'month': assessment_month,
            'year': assessment_year,
            'regcat': regcat,
            'occupation_id': occupation_id,
            'occupation_name': occupation.name,
            'level_id': level_id,
            'level_name': level.name if level else None,
            'center_id': center_id,
        },
        'assessment_month': calendar.month_name[int(assessment_month)],
        'assessment_year': assessment_year,
        'registration_category': regcat,
        'occupation': occupation,
        'level': level,
        'assessment_center': center,
        'papers_list': papers_list,
    }
    


    # Create custom PDF using ReportLab to match the preview template exactly
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from io import BytesIO
    import os
    
    # Create PDF buffer
    buffer = BytesIO()
    
    # Create PDF document in landscape orientation
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    
    # Get styles
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Title'],
        fontSize=16,
        spaceAfter=12,
        alignment=TA_CENTER,
        textColor=colors.black
    )
    
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=6,
        alignment=TA_CENTER,
        textColor=colors.black
    )
    
    # Build PDF content
    elements = []
    
    # Add logo if available
    logo_path = None
    possible_paths = [
        os.path.join(settings.BASE_DIR, 'eims', 'static', 'images', 'uvtab_logo.png'),
        os.path.join(settings.BASE_DIR, 'static', 'images', 'uvtab_logo.png'),
        os.path.join(settings.BASE_DIR, 'emis', 'static', 'images', 'uvtab_logo.png'),
    ]
    for path in possible_paths:
        if path and os.path.exists(path):
            logo_path = path
            break
            
    # Add general header for modular and formal categories (informal has its own center-specific headers)
    if regcat.lower() in ['modular', 'formal']:
        if logo_path:
            try:
                logo = Image(logo_path, width=1*inch, height=1*inch)
                elements.append(logo)
                elements.append(Spacer(1, 12))
            except:
                pass
        
        # Add header information
        elements.append(Paragraph("UGANDA VOCATIONAL AND TECHNICAL ASSESSMENT BOARD", title_style))
        elements.append(Paragraph("PROVISIONAL ASSESSMENT RESULT LIST FOR", header_style))
        elements.append(Paragraph("ASSESSMENT PERIOD:", header_style))
        elements.append(Paragraph(f"{calendar.month_name[int(assessment_month)]} {assessment_year}", header_style))
        elements.append(Spacer(1, 12))
        
        # Add occupation and category info
        elements.append(Paragraph(f"Category: {regcat.title()}", header_style))
        elements.append(Paragraph(f"Occupation: {occupation.name}", header_style))
        if level:
            elements.append(Paragraph(f"Level: {level.name}", header_style))
        elements.append(Spacer(1, 20))

    # Handle modular category with module grouping
    if regcat.lower() == 'modular' and module_result_data:
        # Process each module
        for module_name, module_entries in module_result_data.items():
            if elements:  # Add page break between modules (except for first)
                elements.append(PageBreak())
                
                # Repeat header for new page
                if logo_path:
                    try:
                        logo = Image(logo_path, width=1*inch, height=1*inch)
                        elements.append(logo)
                        elements.append(Spacer(1, 12))
                    except:
                        pass
                
                elements.append(Paragraph("UGANDA VOCATIONAL AND TECHNICAL ASSESSMENT BOARD", title_style))
                elements.append(Paragraph("PROVISIONAL ASSESSMENT RESULT LIST FOR", header_style))
                elements.append(Paragraph("ASSESSMENT PERIOD:", header_style))
                elements.append(Paragraph(f"{calendar.month_name[int(assessment_month)]} {assessment_year}", header_style))
                elements.append(Spacer(1, 12))
                elements.append(Paragraph(f"Category: {regcat.title()}", header_style))
                elements.append(Paragraph(f"Occupation: {occupation.name}", header_style))
                if level:
                    elements.append(Paragraph(f"Level: {level.name}", header_style))
                elements.append(Spacer(1, 20))
            
            # Module header
            elements.append(Paragraph(f"Module: {module_name}", header_style))
            elements.append(Spacer(1, 12))
            
            # Create table data for this module
            table_data = []
            headers = ['S/N', 'Photo', 'Reg No', 'Name', 'Gender', 'Practical', 'Comment']
            table_data.append(headers)
            
            # Add candidate data for this module
            sn = 1
            for entry in module_entries:
                candidate = entry['candidate']
                results = entry['results']
                
                row = [str(sn)]
                
                # Photo cell with fallback logic
                photo_cell = "No Photo"
                # For module entries, we need to get the actual candidate object for photo
                try:
                    actual_candidate = Candidate.objects.get(id=candidate['id'])
                    photo_path = None
                    
                    if hasattr(actual_candidate, 'passport_photo_with_regno') and actual_candidate.passport_photo_with_regno and hasattr(actual_candidate.passport_photo_with_regno, 'path') and os.path.exists(actual_candidate.passport_photo_with_regno.path):
                        photo_path = actual_candidate.passport_photo_with_regno.path
                    elif hasattr(actual_candidate, 'passport_photo') and actual_candidate.passport_photo and hasattr(actual_candidate.passport_photo, 'path') and os.path.exists(actual_candidate.passport_photo.path):
                        photo_path = actual_candidate.passport_photo.path
                    
                    if photo_path:
                        try:
                            photo_img = Image(photo_path, width=0.8*inch, height=0.8*inch)
                            photo_cell = photo_img
                        except Exception as e:
                            photo_cell = "No Photo"
                except:
                    photo_cell = "No Photo"
                
                row.append(photo_cell)
                row.append(candidate['reg_number'] or 'N/A')
                row.append(candidate['full_name'] or 'N/A')
                row.append(candidate['gender'] or 'N/A')
                
                # Practical grade with styling
                practical_grade = 'N/A'
                grade_style = None
                if results and len(results) > 0:
                    grade_value = results[0]['grade'] if results[0]['grade'] else 'N/A'
                    practical_grade = str(grade_value)
                    # Apply red styling for Ms and F grades
                    if grade_value in ['Ms', 'F']:
                        grade_style = 'red'
                
                # Comment: Use actual comment text with styling
                comment = 'Successful'
                comment_style = None
                if results and len(results) > 0:
                    actual_comment = results[0].get('comment', '')
                    if actual_comment in ['Missing', 'Fail']:
                        comment = actual_comment
                        comment_style = 'red_bold'
                    elif actual_comment == 'CTR':
                        comment = 'Fail'
                        comment_style = 'red_bold'
                    elif actual_comment:
                        comment = actual_comment
                
                row.extend([practical_grade, comment])
                table_data.append(row)
                sn += 1
            
            # Create and style the table for this module
            if len(table_data) > 1:  # Only create table if there are candidates
                table = Table(table_data)
                
                # Base table styling
                table_style = [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]
                
                # Apply conditional styling for grades and comments
                for row_idx, entry in enumerate(module_entries, start=1):
                    results = entry['results']
                    if results and len(results) > 0:
                        grade_value = results[0]['grade'] if results[0]['grade'] else 'N/A'
                        actual_comment = results[0].get('comment', '')
                        
                        # Style grade column (column 5 - Practical)
                        if grade_value in ['Ms', 'F']:
                            table_style.extend([
                                ('TEXTCOLOR', (5, row_idx), (5, row_idx), colors.red),
                                ('FONTNAME', (5, row_idx), (5, row_idx), 'Helvetica-Bold')
                            ])
                        
                        # Style comment column (column 6 - Comment)
                        if actual_comment in ['Missing', 'Fail', 'CTR']:
                            table_style.extend([
                                ('TEXTCOLOR', (6, row_idx), (6, row_idx), colors.red),
                                ('FONTNAME', (6, row_idx), (6, row_idx), 'Helvetica-Bold')
                            ])
                
                table.setStyle(TableStyle(table_style))
                elements.append(table)
                elements.append(Spacer(1, 20))
    # Handle informal category with module and paper grouping
        # Handle informal category with module and paper grouping
    elif regcat.lower() in ['informal', "worker's pas", "workers pas"] and informal_module_data:
        # Process each center
        first_center = True
        for center_name, center_modules in informal_module_data.items():
            if not first_center:  # Add page break between centers (except for first)
                elements.append(PageBreak())
            
            # Center header for each center
            if logo_path:
                try:
                    logo = Image(logo_path, width=1*inch, height=1*inch)
                    elements.append(logo)
                    elements.append(Spacer(1, 12))
                except:
                    pass
            
            elements.append(Paragraph("UGANDA VOCATIONAL AND TECHNICAL ASSESSMENT BOARD", title_style))
            elements.append(Paragraph("PROVISIONAL ASSESSMENT RESULT LIST FOR", header_style))
            elements.append(Paragraph("ASSESSMENT PERIOD:", header_style))
            elements.append(Paragraph(f"{calendar.month_name[int(assessment_month)]} {assessment_year}", header_style))
            elements.append(Spacer(1, 12))
            elements.append(Paragraph(f"Category: {regcat.title()}", header_style))
            elements.append(Paragraph(f"Occupation: {occupation.name}", header_style))
            if level:
                elements.append(Paragraph(f"Level: {level.name}", header_style))
            elements.append(Paragraph(f"Assessment Center: {center_name}", header_style))
            elements.append(Spacer(1, 20))
            
            first_center = False
            
            # Process each module within this center
            first_module = True
            for module_name, module_info in center_modules.items():
                # Initialize table_data for each module to prevent UnboundLocalError
                table_data = []
                
                if module_info['candidates']:
                    if not first_module:  # Add page break between modules within same center
                        elements.append(PageBreak())
                        
                        # Add header for new module page
                        if logo_path:
                            try:
                                logo = Image(logo_path, width=1*inch, height=1*inch)
                                elements.append(logo)
                                elements.append(Spacer(1, 12))
                            except:
                                pass
    
                        elements.append(Paragraph("UGANDA VOCATIONAL AND TECHNICAL ASSESSMENT BOARD", title_style))
                        elements.append(Paragraph("PROVISIONAL ASSESSMENT RESULT LIST FOR", header_style))
                        elements.append(Paragraph("ASSESSMENT PERIOD:", header_style))
                        elements.append(Paragraph(f"{calendar.month_name[int(assessment_month)]} {assessment_year}", header_style))
                        elements.append(Spacer(1, 12))
                        elements.append(Paragraph(f"Category: {regcat.title()}", header_style))
                        elements.append(Paragraph(f"Occupation: {occupation.name}", header_style))
                        if level:
                            elements.append(Paragraph(f"Level: {level.name}", header_style))
                        elements.append(Paragraph(f"Assessment Center: {center_name}", header_style))
                        elements.append(Spacer(1, 20))
                    
                    
                    first_module = False
                    
                    # Module header
                    elements.append(Paragraph(f"MODULE: {module_name}", header_style))
                    elements.append(Spacer(1, 8))
                    
                    # Add paper details row below module header
                    if module_info['papers']:
                        paper_details = []
                        for paper in module_info['papers']:
                            paper_detail = f"{paper['code']}: {paper['name']}"
                            paper_details.append(paper_detail)
                        
                        # Create a styled paragraph for paper details
                        paper_info_style = ParagraphStyle(
                            'PaperInfo',
                            parent=getSampleStyleSheet()['Normal'],
                            fontSize=9,
                            textColor=colors.darkblue,
                            leftIndent=20,
                            spaceAfter=8
                        )
                        
                        for paper_detail in paper_details:
                            elements.append(Paragraph(paper_detail, paper_info_style))
                        
                        elements.append(Spacer(1, 8))
                    
                    # Create table data for this module
                    table_data = []
                    
                    # Table headers - dynamic based on papers in this module
                    headers = ['S/N', 'Photo', 'Reg No', 'Name', 'Gender']
                    for paper in module_info['papers']:
                        headers.append(paper['code'])
                    headers.append('Comment')
                    
                    table_data.append(headers)
                    
                    # Add candidate data for this module
                    sn = 1

                for candidate in module_info['candidates']:
                    row = [str(sn)]

                    # Add photo cell with actual image
                    photo_cell = None
                    if candidate.get('passport_photo_with_regno'):
                        try:
                            photo_path = candidate['passport_photo_with_regno']
                            if photo_path.startswith('/media/'):
                                # Convert URL to file path
                                photo_file_path = os.path.join(settings.MEDIA_ROOT, photo_path[7:])  # Remove /media/ prefix
                                if os.path.exists(photo_file_path):
                                    photo_cell = Image(photo_file_path, width=0.8*inch, height=0.8*inch)
                        except:
                            pass

                    if not photo_cell and candidate.get('passport_photo'):
                        try:
                            photo_path = candidate['passport_photo']
                            if photo_path.startswith('/media/'):
                                # Convert URL to file path
                                photo_file_path = os.path.join(settings.MEDIA_ROOT, photo_path[7:])  # Remove /media/ prefix
                                if os.path.exists(photo_file_path):
                                    photo_cell = Image(photo_file_path, width=0.8*inch, height=0.8*inch)
                        except:
                            pass

                    if not photo_cell:
                            photo_cell = 'No Photo'

                    row.append(photo_cell)
                    
                    # Add candidate info
                    row.append(candidate['reg_number'])
                    row.append(candidate['full_name'])
                    row.append(candidate['gender'])
                    
                    # Add grades for each paper
                    for paper in module_info['papers']:
                        paper_result = candidate['paper_results'].get(paper['id'])
                        if paper_result:
                            row.append(paper_result['grade'])
                        else:
                            row.append('N/A')
                    
                    # Add comment
                    comment = 'Fail' if candidate['has_ctr'] else 'Successful'
                    row.append(comment)
                    
                    table_data.append(row)
                    sn += 1
                
                # Create and style the table with enhanced styling
                if len(table_data) > 1:  # Only create table if there are candidates
                    table = Table(table_data)
                    
                    # Enhanced table styling to match preview
                    table_style = [
                        # Header row styling
                        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.4, 0.4, 0.4)),  # Dark grey header
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                        ('TOPPADDING', (0, 0), (-1, 0), 8),
                        
                        # Data rows styling
                        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 1), (-1, -1), 9),
                        ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                        ('TOPPADDING', (0, 1), (-1, -1), 6),
                        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                        
                        # Alternating row colors for better readability
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.Color(0.95, 0.95, 0.85), colors.white]),
                        
                        # Grid and borders
                        ('GRID', (0, 0), (-1, -1), 1, colors.Color(0.3, 0.3, 0.3)),
                        ('BOX', (0, 0), (-1, -1), 2, colors.Color(0.2, 0.2, 0.2)),
                        
                        # Special styling for specific columns
                        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # S/N column
                        ('ALIGN', (2, 1), (2, -1), 'CENTER'),  # Reg No column
                        ('ALIGN', (3, 1), (3, -1), 'LEFT'),    # Name column
                        ('FONTSIZE', (3, 1), (3, -1), 8),      # Smaller font for names
                    ]
                    
                    table.setStyle(TableStyle(table_style))
                    elements.append(table)
                    elements.append(Spacer(1, 20))
    else:
    
        # Process each center (non-modular or fallback)
        for center_name, candidates_data in centered_result_data.items():
            if elements:  # Add page break between centers (except for first)
                elements.append(PageBreak())
                
                # Repeat header for new page
                if logo_path:
                    try:
                        logo = Image(logo_path, width=1*inch, height=1*inch)
                        elements.append(logo)
                        elements.append(Spacer(1, 12))
                    except:
                        pass
                
                elements.append(Paragraph("UGANDA VOCATIONAL AND TECHNICAL ASSESSMENT BOARD", title_style))
                elements.append(Paragraph("PROVISIONAL ASSESSMENT RESULT LIST FOR", header_style))
                elements.append(Paragraph("ASSESSMENT PERIOD:", header_style))
                elements.append(Paragraph(f"{calendar.month_name[int(assessment_month)]} {assessment_year}", header_style))
                elements.append(Spacer(1, 12))
                elements.append(Paragraph(f"Category: {regcat.title()}", header_style))
                elements.append(Paragraph(f"Occupation: {occupation.name}", header_style))
                if level:
                    elements.append(Paragraph(f"Level: {level.name}", header_style))
                elements.append(Spacer(1, 20))
            
            # Center header
            elements.append(Paragraph(f"Assessment Center: {center_name}", header_style))
            elements.append(Spacer(1, 12))
            
            # Create table data for this center
            table_data = []
            
            # Table headers
            if papers_list and len(papers_list) > 0:
                # Paper-based formal occupation
                headers = ['S/N', 'Photo', 'Reg No', 'Name', 'Gender']
                headers.extend(papers_list)  # Add paper codes
                headers.append('Comment')
            elif regcat.lower() == 'modular':
                # Modular: only practical, no theory
                headers = ['S/N', 'Photo', 'Reg No', 'Name', 'Gender', 'Practical', 'Comment']
            else:
                # Formal module-based: theory + practical
                headers = ['S/N', 'Photo', 'Reg No', 'Name', 'Gender', 'Theory', 'Practical', 'Comment']
            
            table_data.append(headers)
            
            # Add candidate data for this center
            sn = 1
            for reg_number, results in candidates_data.items():
                if not results:
                    continue
                    
                candidate = results[0].candidate
                row = [str(sn)]
                
                # Photo cell with fallback logic (copying working formal photo logic)
                photo_cell = "No Photo"
                photo_path = None
                
                # Use the same logic that works for formal PDFs - check .path instead of .url
                if hasattr(candidate, 'passport_photo_with_regno') and candidate.passport_photo_with_regno and hasattr(candidate.passport_photo_with_regno, 'path') and os.path.exists(candidate.passport_photo_with_regno.path):
                    photo_path = candidate.passport_photo_with_regno.path
                elif hasattr(candidate, 'passport_photo') and candidate.passport_photo and hasattr(candidate.passport_photo, 'path') and os.path.exists(candidate.passport_photo.path):
                    photo_path = candidate.passport_photo.path
                
                if photo_path:
                    try:
                        photo_img = Image(photo_path, width=0.8*inch, height=0.8*inch)
                        photo_cell = photo_img
                    except Exception as e:
                        photo_cell = "No Photo"
                
                row.append(photo_cell)
                row.append(candidate.reg_number or 'N/A')
                row.append(candidate.full_name if candidate.full_name else 'N/A')
                row.append(candidate.gender or 'N/A')
                
                if papers_list and len(papers_list) > 0:
                    # Paper-based: add grade for each paper
                    result_dict = {r.paper.code: r for r in results if r.paper}
                    for paper_code in papers_list:
                        if paper_code in result_dict:
                            result = result_dict[paper_code]
                            grade = getattr(result, 'grade', None) or getattr(result, 'marks', None) or 'N/A'
                            row.append(str(grade))
                        else:
                            row.append('N/A')
                    
                    # Comment: "Fail" if any result has comment "CTR", otherwise "Successful"
                    has_ctr = any(r.comment == 'CTR' for r in results if hasattr(r, 'comment'))
                    comment = "Fail" if has_ctr else "Successful"
                    row.append(comment)
                elif regcat.lower() == 'modular':
                    # Modular: only practical grade, no theory
                    practical_grade = 'N/A'
                    
                    for result in results:
                        # For modular, just get the grade (which is practical)
                        if hasattr(result, 'grade') and result.grade:
                            practical_grade = str(result.grade)
                            break  # Take the first grade found
                    
                    # Comment: "Fail" if any result has comment "CTR", otherwise "Successful"
                    has_ctr = any(getattr(r, 'comment', '') == 'CTR' for r in results)
                    comment = "Fail" if has_ctr else "Successful"
                    
                    row.extend([practical_grade, comment])
                else:
                    # Formal module-based: show theory/practical with proper grade extraction
                    theory_grade = 'N/A'
                    practical_grade = 'N/A'
                    
                    for result in results:
                        # Check for theory results
                        if hasattr(result, 'assessment_type'):
                            if result.assessment_type == 'theory':
                                if hasattr(result, 'grade') and result.grade:
                                    theory_grade = str(result.grade)
                            elif result.assessment_type == 'practical':
                                if hasattr(result, 'grade') and result.grade:
                                    practical_grade = str(result.grade)
                        else:
                            # Fallback: try different grade fields
                            if hasattr(result, 'theory_grade') and result.theory_grade:
                                theory_grade = str(result.theory_grade)
                            if hasattr(result, 'practical_grade') and result.practical_grade:
                                practical_grade = str(result.practical_grade)
                            if hasattr(result, 'grade') and result.grade and theory_grade == 'N/A':
                                theory_grade = str(result.grade)
                    
                    # Comment: "Fail" if any result has comment "CTR", otherwise "Successful"
                    has_ctr = any(getattr(r, 'comment', '') == 'CTR' for r in results)
                    comment = "Fail" if has_ctr else "Successful"
                    
                    row.extend([theory_grade, practical_grade, comment])
                
                table_data.append(row)
                sn += 1
            
            # Create and style table for this center with enhanced styling
            if len(table_data) > 1:  # Only create table if there's data
                table = Table(table_data)
                
                # Enhanced table styling to match preview (consistent with modular section)
                table_style = [
                    # Header row styling
                    ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.4, 0.4, 0.4)),  # Dark grey header
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('TOPPADDING', (0, 0), (-1, 0), 8),
                    
                    # Data rows styling
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
                    ('TOPPADDING', (0, 1), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                    
                    # Alternating row colors for better readability
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.Color(0.95, 0.95, 0.85), colors.white]),
                    
                    # Grid and borders
                    ('GRID', (0, 0), (-1, -1), 1, colors.Color(0.3, 0.3, 0.3)),
                    ('BOX', (0, 0), (-1, -1), 2, colors.Color(0.2, 0.2, 0.2)),
                    
                    # Special styling for specific columns
                    ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # S/N column
                    ('ALIGN', (2, 1), (2, -1), 'CENTER'),  # Reg No column
                    ('ALIGN', (3, 1), (3, -1), 'LEFT'),    # Name column
                    ('FONTSIZE', (3, 1), (3, -1), 8),      # Smaller font for names
                ]
                
                table.setStyle(TableStyle(table_style))
                elements.append(table)
                elements.append(Spacer(1, 20))
            else:
                elements.append(Paragraph("No results found for this center.", styles['Normal']))
                elements.append(Spacer(1, 20))
    
    # Build PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    
    # Return PDF response
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    logger.info(f"Generated custom PDF: {filename}")
    return response


def _create_photo_cell_content(candidate, styles, photo_width=0.8*inch, photo_height=0.8*inch):
    """Creates a list of flowables for the candidate photo cell, including image, name, and details."""
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus.flowables import Image
    from reportlab.platypus import Paragraph
    
    cell_elements = []
    photo_image = None

    # Photo Styles
    photo_name_style = ParagraphStyle(
        'PhotoName',
        parent=styles['Normal'],
        fontSize=6,
        alignment=TA_CENTER,
        spaceBefore=1,
        leading=6
    )
    photo_detail_style = ParagraphStyle(
        'PhotoDetail',
        parent=styles['Normal'],
        fontSize=5,
        alignment=TA_CENTER,
        spaceBefore=1,
        leading=5
    )
    
    # Handle photo with fallback logic
    if hasattr(candidate, 'passport_photo_with_regno') and candidate.passport_photo_with_regno:
        try:
            photo_image = Image(candidate.passport_photo_with_regno.path, width=photo_width, height=photo_height)
        except:
            pass
    elif hasattr(candidate, 'passport_photo') and candidate.passport_photo:
        try:
            photo_image = Image(candidate.passport_photo.path, width=photo_width, height=photo_height)
        except:
            pass
    
    if not photo_image:
        photo_image = Paragraph('No Photo', ParagraphStyle('NoPhoto', fontSize=6, alignment=TA_CENTER))
    
    cell_elements.append(photo_image)
    
    occupation_code = candidate.occupation.code if candidate.occupation else 'N/A'
    reg_category_short = candidate.registration_category.upper() if candidate.registration_category else 'N/A'
    
    return cell_elements
    
    return render(request, 'statistics/home.html', context)
@login_required
def statistics_home(request):
    """Enhanced statistics dashboard showing system overview and detailed metrics including assessment series"""
    
    # Get basic counts
    total_candidates = Candidate.objects.count()
    total_occupations = Occupation.objects.count()
    total_centers = AssessmentCenter.objects.count()
    total_results = Result.objects.count()
    
    # Gender breakdown
    gender_data = Candidate.objects.values('gender').annotate(
        count=Count('id')
    ).order_by('-count')
    
    gender_breakdown = []
    gender_colors = {'M': '#3B82F6', 'F': '#EC4899'}  # Blue for Male, Pink for Female
    total_candidates_for_percentage = total_candidates if total_candidates > 0 else 1
    
    for gender in gender_data:
        gender_code = gender['gender']
        count = gender['count']
        percentage = (count / total_candidates_for_percentage) * 100
        
        gender_breakdown.append({
            'code': gender_code,
            'name': 'Male' if gender_code == 'M' else 'Female' if gender_code == 'F' else 'Unknown',
            'count': count,
            'percentage': round(percentage, 1),
            'color': gender_colors.get(gender_code, '#6B7280')
        })
    
    # Registration categories breakdown
    registration_categories = []
    reg_cat_data = Candidate.objects.values('registration_category').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Define colors for different registration categories
    reg_colors = {
        'Formal': '#3B82F6',      # Blue
        'Modular': '#10B981',     # Green
        'Informal': '#F59E0B',    # Yellow
        'Worker\'s PAS': '#EF4444', # Red
    }
    
    for category in reg_cat_data:
        reg_cat = category['registration_category']
        count = category['count']
        percentage = (count / total_candidates_for_percentage) * 100
        
        registration_categories.append({
            'name': reg_cat or 'Unknown',
            'count': count,
            'percentage': round(percentage, 1),
            'color': reg_colors.get(reg_cat, '#6B7280')  # Default gray
        })
    
    # Special needs breakdown
    total_with_special_needs = Candidate.objects.filter(disability=True).count()
    total_without_special_needs = Candidate.objects.filter(disability=False).count()
    
    special_needs_breakdown = [
        {
            'name': 'With Special Needs',
            'count': total_with_special_needs,
            'percentage': round((total_with_special_needs / total_candidates_for_percentage) * 100, 1),
            'color': '#EF4444'  # Red
        },
        {
            'name': 'Without Special Needs',
            'count': total_without_special_needs,
            'percentage': round((total_without_special_needs / total_candidates_for_percentage) * 100, 1),
            'color': '#10B981'  # Green
        }
    ]
    
    # Special needs by Gender breakdown
    special_needs_by_gender = []
    
    # Male with special needs
    male_with_special_needs = Candidate.objects.filter(gender='M', disability=True).count()
    # Female with special needs
    female_with_special_needs = Candidate.objects.filter(gender='F', disability=True).count()
    # Male without special needs
    male_without_special_needs = Candidate.objects.filter(gender='M', disability=False).count()
    # Female without special needs
    female_without_special_needs = Candidate.objects.filter(gender='F', disability=False).count()
    
    special_needs_by_gender = [
        {
            'category': 'Male with Special Needs',
            'count': male_with_special_needs,
            'percentage': round((male_with_special_needs / total_candidates_for_percentage) * 100, 1),
            'color': '#DC2626'  # Dark red
        },
        {
            'category': 'Female with Special Needs',
            'count': female_with_special_needs,
            'percentage': round((female_with_special_needs / total_candidates_for_percentage) * 100, 1),
            'color': '#BE185D'  # Dark pink
        }
    ]
    
    # Registration Category by Gender
    reg_cat_by_gender = []
    for category in reg_cat_data:
        reg_cat = category['registration_category']
        if reg_cat:
            male_count = Candidate.objects.filter(registration_category=reg_cat, gender='M').count()
            female_count = Candidate.objects.filter(registration_category=reg_cat, gender='F').count()
            
            reg_cat_by_gender.append({
                'category': reg_cat,
                'male_count': male_count,
                'female_count': female_count,
                'total_count': male_count + female_count,
                'male_percentage': round((male_count / total_candidates_for_percentage) * 100, 1),
                'female_percentage': round((female_count / total_candidates_for_percentage) * 100, 1),
                'color': reg_colors.get(reg_cat, '#6B7280')
            })
    
    # Assessment Series breakdown (using actual assessment series relationships)
    assessment_series = []

    # Get assessment series with candidate counts using proper relationships
    from django.db.models import Q
    series_with_candidates = AssessmentSeries.objects.annotate(
        total_candidates=Count('candidate'),
        male_count=Count('candidate', filter=Q(candidate__gender='M')),
        female_count=Count('candidate', filter=Q(candidate__gender='F')),
        special_needs_count=Count('candidate', filter=Q(candidate__disability=True)),
        occupation_count=Count('candidate__occupation', distinct=True)
    ).filter(total_candidates__gt=0).order_by('-start_date')

    for series in series_with_candidates:
        assessment_series.append({
            'year': series.start_date.year,
            'month': series.start_date.month,
            'period_name': series.name,
            'total_candidates': series.total_candidates,
            'male_count': series.male_count,
            'female_count': series.female_count,
            'special_needs_count': series.special_needs_count,
            'occupation_count': series.occupation_count
        })
    
    context = {
        'total_candidates': total_candidates,
        'total_occupations': total_occupations,
        'total_centers': total_centers,
        'total_results': total_results,
        'gender_breakdown': gender_breakdown,
        'registration_categories': registration_categories,
        'special_needs_breakdown': special_needs_breakdown,
        'special_needs_by_gender': special_needs_by_gender,
        'reg_cat_by_gender': reg_cat_by_gender,
        'assessment_series': assessment_series,
    }
    
    return render(request, 'statistics/home.html', context)


@login_required
def assessment_series_detail(request, year, month):
    """Detailed breakdown for a specific assessment series using actual series relationships"""
    
    # Find the assessment series based on year/month
    # We'll look for a series that starts in the given year/month
    try:
        assessment_series = AssessmentSeries.objects.filter(
            start_date__year=year,
            start_date__month=month
        ).first()
        
        if not assessment_series:
            # If no exact match, try to find by name pattern
            month_name = calendar.month_name[int(month)]
            series_name = f"{month_name} {year} Series"
            assessment_series = AssessmentSeries.objects.filter(name=series_name).first()
        
        if not assessment_series:
            # Create a fallback if no series found
            month_name = calendar.month_name[int(month)]
            period_name = f"{month_name} {year}"
            candidates_in_period = Candidate.objects.none()
        else:
            period_name = assessment_series.name
            # Get candidates enrolled in this specific assessment series
            candidates_in_period = Candidate.objects.filter(assessment_series=assessment_series)
            
    except (ValueError, AssessmentSeries.DoesNotExist):
        month_name = calendar.month_name[int(month)]
        period_name = f"{month_name} {year}"
        candidates_in_period = Candidate.objects.none()
    
    total_candidates = candidates_in_period.count()
    total_candidates_for_percentage = total_candidates if total_candidates > 0 else 1
    
    # Gender breakdown
    gender_breakdown = []
    gender_colors = {'M': '#3B82F6', 'F': '#EC4899'}
    
    gender_data = candidates_in_period.values('gender').annotate(count=Count('id'))
    for gender in gender_data:
        gender_code = gender['gender']
        count = gender['count']
        percentage = (count / total_candidates_for_percentage) * 100
        
        gender_breakdown.append({
            'code': gender_code,
            'name': 'Male' if gender_code == 'M' else 'Female' if gender_code == 'F' else 'Unknown',
            'count': count,
            'percentage': round(percentage, 1),
            'color': gender_colors.get(gender_code, '#6B7280')
        })
    
    # Registration category breakdown
    reg_cat_breakdown = []
    reg_colors = {
        'Formal': '#3B82F6',
        'Modular': '#10B981',
        'Informal': '#F59E0B',
        'Worker\'s PAS': '#EF4444',
    }
    
    reg_cat_data = candidates_in_period.values('registration_category').annotate(count=Count('id'))
    for category in reg_cat_data:
        reg_cat = category['registration_category']
        count = category['count']
        percentage = (count / total_candidates_for_percentage) * 100
        
        reg_cat_breakdown.append({
            'name': reg_cat or 'Unknown',
            'count': count,
            'percentage': round(percentage, 1),
            'color': reg_colors.get(reg_cat, '#6B7280')
        })
    
    # Registration category by gender breakdown
    reg_cat_by_gender = []
    for category in reg_cat_data:
        reg_cat = category['registration_category']
        if reg_cat:
            male_count = candidates_in_period.filter(registration_category=reg_cat, gender='M').count()
            female_count = candidates_in_period.filter(registration_category=reg_cat, gender='F').count()
            
            reg_cat_by_gender.append({
                'category': reg_cat,
                'male_count': male_count,
                'female_count': female_count,
                'total_count': male_count + female_count,
                'male_percentage': round((male_count / total_candidates_for_percentage) * 100, 1),
                'female_percentage': round((female_count / total_candidates_for_percentage) * 100, 1),
                'color': reg_colors.get(reg_cat, '#6B7280')
            })
    
    # Special needs breakdown
    special_needs_breakdown = []
    special_needs_data = candidates_in_period.values('disability').annotate(count=Count('id'))
    for special_need in special_needs_data:
        disability = special_need['disability']
        count = special_need['count']
        percentage = (count / total_candidates_for_percentage) * 100
        
        special_needs_breakdown.append({
            'name': 'With Special Needs' if disability else 'Without Special Needs',
            'count': count,
            'percentage': round(percentage, 1),
            'color': '#EF4444' if disability else '#10B981'
        })
    
    # Occupation breakdown
    occupation_breakdown = []
    occupation_data = candidates_in_period.values('occupation__name').annotate(count=Count('id')).order_by('-count')
    for occupation in occupation_data:
        occupation_name = occupation['occupation__name']
        count = occupation['count']
        percentage = (count / total_candidates_for_percentage) * 100
        
        occupation_breakdown.append({
            'name': occupation_name or 'Unknown',
            'count': count,
            'percentage': round(percentage, 1)
        })
    # Sector-based analytics
    # 1. Occupation by Sector analysis
    sector_occupation_analysis = []
    sectors_with_candidates = candidates_in_period.values('occupation__sector__name').annotate(
        total_candidates=Count('id'),
        occupation_count=Count('occupation', distinct=True)
    ).filter(total_candidates__gt=0).order_by('-total_candidates')
    
    for sector_data in sectors_with_candidates:
        sector_name = sector_data['occupation__sector__name'] or 'Unknown Sector'
        sector_candidates = candidates_in_period.filter(occupation__sector__name=sector_name)
        
        # Get occupations within this sector
        sector_occupations = sector_candidates.values('occupation__name').annotate(
            count=Count('id')
        ).order_by('-count')
        
        sector_occupation_analysis.append({
            'sector_name': sector_name,
            'total_candidates': sector_data['total_candidates'],
            'occupation_count': sector_data['occupation_count'],
            'percentage': round((sector_data['total_candidates'] / total_candidates_for_percentage) * 100, 1),
            'occupations': list(sector_occupations)
        })
    
    # 2. Gender by Occupation by Sector analysis
    gender_occupation_sector_analysis = []
    for sector_data in sectors_with_candidates:
        sector_name = sector_data['occupation__sector__name'] or 'Unknown Sector'
        sector_candidates = candidates_in_period.filter(occupation__sector__name=sector_name)
        
        sector_occupations = sector_candidates.values('occupation__name').annotate(
            total_count=Count('id'),
            male_count=Count('id', filter=Q(gender='M')),
            female_count=Count('id', filter=Q(gender='F'))
        ).order_by('-total_count')
        
        gender_occupation_sector_analysis.append({
            'sector_name': sector_name,
            'occupations': list(sector_occupations)
        })
    
    # 3. Special Needs by Occupation by Sector analysis
    special_needs_occupation_sector_analysis = []
    for sector_data in sectors_with_candidates:
        sector_name = sector_data['occupation__sector__name'] or 'Unknown Sector'
        sector_candidates = candidates_in_period.filter(occupation__sector__name=sector_name)
        
        sector_occupations = sector_candidates.values('occupation__name').annotate(
            total_count=Count('id'),
            with_special_needs=Count('id', filter=Q(disability=True)),
            without_special_needs=Count('id', filter=Q(disability=False))
        ).order_by('-total_count')
        
        special_needs_occupation_sector_analysis.append({
            'sector_name': sector_name,
            'occupations': list(sector_occupations)
        })
    
    # 4. Most granular: Special Needs by Gender by Occupation by Sector
    granular_sector_analysis = []
    for sector_data in sectors_with_candidates:
        sector_name = sector_data['occupation__sector__name'] or 'Unknown Sector'
        sector_candidates = candidates_in_period.filter(occupation__sector__name=sector_name)
        
        sector_occupations = sector_candidates.values('occupation__name').annotate(
            total_count=Count('id'),
            male_with_special_needs=Count('id', filter=Q(gender='M', disability=True)),
            male_without_special_needs=Count('id', filter=Q(gender='M', disability=False)),
            female_with_special_needs=Count('id', filter=Q(gender='F', disability=True)),
            female_without_special_needs=Count('id', filter=Q(gender='F', disability=False))
        ).order_by('-total_count')
        
        granular_sector_analysis.append({
            'sector_name': sector_name,
            'occupations': list(sector_occupations)
        })
    
    # Add this debug line right before the context = { line
    print(f"DEBUG: reg_cat_breakdown = {reg_cat_breakdown}")
    
    context = {
        'period_name': period_name,
        'year': year,
        'month': month,
        'total_candidates': total_candidates,
        'gender_breakdown': gender_breakdown,
        'reg_cat_breakdown': reg_cat_breakdown,
        'reg_cat_by_gender': reg_cat_by_gender,
        'special_needs_breakdown': special_needs_breakdown,
        'occupation_breakdown': occupation_breakdown,
        'assessment_series': assessment_series,  # Pass the series object for template use
        # Sector-based analytics
        'sector_occupation_analysis': sector_occupation_analysis,
        'gender_occupation_sector_analysis': gender_occupation_sector_analysis,
        'special_needs_occupation_sector_analysis': special_needs_occupation_sector_analysis,
        'granular_sector_analysis': granular_sector_analysis,
    }
    
    return render(request, 'statistics/assessment_series_detail.html', context)

@login_required
def generate_performance_report(request, year, month):
    """Generate performance report PDF for assessment series - OPTIMIZED VERSION"""
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.units import inch
    from django.http import HttpResponse
    from django.conf import settings
    from django.db.models import Q, Count
    from io import BytesIO
    import os
    import time
    from django.db import connection
    
    # Performance monitoring
    start_time = time.time()
    print(f"[PERFORMANCE] Report generation started at {time.strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Get parameters
    category = request.GET.get('category')
    level = request.GET.get('level')
    
    # Get assessment series
    assessment_series = AssessmentSeries.objects.filter(
        start_date__year=year,
        start_date__month=month
    ).first()
    
    if not assessment_series:
        return HttpResponse("Assessment series not found", status=404)
    
    # Map category parameter to database registration_category values
    # Handle all possible variations of Informal/Worker's PAS
    category_mapping = {
        'Modular': 'Modular',
        'Formal': 'Formal', 
        "Worker's PAS": 'Informal',  # Frontend sends "Worker's PAS" but DB stores "Informal"
        'Workers PAS': 'Informal',   # Handle without apostrophe
        'Worker PAS': 'Informal',    # Handle without 's
        'Informal': 'Informal',      # Handle direct informal
        'informal': 'Informal',      # Handle lowercase
        "worker's pas": 'Informal',  # Handle lowercase
        'workers pas': 'Informal',   # Handle lowercase without apostrophe
        'worker pas': 'Informal'     # Handle lowercase without 's
    }
    
    # Get the correct database category value
    db_category = category_mapping.get(category, category)
    print(f"[DEBUG] Frontend category: '{category}', Database category: '{db_category}'")
    
    # Additional debugging: Check what registration categories actually exist in the database
    existing_categories = Candidate.objects.values_list('registration_category', flat=True).distinct()
    print(f"[DEBUG] Existing registration categories in database: {list(existing_categories)}")
    
    # Try both the mapped category and original category to be extra safe
    candidate_filters = [
        Q(registration_category=db_category),
        Q(registration_category=category)  # Fallback to original
    ]
    
    # If it's an informal/worker's pas related category, also try common variations
    if any(term in category.lower() for term in ['informal', 'worker', 'pas']):
        candidate_filters.extend([
            Q(registration_category='Informal'),
            Q(registration_category='informal'),
            Q(registration_category="Worker's PAS"),
            Q(registration_category='Workers PAS'),
            Q(registration_category='Worker PAS')
        ])
    
    # Combine all filters with OR
    combined_filter = candidate_filters[0]
    for filter_q in candidate_filters[1:]:
        combined_filter |= filter_q
    
    # Filter candidates with optimized queries to prevent N+1 problems
    # Use the combined filter to handle all category variations
    filtered_candidates = Candidate.objects.filter(
        assessment_series=assessment_series
    ).filter(
        combined_filter  # Apply the combined category filter
    ).select_related(
        'occupation',
        'occupation__sector',
        'assessment_center'
    ).prefetch_related(
        'result_set',  # Prefetch all results to avoid N+1 queries
        'candidatelevel_set__level'
    )
    
    # Step-by-step debugging to identify where candidates are lost
    print(f"[DEBUG] === FILTERING BREAKDOWN ===")
    
    # Step 1: Check total candidates in assessment series
    series_candidates = Candidate.objects.filter(assessment_series=assessment_series)
    print(f"[DEBUG] Step 1 - Total candidates in assessment series: {series_candidates.count()}")
    
    # Step 2: Check candidates after category filter
    category_candidates = series_candidates.filter(combined_filter)
    print(f"[DEBUG] Step 2 - Candidates after category filter: {category_candidates.count()}")
    
    # Step 3: Apply the full filter
    filtered_candidates = Candidate.objects.filter(
        assessment_series=assessment_series
    ).filter(
        combined_filter
    ).select_related(
        'occupation',
        'occupation__sector',
        'assessment_center'
    ).prefetch_related(
        'result_set',
        'candidatelevel_set__level'
    )
    
    print(f"[DEBUG] Step 3 - After full filtering: {filtered_candidates.count()}")
    print(f"[DEBUG] Category: {category}, Level: {level}")
    
    # Step 4: Level filtering (if applicable)
    # IMPORTANT: Only apply level filtering for Formal category
    # Informal/Worker's PAS should show ALL candidates regardless of level
    if level and category == 'Formal':
        print(f"[DEBUG] Step 4 - Applying level filter for category: {category}")
        pre_level_count = filtered_candidates.count()
        
        try:
            from .models import Level
            level_obj = Level.objects.get(id=level)
            print(f"[DEBUG] Found level object: {level_obj.name} (ID: {level_obj.id})")
            
            # Check how many candidates have this level enrollment
            candidates_with_level = filtered_candidates.filter(candidatelevel__level=level_obj)
            print(f"[DEBUG] Candidates with level {level_obj.name}: {candidates_with_level.count()}")
            
            filtered_candidates = candidates_with_level
            print(f"[DEBUG] After level filter: {pre_level_count} -> {filtered_candidates.count()}")
            
        except (Level.DoesNotExist, ValueError) as e:
            print(f"[DEBUG] Level filtering error: {e}")
            # If level not found, don't filter by level
            pass
    else:
        print(f"[DEBUG] Step 4 - No level filtering applied for category: {category} (showing all candidates)")
    
    print(f"[DEBUG] === FINAL RESULT ===")
    print(f"[DEBUG] Final filtered candidates count: {filtered_candidates.count()}")
    
    # Additional debugging: Show some sample candidate data if any exist
    if filtered_candidates.exists():
        sample_candidates = filtered_candidates[:3]
        for i, candidate in enumerate(sample_candidates):
            print(f"[DEBUG] Sample candidate {i+1}: {candidate.reg_number} - {candidate.full_name} - {candidate.registration_category}")
    else:
        print(f"[DEBUG] No candidates found after all filters!")
        
        # Debug: Check if there are any informal candidates at all in this series
        informal_in_series = Candidate.objects.filter(
            assessment_series=assessment_series,
            registration_category='Informal'
        ).count()
        print(f"[DEBUG] Total Informal candidates in this series: {informal_in_series}")
        
        # Debug: Check if the assessment series exists and has candidates
        if assessment_series:
            print(f"[DEBUG] Assessment series: {assessment_series.name} (ID: {assessment_series.id})")
            total_in_series = Candidate.objects.filter(assessment_series=assessment_series).count()
            print(f"[DEBUG] Total candidates in series: {total_in_series}")
        else:
            print(f"[DEBUG] No assessment series found!")
    
    print(f"[DEBUG] === END DEBUGGING ===")
    
    # Debug logging (original)
    print(f"[DEBUG] Initial candidates count: {filtered_candidates.count()}")
    print(f"[DEBUG] Category: {category}, Level: {level}")
    
    # Create PDF document
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=0.5*inch)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=14,
        spaceAfter=10,
        alignment=1,  # Center alignment
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=12,
        spaceAfter=20,
        alignment=1,  # Center alignment
        fontName='Helvetica-Bold'
    )
    
    # Add logo if available
    logo_path = os.path.join(settings.STATIC_ROOT or settings.BASE_DIR, 'eims', 'static', 'images', 'uvtab_logo.png')
    if not os.path.exists(logo_path):
        logo_path = os.path.join(settings.BASE_DIR, 'eims', 'static', 'images', 'uvtab_logo.png')
    
    if os.path.exists(logo_path):
        logo = Image(logo_path, width=1*inch, height=1*inch)
        logo.hAlign = 'CENTER'
        elements.append(logo)
        elements.append(Spacer(1, 10))
    
    # Header
    elements.append(Paragraph("UGANDA VOCATIONAL AND TECHNICAL ASSESSMENT BOARD (UVTAB)", title_style))
    elements.append(Paragraph("PERFORMANCE SUMMARY, BY PERFORMANCE AND GENDER PER OCCUPATION", subtitle_style))
    elements.append(Paragraph(f"ASSESSMENT PERIOD: {assessment_series.name}", styles['Heading3']))
    elements.append(Paragraph(f"Category: {category}", styles['Normal']))
    if level:
        elements.append(Paragraph(f"Level: {level}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Page 1: Candidate breakdown by occupation
    elements.append(Paragraph("Candidates Performance Details", styles['Heading3']))
    
    # Get occupation data with results using occupation codes
    occupation_data = []
    occupations = filtered_candidates.values('occupation__code', 'occupation__name').annotate(
        registered=Count('id'),
        # Add result-based counts - adjust field names based on your Result model
        # missing=Count('id', filter=Q(result__isnull=True)),
        # normal_progress=Count('id', filter=Q(result__status='passed')),
        # probationary=Count('id', filter=Q(result__status='failed'))
    ).order_by('occupation__code')
    
    # Create table with gender breakdown columns
    table_data = [[
        'S/N', 'Occupation code', 
        # Registered columns
        'Registered', '', '', '',
        # Absent columns  
        'Absent', '', '', '',
        # Assessed columns
        'Assessed', '', '', '',
        # Completed Successfully columns
        'Completed Successfully', '', '', '',
        # Unsuccessful columns
        'Unsuccessful', '', '', ''
    ]]
    
    # Add sub-header row for gender breakdown
    table_data.append([
        '', '',
        'F', 'M', 'TT', '%',  # Registered
        'F', 'M', 'TT', '%',  # Absent
        'F', 'M', 'TT', '%',  # Assessed
        'F', 'M', 'TT', '%',  # Completed Successfully
        'F', 'M', 'TT', '%'   # Unsuccessful
    ])
    
    for i, occ in enumerate(occupations, 1):
        # Get candidates for this occupation
        occupation_candidates = filtered_candidates.filter(
            occupation__code=occ['occupation__code']
        )
        
        # Calculate gender breakdowns for registered
        registered_total = occupation_candidates.count()
        registered_female = occupation_candidates.filter(gender='F').count()
        registered_male = occupation_candidates.filter(gender='M').count()
        registered_pct = 100.0 if registered_total > 0 else 0.0
        
        # Calculate missing (enrolled but no results) by gender
        candidates_with_results = occupation_candidates.filter(
            result__isnull=False
        ).distinct()
        candidates_without_results = occupation_candidates.exclude(
            id__in=candidates_with_results.values_list('id', flat=True)
        )
        
        missing_total = candidates_without_results.count()
        missing_female = candidates_without_results.filter(gender='F').count()
        missing_male = candidates_without_results.filter(gender='M').count()
        missing_pct = (missing_total / registered_total * 100) if registered_total > 0 else 0.0
        
        # Calculate sat for exams by gender
        sat_total = candidates_with_results.count()
        sat_female = candidates_with_results.filter(gender='F').count()
        sat_male = candidates_with_results.filter(gender='M').count()
        sat_pct = (sat_total / registered_total * 100) if registered_total > 0 else 0.0
        
        # Calculate Completed Successfully and Unsuccessful by gender
        normal_progress_candidates = []
        probationary_candidates = []
        
        for candidate in candidates_with_results:
            # Get all results for this candidate
            candidate_results = candidate.result_set.all()
            
            if not candidate_results.exists():
                continue
            
            # Check if candidate has multiple papers
            practical_results = candidate_results.filter(assessment_type='practical')
            
            if practical_results.exists():
                # For candidates with multiple papers, practical is important
                practical_comments = [r.comment.lower() for r in practical_results]
                
                # If any practical contains "unsuccessful", "ctr", or "fail" -> probationary
                if any(comment in ['unsuccessful', 'ctr', 'fail'] for comment in practical_comments):
                    probationary_candidates.append(candidate)
                elif all(comment == 'successful' for comment in practical_comments):
                    normal_progress_candidates.append(candidate)
                else:
                    probationary_candidates.append(candidate)
            else:
                # No practical results, check all results
                all_comments = [r.comment.lower() for r in candidate_results]
                
                if all(comment == 'successful' for comment in all_comments):
                    normal_progress_candidates.append(candidate)
                else:
                    probationary_candidates.append(candidate)
        
        # Calculate gender breakdowns for Completed Successfully
        np_total = len(normal_progress_candidates)
        np_female = sum(1 for c in normal_progress_candidates if c.gender == 'F')
        np_male = sum(1 for c in normal_progress_candidates if c.gender == 'M')
        np_pct = (np_total / registered_total * 100) if registered_total > 0 else 0.0
        
        # Calculate gender breakdowns for Probationary Pass
        pp_total = len(probationary_candidates)
        pp_female = sum(1 for c in probationary_candidates if c.gender == 'F')
        pp_male = sum(1 for c in probationary_candidates if c.gender == 'M')
        pp_pct = (pp_total / registered_total * 100) if registered_total > 0 else 0.0
        
        table_data.append([
            str(i),
            occ['occupation__code'] or 'Unknown',
            # Registered: F, M, TT, %
            str(registered_female),
            str(registered_male), 
            str(registered_total),
            f"{registered_pct:.1f}",
            # Missing: F, M, TT, %
            str(missing_female),
            str(missing_male),
            str(missing_total),
            f"{missing_pct:.1f}",
            # Assessed: F, M, TT, %
            str(sat_female),
            str(sat_male),
            str(sat_total),
            f"{sat_pct:.1f}",
            # Completed Successfully: F, M, TT, %
            str(np_female),
            str(np_male),
            str(np_total),
            f"{np_pct:.1f}",
            # Unsuccessful: F, M, TT, %
            str(pp_female),
            str(pp_male),
            str(pp_total),
            f"{pp_pct:.1f}"
        ])
    
    # Add totals row with actual calculations and gender breakdown
    total_registered = filtered_candidates.count()
    total_registered_female = filtered_candidates.filter(gender='F').count()
    total_registered_male = filtered_candidates.filter(gender='M').count()
    total_registered_pct = 100.0 if total_registered > 0 else 0.0
    
    # Calculate totals across all occupations
    total_candidates_with_results = filtered_candidates.filter(
        result__isnull=False
    ).distinct()
    total_candidates_without_results = filtered_candidates.exclude(
        id__in=total_candidates_with_results.values_list('id', flat=True)
    )
    
    total_missing = total_candidates_without_results.count()
    total_missing_female = total_candidates_without_results.filter(gender='F').count()
    total_missing_male = total_candidates_without_results.filter(gender='M').count()
    total_missing_pct = (total_missing / total_registered * 100) if total_registered > 0 else 0.0
    
    total_sat_for_exams = total_candidates_with_results.count()
    total_sat_female = total_candidates_with_results.filter(gender='F').count()
    total_sat_male = total_candidates_with_results.filter(gender='M').count()
    total_sat_pct = (total_sat_for_exams / total_registered * 100) if total_registered > 0 else 0.0
    
    # Calculate total Normal Progress (NP) and Probationary Pass (PP) with gender breakdown
    total_normal_progress_candidates = []
    total_probationary_candidates = []
    
    for candidate in total_candidates_with_results:
        candidate_results = candidate.result_set.all()
        
        if not candidate_results.exists():
            continue
        
        practical_results = candidate_results.filter(assessment_type='practical')
        
        if practical_results.exists():
            practical_comments = [r.comment.lower() for r in practical_results]
            
            if any(comment in ['unsuccessful', 'ctr', 'fail'] for comment in practical_comments):
                total_probationary_candidates.append(candidate)
            elif all(comment == 'successful' for comment in practical_comments):
                total_normal_progress_candidates.append(candidate)
            else:
                total_probationary_candidates.append(candidate)
        else:
            all_comments = [r.comment.lower() for r in candidate_results]
            
            if all(comment == 'successful' for comment in all_comments):
                total_normal_progress_candidates.append(candidate)
            else:
                total_probationary_candidates.append(candidate)
    
    # Calculate gender breakdowns for totals
    total_np = len(total_normal_progress_candidates)
    total_np_female = sum(1 for c in total_normal_progress_candidates if c.gender == 'F')
    total_np_male = sum(1 for c in total_normal_progress_candidates if c.gender == 'M')
    total_np_pct = (total_np / total_registered * 100) if total_registered > 0 else 0.0
    
    total_pp = len(total_probationary_candidates)
    total_pp_female = sum(1 for c in total_probationary_candidates if c.gender == 'F')
    total_pp_male = sum(1 for c in total_probationary_candidates if c.gender == 'M')
    total_pp_pct = (total_pp / total_registered * 100) if total_registered > 0 else 0.0
    
    table_data.append([
        'Total',
        '',
        # Registered: F, M, TT, %
        str(total_registered_female),
        str(total_registered_male),
        str(total_registered),
        f"{total_registered_pct:.1f}",
        # Missing: F, M, TT, %
        str(total_missing_female),
        str(total_missing_male),
        str(total_missing),
        f"{total_missing_pct:.1f}",
        # Sat For Exams: F, M, TT, %
        str(total_sat_female),
        str(total_sat_male),
        str(total_sat_for_exams),
        f"{total_sat_pct:.1f}",
        # Normal Progress: F, M, TT, %
        str(total_np_female),
        str(total_np_male),
        str(total_np),
        f"{total_np_pct:.1f}",
        # Probationary Pass: F, M, TT, %
        str(total_pp_female),
        str(total_pp_male),
        str(total_pp),
        f"{total_pp_pct:.1f}"
    ])
    
    # Create table with blue color scheme and proper column widths
    # Column widths: S/N, Program, then 4 columns each for 5 metrics = 22 total columns
    col_widths = [
        0.8*inch,  # S/N
        1.2*inch,  # Program code
        # Registered: F, M, TT, %
        0.4*inch, 0.4*inch, 0.5*inch, 0.5*inch,
        # Missing: F, M, TT, %
        0.4*inch, 0.4*inch, 0.5*inch, 0.5*inch,
        # Sat For Exams: F, M, TT, %
        0.4*inch, 0.4*inch, 0.5*inch, 0.5*inch,
        # Normal Progress: F, M, TT, %
        0.4*inch, 0.4*inch, 0.5*inch, 0.5*inch,
        # Probationary Pass: F, M, TT, %
        0.4*inch, 0.4*inch, 0.5*inch, 0.5*inch
    ]
    
    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.8)),  # Blue header
        ('BACKGROUND', (0, 1), (-1, 1), colors.Color(0.3, 0.5, 0.9)),  # Blue sub-header
        ('TEXTCOLOR', (0, 0), (-1, 1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, 1), 8),
        ('FONTSIZE', (0, 2), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 1), 8),
        # Column spanning for main headers to center them over sub-columns
        ('SPAN', (2, 0), (5, 0)),   # Registered spans 4 columns
        ('SPAN', (6, 0), (9, 0)),   # Absent spans 4 columns
        ('SPAN', (10, 0), (13, 0)), # Assessed spans 4 columns
        ('SPAN', (14, 0), (17, 0)), # Completed Successfully spans 4 columns
        ('SPAN', (18, 0), (21, 0)), # Unsuccessful spans 4 columns
        ('BACKGROUND', (0, 2), (-1, -2), colors.Color(0.9, 0.95, 1.0)),  # Light blue rows
        ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.7, 0.85, 0.95)),  # Darker blue for totals
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),  # Bold totals row
        ('BOX', (0, 0), (-1, -1), 1, colors.Color(0.3, 0.5, 0.9)),  # Blue outer border only
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 2), (-1, -2), [colors.Color(0.95, 0.98, 1.0), colors.Color(0.9, 0.95, 1.0)])
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 30))
    
    # Add Occupation Summary Table
    elements.append(Paragraph("Occupation Summary", styles['Heading3']))
    elements.append(Spacer(1, 10))
    
    # Get unique occupations with their sectors from the filtered candidates
    occupation_summary_data = filtered_candidates.values(
        'occupation__code', 
        'occupation__name', 
        'occupation__sector__name'
    ).distinct().order_by('occupation__code')
    
    # Create occupation summary table
    summary_table_data = [[
        'S/N', 'Occupation Code', 'Occupation Name', 'Sector'
    ]]
    
    for idx, occ in enumerate(occupation_summary_data, 1):
        sector_name = occ['occupation__sector__name'] or 'Unknown Sector'
        summary_table_data.append([
            str(idx),
            occ['occupation__code'] or 'N/A',
            occ['occupation__name'] or 'N/A',
            sector_name
        ])
    
    # Create summary table with appropriate column widths
    summary_col_widths = [
        0.6*inch,  # S/N
        1.2*inch,  # Occupation Code
        3.5*inch,  # Occupation Name
        2.0*inch   # Sector
    ]
    
    summary_table = Table(summary_table_data, colWidths=summary_col_widths)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.6, 0.2)),  # Green header
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # Center S/N column
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('BACKGROUND', (0, 1), (-1, -1), colors.Color(0.95, 1.0, 0.95)),  # Light green rows
        ('BOX', (0, 0), (-1, -1), 1, colors.Color(0.2, 0.6, 0.2)),  # Green border
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.Color(0.4, 0.7, 0.4)),  # Light green grid
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.Color(0.98, 1.0, 0.98), colors.Color(0.95, 1.0, 0.95)])
    ]))
    
    elements.append(summary_table)
    elements.append(PageBreak())
    
    # Page 2: Stats by occupation by sector
    elements.append(Paragraph("Performance by Occupation by Sector", title_style))
    elements.append(Spacer(1, 20))
    
    # Get sector data with gender breakdown - OPTIMIZED with batch processing
    # Fix sector query to handle NULL sectors properly
    sectors_raw = filtered_candidates.values('occupation__sector__name').distinct()
    
    # Debug sector information to identify the issue
    sector_list = list(sectors_raw)
    print(f"[DEBUG] Raw sectors found: {sector_list[:10]}...")  # Show first 10 sectors
    print(f"[DEBUG] Total unique sectors: {len(sector_list)}")
    
    # Check if we have NULL sectors causing the issue
    null_sector_count = filtered_candidates.filter(occupation__sector__name__isnull=True).count()
    print(f"[DEBUG] Candidates with NULL sectors: {null_sector_count}")
    
    # Get actual unique sector names, handling NULLs
    unique_sectors = set()
    for sector in sector_list:
        sector_name = sector['occupation__sector__name'] or 'Unknown Sector'
        unique_sectors.add(sector_name)
    
    print(f"[DEBUG] Actual unique sectors: {list(unique_sectors)}")
    
    # Add batch processing for large datasets
    candidate_count = filtered_candidates.count()
    print(f"[PERFORMANCE] Processing {candidate_count} candidates across {len(unique_sectors)} sectors")
    
    # If dataset is very large, implement chunking to prevent memory issues
    BATCH_SIZE = 1000  # Process candidates in batches of 1000
    if candidate_count > BATCH_SIZE:
        print(f"[PERFORMANCE] Large dataset detected, using batch processing with size {BATCH_SIZE}")
    
    # Create sector table with gender breakdown columns
    sector_table_data = [[
        'Sector',
        # Total Candidates columns
        'Total Candidates', '', '', '',
        # Successful columns
        'Successful', '', '', '',
        # Unsuccessful columns
        'Unsuccessful', '', '', ''
    ]]
    
    # Add sub-header row for gender breakdown
    sector_table_data.append([
        '',
        'F', 'M', 'TT', '%',  # Total Candidates
        'F', 'M', 'TT', '%',  # Successful
        'F', 'M', 'TT', '%'   # Unsuccessful
    ])
    
    for sector_name in unique_sectors:
        print(f"[DEBUG] Processing sector: {sector_name}")
        
        # Get candidates for this sector - handle NULL sectors properly
        if sector_name == 'Unknown Sector':
            sector_candidates = filtered_candidates.filter(
                occupation__sector__name__isnull=True
            )
        else:
            sector_candidates = filtered_candidates.filter(
                occupation__sector__name=sector_name
            )
        
        # Calculate total candidates by gender
        total_candidates = sector_candidates.count()
        total_female = sector_candidates.filter(gender='F').count()
        total_male = sector_candidates.filter(gender='M').count()
        total_pct = 100.0 if total_candidates > 0 else 0.0
        
        # Calculate successful candidates (Normal Progress) by gender - OPTIMIZED VERSION
        successful_candidates = []
        unsuccessful_candidates = []
        
        # Get all candidates with results for this sector (already prefetched)
        candidates_with_results = [c for c in sector_candidates if hasattr(c, 'result_set') and c.result_set.exists()]
        
        # Process candidates using prefetched data (no additional DB queries)
        for candidate in candidates_with_results:
            # Use prefetched results - no DB query here
            candidate_results = list(candidate.result_set.all())
            
            if not candidate_results:
                continue
            
            # Filter practical results from prefetched data
            practical_results = [r for r in candidate_results if r.assessment_type == 'practical']
            
            if practical_results:
                practical_comments = [r.comment.lower() if r.comment else '' for r in practical_results]
                
                if any(comment in ['unsuccessful', 'ctr', 'fail'] for comment in practical_comments):
                    unsuccessful_candidates.append(candidate)
                elif all(comment == 'successful' for comment in practical_comments):
                    successful_candidates.append(candidate)
                else:
                    unsuccessful_candidates.append(candidate)
            else:
                all_comments = [r.comment.lower() if r.comment else '' for r in candidate_results]
                
                if all(comment == 'successful' for comment in all_comments):
                    successful_candidates.append(candidate)
                else:
                    unsuccessful_candidates.append(candidate)
        
        # Calculate successful by gender
        successful_total = len(successful_candidates)
        successful_female = sum(1 for c in successful_candidates if c.gender == 'F')
        successful_male = sum(1 for c in successful_candidates if c.gender == 'M')
        successful_pct = (successful_total / total_candidates * 100) if total_candidates > 0 else 0.0
        
        # Calculate unsuccessful by gender
        unsuccessful_total = len(unsuccessful_candidates)
        unsuccessful_female = sum(1 for c in unsuccessful_candidates if c.gender == 'F')
        unsuccessful_male = sum(1 for c in unsuccessful_candidates if c.gender == 'M')
        unsuccessful_pct = (unsuccessful_total / total_candidates * 100) if total_candidates > 0 else 0.0
        
        sector_table_data.append([
            sector_name,
            # Total Candidates: F, M, TT, %
            str(total_female),
            str(total_male),
            str(total_candidates),
            f"{total_pct:.1f}",
            # Successful: F, M, TT, %
            str(successful_female),
            str(successful_male),
            str(successful_total),
            f"{successful_pct:.1f}",
            # Unsuccessful: F, M, TT, %
            str(unsuccessful_female),
            str(unsuccessful_male),
            str(unsuccessful_total),
            f"{unsuccessful_pct:.1f}"
        ])
    
    # Column widths for sector table: Sector name + 3 metrics × 4 columns each = 13 total columns
    sector_col_widths = [
        2.0*inch,  # Sector name
        # Total Candidates: F, M, TT, %
        0.5*inch, 0.5*inch, 0.6*inch, 0.6*inch,
        # Successful: F, M, TT, %
        0.5*inch, 0.5*inch, 0.6*inch, 0.6*inch,
        # Unsuccessful: F, M, TT, %
        0.5*inch, 0.5*inch, 0.6*inch, 0.6*inch
    ]
    
    sector_table = Table(sector_table_data, colWidths=sector_col_widths)
    sector_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.8)),  # Blue header
        ('BACKGROUND', (0, 1), (-1, 1), colors.Color(0.3, 0.5, 0.9)),  # Blue sub-header
        ('TEXTCOLOR', (0, 0), (-1, 1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, 1), 8),
        ('FONTSIZE', (0, 2), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 1), 8),
        # Column spanning for sector table headers
        ('SPAN', (1, 0), (4, 0)),   # Total Candidates spans 4 columns
        ('SPAN', (5, 0), (8, 0)),   # Successful spans 4 columns
        ('SPAN', (9, 0), (12, 0)),  # Unsuccessful spans 4 columns
        ('BACKGROUND', (0, 2), (-1, -1), colors.Color(0.9, 0.95, 1.0)),  # Light blue rows
        ('BOX', (0, 0), (-1, -1), 1, colors.Color(0.3, 0.5, 0.9)),  # Blue outer border only
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 2), (-1, -1), [colors.Color(0.95, 0.98, 1.0), colors.Color(0.9, 0.95, 1.0)])
    ]))
    
    elements.append(sector_table)
    elements.append(PageBreak())
    
    # Page 3+: Center summary pages
    centers = filtered_candidates.values('assessment_center__center_number', 'assessment_center__center_name').annotate(
        total_candidates=Count('id')
    ).filter(total_candidates__gt=0).order_by('assessment_center__center_number')
    
    for center_data in centers:
        center_number = center_data['assessment_center__center_number'] or 'Unknown'
        center_name = center_data['assessment_center__center_name'] or 'Unknown Center'
        center_candidates = filtered_candidates.filter(assessment_center__center_number=center_number)
        
        elements.append(Paragraph(f"Performance by Assessment Center", title_style))
        elements.append(Paragraph(f"Center: {center_number} - {center_name}", styles['Heading3']))
        elements.append(Spacer(1, 20))
        
        # Center-specific occupation breakdown
        center_occupations = center_candidates.values('occupation__code', 'occupation__name').annotate(
            registered=Count('id')
        ).order_by('occupation__code')
        
        # Create center table with gender breakdown columns
        center_table_data = [[
            'S/N', 'Occupation',
            # Registered columns
            'Registered', '', '', '',
            # Absent columns  
            'Absent', '', '', '',
            # Assessed columns
            'Assessed', '', '', '',
            # Completed Successfully columns
            'Completed Successfully', '', '', '',
            # Unsuccessful columns
            'Unsuccessful', '', '', ''
        ]]
        
        # Add sub-header row for gender breakdown
        center_table_data.append([
            '', '',
            'F', 'M', 'TT', '%',  # Registered
            'F', 'M', 'TT', '%',  # Absent
            'F', 'M', 'TT', '%',  # Assessed
            'F', 'M', 'TT', '%',  # Completed Successfully
            'F', 'M', 'TT', '%'   # Unsuccessful
        ])
        
        for i, occ in enumerate(center_occupations, 1):
            # Get candidates for this occupation in this center
            center_occupation_candidates = center_candidates.filter(
                occupation__code=occ['occupation__code']
            )
            
            # Calculate gender breakdowns for registered
            registered_total = center_occupation_candidates.count()
            registered_female = center_occupation_candidates.filter(gender='F').count()
            registered_male = center_occupation_candidates.filter(gender='M').count()
            registered_pct = 100.0 if registered_total > 0 else 0.0
            
            # Calculate missing (enrolled but no results) by gender
            center_candidates_with_results = center_occupation_candidates.filter(
                result__isnull=False
            ).distinct()
            center_candidates_without_results = center_occupation_candidates.exclude(
                id__in=center_candidates_with_results.values_list('id', flat=True)
            )
            
            missing_total = center_candidates_without_results.count()
            missing_female = center_candidates_without_results.filter(gender='F').count()
            missing_male = center_candidates_without_results.filter(gender='M').count()
            missing_pct = (missing_total / registered_total * 100) if registered_total > 0 else 0.0
            
            # Calculate sat for exams by gender
            sat_total = center_candidates_with_results.count()
            sat_female = center_candidates_with_results.filter(gender='F').count()
            sat_male = center_candidates_with_results.filter(gender='M').count()
            sat_pct = (sat_total / registered_total * 100) if registered_total > 0 else 0.0
            
            # Calculate Completed Successfully and Unsuccessful by gender
            normal_progress_candidates = []
            probationary_candidates = []
            
            for candidate in center_candidates_with_results:
                candidate_results = candidate.result_set.all()
                
                if not candidate_results.exists():
                    continue
                
                practical_results = candidate_results.filter(assessment_type='practical')
                
                if practical_results.exists():
                    practical_comments = [r.comment.lower() for r in practical_results]
                    
                    if any(comment in ['unsuccessful', 'ctr', 'fail'] for comment in practical_comments):
                        probationary_candidates.append(candidate)
                    elif all(comment == 'successful' for comment in practical_comments):
                        normal_progress_candidates.append(candidate)
                    else:
                        probationary_candidates.append(candidate)
                else:
                    all_comments = [r.comment.lower() for r in candidate_results]
                    
                    if all(comment == 'successful' for comment in all_comments):
                        normal_progress_candidates.append(candidate)
                    else:
                        probationary_candidates.append(candidate)
            
            # Calculate gender breakdowns for Completed Successfully
            np_total = len(normal_progress_candidates)
            np_female = sum(1 for c in normal_progress_candidates if c.gender == 'F')
            np_male = sum(1 for c in normal_progress_candidates if c.gender == 'M')
            np_pct = (np_total / registered_total * 100) if registered_total > 0 else 0.0
            
            # Calculate gender breakdowns for Probationary Pass
            pp_total = len(probationary_candidates)
            pp_female = sum(1 for c in probationary_candidates if c.gender == 'F')
            pp_male = sum(1 for c in probationary_candidates if c.gender == 'M')
            pp_pct = (pp_total / registered_total * 100) if registered_total > 0 else 0.0
            
            center_table_data.append([
                str(i),
                occ['occupation__code'] or 'Unknown',
                # Registered: F, M, TT, %
                str(registered_female),
                str(registered_male),
                str(registered_total),
                f"{registered_pct:.1f}",
                # Missing: F, M, TT, %
                str(missing_female),
                str(missing_male),
                str(missing_total),
                f"{missing_pct:.1f}",
                # Assessed: F, M, TT, %
                str(sat_female),
                str(sat_male),
                str(sat_total),
                f"{sat_pct:.1f}",
                # Completed Successfully: F, M, TT, %
                str(np_female),
                str(np_male),
                str(np_total),
                f"{np_pct:.1f}",
                # Unsuccessful: F, M, TT, %
                str(pp_female),
                str(pp_male),
                str(pp_total),
                f"{pp_pct:.1f}"
            ])
        
        # Center totals with actual calculations and gender breakdown
        center_total = center_candidates.count()
        center_total_female = center_candidates.filter(gender='F').count()
        center_total_male = center_candidates.filter(gender='M').count()
        center_total_pct = 100.0 if center_total > 0 else 0.0
        
        # Calculate center totals
        center_total_with_results = center_candidates.filter(
            result__isnull=False
        ).distinct()
        center_total_without_results = center_candidates.exclude(
            id__in=center_total_with_results.values_list('id', flat=True)
        )
        
        center_total_missing = center_total_without_results.count()
        center_total_missing_female = center_total_without_results.filter(gender='F').count()
        center_total_missing_male = center_total_without_results.filter(gender='M').count()
        center_total_missing_pct = (center_total_missing / center_total * 100) if center_total > 0 else 0.0
        
        center_total_sat_for_exams = center_total_with_results.count()
        center_total_sat_female = center_total_with_results.filter(gender='F').count()
        center_total_sat_male = center_total_with_results.filter(gender='M').count()
        center_total_sat_pct = (center_total_sat_for_exams / center_total * 100) if center_total > 0 else 0.0
        
        # Calculate center total Normal Progress (NP) and Probationary Pass (PP) with gender breakdown
        center_total_normal_progress_candidates = []
        center_total_probationary_candidates = []
        
        for candidate in center_total_with_results:
            candidate_results = candidate.result_set.all()
            
            if not candidate_results.exists():
                continue
            
            practical_results = candidate_results.filter(assessment_type='practical')
            
            if practical_results.exists():
                practical_comments = [r.comment.lower() for r in practical_results]
                
                if any(comment in ['unsuccessful', 'ctr', 'fail'] for comment in practical_comments):
                    center_total_probationary_candidates.append(candidate)
                elif all(comment == 'successful' for comment in practical_comments):
                    center_total_normal_progress_candidates.append(candidate)
                else:
                    center_total_probationary_candidates.append(candidate)
            else:
                all_comments = [r.comment.lower() for r in candidate_results]
                
                if all(comment == 'successful' for comment in all_comments):
                    center_total_normal_progress_candidates.append(candidate)
                else:
                    center_total_probationary_candidates.append(candidate)
        
        # Calculate gender breakdowns for center totals
        center_total_np = len(center_total_normal_progress_candidates)
        center_total_np_female = sum(1 for c in center_total_normal_progress_candidates if c.gender == 'F')
        center_total_np_male = sum(1 for c in center_total_normal_progress_candidates if c.gender == 'M')
        center_total_np_pct = (center_total_np / center_total * 100) if center_total > 0 else 0.0
        
        center_total_pp = len(center_total_probationary_candidates)
        center_total_pp_female = sum(1 for c in center_total_probationary_candidates if c.gender == 'F')
        center_total_pp_male = sum(1 for c in center_total_probationary_candidates if c.gender == 'M')
        center_total_pp_pct = (center_total_pp / center_total * 100) if center_total > 0 else 0.0
        
        center_table_data.append([
            'TOTAL',
            '',
            # Registered: F, M, TT, %
            str(center_total_female),
            str(center_total_male),
            str(center_total),
            f"{center_total_pct:.1f}",
            # Missing: F, M, TT, %
            str(center_total_missing_female),
            str(center_total_missing_male),
            str(center_total_missing),
            f"{center_total_missing_pct:.1f}",
            # Assessed: F, M, TT, %
            str(center_total_sat_female),
            str(center_total_sat_male),
            str(center_total_sat_for_exams),
            f"{center_total_sat_pct:.1f}",
            # Completed Successfully: F, M, TT, %
            str(center_total_np_female),
            str(center_total_np_male),
            str(center_total_np),
            f"{center_total_np_pct:.1f}",
            # Unsuccessful: F, M, TT, %
            str(center_total_pp_female),
            str(center_total_pp_male),
            str(center_total_pp),
            f"{center_total_pp_pct:.1f}"
        ])
        
        # Column widths for center table: S/N, Program, then 4 columns each for 5 metrics = 22 total columns
        center_col_widths = [
            0.8*inch,  # S/N
            1.2*inch,  # Program code
            # Registered: F, M, TT, %
            0.4*inch, 0.4*inch, 0.5*inch, 0.5*inch,
            # Missing: F, M, TT, %
            0.4*inch, 0.4*inch, 0.5*inch, 0.5*inch,
            # Assessed: F, M, TT, %
            0.4*inch, 0.4*inch, 0.5*inch, 0.5*inch,
            # Completed Successfully: F, M, TT, %
            0.4*inch, 0.4*inch, 0.5*inch, 0.5*inch,
            # Unsuccessful: F, M, TT, %
            0.4*inch, 0.4*inch, 0.5*inch, 0.5*inch
        ]
        
        center_table = Table(center_table_data, colWidths=center_col_widths)
        center_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.8)),  # Blue header
            ('BACKGROUND', (0, 1), (-1, 1), colors.Color(0.3, 0.5, 0.9)),  # Blue sub-header
            ('TEXTCOLOR', (0, 0), (-1, 1), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, 1), 8),
            ('FONTSIZE', (0, 2), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 1), 8),
            # Column spanning for center table headers
            ('SPAN', (2, 0), (5, 0)),   # Registered spans 4 columns
            ('SPAN', (6, 0), (9, 0)),   # Absent spans 4 columns
            ('SPAN', (10, 0), (13, 0)), # Assessed spans 4 columns
            ('SPAN', (14, 0), (17, 0)), # Completed Successfully spans 4 columns
            ('SPAN', (18, 0), (21, 0)), # Unsuccessful spans 4 columns
            ('BACKGROUND', (0, 2), (-1, -2), colors.Color(0.9, 0.95, 1.0)),  # Light blue rows
            ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.7, 0.85, 0.95)),  # Darker blue for totals
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),  # Bold totals row
            ('BOX', (0, 0), (-1, -1), 1, colors.Color(0.3, 0.5, 0.9)),  # Blue outer border only
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 2), (-1, -2), [colors.Color(0.95, 0.98, 1.0), colors.Color(0.9, 0.95, 1.0)])
        ]))
        
        elements.append(center_table)
        
        # Add page break if not the last center
        if center_data != list(centers)[-1]:
            elements.append(PageBreak())
    
    # Build PDF
    doc.build(elements)
    
    # Get the value of the BytesIO buffer and create response
    pdf = buffer.getvalue()
    buffer.close()
    
    # Performance monitoring
    end_time = time.time()
    execution_time = end_time - start_time
    query_count = len(connection.queries)
    print(f"[PERFORMANCE] Report generation completed in {execution_time:.2f} seconds")
    print(f"[PERFORMANCE] Total database queries: {query_count}")
    print(f"[PERFORMANCE] Report generated at {time.strftime('%Y-%m-%d %H:%M:%S')}")
    
    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="performance_report_{category}_{year}_{month}.pdf"'
    response.write(pdf)
    
    return response

    def get_candidate_filter_urls(assessment_year=None, assessment_month=None):
        """Generate common filter URLs for statistics templates"""
        base_filters = {}
        if assessment_year:
            base_filters['assessment_year'] = assessment_year
        if assessment_month:
            base_filters['assessment_month'] = assessment_month
    
        return {
            'male_url': build_candidate_filter_url(base_filters, gender='M'),
            'female_url': build_candidate_filter_url(base_filters, gender='F'),
            'special_needs_url': build_candidate_filter_url(base_filters, disability='true'),
            'no_special_needs_url': build_candidate_filter_url(base_filters, disability='false'),
            'male_special_needs_url': build_candidate_filter_url(base_filters, gender='M', disability='true'),
            'female_special_needs_url': build_candidate_filter_url(base_filters, gender='F', disability='true'),
            'formal_url': build_candidate_filter_url(base_filters, registration_category='Formal'),
            'modular_url': build_candidate_filter_url(base_filters, registration_category='Modular'),
            'informal_url': build_candidate_filter_url(base_filters, registration_category='Informal'),
            'workers_pas_url': build_candidate_filter_url(base_filters, registration_category="Worker's PAS"),
        }

@login_required
def assessment_series_list(request):
    """Display list of all assessment series with pagination"""
    series_list = AssessmentSeries.objects.all().order_by('-is_current', '-start_date')
    
    # Add pagination
    paginator = Paginator(series_list, 10)  # Show 10 series per page
    page_number = request.GET.get('page')
    series = paginator.get_page(page_number)
    
    context = {
        'series': series,
        'total_count': series_list.count(),
        'current_series': series_list.filter(is_current=True).first()
    }
    
    return render(request, 'Assessment_series/list.html', context)


@login_required
def assessment_series_create(request):
    """Create a new assessment series"""
    if request.method == 'POST':
        form = AssessmentSeriesForm(request.POST)
        if form.is_valid():
            series = form.save()
            messages.success(request, f'Assessment Series "{series.name}" created successfully!')
            return redirect('assessment_series_view', pk=series.pk)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AssessmentSeriesForm()
    
    context = {
        'form': form,
        'title': 'Create Assessment Series'
    }
    
    return render(request, 'Assessment_series/create.html', context)


@login_required
def assessment_series_view(request, pk):
    """View details of a specific assessment series"""
    series = get_object_or_404(AssessmentSeries, pk=pk)
    
    # Calculate some statistics for this series
    from datetime import datetime
    
    # Get candidates enrolled during this series period
    candidates_in_period = Candidate.objects.filter(
        assessment_date__gte=series.start_date,
        assessment_date__lte=series.end_date
    )
    
    # Get results for this period
    results_in_period = Result.objects.filter(
        assessment_date__gte=series.start_date,
        assessment_date__lte=series.end_date
    )
    
    # Calculate statistics
    total_candidates = candidates_in_period.count()
    total_results = results_in_period.count()
    male_candidates = candidates_in_period.filter(gender='M').count()
    female_candidates = candidates_in_period.filter(gender='F').count()
    special_needs_candidates = candidates_in_period.filter(disability=True).count()
    
    # Check if series is active (current date is within series period)
    today = datetime.now().date()
    is_active = series.start_date <= today <= series.end_date
    
    # Check if results can be released (current date is past release date)
    can_release_results = today >= series.date_of_release
    
    context = {
        'series': series,
        'total_candidates': total_candidates,
        'total_results': total_results,
        'male_candidates': male_candidates,
        'female_candidates': female_candidates,
        'special_needs_candidates': special_needs_candidates,
        'is_active': is_active,
        'can_release_results': can_release_results,
        'today': today
    }
    
    return render(request, 'Assessment_series/view.html', context)


@login_required
def assessment_series_edit(request, pk):
    """Edit an existing assessment series"""
    series = get_object_or_404(AssessmentSeries, pk=pk)
    
    if request.method == 'POST':
        form = AssessmentSeriesForm(request.POST, instance=series)
        if form.is_valid():
            updated_series = form.save()
            messages.success(request, f'Assessment Series "{updated_series.name}" updated successfully!')
            return redirect('assessment_series_view', pk=updated_series.pk)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AssessmentSeriesForm(instance=series)
    
    context = {
        'form': form,
        'series': series,
        'title': f'Edit Assessment Series: {series.name}'
    }
    
    return render(request, 'Assessment_series/create.html', context)


@login_required
def assessment_series_delete(request, pk):
    """Delete an assessment series"""
    series = get_object_or_404(AssessmentSeries, pk=pk)
    
    if request.method == 'POST':
        series_name = series.name
        series.delete()
        messages.success(request, f'Assessment Series "{series_name}" deleted successfully!')
        return redirect('assessment_series_list')
    
    context = {
        'series': series,
        'title': f'Delete Assessment Series: {series.name}'
    }
    
    return render(request, 'Assessment_series/delete.html', context)


@login_required
def assessment_series_set_current(request, pk):
    """Set a specific assessment series as current"""
    series = get_object_or_404(AssessmentSeries, pk=pk)
    
    if request.method == 'POST':
        # Set this series as current (the model's save method will handle unsetting others)
        series.is_current = True
        series.save()
        messages.success(request, f'Assessment Series "{series.name}" is now set as current!')
        return redirect('assessment_series_view', pk=series.pk)
    
    context = {
        'series': series,
        'title': f'Set as Current: {series.name}'
    }
    
    return render(request, 'Assessment_series/set_current.html', context)


@login_required
def assessment_series_years(request):
    """Display available years with series counts"""
    # Get all years that have assessment series
    years_data = []
    
    # Get distinct years from start_date field
    series_years = AssessmentSeries.objects.dates('start_date', 'year', order='DESC')
    
    for year_date in series_years:
        year = year_date.year
        series_count = AssessmentSeries.objects.filter(
            start_date__year=year
        ).count()
        
        current_series_count = AssessmentSeries.objects.filter(
            start_date__year=year,
            is_current=True
        ).count()
        
        years_data.append({
            'year': year,
            'series_count': series_count,
            'current_series_count': current_series_count,
            'has_current': current_series_count > 0
        })
    
    # If no years exist, show current year
    if not years_data:
        current_year = datetime.now().year
        years_data.append({
            'year': current_year,
            'series_count': 0,
            'current_series_count': 0,
            'has_current': False
        })
    
    context = {
        'years_data': years_data,
        'total_series': AssessmentSeries.objects.count(),
        'current_year': datetime.now().year
    }
    
    return render(request, 'Assessment_series/years.html', context)


@login_required
def assessment_series_year_detail(request, year):
    """Display all assessment series for a specific year"""
    # Get all series for the specified year
    series_list = AssessmentSeries.objects.filter(
        start_date__year=year
    ).order_by('-is_current', 'start_date')
    
    # Add pagination
    paginator = Paginator(series_list, 12)  # Show 12 series per page (monthly)
    page_number = request.GET.get('page')
    series = paginator.get_page(page_number)
    
    # Get year statistics
    current_series = series_list.filter(is_current=True).first()
    total_count = series_list.count()
    
    # Get months with series for this year
    months_with_series = series_list.dates('start_date', 'month')
    month_names = [date.strftime('%B') for date in months_with_series]
    
    context = {
        'series': series,
        'year': year,
        'total_count': total_count,
        'current_series': current_series,
        'months_with_series': month_names,
        'prev_year': year - 1,
        'next_year': year + 1,
        'current_year': datetime.now().year
    }
    
    return render(request, 'Assessment_series/year_detail.html', context)


@login_required
def assessment_series_create_for_year(request, year=None):
    """Create a new assessment series, optionally pre-filled for a specific year"""
    if request.method == 'POST':
        form = AssessmentSeriesForm(request.POST)
        if form.is_valid():
            series = form.save()
            messages.success(request, f'Assessment Series "{series.name}" created successfully!')
            # Redirect to the year detail page
            series_year = series.start_date.year
            return redirect('assessment_series_year_detail', year=series_year)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = AssessmentSeriesForm()
        # If year is provided, pre-fill the form with January 1st of that year
        if year:
            from datetime import date
            form.initial = {
                'start_date': date(year, 1, 1),
                'end_date': date(year, 1, 31),
                'date_of_release': date(year, 2, 1)
            }
    
    context = {
        'form': form,
        'title': f'Create Assessment Series{f" for {year}" if year else ""}',
        'year': year
    }
    
    return render(request, 'Assessment_series/create.html', context)

@login_required
def assessment_series_toggle_results(request, pk):
    """Toggle the results release status for an assessment series"""
    series = get_object_or_404(AssessmentSeries, pk=pk)
    
    if request.method == 'POST':
        # Toggle the results_released status
        series.results_released = not series.results_released
        series.save()
        
        status = "released" if series.results_released else "hidden"
        messages.success(request, f'Results for "{series.name}" have been {status}.')
        
        return redirect('assessment_series_view', pk=pk)
    
    # If not POST, redirect back to view
    return redirect('assessment_series_view', pk=pk)

@login_required
def assessment_series_statistical_report(request, pk):
    """Generate a statistical PDF report for a specific Assessment Series"""
    series = get_object_or_404(AssessmentSeries, pk=pk)
    
    # Get candidates enrolled during this series period
    candidates = Candidate.objects.filter(
        assessment_date__gte=series.start_date,
        assessment_date__lte=series.end_date
    )
    
    # Create the HttpResponse object with PDF headers
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Assessment_Series_Report_{series.name.replace(" ", "_")}.pdf"'
    
    # Create the PDF object
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
    
    # Container for the 'Flowable' objects
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.HexColor('#4A5568')
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        spaceBefore=20,
        textColor=colors.HexColor('#2D3748')
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=6,
    )
    
    # Title
    title = Paragraph(f"Assessment Series Statistical Report<br/>{series.name}", title_style)
    elements.append(title)
    
    # Series Information
    series_info = f"""
    <b>Series Period:</b> {series.start_date.strftime('%B %d, %Y')} - {series.end_date.strftime('%B %d, %Y')}<br/>
    <b>Results Release Date:</b> {series.date_of_release.strftime('%B %d, %Y')}<br/>
    <b>Report Generated:</b> {datetime.now().strftime('%B %d, %Y at %I:%M %p')}<br/>
    <b>Total Enrolled Candidates:</b> {candidates.count()}
    """
    elements.append(Paragraph(series_info, normal_style))
    elements.append(Spacer(1, 20))
    
    # 1. GENDER DISTRIBUTION
    elements.append(Paragraph("1. GENDER DISTRIBUTION", heading_style))
    
    gender_stats = candidates.values('gender').annotate(count=Count('gender')).order_by('gender')
    gender_data = [['Gender', 'Count', 'Percentage']]
    
    total_candidates = candidates.count()
    for stat in gender_stats:
        gender_name = 'Male' if stat['gender'] == 'M' else 'Female'
        percentage = (stat['count'] / total_candidates * 100) if total_candidates > 0 else 0
        gender_data.append([gender_name, str(stat['count']), f"{percentage:.1f}%"])
    
    if total_candidates == 0:
        gender_data.append(['No candidates enrolled', '-', '-'])
    
    gender_table = Table(gender_data, colWidths=[2*inch, 1*inch, 1*inch])
    gender_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A5568')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(gender_table)
    elements.append(Spacer(1, 20))
    
    # 2. SPECIAL NEEDS DISTRIBUTION
    elements.append(Paragraph("2. SPECIAL NEEDS DISTRIBUTION", heading_style))
    
    special_needs_count = candidates.filter(disability=True).count()
    no_special_needs_count = candidates.filter(disability=False).count()
    
    special_needs_data = [['Category', 'Count', 'Percentage']]
    
    if total_candidates > 0:
        special_needs_percentage = (special_needs_count / total_candidates * 100)
        no_special_needs_percentage = (no_special_needs_count / total_candidates * 100)
        
        special_needs_data.append(['Candidates with Special Needs', str(special_needs_count), f"{special_needs_percentage:.1f}%"])
        special_needs_data.append(['Candidates without Special Needs', str(no_special_needs_count), f"{no_special_needs_percentage:.1f}%"])
    else:
        special_needs_data.append(['No candidates enrolled', '-', '-'])
    
    special_needs_table = Table(special_needs_data, colWidths=[3*inch, 1*inch, 1*inch])
    special_needs_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A5568')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(special_needs_table)
    
    # Nature of Disabilities breakdown (if any special needs candidates exist)
    if special_needs_count > 0:
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("2.1 Nature of Disabilities", ParagraphStyle('SubHeading', parent=styles['Heading3'], fontSize=12, spaceAfter=8)))
        
        # Get all disability natures for candidates in this series
        disability_stats = []
        special_needs_candidates = candidates.filter(disability=True)
        
        for candidate in special_needs_candidates:
            for disability in candidate.nature_of_disability.all():
                disability_stats.append(disability.name)
        
        # Count occurrences
        disability_counts = Counter(disability_stats)
        
        disability_data = [['Nature of Disability', 'Count']]
        for disability_name, count in disability_counts.most_common():
            disability_data.append([disability_name, str(count)])
        
        if not disability_counts:
            disability_data.append(['No specific disabilities recorded', '-'])
        
        disability_table = Table(disability_data, colWidths=[3*inch, 1*inch])
        disability_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A5568')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(disability_table)
    
    elements.append(Spacer(1, 20))
    
    # 3. OCCUPATIONS ENROLLED
    elements.append(Paragraph("3. OCCUPATIONS ENROLLED", heading_style))
    
    occupation_stats = candidates.values(
        'occupation__name', 
        'occupation__category__name'
    ).annotate(
        count=Count('occupation')
    ).order_by('occupation__category__name', 'occupation__name')
    
    occupation_data = [['Occupation Category', 'Occupation', 'Count', 'Percentage']]
    
    for stat in occupation_stats:
        if stat['occupation__name']:  # Only include if occupation is not null
            percentage = (stat['count'] / total_candidates * 100) if total_candidates > 0 else 0
            occupation_data.append([
                stat['occupation__category__name'] or 'Uncategorized',
                stat['occupation__name'],
                str(stat['count']),
                f"{percentage:.1f}%"
            ])
    
    if len(occupation_data) == 1:  # Only header row
        occupation_data.append(['No occupations recorded', '-', '-', '-'])
    
    occupation_table = Table(occupation_data, colWidths=[1.5*inch, 2.5*inch, 0.8*inch, 0.8*inch])
    occupation_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4A5568')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(occupation_table)
    
    # Summary section
    elements.append(Spacer(1, 30))
    elements.append(Paragraph("SUMMARY", heading_style))
    
    # Calculate special needs percentage safely
    special_needs_percentage = (special_needs_count/total_candidates*100) if total_candidates > 0 else 0
    
    summary_text = f"""
    This report provides a comprehensive statistical analysis of the {series.name} assessment series.
    The data shows enrollment patterns across gender, special needs requirements, and occupational categories.
    
    <b>Key Highlights:</b><br/>
    - Total candidates enrolled: {total_candidates}<br/>
    - Candidates with special needs: {special_needs_count} ({special_needs_percentage:.1f}% of total)<br/>
    - Number of different occupations: {len([stat for stat in occupation_stats if stat['occupation__name']])}<br/>
    - Series duration: {(series.end_date - series.start_date).days + 1} days
    """
    
    elements.append(Paragraph(summary_text, normal_style))
    
    # Build PDF
    doc.build(elements)
    
    # Get the value of the BytesIO buffer and write it to the response
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response


@login_required
def statistical_reports_home(request):
    """Home page for statistical reports - shows available assessment series"""
    series_list = AssessmentSeries.objects.all().order_by('-start_date')
    
    context = {
        'series_list': series_list,
        'title': 'Statistical Reports'
    }
    
    return render(request, 'statistical_reports/home.html', context)

@login_required
def api_assessment_series(request):
    """API endpoint to get Assessment Series data for bulk enrollment"""
    try:
        # Get all assessment series, ordered by most recent first
        series_list = AssessmentSeries.objects.all().order_by('-start_date')
        
        series_data = []
        for series in series_list:
            series_data.append({
                'id': series.id,
                'name': series.name,
                'start_date': series.start_date.strftime('%Y-%m-%d'),
                'end_date': series.end_date.strftime('%Y-%m-%d'),
                'is_current': series.is_current,
                'results_released': series.results_released
            })
        
        return JsonResponse({
            'success': True,
            'series': series_data
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
def check_session_status(request):
    """API endpoint to check if user session is still valid and return user role info"""
    try:
        # Check if user is authenticated (this will be False if session expired)
        if not request.user.is_authenticated:
            return JsonResponse({
                'authenticated': False,
                'is_center_rep': False,
                'is_admin_or_staff': False
            })
        
        # Determine user role
        is_center_rep = hasattr(request.user, 'centerrep')
        is_admin_or_staff = (
            request.user.is_superuser or 
            request.user.is_staff or 
            hasattr(request.user, 'staff_profile') or 
            hasattr(request.user, 'supportstaff')
        )
        
        return JsonResponse({
            'authenticated': True,
            'is_center_rep': is_center_rep,
            'is_admin_or_staff': is_admin_or_staff,
            'username': request.user.username,
            'full_name': request.user.get_full_name() or request.user.username
        })
    except Exception as e:
        return JsonResponse({
            'authenticated': False,
            'error': str(e),
            'is_center_rep': False,
            'is_admin_or_staff': False
        })

# Candidate Verification System
from django.utils import timezone
from django.contrib.auth.decorators import user_passes_test
from django.views.decorators.http import require_POST

def is_admin_or_staff(user):
    """Check if user is admin or staff (not center representative)"""
    return user.is_superuser or user.is_staff or hasattr(user, 'staff_profile') or hasattr(user, 'supportstaff')

@login_required
@user_passes_test(is_admin_or_staff)
@require_POST
def verify_candidate(request, id):
    """Verify a candidate - admin/staff only"""
    candidate = get_object_or_404(Candidate, id=id)
    
    try:
        # Update candidate verification status
        candidate.verification_status = 'verified'
        candidate.verification_date = timezone.now()
        candidate.verified_by = request.user
        candidate.decline_reason = None  # Clear any previous decline reason
        candidate.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Candidate verified successfully',
            'status': 'verified'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
@user_passes_test(is_admin_or_staff)
@require_POST
def decline_candidate(request, id):
    """Decline a candidate with reason - admin/staff only"""
    candidate = get_object_or_404(Candidate, id=id)
    
    try:
        decline_reason = request.POST.get('decline_reason', '').strip()
        
        if not decline_reason:
            return JsonResponse({
                'success': False,
                'error': 'Decline reason is required'
            })
        
        # Update candidate verification status
        candidate.verification_status = 'declined'
        candidate.verification_date = timezone.now()
        candidate.verified_by = request.user
        candidate.decline_reason = decline_reason
        candidate.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Candidate declined successfully',
            'status': 'declined',
            'decline_reason': decline_reason
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
def api_occupations_by_category(request):
    """
    AJAX endpoint to get occupations filtered by registration category
    """
    from django.http import JsonResponse
    from .models import Occupation, OccupationCategory
    
    registration_category = request.GET.get('registration_category', '').lower()
    
    if not registration_category:
        return JsonResponse({'occupations': []})
    
    # Filter occupations based on registration category
    if registration_category == 'modular':
        # For modular, only show occupations that allow modular registration
        occupations = Occupation.objects.filter(has_modular=True).order_by('code')
    elif registration_category == 'formal':
        # For formal, only show occupations in the "Formal" category
        try:
            formal_category = OccupationCategory.objects.get(name__iexact='Formal')
            occupations = Occupation.objects.filter(category=formal_category).order_by('code')
        except OccupationCategory.DoesNotExist:
            # Fallback: show all non-modular occupations
            occupations = Occupation.objects.filter(has_modular=False).order_by('code')
    elif registration_category == 'informal':
        # For informal/worker's PAS, only show occupations in the "Worker's PAS" category
        try:
            workers_pas_category = OccupationCategory.objects.filter(name__iregex=r"worker('?s)? pas").first()
            if workers_pas_category:
                occupations = Occupation.objects.filter(category=workers_pas_category).order_by('code')
            else:
                # Fallback: try alternative names
                workers_pas_category = OccupationCategory.objects.filter(
                    name__icontains='worker'
                ).first()
                if workers_pas_category:
                    occupations = Occupation.objects.filter(category=workers_pas_category).order_by('code')
                else:
                    occupations = Occupation.objects.none()
        except Exception:
            occupations = Occupation.objects.none()
    else:
        occupations = Occupation.objects.none()
    
    # Format occupations for JSON response
    occupation_data = []
    for occ in occupations:
        occupation_data.append({
            'id': occ.id,
            'code': occ.code,
            'name': occ.name,
            'display_name': f"{occ.code} - {occ.name}"
        })
    
    return JsonResponse({'occupations': occupation_data})

@login_required
def api_all_levels_modules_papers(request):
    """API endpoint to get all levels, modules, and papers for cross-level Worker's PAS/Informal enrollment"""
    occupation_id = request.GET.get('occupation_id')
    
    if not occupation_id:
        return JsonResponse({'error': 'occupation_id parameter is required'}, status=400)
    
    try:
        occupation = Occupation.objects.get(id=occupation_id)
        
        # Get all levels for this occupation
        from .models import OccupationLevel
        occupation_levels = OccupationLevel.objects.filter(occupation=occupation).select_related('level')
        all_levels = [ol.level for ol in occupation_levels]
        
        # Build level_module_data structure similar to EnrollmentForm
        level_module_data = []
        
        for level in all_levels:
            # Get all modules for this occupation/level
            modules = Module.objects.filter(occupation=occupation, level=level)
            
            level_data = {
                'level': {
                    'id': level.id,
                    'name': level.name
                },
                'modules': []
            }
            
            for module in modules:
                papers = Paper.objects.filter(module=module, occupation=occupation, level=level)
                
                module_data = {
                    'module': {
                        'id': module.id,
                        'name': module.name,
                        'code': module.code
                    },
                    'papers': []
                }
                
                for paper in papers:
                    paper_info = {
                        'paper': {
                            'id': paper.id,
                            'name': paper.name,
                            'code': paper.code
                        }
                    }
                    module_data['papers'].append(paper_info)
                
                if module_data['papers']:  # Only add modules that have papers
                    level_data['modules'].append(module_data)
            
            if level_data['modules']:  # Only add levels that have modules with papers
                level_module_data.append(level_data)
        
        return JsonResponse({
            'success': True,
            'level_module_data': level_module_data
        })
        
    except Occupation.DoesNotExist:
        return JsonResponse({'error': 'Occupation not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ===== CANDIDATE PORTAL VIEWS =====

def candidate_portal_login(request):
    """Candidate portal login using registration number as both username and password"""
    from django.contrib import messages
    
    if request.method == 'POST':
        reg_number = request.POST.get('reg_number', '').strip()
        
        if not reg_number:
            messages.error(request, 'Please enter your registration number.')
            return render(request, 'candidate_portal/login.html')
        
        try:
            # Find candidate by registration number
            candidate = Candidate.objects.get(reg_number__iexact=reg_number)
            
            # Store candidate info in session for authentication
            request.session['candidate_id'] = candidate.id
            request.session['candidate_portal'] = True
            request.session['candidate_reg_number'] = candidate.reg_number
            
            # Redirect to candidate view
            return redirect('candidate_portal_view', id=candidate.id)
            
        except Candidate.DoesNotExist:
            messages.error(request, 'Invalid registration number. Please check and try again.')
    
    return render(request, 'candidate_portal/login.html')


def candidate_portal_logout(request):
    """Logout from candidate portal"""
    # Clear candidate portal session data
    if 'candidate_id' in request.session:
        del request.session['candidate_id']
    if 'candidate_portal' in request.session:
        del request.session['candidate_portal']
    if 'candidate_reg_number' in request.session:
        del request.session['candidate_reg_number']
    
    from django.contrib import messages
    messages.success(request, 'You have been logged out successfully.')
    return redirect('candidate_portal_login')


def candidate_portal_view(request, id):
    """Candidate portal view - same as candidate_view but without edit/enroll/testimonial buttons"""
    # Check if candidate is authenticated via portal
    candidate_id = request.session.get('candidate_id')
    if not candidate_id or not request.session.get('candidate_portal'):
        return redirect('candidate_portal_login')
    
    # Ensure candidate can only view their own record
    if int(candidate_id) != int(id):
        from django.contrib import messages
        messages.error(request, 'Access denied. You can only view your own results.')
        return redirect('candidate_portal_view', id=candidate_id)
    
    print('DEBUG: candidate_portal_view called for candidate', id)
    from .models import AssessmentCenter, Occupation, Result, CandidateLevel, CandidateModule, Paper, Module, CandidatePaper, AssessmentSeries
    candidate = get_object_or_404(Candidate, id=id)

    # Check if results have been released for the candidate's assessment series
    # Use same logic as regular candidate view - candidates should see results like non-center-reps
    candidate_series = candidate.assessment_series
    results_released = candidate_series and candidate_series.results_released or True  # Candidates can always see their results
    
    # Get candidate results
    results = Result.objects.filter(candidate=candidate).select_related('level', 'module', 'paper').order_by('assessment_date')
    
    # Only show results if they have been released
    if not results_released:
        results = []
    
    # Get enrollment information
    enrolled_levels = CandidateLevel.objects.filter(candidate=candidate).select_related('level')
    enrolled_modules = CandidateModule.objects.filter(candidate=candidate).select_related('module')
    enrolled_papers = CandidatePaper.objects.filter(candidate=candidate).select_related('paper', 'module')
    
    # Organize results by category
    reg_cat = getattr(candidate, 'registration_category', '').lower()
    
    # Get results organized by structure
    results_by_level = {}
    results_by_module = {}
    results_by_paper = {}
    
    for result in results:
        if result.level:
            if result.level.id not in results_by_level:
                results_by_level[result.level.id] = []
            results_by_level[result.level.id].append(result)
        
        if result.module:
            if result.module.id not in results_by_module:
                results_by_module[result.module.id] = []
            results_by_module[result.module.id].append(result)
        
        if result.paper:
            if result.paper.id not in results_by_paper:
                results_by_paper[result.paper.id] = []
            results_by_paper[result.paper.id].append(result)
    
    # Determine if candidate has any results
    has_results = results.exists() if results_released else False
    
    # Create results_summary for template (required for Worker's PAS/Informal display)
    from .models import Level
    results_summary_for_display = []
    
    if results.exists():
        # Get all levels that have results
        all_results_levels = results.values('level_id', 'level__name').distinct()
        
        # Create comprehensive results summary including historical results
        levels_processed = set()
        for level_data in all_results_levels:
            level_id = level_data['level_id']
            
            # Skip if we've already processed this level
            if level_id in levels_processed:
                continue
            levels_processed.add(level_id)
            
            level = Level.objects.get(id=level_id)
            
            # Get all results for this level
            level_results = results.filter(level_id=level_id)
            
            # Group results by module, then by paper
            modules_with_results = {}
            for result in level_results:
                module_id = result.module.id
                if module_id not in modules_with_results:
                    modules_with_results[module_id] = {
                        'module': result.module,
                        'papers': set(),  # Use set to avoid duplicate papers
                        'results': []
                    }
                modules_with_results[module_id]['papers'].add(result.paper)
                modules_with_results[module_id]['results'].append(result)
            
            # Convert paper sets back to lists
            for module_data in modules_with_results.values():
                module_data['papers'] = list(module_data['papers'])
            
            results_summary_for_display.append({
                'level': level,
                'modules': list(modules_with_results.values())
            })
    
    context = {
        'candidate': candidate,
        'results': results,
        'results_released': results_released,
        'enrolled_levels': enrolled_levels,
        'enrolled_modules': enrolled_modules,
        'enrolled_papers': enrolled_papers,
        'results_by_level': results_by_level,
        'results_by_module': results_by_module,
        'results_by_paper': results_by_paper,
        'has_results': has_results,
        'results_summary': results_summary_for_display,  # Required for template display
        'is_candidate_portal': True,  # Flag to indicate this is candidate portal view
        'is_center_rep': False,  # Candidates are not center reps
    }
    
    return render(request, 'candidates/view.html', context)
